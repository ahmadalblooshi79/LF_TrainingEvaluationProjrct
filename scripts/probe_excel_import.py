"""Probe Excel file for import mapping (no Flask)."""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

path = Path(r"c:\Users\USER\Documents\GitHub\LF_Project\06 تقييم التنقل الإداري.xlsx")
wb = load_workbook(path, data_only=True)
ws = wb.active
print("sheet", ws.title, "rows", ws.max_row, "cols", ws.max_column)
for r in range(1, min(45, ws.max_row + 1)):
    vals = []
    for c in range(1, 11):
        v = ws.cell(r, c).value
        vals.append("" if v is None else str(v).replace("\n", " ")[:40])
    print(f"{r:2}", "|", " | ".join(vals))
print("merged:", len(list(ws.merged_cells.ranges)))

"""
قراءة ورقة قائمة تقييم Excel مع دعم القوالب التي تبدأ بصفوف تعريف (وحدة، تاريخ…)
ثم صف عناوين «القصوى / المكتسبة» — كما في النماذج العسكرية الموحدة.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from app.evaluation_list_columns import (
    annotate_evaluation_row_kinds,
    build_structured_rows,
    import_body_end_row_index,
    military_template_column_indices,
    normalize_ar_header,
    parse_max_cell,
    resolve_evaluation_column_indices,
    resolve_rubric_subheader_indices,
    try_named_eval_column_indices,
)
from app.xlsx_grid_preview import read_xlsx_sheet_preview


def _pad_grid(grid: list[list[str]], ncol: int) -> list[list[str]]:
    out: list[list[str]] = []
    for r in grid:
        cells = list(r)
        while len(cells) < ncol:
            cells.append("")
        out.append(cells[:ncol])
    return out


def _find_rubric_subheader_row_index(grid: list[list[str]]) -> int | None:
    for ri, row in enumerate(grid):
        parts = [normalize_ar_header(c) for c in row]
        joined = " ".join(parts)
        if "قصوى" in joined and "مكتسب" in joined:
            return ri
    return None


def _pad_row_to_ncol(row: list[str], ncol: int) -> list[str]:
    cells = list(row)
    while len(cells) < ncol:
        cells.append("")
    return cells


def _load_evaluation_import_grid(path: Path, *, sheet_index: int = 0) -> dict[str, Any]:
    """
    قراءة الورقة مع توسيع الخلايا المدمجة (لعمود عناصر التقييم وغيره).
    يعيد نفس مفاتيح المعاينة الأساسية أو ``error``.
    """
    from app.xlsx_grid_preview import MAX_PREVIEW_COLS, MAX_PREVIEW_ROWS, _cell_to_str

    path = Path(path)
    out: dict[str, Any] = {
        "sheet_title": "",
        "grid_rows": [],
        "ncol": 0,
        "error": None,
    }
    if not path.is_file():
        out["error"] = "الملف غير موجود."
        return out
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        out["error"] = "معاينة الجدول متاحة لملفات Excel (.xlsx) فقط."
        return out
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError:
        out["error"] = "تعذر فتح ملف Excel لأن مكتبة openpyxl غير مثبتة."
        return out
    wb = None
    try:
        wb = load_workbook(filename=str(path), data_only=True)
        names = wb.sheetnames
        if not names:
            out["error"] = "لا توجد أوراق في الملف."
            return out
        si = max(0, min(sheet_index, len(names) - 1))
        ws = wb[names[si]]
        out["sheet_title"] = ws.title or ""
        mr = int(getattr(ws, "max_row", None) or 1)
        mc = int(getattr(ws, "max_column", None) or 1)
        max_row = min(mr, MAX_PREVIEW_ROWS)
        max_col = min(mc, MAX_PREVIEW_COLS)
        grid: list[list[str]] = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            grid.append([_cell_to_str(c) for c in row])
        ncol = max_col
        for r in grid:
            while len(r) < ncol:
                r.append("")
        for merged in ws.merged_cells.ranges:
            anchor = _cell_to_str(ws.cell(merged.min_row, merged.min_col).value)
            if not anchor:
                continue
            # توسيع الدمج في أعمدة العناوين (ب–د) فقط — لا ننسخ التذييل إلى أعمدة العلامات
            col_lo = max(1, merged.min_col)
            col_hi = min(max_col, merged.max_col, 4)
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(col_lo, col_hi + 1):
                    if 1 <= r <= len(grid) and 1 <= c <= ncol:
                        grid[r - 1][c - 1] = anchor
        out["grid_rows"] = grid
        out["ncol"] = ncol
    except Exception as exc:
        out["error"] = f"تعذر قراءة ملف Excel: {exc}"
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return out


def _forward_fill_score_elements(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """عند دمج عمود (ب) تبقى بعض صفوف العلامات بلا نص — نملأ من آخر عنوان بند."""
    last_el = ""
    out: list[dict[str, Any]] = []
    for r in raw_rows:
        el = (r.get("element") or "").strip()
        mx = parse_max_cell(r.get("max_val"))
        row = dict(r)
        if el:
            last_el = el
        elif last_el and mx is not None and mx > 0:
            row["element"] = last_el
        out.append(row)
    return out


def _body_rows_before_footer(
    grid: list[list[str]], *, start_index: int, ncol: int
) -> list[list[str]]:
    end = import_body_end_row_index(grid, rubric_row_index=start_index - 1, ncol=ncol)
    return grid[start_index:end]


def _scan_first_named_triple_header(
    grid: list[list[str]], ncol: int, *, max_scan: int = 36
) -> tuple[int, tuple[int, int, int], str] | None:
    """أول صف يحتوي عناوين الأعمدة الثلاثة بالأسماء (مفيد عند صفوف تعريف قبل جدول التقييم)."""
    for hi in range(min(max_scan, len(grid))):
        row = _pad_row_to_ncol(grid[hi], ncol)
        hit = try_named_eval_column_indices(row)
        if hit is not None:
            triple, src = hit
            return hi, triple, src
    return None


def read_evaluation_list_sheet(path: Path, *, sheet_index: int = 0) -> dict[str, Any]:
    """
    يعيد نفس مفاتيح ``read_xlsx_sheet_preview`` تقريبًا، مع:
    - ``eval_layout``: ``rubric`` | ``legacy``
    - ``eval_structured``, ``eval_rows``, ``eval_column_source``, ``eval_input_mode``

    عمود «عناصر التقييم» يُستنتج من عنوان العمود في الصف (وليس دائمًا العمود A)،
    ويُبحث في أول صفوف الورقة عن صف العناوين الكامل عند وجود صفوف تعريف أعلاه.
    """
    base = _load_evaluation_import_grid(path, sheet_index=sheet_index)
    if base.get("error"):
        fallback = read_xlsx_sheet_preview(path, sheet_index=sheet_index)
        if not fallback.get("error"):
            base = fallback
    out: dict[str, Any] = {
        "sheet_title": base.get("sheet_title") or "",
        "grid_rows": base.get("grid_rows") or [],
        "header_row": [],
        "body_rows": [],
        "col_labels": [],
        "ncol": base.get("ncol") or 0,
        "error": base.get("error"),
    }
    out["eval_layout"] = "legacy"
    out["eval_structured"] = False
    out["eval_rows"] = []
    out["eval_column_source"] = None
    out["eval_input_mode"] = "scale5"

    if base.get("error"):
        return out

    grid = base.get("grid_rows") or []
    if not grid:
        return out

    ncol = max(int(base.get("ncol") or 0), max((len(r) for r in grid), default=0))
    ncol = max(ncol, 1)
    grid = _pad_grid(grid, ncol)
    out["ncol"] = ncol
    out["header_row"] = grid[0]
    out["body_rows"] = grid[1:] if len(grid) > 1 else []

    rubric_i = _find_rubric_subheader_row_index(grid)

    if rubric_i is not None:
        sub_header = grid[rubric_i]
        mil = military_template_column_indices(ncol)
        i_pct = i_grade = i_notes = None
        if mil is not None:
            i_el, i_mx, i_aq, i_pct, i_grade, i_notes = mil
            src = "military_template_befghi"
            acquired_cap_five = False
        else:
            triple_r = resolve_rubric_subheader_indices(sub_header)
            if triple_r is not None:
                i_el, i_mx, i_aq = triple_r
                src = "rubric_subheader"
                acquired_cap_five = False
            else:
                resolved = resolve_evaluation_column_indices(sub_header, ncol)
                if resolved is None:
                    return out
                (i_el, i_mx, i_aq), src = resolved
                acquired_cap_five = True

        body_rows = _body_rows_before_footer(grid, start_index=rubric_i + 1, ncol=ncol)
        body_start = rubric_i + 2  # رقم صف Excel الأول بعد صف العناوين
        raw_rows = build_structured_rows(
            body_rows,
            ncol,
            i_el,
            i_mx,
            i_aq,
            acquired_cap_five=acquired_cap_five,
            i_pct=i_pct,
            i_grade=i_grade,
            i_notes=i_notes,
            body_start_row_1based=body_start,
        )
        raw_rows = _forward_fill_score_elements(raw_rows)
        eval_rows = annotate_evaluation_row_kinds(raw_rows)
        if not eval_rows and raw_rows:
            eval_rows = [
                {
                    **r,
                    "row_kind": "score",
                    "max_num": parse_max_cell(r.get("max_val")),
                }
                for r in raw_rows
            ]
        out["eval_rows"] = eval_rows
        out["eval_structured"] = len(eval_rows) > 0
        out["eval_column_source"] = src
        out["eval_layout"] = "rubric"
        out["eval_input_mode"] = "variable"
        out["body_rows"] = body_rows
        return out

    named = _scan_first_named_triple_header(grid, ncol)
    if named is not None:
        hi, (i_el, i_mx, i_aq), src = named
        header_row = _pad_row_to_ncol(grid[hi], ncol)
        body_rows = grid[hi + 1 :]
        raw_rows = build_structured_rows(
            body_rows, ncol, i_el, i_mx, i_aq, acquired_cap_five=True
        )
        eval_rows = annotate_evaluation_row_kinds(raw_rows)
        if not eval_rows and raw_rows:
            eval_rows = [
                {
                    **r,
                    "row_kind": "score",
                    "max_num": parse_max_cell(r.get("max_val")),
                }
                for r in raw_rows
            ]
        out["eval_rows"] = eval_rows
        out["eval_structured"] = len(eval_rows) > 0
        out["eval_column_source"] = src
        out["eval_layout"] = "legacy"
        out["eval_input_mode"] = "scale5"
        out["header_row"] = header_row
        out["body_rows"] = body_rows
        return out

    header_row = grid[0]
    resolved = resolve_evaluation_column_indices(header_row, ncol)
    if resolved is None:
        return out
    (i_el, i_mx, i_aq), src = resolved
    body_rows = grid[1:]
    raw_rows = build_structured_rows(
        body_rows, ncol, i_el, i_mx, i_aq, acquired_cap_five=True
    )
    eval_rows = [
        {
            **r,
            "row_kind": "score",
            "max_num": parse_max_cell(r.get("max_val")),
        }
        for r in raw_rows
    ]
    out["eval_rows"] = eval_rows
    out["eval_structured"] = len(eval_rows) > 0
    out["eval_column_source"] = src
    out["eval_layout"] = "legacy"
    out["eval_input_mode"] = "scale5"
    out["body_rows"] = body_rows
    return out


def ratio_to_performance_band(ratio: float) -> str:
    """نطاق أداء البند من النسبة المكتسبة/القصوى (0..1) — يطابق grade_label_from_percent."""
    pct = max(0.0, min(100.0, float(ratio) * 100.0))
    if pct < 60:
        return "راسب"
    if pct < 70:
        return "مقبول"
    if pct < 80:
        return "جيد"
    if pct < 90:
        return "جيد_جدا"
    return "ممتاز"


def extract_rubric_classification_dataset(
    paths: list[Path],
    *,
    sheet_index: int = 0,
) -> tuple[list[list[Any]], list[str], list[str], str]:
    """
    يستخرج من كل ملف بنود التقييم ذات قصوى رقمية وعلامة مكتسبة مدخلة.
    الميزات: [نص البند، القصوى كرقم]. الهدف: نطاق الأداء من نسبة (مكتسب÷قصوى).

    يعيد (صفوف_X، قائمة_y، أسماء_الميزات، رسالة_خطأ).
    """
    xs: list[list[Any]] = []
    ys: list[str] = []
    feat_names = ["بند", "القصوى"]

    for p in paths:
        info = read_evaluation_list_sheet(p, sheet_index=sheet_index)
        if info.get("error") or not info.get("eval_structured"):
            continue
        for row in info.get("eval_rows") or []:
            if row.get("row_kind") != "score":
                continue
            mx = row.get("max_num")
            if mx is None:
                mx = parse_max_cell(row.get("max_val"))
            if mx is None or float(mx) <= 0:
                continue
            aq_raw = (row.get("acquired_initial") or "").strip()
            if not aq_raw or aq_raw == "na":
                continue
            try:
                aq_s = aq_raw.replace(",", ".").replace("٫", ".")
                aq = float(aq_s)
            except ValueError:
                continue
            mx_f = float(mx)
            aq = max(0.0, min(mx_f, aq))
            ratio = aq / mx_f if mx_f > 0 else 0.0
            el = (row.get("element") or "").strip() or "___فارغ___"
            xs.append([el, float(mx)])
            ys.append(ratio_to_performance_band(ratio))

    if not ys:
        return (
            [],
            [],
            feat_names,
            "لم يُستخرج أي بند فيه علامة مكتسبة من ملفات الإكسل (إن كانت القوالب فارغة عندك، جرّب التدريب من نتائج النظام المحفوظة: --format saved).",
        )
    return xs, ys, feat_names, ""

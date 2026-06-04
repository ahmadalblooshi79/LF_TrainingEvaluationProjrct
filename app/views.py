import hashlib
import io
import json
import logging
import mimetypes
import re
import sys
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, unquote

from flask import (
    Blueprint,
    abort,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename
from sqlalchemy import case, delete, desc, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

from app.config import (
    CHAT_UPLOAD_DIR,
    DILEMMA_PDF_DIR,
    EVALUATION_LIST_XLSX_DIR,
    INFO_BANK_DIR,
    PLANNER_FLOW_BUNDLE_DIR,
    VISUAL_DOC_DIR,
)
from app.auth import get_current_user_optional, hash_password, verify_password
from app.models import (
    ChatMessage,
    ChatMessageRead,
    ChatRoom,
    ChatRoomKind,
    ChatRoomMember,
    Exercise,
    ExerciseBattleUnitPersonnel,
    ExerciseObjective,
    ExercisePhase,
    ExerciseRosterKind,
    ExerciseRosterRow,
    ExercisePlannerFlowBundle,
    ExercisePlannerFlowBundleEventFlow,
    ExercisePlannerFlowBundleActionEval,
    PlannerFlowBundleEvalSavedResult,
    ExerciseStatus,
    DilemmaItem,
    EvaluationCriterionMedia,
    EvaluationListPdfItem,
    EvaluationListSavedResult,
    AnalystEvaluationCriteriaResult,
    AnalystEvaluationCriteriaUnit,
    AnalystEvaluationCriteriaPhaseItem,
    AnalystFinalEvaluationAllocatedMax,
    AnalystFinalEvaluationPhaseAllocatedMax,
    ExerciseNotification,
    VisualDocument,
    InformationBankPhaseNote,
    InformationBankUnitNote,
    InformationBankTrainingPhase,
    InformationBankUnitLevel,
    InformationBankTreeNode,
    InfoBankEventFlowPdf,
    InfoBankActionEvalXlsx,
    InfoBankDilemmaEvalXlsx,
    JudgeTraineeAssignment,
    Reference,
    RefType,
    RoleDef,
    RoleKey,
    User,
)
from app.eval_criterion_media import (
    criterion_media_absolute_path,
    group_media_rows,
    persist_criterion_medium,
)
from app.evaluation_workflow import (
    apply_chief_approve,
    apply_chief_reopen,
    apply_judge_save_after_reopen,
    apply_judge_approve,
    build_evaluation_list_row,
    build_planner_flow_eval_row,
    evaluation_unit_home_rows,
    evaluation_unit_home_totals,
    eval_status_done,
    eval_chief_approved,
    eval_chief_can_approve,
    eval_chief_can_reopen,
    eval_control_approved,
    eval_judge_approved,
    eval_judge_can_approve,
    eval_judge_can_edit,
    eval_reopened_for_judge,
    eval_workflow_label_ar,
)
from app.permissions import (
    can_access_analyst_hub,
    can_access_chief_judge_hub,
    can_access_control_hub,
    can_access_judge_hub,
    can_access_planner_hub,
    can_approve_evaluation_results,
    can_chief_approve_evaluation_results,
    can_chief_reopen_evaluation_for_judge,
    can_edit_references,
    can_judge_exercise,
    can_manage_chat_rooms,
    can_manage_information_bank,
    can_oversee_judge_planner_flow_materials,
    can_manage_users,
    can_plan_exercises,
    can_save_evaluation_results,
    can_use_chat_rooms,
    can_view_information_bank,
    can_view_notifications_log,
    can_use_visual_documentation,
    is_analyst,
    is_chief_judge,
    is_control,
    is_judge,
    is_planner,
    is_system_admin,
)
from app import exercise_options as ex_opts
from app.exercise_phase_catalog import (
    DEFAULT_EXERCISE_PHASE,
    EXERCISE_PHASE_OPTIONS,
    default_exercise_phase_key,
    exercise_phase_keys,
    exercise_phase_label,
    normalize_exercise_phase,
)
from app.unit_levels_catalog import (
    UNIT_LEVELS,
    coerce_roster_import_position_cell,
    default_unit_level_key,
    label_for_unit_level_key,
    normalize_unit_level_key,
    planning_included_unit_keys,
    unit_level_row,
)
from app.information_bank_catalog import (
    INFO_BANK_UNIT_LEVEL_TEMPLATES,
    INFO_BANK_UNIT_LEVELS,
    brigade_group_for_tab,
    brigade_tab_for_group,
    unit_catalog_key_for_brigade,
    TRAINING_PHASES,
    info_bank_unit_label,
    training_phase_label,
)
from app.evaluation_list_columns import (
    acquired_select_options,
    display_grade_label,
    grade_allows_judge_approve,
    grade_label_from_percent,
    parse_max_cell,
)
from app.evaluation_sheet_parser import read_evaluation_list_sheet
from app.roster_import import parse_roster_rows_from_upload
from app.exercise_store import (
    archive_and_clear_current_exercise,
    wipe_exercise_from_system,
    export_directory,
    extract_create_form_prefill_from_export_json,
    import_exercise_bundle_from_dict,
    list_export_json_files,
    open_export_directory_in_os,
    purge_all_exercises_and_dilemmas,
    read_exercise_id_from_json_bytes,
    read_exercise_id_from_json_path,
    write_exercise_json_file,
)
from app.seed import DEMO_PASSWORD
from app.battle_organization import BATTLE_ORG_DEMO_ROOT
from app.ai_service import suggest_instructions_or_notes
from app.info_bank_access import (
    INFO_BANK_GATE_SESSION_KEY,
    clear_information_bank_gate,
    information_bank_gate_ok,
    is_ibank_included_save_request,
    is_information_bank_path,
)

bp = Blueprint("views", __name__)

def _dilemma_pdf_abspath(relpath: str) -> Path | None:
    """مسار ملف PDF تحت DILEMMA_PDF_DIR دون تجاوز الجذر."""
    if not relpath or not isinstance(relpath, str):
        return None
    norm = relpath.replace("\\", "/").strip()
    if not norm or any(part == ".." for part in norm.split("/")):
        return None
    root = DILEMMA_PDF_DIR.resolve()
    out = (root / norm).resolve()
    try:
        out.relative_to(root)
    except ValueError:
        return None
    return out if out.is_file() else None


def _unlink_dilemma_stored_pdf(relpath: str) -> None:
    p = _dilemma_pdf_abspath(relpath)
    if p is None:
        return
    try:
        p.unlink()
    except OSError:
        pass


def _evaluation_list_file_abspath(relpath: str) -> Path | None:
    if not relpath or not isinstance(relpath, str):
        return None
    norm = relpath.replace("\\", "/").strip()
    if not norm or any(part == ".." for part in norm.split("/")):
        return None
    root = EVALUATION_LIST_XLSX_DIR.resolve()
    out = (root / norm).resolve()
    try:
        out.relative_to(root)
    except ValueError:
        return None
    return out if out.is_file() else None


def _evaluation_sheet_view_context(fspath: Path) -> dict:
    """قراءة ملف قائمة التقييم مع اكتشاف قوالب الصفوف (القصوى/المكتسبة) تلقائيًا."""
    from app.evaluation_element_display import enrich_eval_rows_element_styles

    sheet = read_evaluation_list_sheet(fspath)
    es = bool(sheet.get("eval_structured"))
    raw_eval = sheet.get("eval_rows") or []
    eval_rows = enrich_eval_rows_element_styles(raw_eval) if es else raw_eval
    return {
        "preview_error": sheet.get("error"),
        "sheet_title": sheet.get("sheet_title") or "",
        "header_row": sheet.get("header_row") or [],
        "body_rows": sheet.get("body_rows") or [],
        "eval_structured": es,
        "eval_column_source": sheet.get("eval_column_source"),
        "eval_rows": eval_rows,
        "eval_input_mode": sheet.get("eval_input_mode") or "scale5",
        "eval_layout": sheet.get("eval_layout") or "legacy",
        "acquired_options": acquired_select_options() if es else [],
    }


def _saved_payload_aligned_with_eval_rows(
    saved_payload: dict | None, eval_rows: list | None
) -> dict:
    """يتجاهل حفظاً قديماً بعدد صفوف لا يطابق القالب الحالي (يمنع انحراف المكتسبة والمجاميع)."""
    if not saved_payload or not isinstance(saved_payload, dict):
        return {}
    rows = saved_payload.get("rows") or []
    template = eval_rows or []
    if not template or len(rows) != len(template):
        return {}
    return saved_payload


def _unlink_evaluation_list_stored_file(relpath: str) -> None:
    p = _evaluation_list_file_abspath(relpath)
    if p is None:
        return
    try:
        p.unlink()
    except OSError:
        pass


def _is_pdf_bytes(data: bytes) -> bool:
    return bool(data) and data[:4] == b"%PDF"


def _is_docx_bytes(data: bytes) -> bool:
    """Word .docx — أرشيف ZIP يحتوي مجلد word/."""
    if not data or len(data) < 64:
        return False
    if data[:2] != b"PK":
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            return any(n.startswith("word/") for n in z.namelist())
    except zipfile.BadZipFile:
        return False


def _is_doc_bytes(data: bytes) -> bool:
    """Word .doc — مركب OLE (CF)."""
    if not data or len(data) < 512:
        return False
    return data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _info_bank_event_flow_sniff_ext(data: bytes) -> str | None:
    """يحدد امتداد الحفظ الآمن (.pdf / .docx / .doc) من محتوى الملف."""
    if _is_pdf_bytes(data):
        return ".pdf"
    if _is_docx_bytes(data):
        return ".docx"
    if _is_doc_bytes(data):
        return ".doc"
    return None


def _mimetype_info_bank_event_flow(path: Path) -> str:
    n = path.name.lower()
    if n.endswith(".pdf"):
        return "application/pdf"
    if n.endswith(".docx"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if n.endswith(".doc"):
        return "application/msword"
    guessed, _ = mimetypes.guess_type(path.name)
    return guessed or "application/octet-stream"


def _is_xlsx_bytes(data: bytes) -> bool:
    """ملف Excel (.xlsx) هو أرشيف ZIP يحتوي [Content_Types].xml."""
    if not data or len(data) < 64:
        return False
    if data[:2] != b"PK":
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            return "[Content_Types].xml" in z.namelist()
    except zipfile.BadZipFile:
        return False


_INFO_BANK_KIND_DIRS = {"event_flow": "event_flow", "action_eval": "action_eval", "dilemma_eval": "dilemma_eval"}


def _info_bank_file_abspath(kind: str, relpath: str) -> Path | None:
    if kind not in _INFO_BANK_KIND_DIRS:
        return None
    if not relpath or not isinstance(relpath, str):
        return None
    norm = relpath.replace("\\", "/").strip()
    if not norm or any(part == ".." for part in norm.split("/")):
        return None
    root = INFO_BANK_DIR.resolve()
    out = (root / norm).resolve()
    try:
        out.relative_to(root)
    except ValueError:
        return None
    return out if out.is_file() else None


def _unlink_info_bank_file(kind: str, relpath: str) -> None:
    p = _info_bank_file_abspath(kind, relpath)
    if p is None:
        return
    try:
        p.unlink()
    except OSError:
        pass


def _mimetype_for_eval_list_file(path: Path) -> str:
    n = path.name.lower()
    if n.endswith(".xlsx"):
        return (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if n.endswith(".pdf"):
        return "application/pdf"
    return "application/octet-stream"


def _planner_flow_bundle_root() -> Path:
    PLANNER_FLOW_BUNDLE_DIR.mkdir(parents=True, exist_ok=True)
    return PLANNER_FLOW_BUNDLE_DIR.resolve()


def _planner_bundle_file_abspath(relpath: str | None) -> Path | None:
    if not relpath or not isinstance(relpath, str):
        return None
    norm = relpath.replace("\\", "/").strip()
    if not norm or any(part == ".." for part in norm.split("/")):
        return None
    root = _planner_flow_bundle_root()
    out = (root / norm).resolve()
    try:
        out.relative_to(root)
    except ValueError:
        return None
    return out if out.is_file() else None


# عناوين ناتجة عن secure_filename لأسماء عربية/غريبة (pdf.1، xlsx_2، …)
_DEGENERATE_UPLOAD_TITLE_RE = re.compile(
    r"^(?:pdf|xlsx|excel|file)(?:[\s._-]+\d*)?$",
    re.IGNORECASE,
)


def _is_degenerate_upload_title(s: str) -> bool:
    t = (s or "").strip()
    if not t or len(t) <= 1:
        return True
    return bool(_DEGENERATE_UPLOAD_TITLE_RE.match(t))


def _planner_blob_display_filename(
    *, stored_title: str, relpath: str, fallback: str
) -> str:
    """اسم عرض مقروء: يفضّل العنوان المحفوظ إن كان معبّراً، وإلا اسم الملف من المسار النسبي."""
    rp = (relpath or "").replace("\\", "/").strip()
    base = Path(rp).name if rp else ""
    st = (stored_title or "").strip()
    if st and not _is_degenerate_upload_title(st):
        return st
    if base:
        return base
    return st or fallback


def _unlink_planner_bundle_file(relpath: str) -> None:
    p = _planner_bundle_file_abspath(relpath)
    if p is None:
        return
    try:
        p.unlink()
    except OSError:
        pass


def _sanitize_planner_bundle_orphan_event_flow(
    db, bundle: ExercisePlannerFlowBundle | None
) -> bool:
    """إزالة مسار مجرى الأحداث من السجل إذا لم يعد الملف موجوداً على القرص."""
    if bundle is None:
        return False
    rel = (bundle.event_flow_file_relpath or "").strip()
    if not rel:
        return False
    if _planner_bundle_file_abspath(rel) is not None:
        return False
    bundle.event_flow_file_relpath = ""
    bundle.event_flow_title = ""
    bundle.linked_at = None
    bundle.updated_at = datetime.utcnow()
    db.commit()
    return True


def _hashes_of_unit_pdfs(
    db,
    model,
    unit_key: str,
    abspath_fn,
    exercise_id: int | None = None,
    exercise_phase: str | None = None,
) -> set[str]:
    """SHA-256 لمحتوى كل ملف محفوظ لهذا المستوى داخل التمرين الحالي (لمنع تكرار الملف)."""
    out: set[str] = set()
    q = db.query(model).filter(model.unit_level_key == unit_key)
    if exercise_id is not None and hasattr(model, "exercise_id"):
        q = q.filter(model.exercise_id == exercise_id)
    if exercise_phase is not None and hasattr(model, "exercise_phase"):
        q = q.filter(model.exercise_phase == exercise_phase)
    for row in q.all():
        rel = (row.pdf_relpath or "").strip()
        if not rel:
            continue
        p = abspath_fn(rel)
        if p is None or not p.is_file():
            continue
        try:
            out.add(hashlib.sha256(p.read_bytes()).hexdigest())
        except OSError:
            pass
    return out


def _normalized_exercise_phase(val: str | None) -> str:
    return normalize_exercise_phase(val)


def _exercise_phase_order_expr(column):
    """ترتيب SQL حسب كتالوج المراحل؛ عند فراغ الكتالوج يُرتب بالقيمة الخام."""
    keys = exercise_phase_keys()
    if not keys:
        return column
    return case(
        {key: idx for idx, key in enumerate(keys)},
        value=column,
        else_=len(keys),
    )


def _unit_level_order_expr(column):
    """ترتيب SQL حسب كتالوج مستويات الوحدة؛ عند فراغ الكتالوج يُرتب بالقيمة الخام."""
    if not UNIT_LEVELS:
        return column
    return case(
        {row["key"]: idx for idx, row in enumerate(UNIT_LEVELS)},
        value=column,
        else_=len(UNIT_LEVELS),
    )


def _require_unit_level_row(unit_key: str | None) -> dict[str, str] | None:
    """صف كتالوج التخطيط؛ ``None`` إذا المفتاح فارغاً؛ 404 إذا المفتاح غير معروف."""
    k = (unit_key or "").strip()
    if not k:
        return None
    row = unit_level_row(k)
    if row is None:
        abort(404)
    return row


def _can_manage_dilemma_eval_catalog(user: User) -> bool:
    return is_system_admin(user) or is_planner(user)


def _require_planner_hub_catalog_access(user: User | None) -> None:
    """رفع وإدارة ملفات المعاضل وقوائم التقييم — مساحة التخطيط فقط (مخطّط أو إدارة النظام)."""
    if not user:
        abort(403)
    if not can_access_planner_hub(user) or not _can_manage_dilemma_eval_catalog(user):
        abort(403)


def _admin_exercise_form_ctx() -> dict:
    return {
        "exercise_types": ex_opts.EXERCISE_TYPES,
        "exercise_levels": ex_opts.EXERCISE_LEVELS,
        "missions": ex_opts.MISSIONS,
    }


def _empty_create_form_prefill() -> dict[str, str]:
    return {
        "trained_unit": "",
        "location_label": "",
        "exercise_name": "",
        "exercise_type": "",
        "exercise_level": "",
        "mission": "",
        "planned_start": "",
        "planned_end": "",
    }


def _dt_for_datetime_local(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt.strftime("%Y-%m-%dT%H:%M")


def _clip_create_text(val: str | None, max_len: int) -> str:
    return (val or "").strip()[:max_len]


def _prefill_create_form_from_exercise(ex: Exercise) -> dict[str, str]:
    """تعبئة النموذج من التمرين الحالي — القيم المخزّنة كما هي (حتى لو خارج القوائم)."""
    out = _empty_create_form_prefill()
    out["trained_unit"] = _clip_create_text(ex.trained_unit, 400)
    out["location_label"] = _clip_create_text(ex.location_label, 400)
    out["exercise_name"] = _clip_create_text(ex.title, 500)
    out["exercise_type"] = _clip_create_text(ex.exercise_type, 200)
    out["exercise_level"] = _clip_create_text(ex.exercise_level, 200)
    out["mission"] = _clip_create_text(ex.mission_label, 400)
    out["planned_start"] = _dt_for_datetime_local(ex.planned_start)
    out["planned_end"] = _dt_for_datetime_local(ex.planned_end)
    return out


def _workspace_exercise_for_admin_form(db, user: User) -> Exercise | None:
    """التمرين المعروض في النموذج — نفس التمرين الظاهر في الشريط العلوي."""
    ex = _current_workspace_exercise(db, user)
    if ex is not None:
        return ex
    return db.query(Exercise).order_by(Exercise.id.desc()).first()


def _prefill_create_form_from_request() -> dict[str, str]:
    out = _empty_create_form_prefill()

    def pick(field: str, allowed: list[str]) -> str:
        v = (request.form.get(field) or "").strip()
        return v if v in allowed else ""

    out["trained_unit"] = _clip_create_text(request.form.get("trained_unit"), 400)
    out["location_label"] = _clip_create_text(request.form.get("location_label"), 400)
    out["exercise_name"] = _clip_create_text(request.form.get("exercise_name"), 500)
    out["exercise_type"] = pick("exercise_type", ex_opts.EXERCISE_TYPES)
    out["exercise_level"] = pick("exercise_level", ex_opts.EXERCISE_LEVELS)
    out["mission"] = pick("mission", ex_opts.MISSIONS)
    for fld in ("planned_start", "planned_end"):
        raw = (request.form.get(fld) or "").strip()
        if raw:
            try:
                datetime.fromisoformat(raw)
                out[fld] = raw
            except Exception:
                out[fld] = ""
    return out


def _wants_import_json_response() -> bool:
    return "application/json" in (request.headers.get("Accept") or "")


def _import_full_json_error(msg: str, *, status: int = 400):
    if _wants_import_json_response():
        return jsonify({"ok": False, "error": msg}), status
    return redirect("/admin/exercises/create?err=" + quote(msg, safe=""))


def _import_full_json_ok(eid: int):
    if _wants_import_json_response():
        return jsonify({"ok": True, "redirect": f"/exercises/{eid}"})
    return redirect(f"/exercises/{eid}")


def _lines_to_objectives(raw: str) -> list[str]:
    items: list[str] = []
    for line in (raw or "").splitlines():
        s = line.strip().lstrip("•*-").strip()
        if not s:
            continue
        if len(s) > 2000:
            s = s[:2000]
        items.append(s)
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        k = x.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out[:200]


def _parse_objectives_file_storage(f) -> list[str]:
    if not f or not getattr(f, "filename", ""):
        return []
    try:
        data = f.read()
    except Exception:
        return []
    if not data:
        return []
    filename = getattr(f, "filename", "") or ""

    from app.objectives_multiformat import extract_objectives_from_file

    parsed = extract_objectives_from_file(filename, data)
    if parsed:
        return parsed

    text = data.decode("utf-8", errors="ignore")
    if "," in text or ";" in text:
        text = text.replace(";", "\n").replace(",", "\n")
    return _lines_to_objectives(text)


def _ctx(user=None, **extra):
    d = {
        "user": user,
        "RoleKey": RoleKey,
        "normalize_exercise_phase": _normalized_exercise_phase,
        "phase_label_ar": _phase_label_ar,
    }
    d.update(extra)
    return d


def _judge_assignment_for_current_exercise(db, user: User, ex: Exercise | None) -> JudgeTraineeAssignment | None:
    """تخصيص هذا المحكم للتمرين الحالي (إن وجد)."""
    if ex is None:
        return None
    judge_id = int(getattr(user, "id", 0) or 0)
    if judge_id <= 0:
        return None
    return (
        db.query(JudgeTraineeAssignment)
        .filter(
            JudgeTraineeAssignment.exercise_id == ex.id,
            JudgeTraineeAssignment.judge_user_id == judge_id,
        )
        .first()
    )


def _enforce_judge_unit_scope(db, user: User, ex: Exercise | None, unit_key: str) -> None:
    """للمحكمين (غير إدارة النظام): السماح فقط بوحدة المتدرب المخصصة لهم."""
    if ex is None:
        return
    if is_system_admin(user):
        return
    if not is_judge(user):
        return
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if assigned and assigned != (unit_key or "").strip():
        abort(403)


def _enforce_judge_has_assignment_for_unit(db, user: User, ex: Exercise | None, unit_key: str) -> None:
    """إذا كان المحكم مرتبطاً بوحدة غير موجودة في ملفات المعاضل/التقييم الحالية، امنع الوصول برسالة واضحة."""
    if ex is None or user is None:
        return
    if is_system_admin(user) or not is_judge(user):
        return
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if not assigned:
        abort(403)
    if assigned != (unit_key or "").strip():
        abort(403)


def _evaluation_canonical_saved_row(db, exercise_id: int, evaluation_item_id: int) -> EvaluationListSavedResult | None:
    """سجل نتيجة التقييم الموحّد لكل عنصر/تمرين (آخر تحديث زمني)."""
    return (
        db.query(EvaluationListSavedResult)
        .filter(
            EvaluationListSavedResult.exercise_id == exercise_id,
            EvaluationListSavedResult.evaluation_item_id == evaluation_item_id,
        )
        .order_by(EvaluationListSavedResult.updated_at.desc(), EvaluationListSavedResult.id.desc())
        .first()
    )


def _evaluation_canonical_map_for_items(db, exercise_id: int, item_ids: list[int]) -> dict[int, EvaluationListSavedResult]:
    """أحدث نتيجة محفوظة لكل عنصر — مع تفضيل السجلات المرتبطة بالتمرين الحالي."""
    if not item_ids:
        return {}
    rows = (
        db.query(EvaluationListSavedResult)
        .filter(EvaluationListSavedResult.evaluation_item_id.in_(item_ids))
        .order_by(EvaluationListSavedResult.updated_at.desc(), EvaluationListSavedResult.id.desc())
        .all()
    )
    out: dict[int, EvaluationListSavedResult] = {}
    for r in rows:
        iid = int(r.evaluation_item_id)
        prev = out.get(iid)
        if prev is None:
            out[iid] = r
            continue
        rid = int(getattr(r, "exercise_id", 0) or 0)
        pid = int(getattr(prev, "exercise_id", 0) or 0)
        if rid == int(exercise_id) and pid != int(exercise_id):
            out[iid] = r
    return out


def _evaluation_saved_total_pct(sr: EvaluationListSavedResult | None) -> float | None:
    if sr is None:
        return None
    v = getattr(sr, "total_pct", None)
    if v is not None:
        return float(v)
    rows = _parse_saved_eval_rows(getattr(sr, "payload_json", None))
    pcts = [float(x) for x in (_eval_row_score_pct(r) for r in rows) if x is not None]
    return (sum(pcts) / len(pcts)) if pcts else None


_CONTROL_REPORT_PHASE_FALLBACK: tuple[tuple[str, str], ...] = (
    ("preparation", "مرحلة التحضير"),
    ("opening", "مرحلة الإنفتاح"),
    ("battle_exposure", "مرحلة المعركة التعرضية"),
    ("reorganization", "مرحلة مسارات التقييم"),
)


def _control_report_catalog_phase_columns() -> list[tuple[str, str]]:
    """ترتيب مراحل التمرين من كتالوج التخطيط (بنك المعلومات — المدرجة في التمرين)."""
    if EXERCISE_PHASE_OPTIONS:
        return list(EXERCISE_PHASE_OPTIONS)
    return list(_CONTROL_REPORT_PHASE_FALLBACK)


def _control_report_effective_phase_key(raw: str | None) -> str:
    """مفتاح مرحلة موحّد للتقرير — يطابق كتالوج التخطيط مع دعم المفاتيح القديمة."""
    ph = _normalized_exercise_phase(raw)
    if not ph:
        return ""
    catalog_keys = {pk for pk, _ in _control_report_catalog_phase_columns()}
    if ph in catalog_keys:
        return ph
    legacy_to_catalog = {
        "main": "battle_exposure",
        "reorg": "reorganization",
    }
    alt = legacy_to_catalog.get(ph)
    if alt and alt in catalog_keys:
        return alt
    catalog_to_legacy = {
        "battle_exposure": "main",
        "reorganization": "reorg",
    }
    alt2 = catalog_to_legacy.get(ph)
    if alt2 and alt2 in catalog_keys:
        return alt2
    return ph


def _control_active_phase_columns(
    dots_by_unit_phase: dict[tuple[str, str], list[dict]],
) -> list[tuple[str, str]]:
    """مراحل لها تقييم محفوظ فعلياً فقط — بنفس ترتيب الكتالوج."""
    active_keys = {
        pk
        for (_uk, pk), dots in dots_by_unit_phase.items()
        if pk and dots
    }
    return [
        (pk, lbl)
        for pk, lbl in _control_report_catalog_phase_columns()
        if pk in active_keys
    ]

# مفتاح ألوان نتائج القوائم — متوافق مع grade_label_from_percent
_CONTROL_REPORT_GRADE_LEGEND: tuple[tuple[str, str, str], ...] = (
    ("راسب", "أقل من 60%", "#ef4444"),
    ("مقبول", "60% – 69%", "#f97316"),
    ("جيد", "70% – 79%", "#eab308"),
    ("جيد جدا", "80% – 89%", "#38bdf8"),
    ("ممتاز", "90% – 100%", "#22c55e"),
)
_CONTROL_REPORT_GRADE_COLORS: dict[str, str] = {
    "fail": "#ef4444",
    "medium": "#f97316",
    "good": "#eab308",
    "very_good": "#38bdf8",
    "excellent": "#22c55e",
}


def _control_report_grade_legend() -> list[dict]:
    return [{"label": lbl, "range": rng, "color": col} for lbl, rng, col in _CONTROL_REPORT_GRADE_LEGEND]


def _control_report_grade_band(pct: float) -> str:
    """مفتاح CSS/البيانات لمستوى النتيجة (يطابق grade_label_from_percent)."""
    p = float(pct)
    if p >= 90:
        return "excellent"
    if p >= 80:
        return "very_good"
    if p >= 70:
        return "good"
    if p >= 60:
        return "medium"
    return "fail"


def _control_report_grade_color(pct: float) -> str:
    return _CONTROL_REPORT_GRADE_COLORS[_control_report_grade_band(pct)]


def _control_report_dot_color(pct: float) -> str:
    return _control_report_grade_color(pct)


def _control_report_approval_location_ar(saved: EvaluationListSavedResult | None) -> str:
    if saved is None or not (getattr(saved, "payload_json", None) or "").strip():
        return "—"
    if eval_control_approved(saved):
        return "موقع الاعتماد: هيئة السيطرة"
    if eval_chief_approved(saved):
        return "موقع الاعتماد: كبير المحكمين"
    if eval_judge_approved(saved):
        return "موقع الاعتماد: المحكم"
    if eval_reopened_for_judge(saved):
        return "معاد للمحكم — بانتظار الاعتماد"
    return "محفوظ — بانتظار اعتماد المحكم"


def _control_report_dot_sort_key(dot: dict) -> tuple:
    """قوائم المحكم أولاً ثم قوائم المجرى/الإجراءات."""
    src = (dot.get("source") or "judge_eval").strip()
    return (0 if src == "judge_eval" else 1, int(dot.get("sort_id") or 0))


def _control_planner_flow_detail_dots(
    db,
    exercise_id: int,
    *,
    users_by_id: dict[int, str],
    judge_roster: dict[str, str],
) -> list[dict]:
    """نقاط تقرير السيطرة من حزم المجرى — قوائم تقييم الإجراءات المحفوظة للمحكم."""
    action_rows = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .join(ExercisePlannerFlowBundle)
        .filter(ExercisePlannerFlowBundle.exercise_id == int(exercise_id))
        .order_by(
            ExercisePlannerFlowBundle.unit_level_key,
            ExercisePlannerFlowBundle.exercise_phase,
            ExercisePlannerFlowBundleActionEval.slot_index,
            ExercisePlannerFlowBundleActionEval.id,
        )
        .all()
    )
    if not action_rows:
        return []
    dots: list[dict] = []
    for action_row in action_rows:
        bundle = action_row.bundle
        if bundle is None:
            continue
        canon = _planner_bundle_eval_canonical_saved(db, int(exercise_id), int(action_row.id))
        if canon is None or not (getattr(canon, "payload_json", None) or "").strip():
            continue
        pct_f = _evaluation_saved_total_pct(canon)
        if pct_f is None:
            continue
        uk = (getattr(bundle, "unit_level_key", None) or "").strip()
        ph = _control_report_effective_phase_key(getattr(bundle, "exercise_phase", None))
        if not uk or not ph:
            continue
        title = _planner_blob_display_filename(
            stored_title=action_row.title or "",
            relpath=action_row.file_relpath or "",
            fallback=f"قائمة تقييم إجراءات {int(action_row.slot_index)}",
        )
        jid = getattr(canon, "saved_by_id", None)
        jname = users_by_id.get(int(jid)) if jid is not None else None
        if not jname:
            jname = judge_roster.get(uk, "—")
        pv = int(round(float(pct_f)))
        dots.append(
            {
                "source": "planner_flow",
                "sort_id": int(action_row.id),
                "unit_key": uk,
                "phase_key": ph,
                "pct": pv,
                "color": _control_report_dot_color(pv),
                "list_title": title,
                "judge_name": jname,
                "approval_location": _control_report_approval_location_ar(canon),
                "view_url": url_for(
                    "views.control_planner_flow_action_view",
                    unit_key=uk,
                    action_eval_id=int(action_row.id),
                ),
            }
        )
    return dots


def _control_build_unit_detail_rows(
    db,
    exercise_id: int,
    eval_items: list,
    saved_by_item: dict[int, EvaluationListSavedResult],
) -> tuple[list[dict], list[tuple[str, str]]]:
    """صفوف جدول أداء الوحدات التفصيلي: نقطة ملونة لكل قائمة تقييم محفوظة ضمن مرحلة."""
    dots_by_unit_phase: dict[tuple[str, str], list[dict]] = {}

    user_ids = {
        int(sr.saved_by_id)
        for sr in saved_by_item.values()
        if getattr(sr, "saved_by_id", None) is not None
    }
    users_by_id: dict[int, str] = {}
    if user_ids:
        for u in db.query(User).filter(User.id.in_(user_ids)).all():
            users_by_id[int(u.id)] = (getattr(u, "full_name", "") or "").strip() or f"مستخدم #{u.id}"

    judge_roster: dict[str, str] = {}
    for jr in (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == exercise_id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .all()
    ):
        uk = (jr.unit_level_key or "").strip()
        if uk and uk not in judge_roster:
            judge_roster[uk] = (jr.full_name or "").strip() or "—"

    for it in eval_items:
        iid = int(getattr(it, "id", 0) or 0)
        sr = saved_by_item.get(iid)
        if sr is None:
            continue
        pct_f = _evaluation_saved_total_pct(sr)
        if pct_f is None:
            continue
        uk = (getattr(it, "unit_level_key", None) or "").strip()
        ph = _control_report_effective_phase_key(getattr(it, "exercise_phase", None))
        if not uk or not ph:
            continue
        title = (getattr(it, "text", None) or "قائمة تقييم").strip()
        jid = getattr(sr, "saved_by_id", None)
        jname = users_by_id.get(int(jid)) if jid is not None else None
        if not jname:
            jname = judge_roster.get(uk, "—")
        approval_loc = _control_report_approval_location_ar(sr)
        pv = int(round(float(pct_f)))
        dot = {
            "source": "judge_eval",
            "sort_id": iid,
            "unit_key": uk,
            "phase_key": ph,
            "pct": pv,
            "color": _control_report_dot_color(pv),
            "list_title": title,
            "judge_name": jname,
            "approval_location": approval_loc,
            "view_url": url_for(
                "views.control_evaluation_list_file_viewer",
                unit_key=uk,
                item_id=iid,
            ),
        }
        dots_by_unit_phase.setdefault((uk, ph), []).append(dot)

    for pdot in _control_planner_flow_detail_dots(
        db,
        exercise_id,
        users_by_id=users_by_id,
        judge_roster=judge_roster,
    ):
        uk = (pdot.get("unit_key") or "").strip()
        ph = (pdot.get("phase_key") or "").strip()
        if uk and ph:
            dots_by_unit_phase.setdefault((uk, ph), []).append(pdot)

    active_phase_columns = _control_active_phase_columns(dots_by_unit_phase)
    units_with_data: list[str] = []
    seen_uk: set[str] = set()
    for ul in UNIT_LEVELS:
        uk = (ul.get("key") or "").strip()
        if not uk or uk in seen_uk:
            continue
        if any(dots_by_unit_phase.get((uk, pk)) for pk, _ in active_phase_columns):
            units_with_data.append(uk)
            seen_uk.add(uk)
    for uk in sorted({k[0] for k in dots_by_unit_phase}):
        if uk not in seen_uk:
            units_with_data.append(uk)

    unit_rows: list[dict] = []
    for uk in units_with_data:
        phases_out = []
        for pk, plbl in active_phase_columns:
            phases_out.append(
                {
                    "key": pk,
                    "label": plbl,
                    "dots": sorted(
                        dots_by_unit_phase.get((uk, pk)) or [],
                        key=_control_report_dot_sort_key,
                    ),
                }
            )
        unit_rows.append(
            {
                "unit_key": uk,
                "unit_label": label_for_unit_level_key(uk) or uk,
                "phases": phases_out,
            }
        )
    return unit_rows, active_phase_columns


def _control_phase_max_dot_counts(unit_detail_rows: list[dict]) -> list[int]:
    """أقصى عدد نقاط تقييم لكل عمود مرحلة — لضبط عرض العمود في التقرير."""
    if not unit_detail_rows:
        return []
    n_phases = len(unit_detail_rows[0].get("phases") or [])
    maxes = [0] * n_phases
    for urow in unit_detail_rows:
        for i, ph in enumerate(urow.get("phases") or []):
            if i < n_phases:
                maxes[i] = max(maxes[i], len(ph.get("dots") or []))
    return [max(1, m) for m in maxes]


def _control_build_list_number_row(phase_max_dots: list[int]) -> list[dict]:
    """صف ترقيم قوائم التقييم — لكل مرحلة تسلسل مستقل يبدأ من 1."""
    out: list[dict] = []
    for max_dots in phase_max_dots:
        count = max(0, int(max_dots or 0))
        out.append({"slots": list(range(1, count + 1)), "max_dots": count})
    return out


def _control_group_label_lines(label: str) -> tuple[str, str]:
    """تقسيم اسم الوحدة لسطرين في رسم «أداء المجموعات»."""
    s = (label or "").strip() or "—"
    parts = [p.strip() for p in re.split(r"\s*/\s*", s) if p.strip()]
    if len(parts) >= 3:
        mid = (len(parts) + 1) // 2
        return (" / ".join(parts[:mid]), " / ".join(parts[mid:]))
    if len(parts) == 2:
        return (parts[0], parts[1])
    words = s.split()
    if len(words) >= 2:
        best_idx = 1
        best_diff = len(s)
        for i in range(1, len(words)):
            diff = abs(len(" ".join(words[:i])) - len(" ".join(words[i:])))
            if diff < best_diff:
                best_diff = diff
                best_idx = i
        return (" ".join(words[:best_idx]), " ".join(words[best_idx:]))
    return (s, "\u00a0")


def _evaluation_delete_duplicate_saves(db, *, exercise_id: int, evaluation_item_id: int, keep_id: int) -> None:
    db.query(EvaluationListSavedResult).filter(
        EvaluationListSavedResult.exercise_id == exercise_id,
        EvaluationListSavedResult.evaluation_item_id == evaluation_item_id,
        EvaluationListSavedResult.id != keep_id,
    ).delete(synchronize_session=False)


def _purge_eval_criterion_media_rows(db, media_rows: list) -> None:
    for m in media_rows:
        abs_p = criterion_media_absolute_path((getattr(m, "file_relpath", None) or "").strip())
        if abs_p is not None and abs_p.is_file():
            try:
                abs_p.unlink()
            except OSError:
                pass
        db.delete(m)


def _sync_evaluation_list_item_phase(
    db, item: EvaluationListPdfItem, phase_key: str
) -> None:
    """تحديث مرحلة التمرين على العنصر وجميع السجلات المرتبطة به في التمرين."""
    from app.models import JudgeIncompleteTaskStatus

    item.exercise_phase = phase_key
    if item.exercise_id is not None:
        db.query(EvaluationListSavedResult).filter(
            EvaluationListSavedResult.exercise_id == item.exercise_id,
            EvaluationListSavedResult.evaluation_item_id == item.id,
        ).update(
            {EvaluationListSavedResult.exercise_phase: phase_key},
            synchronize_session=False,
        )
        db.query(JudgeIncompleteTaskStatus).filter(
            JudgeIncompleteTaskStatus.exercise_id == item.exercise_id,
            JudgeIncompleteTaskStatus.evaluation_item_id == item.id,
        ).update(
            {JudgeIncompleteTaskStatus.exercise_phase: phase_key},
            synchronize_session=False,
        )


def _evaluation_grade_from_payload_rows(rows: list) -> tuple[float | None, str]:
    """
    نسبة إجمالية واقعية: مجموع المكتسبة ÷ مجموع القصوى لكل بنود التقييم (عدا «لا ينطبق»)،
    مع وزن 5 لكل بند بلا قصوى رقمية — يطابق حساب الواجهة عند وجود row_kind في الحمولة.
    للحمولات القديمة دون row_kind يُحتفظ بمتوسط نسب الصفوف كسلوك سابق.
    """
    safe = [r for r in rows[:2000] if isinstance(r, dict)]
    if not safe:
        return None, ""
    has_row_kind = any((str(r.get("row_kind") or "").strip()) for r in safe)
    if not has_row_kind:
        pcts: list[float] = []
        for r in safe:
            p = _eval_row_score_pct(r)
            if p is not None:
                pcts.append(float(p))
        if not pcts:
            return None, ""
        total_pct = sum(pcts) / len(pcts)
        return total_pct, grade_label_from_percent(total_pct)
    sum_acq = 0.0
    sum_den = 0.0
    for r in safe:
        if str(r.get("row_kind") or "").strip().lower() == "section":
            continue
        acq = r.get("acquired")
        acq_s = ("" if acq is None else str(acq)).strip().lower()
        if acq_s == "na":
            continue
        sum_den += _eval_row_effective_max(r)
        if acq_s:
            try:
                sum_acq += float(str(acq).replace(",", "."))
            except (TypeError, ValueError):
                pass
    if sum_den <= 0:
        return None, ""
    total_pct = (sum_acq / sum_den) * 100.0
    return total_pct, grade_label_from_percent(total_pct)


def _evaluation_payload_mark_totals(rows: list) -> tuple[float, float]:
    """مجموع القصوى والمكتسبة لحمولة تقييم محفوظة، بنفس منطق حساب النسبة."""
    safe = [r for r in rows[:2000] if isinstance(r, dict)]
    if not safe:
        return 0.0, 0.0
    sum_max = 0.0
    sum_acquired = 0.0
    for r in safe:
        if str(r.get("row_kind") or "").strip().lower() == "section":
            continue
        acq = r.get("acquired")
        acq_s = ("" if acq is None else str(acq)).strip().lower()
        if acq_s == "na":
            continue
        sum_max += _eval_row_effective_max(r)
        if acq_s:
            try:
                sum_acquired += float(str(acq).replace(",", "."))
            except (TypeError, ValueError):
                pass
    return sum_max, sum_acquired


def _evaluation_list_template_rows(item) -> list[dict]:
    """صفوف قالب Excel لقائمة التقييم (كما في واجهة المحكم)."""
    relpath = (getattr(item, "pdf_relpath", None) or "").strip()
    if not relpath:
        return []
    fspath = _evaluation_list_file_abspath(relpath)
    if fspath is None:
        return []
    sheet = read_evaluation_list_sheet(fspath)
    return list(sheet.get("eval_rows") or [])


def _evaluation_list_judge_sum_totals(
    saved_rows: list | None,
    template_rows: list | None = None,
) -> tuple[float, float]:
    """
    مجموع القصوى والمكتسبة كما في صفحة المحكم (eval-sum-max / eval-sum-acquired):
    - القصوى: مجموع max_num لكل بند score من قالب Excel.
    - المكتسبة: مجموع acquired المحفوظة (أو الأولية من Excel) لكل بند غير «لا ينطبق».
    """
    saved_rows = [r for r in (saved_rows or [])[:2000] if isinstance(r, dict)]
    template_rows = [r for r in (template_rows or [])[:2000] if isinstance(r, dict)]

    def _round_mark(value: float) -> float:
        return round(float(value) * 100.0) / 100.0

    if template_rows:
        sum_max = 0.0
        sum_acquired = 0.0
        any_acquired = False
        for idx, trow in enumerate(template_rows):
            rk = str(trow.get("row_kind") or "score").strip().lower()
            if rk == "section":
                continue
            mx = trow.get("max_num")
            if mx is None:
                mx = parse_max_cell(trow.get("max_val"))
            if mx is not None and float(mx) > 0:
                sum_max += float(mx)
            saved_row = saved_rows[idx] if idx < len(saved_rows) else {}
            acq = saved_row.get("acquired") if isinstance(saved_row, dict) else None
            if isinstance(saved_row, dict) and "acquired" in saved_row:
                acq_s = ("" if acq is None else str(acq)).strip().lower()
            else:
                acq = trow.get("acquired_initial")
                acq_s = ("" if acq is None else str(acq)).strip().lower()
            if acq_s and acq_s != "na":
                try:
                    sum_acquired += float(str(acq).replace(",", "."))
                    any_acquired = True
                except (TypeError, ValueError):
                    pass
        sum_max = _round_mark(sum_max) if sum_max > 0 else 0.0
        if not any_acquired:
            return sum_max, 0.0
        return sum_max, _round_mark(sum_acquired)

    sum_max = 0.0
    sum_acquired = 0.0
    any_acquired = False
    for r in saved_rows:
        if str(r.get("row_kind") or "").strip().lower() == "section":
            continue
        mx = parse_max_cell(r.get("max_val"))
        if mx is not None and mx > 0:
            sum_max += float(mx)
        acq = r.get("acquired")
        acq_s = ("" if acq is None else str(acq)).strip().lower()
        if acq_s and acq_s != "na":
            try:
                sum_acquired += float(str(acq).replace(",", "."))
                any_acquired = True
            except (TypeError, ValueError):
                pass
    sum_max = _round_mark(sum_max) if sum_max > 0 else 0.0
    if not any_acquired:
        return sum_max, 0.0
    return sum_max, _round_mark(sum_acquired)


def _evaluation_list_page_footer_stats(
    saved_rows: list | None,
    template_rows: list | None = None,
    *,
    saved_total_pct: float | None = None,
    saved_grade: str | None = None,
) -> dict:
    """إحصائيات تذييل قائمة التقييم: مجموع القصوى، المكتسبة، النسبة الفعلية، التقدير."""
    sum_max, sum_acquired = _evaluation_list_judge_sum_totals(saved_rows, template_rows)
    actual_pct = saved_total_pct
    grade = (saved_grade or "").strip()
    if actual_pct is None and saved_rows:
        actual_pct, grade_computed = _evaluation_grade_from_payload_rows(saved_rows)
        if not grade:
            grade = grade_computed or ""
    elif actual_pct is not None and not grade:
        grade = grade_label_from_percent(actual_pct) or ""
    return {
        "sum_max": sum_max,
        "sum_acquired": sum_acquired,
        "actual_pct": actual_pct,
        "grade": grade or "—",
    }


_FINAL_EVAL_PHASE_ALIASES: dict[str, tuple[str, ...]] = {
    "preparation": ("preparation",),
    "opening": ("opening",),
    "main": ("main", "battle_exposure"),
    "reorg": ("reorg", "reorganization", "evaluation_tracks"),
    "evaluation_tracks": ("evaluation_tracks", "reorg", "reorganization"),
}


def _final_eval_logical_phase_key(phase_key: str) -> str:
    pk = (phase_key or "").strip()
    if pk in ("main", "battle_exposure"):
        return "main"
    if pk in ("reorg", "reorganization", "evaluation_tracks"):
        return "reorg"
    return pk


def _final_eval_detail_rows_for_unit_phase(
    details_by_unit_phase: dict[tuple[str, str], list[dict]],
    unit_key: str,
    phase_key: str,
) -> list[dict]:
    """صفوف قوائم التقييم لوحدة ومرحلة مع دمج مفاتيح المراحل المرادفة."""
    keys = _FINAL_EVAL_PHASE_ALIASES.get(
        (phase_key or "").strip(),
        ((phase_key or "").strip(),),
    )
    merged: list[dict] = []
    seen_ids: set[int] = set()
    for pk in keys:
        if not pk:
            continue
        for row in details_by_unit_phase.get((unit_key, pk), []):
            item_id = row.get("item_id")
            if item_id is not None:
                iid = int(item_id)
                if iid in seen_ids:
                    continue
                seen_ids.add(iid)
            merged.append(row)
    merged.sort(key=lambda x: (x.get("item_sort_order", 0), x.get("item_id", 0)))
    return merged


def _final_report_phase_summary(rows: list[dict]) -> dict:
    max_mark = sum(float(r.get("max_mark") or 0.0) for r in rows)
    acquired_mark = sum(float(r.get("acquired_mark") or 0.0) for r in rows)
    phase_pcts = [
        float(r["phase_pct"]) for r in rows if r.get("phase_pct") is not None
    ]
    if phase_pcts:
        pct = sum(phase_pcts) / len(phase_pcts)
    else:
        pct = (acquired_mark / max_mark) * 100.0 if max_mark > 0 else None
    return {
        "max_mark": max_mark,
        "acquired_mark": acquired_mark,
        "pct": pct,
        "grade": grade_label_from_percent(pct) if pct is not None else "—",
    }


def _apply_criteria_totals_to_phase_rows(
    phase_rows: list[dict],
    *,
    criteria_max_by_phase: dict[str, float],
    criteria_acquired_by_phase: dict[str, float],
    criteria_pct_by_phase: dict[str, float | None],
    criteria_grade_by_phase: dict[str, str] | None = None,
) -> None:
    """جدول جميع المراحل: القصوى/المكتسبة/النسبة/التقدير من صف الإجمالي في جدول تفاصيل المرحلة."""
    grades = criteria_grade_by_phase or {}
    for row in phase_rows:
        phase_key = _final_eval_logical_phase_key(row.get("phase_key") or "")
        if phase_key not in criteria_max_by_phase:
            continue
        max_mark = float(criteria_max_by_phase.get(phase_key) or 0.0)
        acquired_mark = float(criteria_acquired_by_phase.get(phase_key, 0.0))
        pct = criteria_pct_by_phase.get(phase_key)
        row["max_mark"] = max_mark
        row["acquired_mark"] = acquired_mark
        if pct is not None:
            row["phase_pct"] = float(pct)
        elif max_mark > 0:
            row["phase_pct"] = (
                (acquired_mark / max_mark) * 100.0 if acquired_mark > 0 else 0.0
            )
        else:
            row["phase_pct"] = None
        if phase_key in grades:
            row["phase_grade"] = grades[phase_key]
        else:
            row["phase_grade"] = (
                grade_label_from_percent(row["phase_pct"])
                if row.get("phase_pct") is not None
                else "—"
            )


def _ensure_unit_phase_rows_for_all_phases(
    unit_key: str,
    unit_label: str,
    existing_rows: list[dict],
) -> list[dict]:
    """صف مرحلة لكل مرحلة تمرين — لعرض جميع مستويات الوحدات في التقرير النهائي."""
    by_phase: dict[str, dict] = {}
    for row in existing_rows or []:
        pk = (row.get("phase_key") or "").strip()
        if pk:
            by_phase[pk] = row
    out: list[dict] = []
    for phase_key in exercise_phase_keys():
        if phase_key in by_phase:
            out.append(by_phase[phase_key])
            continue
        out.append(
            {
                "unit_key": unit_key,
                "unit_label": unit_label,
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "max_mark": 0.0,
                "acquired_mark": 0.0,
                "phase_pct": None,
                "phase_grade": "—",
                "unit_total_pct": None,
                "unit_grade": "—",
            }
        )
    return out


def _recompute_unit_total_pct_on_phase_rows(unit_phase_rows: list[dict]) -> None:
    phase_pcts = [
        float(r["phase_pct"])
        for r in unit_phase_rows
        if r.get("phase_key") and r.get("phase_pct") is not None
    ]
    unit_total_pct = sum(phase_pcts) / len(phase_pcts) if phase_pcts else None
    unit_grade = (
        grade_label_from_percent(unit_total_pct) if unit_total_pct is not None else "—"
    )
    for row in unit_phase_rows:
        row["unit_total_pct"] = unit_total_pct
        row["unit_grade"] = unit_grade


_FINAL_EVAL_REPORT_PHASE_MAX_PREFIX = "report_phase_max__"


def _report_phase_max_field_name(unit_key: str, phase_key: str) -> str:
    """اسم حقل فريد لكل وحدة+مرحلة في نموذج التقرير النهائي."""
    return f"{_FINAL_EVAL_REPORT_PHASE_MAX_PREFIX}{(unit_key or '').strip()}|{_normalized_exercise_phase(phase_key)}"


def _parse_report_phase_max_field_name(name: str) -> tuple[str, str] | None:
    if not name.startswith(_FINAL_EVAL_REPORT_PHASE_MAX_PREFIX):
        return None
    rest = name[len(_FINAL_EVAL_REPORT_PHASE_MAX_PREFIX) :]
    if "|" not in rest:
        return None
    unit_key, phase_key = rest.split("|", 1)
    unit_key = (unit_key or "").strip()
    phase_key = _normalized_exercise_phase(phase_key)
    if not unit_key or not phase_key:
        return None
    return unit_key, phase_key


def _final_eval_phase_manual_max_map(db, exercise_id: int) -> dict[tuple[str, str], float]:
    rows = (
        db.query(AnalystFinalEvaluationPhaseAllocatedMax)
        .filter(AnalystFinalEvaluationPhaseAllocatedMax.exercise_id == int(exercise_id))
        .all()
    )
    out: dict[tuple[str, str], float] = {}
    for row in rows:
        if row.max_mark is None:
            continue
        uk = (row.unit_level_key or "").strip()
        pk = _normalized_exercise_phase(row.phase_key)
        if uk and pk:
            out[(uk, pk)] = float(row.max_mark)
    return out


def _analyst_criteria_phase_max_map(db, exercise_id: int) -> dict[tuple[str, str], float]:
    """يبني خريطة (unit_level_key, phase_key) → مجموع 'القصوى' المُدخل في
    «مساحة المحللين / معايير التقييم / جدول توزيع النسبة المئوية الإجمالية للتقييم».

    تُستخدم في التقرير النهائي لتعبئة عمود «القصوى» تلقائياً عندما لا تتوفر قيمة
    يدوية محفوظة في AnalystFinalEvaluationPhaseAllocatedMax.
    """
    criteria_units = (
        db.query(AnalystEvaluationCriteriaUnit)
        .filter(AnalystEvaluationCriteriaUnit.exercise_id == int(exercise_id))
        .all()
    )
    unit_id_to_level: dict[int, str] = {}
    for cu in criteria_units:
        uk = _resolve_unit_level_key_for_criteria_label(cu.label or "")
        if uk:
            unit_id_to_level[int(cu.id)] = uk
    if not unit_id_to_level:
        return {}

    phase_items = (
        db.query(AnalystEvaluationCriteriaPhaseItem)
        .filter(AnalystEvaluationCriteriaPhaseItem.exercise_id == int(exercise_id))
        .all()
    )
    out: dict[tuple[str, str], float] = {}
    for item in phase_items:
        if item.allocated_mark is None:
            continue
        cu_id = int(item.criteria_unit_id or 0)
        uk = unit_id_to_level.get(cu_id, "")
        pk = _normalized_exercise_phase(item.phase_key or "")
        if not uk or not pk:
            continue
        out[(uk, pk)] = out.get((uk, pk), 0.0) + float(item.allocated_mark)
    return out


def _upsert_final_eval_report_phase_max(
    db,
    *,
    exercise_id: int,
    unit_key: str,
    phase_key: str,
    mark_raw: str | None,
) -> float | None:
    """حفظ/حذف علامة قصوى يدوية لصف وحدة+مرحلة؛ يُرجع القيمة المحفوظة أو None."""
    unit_key = (unit_key or "").strip()
    phase_key = _normalized_exercise_phase(phase_key)
    if not unit_key or not phase_key:
        return None
    raw = (mark_raw or "").strip().replace(",", ".")
    mark = _parse_mark_form_value(raw) if raw else None
    row = (
        db.query(AnalystFinalEvaluationPhaseAllocatedMax)
        .filter(
            AnalystFinalEvaluationPhaseAllocatedMax.exercise_id == int(exercise_id),
            AnalystFinalEvaluationPhaseAllocatedMax.unit_level_key == unit_key,
            AnalystFinalEvaluationPhaseAllocatedMax.phase_key == phase_key,
        )
        .first()
    )
    if mark is None:
        if row is not None:
            db.delete(row)
        return None
    if row is None:
        row = AnalystFinalEvaluationPhaseAllocatedMax(
            exercise_id=int(exercise_id),
            unit_level_key=unit_key,
            phase_key=phase_key,
        )
        db.add(row)
    row.max_mark = float(mark)
    row.unit_level_key = unit_key
    row.phase_key = phase_key
    return float(mark)


def _final_eval_phase_acquired_total(
    db, *, exercise_id: int, unit_key: str, phase_key: str
) -> float:
    """مجموع المكتسبة من قوائم التقييم المحفوظة/المعتمدة لوحدة ومرحلة."""
    unit_key = (unit_key or "").strip()
    phase_key = _normalized_exercise_phase(phase_key)
    if not unit_key or not phase_key:
        return 0.0
    items = (
        db.query(EvaluationListPdfItem)
        .filter(
            EvaluationListPdfItem.exercise_id == int(exercise_id),
            EvaluationListPdfItem.unit_level_key == unit_key,
        )
        .all()
    )
    item_ids = [
        int(it.id)
        for it in items
        if _normalized_exercise_phase(getattr(it, "exercise_phase", None)) == phase_key
    ]
    if not item_ids:
        return 0.0
    latest_by_item = _evaluation_canonical_map_for_items(db, int(exercise_id), item_ids)
    acquired = 0.0
    for item_id in item_ids:
        saved = latest_by_item.get(int(item_id))
        if saved is None or not (getattr(saved, "payload_json", "") or "").strip():
            continue
        rows = _parse_saved_eval_rows(saved.payload_json)
        _payload_max, item_acquired = _evaluation_payload_mark_totals(rows)
        if item_acquired > 0:
            acquired += item_acquired
    return acquired


def _final_eval_phase_row_metrics(
    db,
    *,
    exercise_id: int,
    unit_key: str,
    phase_key: str,
    manual_max: float | None = None,
) -> dict:
    """مقاييس صف مرحلة بعد إدخال القصوى اليدوية."""
    unit_key = (unit_key or "").strip()
    phase_key = _normalized_exercise_phase(phase_key)
    acquired_mark = _final_eval_phase_acquired_total(
        db, exercise_id=int(exercise_id), unit_key=unit_key, phase_key=phase_key
    )
    if manual_max is None:
        stored = _final_eval_phase_manual_max_map(db, int(exercise_id)).get((unit_key, phase_key))
        manual_max = float(stored) if stored is not None else None
    if manual_max is not None:
        max_mark = float(manual_max)
    else:
        # «القصوى» من معايير التقييم كقيمة افتراضية عند غياب الإدخال اليدوي
        criteria_max = _analyst_criteria_phase_max_map(db, int(exercise_id)).get(
            (unit_key, phase_key)
        )
        max_mark = float(criteria_max) if criteria_max is not None else 0.0
    phase_pct: float | None = None
    if max_mark > 0 and acquired_mark > 0:
        phase_pct = (acquired_mark / max_mark) * 100.0
    elif max_mark > 0 and acquired_mark <= 0:
        phase_pct = 0.0
    return {
        "acquired_mark": acquired_mark,
        "manual_max_mark": float(manual_max) if manual_max is not None else None,
        "phase_pct": phase_pct,
        "phase_grade": grade_label_from_percent(phase_pct) if phase_pct is not None else "—",
    }


def _save_final_eval_report_phase_maxes(db, *, exercise_id: int) -> None:
    """حفظ علامات القصوى اليدوية لجدول التقرير النهائي (حقل مستقل لكل وحدة ومرحلة)."""
    for field_name in request.form:
        parsed = _parse_report_phase_max_field_name(field_name)
        if parsed is None:
            continue
        unit_key, phase_key = parsed
        _upsert_final_eval_report_phase_max(
            db,
            exercise_id=int(exercise_id),
            unit_key=unit_key,
            phase_key=phase_key,
            mark_raw=request.form.get(field_name),
        )
    db.commit()


def _final_eval_manual_max_map(db, exercise_id: int) -> dict[int, float]:
    rows = (
        db.query(AnalystFinalEvaluationAllocatedMax)
        .filter(AnalystFinalEvaluationAllocatedMax.exercise_id == int(exercise_id))
        .all()
    )
    out: dict[int, float] = {}
    for row in rows:
        if row.max_mark is None:
            continue
        out[int(row.evaluation_item_id)] = float(row.max_mark)
    return out


def _upsert_final_eval_item_allocated_max(
    db,
    *,
    exercise_id: int,
    unit_key: str,
    item_id: int,
    mark_raw: str | None,
) -> float | None:
    """حفظ/حذف علامة قصوى مخصصة لقائمة تقييم واحدة."""
    unit_key = (unit_key or "").strip()
    item = db.get(EvaluationListPdfItem, int(item_id))
    if item is None or int(item.exercise_id or 0) != int(exercise_id):
        return None
    if (item.unit_level_key or "").strip() != unit_key:
        return None
    raw = (mark_raw or "").strip().replace(",", ".")
    mark = _parse_mark_form_value(raw) if raw else None
    phase_key = _normalized_exercise_phase(item.exercise_phase)
    row = (
        db.query(AnalystFinalEvaluationAllocatedMax)
        .filter(
            AnalystFinalEvaluationAllocatedMax.exercise_id == int(exercise_id),
            AnalystFinalEvaluationAllocatedMax.evaluation_item_id == int(item_id),
        )
        .first()
    )
    if mark is None:
        if row is not None:
            db.delete(row)
        return None
    if row is None:
        row = AnalystFinalEvaluationAllocatedMax(
            exercise_id=int(exercise_id),
            evaluation_item_id=int(item_id),
            unit_level_key=unit_key,
            phase_key=phase_key,
        )
        db.add(row)
    row.max_mark = float(mark)
    row.unit_level_key = unit_key
    row.phase_key = phase_key
    return float(mark)


def _save_final_eval_manual_maxes(
    db,
    *,
    exercise_id: int,
    unit_key: str,
    item_ids: list[int],
) -> None:
    """حفظ علامات القصوى المخصصة من نموذج التقييم النهائي (حقول allocated_max__{item_id})."""
    unit_key = (unit_key or "").strip()
    for item_id in item_ids:
        _upsert_final_eval_item_allocated_max(
            db,
            exercise_id=int(exercise_id),
            unit_key=unit_key,
            item_id=int(item_id),
            mark_raw=request.form.get(f"allocated_max__{item_id}"),
        )
    db.commit()


def _allocated_pct_from_marks(acquired_mark: float | None, max_mark: float | None) -> float | None:
    if max_mark is None or float(max_mark) <= 0:
        return None
    if acquired_mark is None:
        return None
    return (float(acquired_mark) / float(max_mark)) * 100.0  # يشمل 0% عند مكتسبة = 0


def _allocated_acquired_from_pct_and_max(
    pct: float | None, max_mark: float | None
) -> float | None:
    """مكتسبة العلامة المخصصة = (النسبة ÷ 100) × القصوى."""
    if pct is None or max_mark is None or float(max_mark) <= 0:
        return None
    return round((float(pct) / 100.0) * float(max_mark), 2)


def _eval_list_marks_by_label(
    detail_rows: list[dict],
) -> tuple[dict[str, float], dict[str, float], dict[str, float | None], dict[str, str]]:
    """قصوى ومكتسبة ونسبة فعلية وتقدير حسب عنوان قائمة التقييم."""
    max_by_label: dict[str, float] = {}
    acq_by_label: dict[str, float] = {}
    pct_by_label: dict[str, float | None] = {}
    grade_by_label: dict[str, str] = {}
    for row in detail_rows or []:
        key = _norm_eval_list_label_for_match(row.get("list_label") or "")
        if not key or not row.get("has_saved_payload"):
            continue
        max_by_label[key] = float(row.get("evaluation_list_max_mark") or 0.0)
        acq_by_label[key] = float(row.get("evaluation_list_acquired_mark") or 0.0)
        pct_by_label[key] = row.get("evaluation_list_actual_pct")
        grade_by_label[key] = (row.get("evaluation_list_grade") or row.get("allocated_grade") or "—")
    return max_by_label, acq_by_label, pct_by_label, grade_by_label


def _enrich_criteria_items_with_acquired(
    criteria_items: list[dict],
    all_detail_rows: list[dict],
) -> list[dict]:
    """ربط علامات قوائم التقييم — النسبة = النسبة المئوية الفعلية من صفحة قائمة التقييم."""
    max_by_label, acq_by_label, pct_by_label, grade_by_label = _eval_list_marks_by_label(
        all_detail_rows
    )
    enriched: list[dict] = []
    for item in criteria_items or []:
        row = {**item}
        key = _norm_eval_list_label_for_match(row.get("criteria_text") or "")
        eval_max = max_by_label.get(key) if key in max_by_label else None
        eval_acq = acq_by_label.get(key) if key in acq_by_label else None
        actual_pct = pct_by_label.get(key) if key in pct_by_label else None
        row["evaluation_list_max_mark"] = eval_max
        row["evaluation_list_acquired_mark"] = eval_acq
        row["evaluation_list_actual_pct"] = actual_pct
        pct = actual_pct
        if pct is None:
            pct = _allocated_pct_from_marks(eval_acq, eval_max)
        row["allocated_pct"] = pct
        alloc_max = row.get("allocated_mark")
        if alloc_max is not None:
            try:
                alloc_max = float(alloc_max)
            except (TypeError, ValueError):
                alloc_max = None
        row["allocated_acquired_mark"] = _allocated_acquired_from_pct_and_max(pct, alloc_max)
        row["allocated_grade"] = (
            grade_by_label.get(key)
            if key in grade_by_label
            else (grade_label_from_percent(pct) if pct is not None else "—")
        )
        enriched.append(row)
    return enriched


def _criteria_items_allocated_acquired_total(items: list[dict]) -> float:
    return sum(
        float(item["allocated_acquired_mark"])
        for item in (items or [])
        if item.get("allocated_acquired_mark") is not None
    )


def _criteria_items_eval_list_max_total(items: list[dict]) -> float:
    return sum(
        float(item["evaluation_list_max_mark"])
        for item in (items or [])
        if item.get("evaluation_list_max_mark") is not None
    )


def _criteria_items_eval_list_acquired_total(items: list[dict]) -> float:
    return sum(
        float(item["evaluation_list_acquired_mark"])
        for item in (items or [])
        if item.get("evaluation_list_acquired_mark") is not None
    )


def _criteria_items_eval_list_footer_pct(items: list[dict]) -> float | None:
    """نسبة إجمالية للمرحلة من النسب الفعلية لقوائم التقييم."""
    if not items:
        return None
    actuals = [
        float(item["evaluation_list_actual_pct"])
        for item in items
        if item.get("evaluation_list_actual_pct") is not None
    ]
    if len(actuals) == len(items) and len(actuals) == 1:
        return actuals[0]
    return _allocated_pct_from_marks(
        _criteria_items_eval_list_acquired_total(items),
        _criteria_items_eval_list_max_total(items),
    )


def _criteria_items_eval_list_footer_grade(items: list[dict], pct: float | None) -> str:
    if pct is None:
        return "—"
    grades = {
        (item.get("allocated_grade") or item.get("evaluation_list_grade") or "").strip()
        for item in (items or [])
        if (item.get("allocated_grade") or item.get("evaluation_list_grade"))
    }
    grades.discard("")
    grades.discard("—")
    if len(grades) == 1:
        return next(iter(grades))
    return grade_label_from_percent(pct) or "—"


def _round_pct_display(value: float) -> float:
    """تقريب النسبة كما في الجدول (منزلتان) قبل تجميع المربعات."""
    return round(float(value), 2)


def _build_final_report_exercise_summary(final_rows: list[dict]) -> dict:
    """نسب المربعات = مجموع نسب صفوف الجدول (لكل مرحلة) ÷ عدد الوحدات؛ المجموع العام = مجموع نسب المراحل ÷ عدد المراحل."""
    by_phase_units: dict[str, dict[str, float]] = {}
    phase_labels: dict[str, str] = {}
    for row in final_rows:
        phase_key = (row.get("phase_key") or "").strip()
        if not phase_key:
            continue
        pct = row.get("phase_pct")
        if pct is None:
            continue
        unit_key = (row.get("unit_key") or "").strip() or "—"
        phase_labels[phase_key] = (row.get("phase_label") or _phase_label_ar(phase_key))
        by_phase_units.setdefault(phase_key, {})[unit_key] = _round_pct_display(float(pct))

    phase_summaries: list[dict] = []
    phase_pcts_for_exercise: list[float] = []
    for phase_key in exercise_phase_keys():
        unit_pcts_map = by_phase_units.get(phase_key, {})
        if unit_pcts_map:
            unit_pcts = list(unit_pcts_map.values())
            phase_pct = _round_pct_display(sum(unit_pcts) / len(unit_pcts))
            phase_pcts_for_exercise.append(phase_pct)
        else:
            phase_pct = None
        phase_summaries.append(
            {
                "phase_key": phase_key,
                "phase_label": phase_labels.get(phase_key) or _phase_label_ar(phase_key),
                "pct": phase_pct,
                "grade": grade_label_from_percent(phase_pct) if phase_pct is not None else "—",
                "unit_count": len(unit_pcts_map),
            }
        )

    exercise_pct = (
        _round_pct_display(sum(phase_pcts_for_exercise) / len(phase_pcts_for_exercise))
        if phase_pcts_for_exercise
        else None
    )
    return {
        "phase_summaries": phase_summaries,
        "exercise_pct": exercise_pct,
        "exercise_grade": grade_label_from_percent(exercise_pct) if exercise_pct is not None else "—",
        "phase_count": len(phase_pcts_for_exercise),
    }


# ألوان مخطط مراحل التمرين — لون ثابت لكل مرحلة (4 ألوان مميزة).
_CONTROL_PHASE_BAR_HEX: dict[str, str] = {
    "preparation": "#3b82f6",
    "opening": "#9333ea",
    "main": "#f59e0b",
    "battle_exposure": "#f59e0b",
    "reorg": "#14b8a6",
    "reorganization": "#14b8a6",
    "evaluation_tracks": "#10b981",
}


def _control_phase_bar_hex(phase_key: str) -> str:
    return _CONTROL_PHASE_BAR_HEX.get((phase_key or "").strip(), "#8b7355")


def _phase_summary_from_eval_items(
    eval_items: list,
    saved_by_item: dict[int, object],
) -> dict:
    """ملخص نسب المراحل (نفس منطق التقييم النهائي للمحللين) من قوائم التقييم المحفوظة."""
    phase_totals: dict[tuple[str, str], dict] = {}
    for item in eval_items:
        saved = saved_by_item.get(int(getattr(item, "id", 0) or 0))
        if saved is None:
            continue
        rows = _parse_saved_eval_rows(getattr(saved, "payload_json", None))
        max_mark, acquired_mark = _evaluation_payload_mark_totals(rows)
        if max_mark <= 0:
            continue
        unit_key = (getattr(item, "unit_level_key", None) or getattr(saved, "unit_level_key", None) or "").strip()
        phase_key = _normalized_exercise_phase(
            getattr(item, "exercise_phase", None) or getattr(saved, "exercise_phase", None)
        )
        block = phase_totals.setdefault(
            (unit_key, phase_key),
            {
                "unit_key": unit_key,
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "max_mark": 0.0,
                "acquired_mark": 0.0,
            },
        )
        block["max_mark"] += max_mark
        block["acquired_mark"] += acquired_mark

    final_rows: list[dict] = []
    for block in phase_totals.values():
        max_mark = float(block.get("max_mark") or 0.0)
        phase_pct = (float(block["acquired_mark"]) / max_mark) * 100.0 if max_mark > 0 else None
        final_rows.append({**block, "phase_pct": phase_pct})
    return _build_final_report_exercise_summary(final_rows)


def _phase_summary_for_control_report(
    db,
    exercise_id: int,
    eval_items: list,
    saved_by_item: dict[int, EvaluationListSavedResult],
) -> dict:
    """ملخص مراحل التقرير من قوائم المحكم + قوائم المجرى/الإجراءات."""
    phase_totals: dict[tuple[str, str], dict] = {}

    def _accumulate(unit_key: str, phase_key: str, rows: list) -> None:
        max_mark, acquired_mark = _evaluation_payload_mark_totals(rows)
        if max_mark <= 0:
            return
        block = phase_totals.setdefault(
            (unit_key, phase_key),
            {
                "unit_key": unit_key,
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "max_mark": 0.0,
                "acquired_mark": 0.0,
            },
        )
        block["max_mark"] += max_mark
        block["acquired_mark"] += acquired_mark

    for it in eval_items:
        saved = saved_by_item.get(int(getattr(it, "id", 0) or 0))
        if saved is None:
            continue
        rows = _parse_saved_eval_rows(getattr(saved, "payload_json", None))
        uk = (
            getattr(it, "unit_level_key", None) or getattr(saved, "unit_level_key", None) or ""
        ).strip()
        ph = _control_report_effective_phase_key(
            getattr(it, "exercise_phase", None) or getattr(saved, "exercise_phase", None)
        )
        if uk and ph:
            _accumulate(uk, ph, rows)

    for action_row in (
        db.query(ExercisePlannerFlowBundleActionEval)
        .join(ExercisePlannerFlowBundle)
        .filter(ExercisePlannerFlowBundle.exercise_id == int(exercise_id))
        .all()
    ):
        bundle = action_row.bundle
        if bundle is None:
            continue
        canon = _planner_bundle_eval_canonical_saved(db, int(exercise_id), int(action_row.id))
        if canon is None or not (getattr(canon, "payload_json", None) or "").strip():
            continue
        rows = _parse_saved_eval_rows(getattr(canon, "payload_json", None))
        uk = (getattr(bundle, "unit_level_key", None) or "").strip()
        ph = _control_report_effective_phase_key(getattr(bundle, "exercise_phase", None))
        if uk and ph:
            _accumulate(uk, ph, rows)

    final_rows: list[dict] = []
    for block in phase_totals.values():
        max_mark = float(block.get("max_mark") or 0.0)
        phase_pct = (float(block["acquired_mark"]) / max_mark) * 100.0 if max_mark > 0 else None
        final_rows.append({**block, "phase_pct": phase_pct})
    return _build_final_report_exercise_summary(final_rows)


def _distribution_from_phase_summary(summary: dict) -> list[dict]:
    """أعمدة مخطط «أداء الوحدات حسب مراحل التمرين»: مراحل مُقيَّمة فقط (لا مرحلة بلا نتيجة)."""
    phase_order = {
        key: idx for idx, key in enumerate(pk for pk, _ in _control_report_catalog_phase_columns())
    }
    items: list[dict] = []
    for ps in summary.get("phase_summaries") or []:
        phase_key = (ps.get("phase_key") or "").strip()
        if not phase_key:
            continue
        pct = ps.get("pct")
        if pct is None:
            continue
        pct_f = float(pct)
        items.append(
            {
                "phase_key": phase_key,
                "label": ps.get("phase_label") or _phase_label_ar(phase_key),
                "pct": pct_f,
                "pct_display": _round_pct_display(pct_f),
                "count": int(ps.get("unit_count") or 0),
                "color": _control_phase_bar_hex(phase_key),
                "_order": phase_order.get(phase_key, 999),
            }
        )
    items.sort(key=lambda x: x["_order"])
    for row in items:
        row.pop("_order", None)
    return items


def _distribution_phase_donut_css(items: list[dict]) -> str:
    """تدرج دائري لمخطط مراحل التمرين — حجم الشريحة يتناسب مع نسبة المرحلة."""
    if not items:
        return "conic-gradient(var(--tint-200, #e8e0d8) 0deg 360deg)"
    weights = [max(0.0, float(x.get("pct") or 0)) for x in items]
    total = sum(weights)
    stops: list[str] = []
    acc = 0.0
    if total <= 0:
        share_each = 100.0 / len(items)
        for item in items:
            col = (item.get("color") or "#e8e0d8").strip()
            nxt = acc + share_each
            stops.append(f"{col} {acc:.2f}% {nxt:.2f}%")
            item["share_pct"] = round(share_each, 1)
            acc = nxt
    else:
        for item, w in zip(items, weights):
            share = (w / total) * 100.0
            col = (item.get("color") or "#e8e0d8").strip()
            nxt = acc + share
            stops.append(f"{col} {acc:.2f}% {nxt:.2f}%")
            item["share_pct"] = round(share, 1)
            acc = nxt
    return f"conic-gradient(from 0.25turn, {', '.join(stops)})"


FINAL_EVALUATION_TRACK_UNIT_KEYS: set[str] = {
    "mech_infantry_bn",
    "mech_infantry_bn_3",
    "mech_infantry_bn_13",
    "tank_bn",
    "tank_bn_4",
}
FINAL_EVALUATION_TRACK_PHASE_KEY = "reorg"
FINAL_EVALUATION_TRACK_PHASE_LABEL = "مرحلة مسارات التقييم"

# ربط مرحلة التقييم النهائي بمرحلة معايير التقييم (جدول التوزيع → تفاصيل المرحلة).
_FINAL_EVAL_TO_CRITERIA_PHASE: dict[str, tuple[str, ...]] = {
    "preparation": ("preparation",),
    "opening": ("opening",),
    "main": ("main", "battle_exposure"),
    "reorg": ("reorg", "reorganization", "evaluation_tracks"),
    "evaluation_tracks": ("evaluation_tracks", "reorg", "reorganization"),
}


def _evaluation_commit_payload_save(
    db,
    *,
    user: User,
    item: EvaluationListPdfItem,
    current_exercise: Exercise,
    raw: str,
) -> None:
    """يحدّث أو ينشئ السجل الموحّد لنتائج التقييم؛ بعد الحفظ يُزال أي سجل قديم مكرّر لنفس العنصر."""
    if not can_save_evaluation_results(user):
        abort(403)
    try:
        payload = json.loads(raw)
    except Exception:
        abort(400)
    if not isinstance(payload, dict):
        abort(400)
    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        abort(400)
    total_pct, grade = _evaluation_grade_from_payload_rows(rows)
    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is not None and not eval_judge_can_edit(saved):
        abort(403)
    was_reopened = eval_reopened_for_judge(saved) if saved is not None else False
    if saved is None:
        saved = EvaluationListSavedResult(
            evaluation_item_id=item.id,
            exercise_id=current_exercise.id,
            exercise_phase=_normalized_exercise_phase(getattr(item, "exercise_phase", None)),
            unit_level_key=item.unit_level_key or "",
            saved_by_id=getattr(user, "id", None),
            is_approved=False,
        )
        db.add(saved)
    saved.payload_json = raw
    saved.total_pct = total_pct
    saved.grade_label = grade
    saved.saved_by_id = getattr(user, "id", None)
    if was_reopened:
        apply_judge_save_after_reopen(saved)
    db.flush()
    _evaluation_delete_duplicate_saves(db, exercise_id=current_exercise.id, evaluation_item_id=item.id, keep_id=saved.id)
    db.commit()


def _planner_bundle_eval_canonical_saved(
    db, exercise_id: int, bundle_action_eval_id: int
) -> PlannerFlowBundleEvalSavedResult | None:
    return (
        db.query(PlannerFlowBundleEvalSavedResult)
        .filter(
            PlannerFlowBundleEvalSavedResult.exercise_id == exercise_id,
            PlannerFlowBundleEvalSavedResult.bundle_action_eval_id
            == int(bundle_action_eval_id),
        )
        .order_by(
            PlannerFlowBundleEvalSavedResult.updated_at.desc(),
            PlannerFlowBundleEvalSavedResult.id.desc(),
        )
        .first()
    )


def _planner_bundle_eval_delete_duplicate_saves(
    db, *, exercise_id: int, bundle_action_eval_id: int, keep_id: int
) -> None:
    db.query(PlannerFlowBundleEvalSavedResult).filter(
        PlannerFlowBundleEvalSavedResult.exercise_id == exercise_id,
        PlannerFlowBundleEvalSavedResult.bundle_action_eval_id
        == int(bundle_action_eval_id),
        PlannerFlowBundleEvalSavedResult.id != keep_id,
    ).delete(synchronize_session=False)


def _planner_bundle_eval_commit_payload_save(
    db,
    *,
    user: User,
    action_row: ExercisePlannerFlowBundleActionEval,
    bundle: ExercisePlannerFlowBundle,
    current_exercise: Exercise,
    raw: str,
) -> None:
    if not can_save_evaluation_results(user):
        abort(403)
    try:
        payload = json.loads(raw)
    except Exception:
        abort(400)
    if not isinstance(payload, dict):
        abort(400)
    rows = payload.get("rows") or []
    if not isinstance(rows, list):
        abort(400)
    total_pct, grade = _evaluation_grade_from_payload_rows(rows)
    saved = _planner_bundle_eval_canonical_saved(
        db, current_exercise.id, action_row.id
    )
    if saved is not None and not eval_judge_can_edit(saved):
        abort(403)
    was_reopened = eval_reopened_for_judge(saved) if saved is not None else False
    if saved is None:
        saved = PlannerFlowBundleEvalSavedResult(
            bundle_action_eval_id=action_row.id,
            exercise_id=current_exercise.id,
            exercise_phase=_normalized_exercise_phase(bundle.exercise_phase),
            unit_level_key=bundle.unit_level_key or "",
            saved_by_id=getattr(user, "id", None),
            is_approved=False,
        )
        db.add(saved)
    saved.payload_json = raw
    saved.total_pct = total_pct
    saved.grade_label = grade
    saved.saved_by_id = getattr(user, "id", None)
    saved.unit_level_key = bundle.unit_level_key or ""
    saved.exercise_phase = _normalized_exercise_phase(bundle.exercise_phase)
    if was_reopened:
        apply_judge_save_after_reopen(saved)
    db.flush()
    _planner_bundle_eval_delete_duplicate_saves(
        db,
        exercise_id=current_exercise.id,
        bundle_action_eval_id=action_row.id,
        keep_id=saved.id,
    )
    db.commit()


def _judge_planner_flow_action_bundle_row(
    db, user: User, ex: Exercise | None, slot: int
) -> tuple[ExercisePlannerFlowBundle, ExercisePlannerFlowBundleActionEval] | None:
    """حزمة المحكم وفتحة الإجراء إن وُجد الملف ومطابقة التخصيص."""
    if ex is None:
        return None
    oversee_jid = (
        _planner_flow_oversee_judge_id_from_request()
        if can_oversee_judge_planner_flow_materials(user)
        else None
    )
    bundle = _judge_assigned_planner_bundle(
        db, user, ex, judge_user_id=oversee_jid
    )
    if bundle is None:
        return None
    if not can_oversee_judge_planner_flow_materials(user):
        _enforce_judge_unit_scope(db, user, ex, bundle.unit_level_key)
        _enforce_judge_has_assignment_for_unit(db, user, ex, bundle.unit_level_key)
    row = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(
            ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id,
            ExercisePlannerFlowBundleActionEval.slot_index == int(slot),
        )
        .first()
    )
    if row is None or not (row.file_relpath or "").strip():
        return None
    return bundle, row


def _phase_label_ar(phase: str | None) -> str:
    return exercise_phase_label(phase)


def _parse_saved_eval_rows(payload_json: str | None) -> list[dict]:
    if not (payload_json or "").strip():
        return []
    try:
        data = json.loads(payload_json)
    except Exception:
        return []
    if not isinstance(data, dict):
        return []
    rows = data.get("rows") or []
    return rows if isinstance(rows, list) else []


def _parse_saved_eval_max_positive(raw) -> float | None:
    if raw in (None, ""):
        return None
    try:
        v = float(str(raw).replace(",", "."))
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _evaluation_payload_has_empty_acquired_for_approve(rows: list) -> bool:
    """
    True إذا وُجد بند تقييم (صف score) بمكتسبة فارغة — لا يُقبل الاعتماد.
    «لا ينطبق» يُعتبر إدخالاً صالحاً. صفوف الأقسام تُستثنى.
    للحمولات القديمة دون row_kind: يُفحص الصف إذا وُجدت له قصوى رقمية ولم تُدخل مكتسبة.
    """
    if not isinstance(rows, list):
        return True
    safe = [r for r in rows if isinstance(r, dict)]
    if not safe:
        return True
    has_row_kind = any(str(r.get("row_kind") or "").strip() for r in safe)
    for r in safe:
        rk = str(r.get("row_kind") or "").strip().lower()
        if rk == "section":
            continue
        acq = r.get("acquired")
        empty_acq = acq is None or str(acq).strip() == ""
        if not empty_acq:
            continue
        if rk == "score":
            return True
        if not has_row_kind and _parse_saved_eval_max_positive(r.get("max_val")) is not None:
            return True
    return False


def _evaluation_saved_allows_judge_approve(saved) -> bool:
    if saved is None:
        return False
    rows = _parse_saved_eval_rows(getattr(saved, "payload_json", None))
    return grade_allows_judge_approve(
        getattr(saved, "grade_label", None),
        total_pct=getattr(saved, "total_pct", None),
        payload_rows=rows,
    )


def _eval_list_viewer_ctx(user: User, saved) -> dict:
    """سياق مشترك لعرض/تعديل قائمة تقييم (محكم أو كبير محكمين)."""
    grade_blocks_approve = bool(
        saved is not None
        and can_approve_evaluation_results(user)
        and eval_judge_can_approve(saved)
        and not _evaluation_saved_allows_judge_approve(saved)
    )
    return {
        "saved_is_approved": eval_judge_approved(saved),
        "saved_approved_at": getattr(saved, "approved_at", None) if saved else None,
        "saved_is_chief_approved": eval_chief_approved(saved),
        "saved_chief_approved_at": getattr(saved, "chief_approved_at", None) if saved else None,
        "saved_reopened_for_judge": eval_reopened_for_judge(saved),
        "eval_workflow_label": eval_workflow_label_ar(saved),
        "eval_can_edit": bool(can_save_evaluation_results(user) and eval_judge_can_edit(saved)),
        "show_eval_approve": bool(
            can_approve_evaluation_results(user)
            and eval_judge_can_approve(saved)
            and _evaluation_saved_allows_judge_approve(saved)
        ),
        "show_eval_approve_form": bool(
            can_approve_evaluation_results(user)
            and eval_judge_can_approve(saved)
            and saved is not None
        ),
        "eval_approve_grade_blocked": grade_blocks_approve,
        "show_chief_approve": bool(
            can_chief_approve_evaluation_results(user) and eval_chief_can_approve(saved)
        ),
        "show_chief_reopen": bool(
            can_chief_reopen_evaluation_for_judge(user) and eval_chief_can_reopen(saved)
        ),
    }


def _eval_crit_media_sheet_ctx(
    db,
    user: User | None,
    *,
    exercise: Exercise | None,
    list_item_id: int | None,
    bundle_action_eval_id: int | None,
    eval_can_edit: bool,
) -> dict:
    ctx = {
        "eval_crit_media_by_row": {},
        "eval_crit_list_item_id": None,
        "eval_crit_bundle_action_id": None,
        "eval_crit_upload_url": "",
    }
    if exercise is None or (list_item_id is None and bundle_action_eval_id is None):
        return ctx
    ex_id = int(exercise.id)
    ctx["eval_crit_media_by_row"] = group_media_rows(
        db,
        ex_id,
        list_item_id=list_item_id,
        bundle_action_eval_id=bundle_action_eval_id,
    )
    ctx["eval_crit_list_item_id"] = int(list_item_id) if list_item_id is not None else None
    ctx["eval_crit_bundle_action_id"] = (
        int(bundle_action_eval_id) if bundle_action_eval_id is not None else None
    )
    if eval_can_edit and user and can_save_evaluation_results(user):
        ctx["eval_crit_upload_url"] = url_for("views.eval_criterion_media_upload")
    return ctx


def _eval_crit_user_can_stream_media(db, user: User | None, m: EvaluationCriterionMedia) -> bool:
    """إطلاع/تشغيل: إدارة النظام، السيطرة، التخطيط، كبير المحكمين، المحللين، والمحكم ضمن النطاق."""
    if user is None:
        return False
    if not db.get(Exercise, m.exercise_id):
        return False
    if is_system_admin(user):
        return True
    ex_id = int(m.exercise_id)
    if can_access_analyst_hub(user):
        ws = _admin_current_workspace_exercise(db, user)
        return ws is not None and int(ws.id) == ex_id
    if can_access_planner_hub(user):
        ws = _admin_current_workspace_exercise(db, user)
        return ws is not None and int(ws.id) == ex_id
    if can_access_control_hub(user):
        ws = _current_workspace_exercise(db, user)
        return ws is not None and int(ws.id) == ex_id
    if can_access_chief_judge_hub(user):
        ws = _current_workspace_exercise(db, user)
        return ws is not None and int(ws.id) == ex_id
    if not can_access_judge_hub(user):
        return False
    ws = _current_workspace_exercise(db, user)
    if ws is None or int(ws.id) != ex_id:
        return False
    if m.evaluation_list_item_id is not None:
        pdf = db.get(EvaluationListPdfItem, int(m.evaluation_list_item_id))
        if pdf is None or int(pdf.exercise_id or 0) != ex_id:
            return False
        _enforce_judge_unit_scope(db, user, ws, pdf.unit_level_key or "")
        return True
    if m.bundle_action_eval_id is not None:
        ar = db.get(ExercisePlannerFlowBundleActionEval, int(m.bundle_action_eval_id))
        if ar is None:
            return False
        b = db.get(ExercisePlannerFlowBundle, int(ar.bundle_id))
        if b is None or int(b.exercise_id) != ex_id:
            return False
        _enforce_judge_unit_scope(db, user, ws, b.unit_level_key or "")
        return True
    return False


def _eval_crit_user_can_upload_media(
    db,
    user: User | None,
    *,
    exercise_id: int,
    unit_level_key: str,
    list_item_id: int | None,
    bundle_action_eval_id: int | None,
    canonical_saved: EvaluationListSavedResult | PlannerFlowBundleEvalSavedResult | None,
) -> bool:
    if user is None or not can_save_evaluation_results(user):
        return False
    if not eval_judge_can_edit(canonical_saved):
        return False
    ex = db.get(Exercise, int(exercise_id))
    if ex is None:
        return False
    if is_system_admin(user):
        return True
    if can_access_planner_hub(user):
        ws = _admin_current_workspace_exercise(db, user)
        return ws is not None and int(ws.id) == int(exercise_id)
    if not can_access_judge_hub(user):
        return False
    ws = _current_workspace_exercise(db, user)
    if ws is None or int(ws.id) != int(exercise_id):
        return False
    _enforce_judge_unit_scope(db, user, ex, unit_level_key or "")
    return True


def _eval_row_canonical_saved_for_crit_upload(
    db, *, exercise_id: int, list_item_id: int | None, bundle_action_eval_id: int | None
) -> EvaluationListSavedResult | PlannerFlowBundleEvalSavedResult | None:
    if list_item_id is not None:
        return _evaluation_canonical_saved_row(db, int(exercise_id), int(list_item_id))
    if bundle_action_eval_id is not None:
        return _planner_bundle_eval_canonical_saved(
            db, int(exercise_id), int(bundle_action_eval_id)
        )
    return None


def _eval_row_effective_max(row: dict) -> float:
    """قصوى الصف لحساب النسبة الإجمالية؛ يطابق effectiveMaxForRow في الواجهة (افتراض 5)."""
    mx_raw = row.get("max_val")
    if mx_raw not in (None, ""):
        try:
            mx = float(str(mx_raw).replace(",", "."))
            if mx > 0:
                return mx
        except (TypeError, ValueError):
            pass
    return 5.0


def _eval_row_score_pct(row: dict) -> float | None:
    """تقريب 0..100 من المكتسبة والقصوى (أو منطق 5 درجات)."""
    if not isinstance(row, dict):
        return None
    acq = row.get("acquired")
    if acq is None or acq == "" or str(acq).strip().lower() == "na":
        return None
    try:
        a = float(str(acq).replace(",", "."))
    except (TypeError, ValueError):
        return None
    mx_raw = row.get("max_val")
    mx = None
    if mx_raw not in (None, ""):
        try:
            mx = float(str(mx_raw).replace(",", "."))
        except (TypeError, ValueError):
            mx = None
    if mx is not None and mx > 0:
        return max(0.0, min(100.0, (a / mx) * 100.0))
    return max(0.0, min(100.0, (a / 5.0) * 100.0))


def _pct_status_band(pct: float | None) -> str:
    """حالة عرض: high | mid | low | na"""
    if pct is None:
        return "na"
    if pct >= 75.0:
        return "high"
    if pct >= 50.0:
        return "mid"
    return "low"


def _trainee_commander_for_unit(db, exercise_id: int, unit_key: str) -> str:
    row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == exercise_id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if row is None:
        return ""
    return (row.full_name or "").strip()


def _build_analyst_evaluation_results_dashboard(
    db,
    user: User,
    *,
    approved_only: bool = True,
    matrix_mode: str = "objectives",
) -> dict:
    """بيانات صفحة محللين / مصفوفة حالة الأهداف مقابل الوحدات + جانبية."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {"has_exercise": False}
    ex = (
        db.query(Exercise)
        .options(joinedload(Exercise.objectives))
        .filter(Exercise.id == ex0.id)
        .first()
    )
    if ex is None:
        return {"has_exercise": False}

    objectives = sorted(
        [o for o in (ex.objectives or []) if o is not None],
        key=lambda o: (int(getattr(o, "sort_order", 0) or 0), int(getattr(o, "id", 0) or 0)),
    )

    eval_items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id)
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    saved_by_item: dict[int, EvaluationListSavedResult] = {}
    if eval_items:
        item_ids = [int(it.id) for it in eval_items if getattr(it, "id", None) is not None]
        saved_query = db.query(EvaluationListSavedResult).filter(
            EvaluationListSavedResult.exercise_id == ex.id,
            EvaluationListSavedResult.evaluation_item_id.in_(item_ids),
        )
        if approved_only:
            saved_query = saved_query.filter(EvaluationListSavedResult.is_approved == True)
        saved_rows = (
            saved_query
            .order_by(
                EvaluationListSavedResult.evaluation_item_id,
                EvaluationListSavedResult.approved_at.desc(),
                EvaluationListSavedResult.updated_at.desc(),
                EvaluationListSavedResult.id.desc(),
            )
            .all()
        )
        for sr in saved_rows:
            iid = int(sr.evaluation_item_id)
            if iid not in saved_by_item:
                saved_by_item[iid] = sr
    by_unit: dict[str, list[EvaluationListPdfItem]] = {}
    for it in eval_items:
        uk = (it.unit_level_key or "").strip()
        if not uk:
            continue
        by_unit.setdefault(uk, []).append(it)

    # رؤوس الأعمدة: صفحة النتائج تستخدم قوائم التقييم، أما التحليل الجانبي فيبقى على الأهداف/البنود.
    if matrix_mode == "evaluation_lists":
        matrix_columns = []
        for i, it in enumerate([x for x in eval_items if int(x.id) in saved_by_item]):
            title = (it.text or "قائمة تقييم").strip()
            unit_label = label_for_unit_level_key(it.unit_level_key or "") or (it.unit_level_key or "")
            matrix_columns.append(
                {
                    "idx": i,
                    "code": f"قائمة {i + 1:02d}",
                    "text": title[:220] + ("…" if len(title) > 220 else ""),
                    "full_text": f"{title} — {unit_label} — {_phase_label_ar(getattr(it, 'exercise_phase', None))}",
                    "item_id": int(it.id),
                    "unit_key": (it.unit_level_key or "").strip(),
                }
            )
        n_cols = len(matrix_columns)
    elif objectives:
        matrix_columns: list[dict] = []
        for i, o in enumerate(objectives):
            txt = (o.text or "").strip()
            matrix_columns.append(
                {
                    "idx": i,
                    "code": f"هدف {i + 1:02d}",
                    "text": txt[:220] + ("…" if len(txt) > 220 else ""),
                    "full_text": txt,
                }
            )
        n_cols = len(matrix_columns)
    else:
        max_rows = 0
        for it in eval_items:
            canon = saved_by_item.get(int(it.id))
            if canon is None:
                continue
            rows = _parse_saved_eval_rows(canon.payload_json)
            max_rows = max(max_rows, len(rows))
        n_cols = max(1, min(max_rows, 14))
        matrix_columns = [
            {
                "idx": i,
                "code": f"بند {i + 1:02d}",
                "text": f"عنصر التقييم رقم {i + 1} (بدون أهداف مُعرَّفة في التمرين)",
                "full_text": "",
            }
            for i in range(n_cols)
        ]

    matrix_rows: list[dict] = []
    for ul_pos, ul in enumerate(UNIT_LEVELS):
        uk = ul.get("key") or ""
        ulab = ul.get("label") or uk
        items_u = by_unit.get(uk) or []
        sub = _trainee_commander_for_unit(db, ex.id, uk)
        cells: list[dict] = []
        for j in range(n_cols):
            if matrix_mode == "evaluation_lists":
                col = matrix_columns[j] if j < len(matrix_columns) else {}
                canon = None
                if (col.get("unit_key") or "") == uk:
                    canon = saved_by_item.get(int(col.get("item_id") or 0))
                best_pct = getattr(canon, "total_pct", None) if canon is not None else None
                if best_pct is None and canon is not None:
                    row_scores = [
                        _eval_row_score_pct(r)
                        for r in _parse_saved_eval_rows(canon.payload_json)
                        if isinstance(r, dict)
                    ]
                    row_scores = [x for x in row_scores if x is not None]
                    best_pct = (sum(row_scores) / len(row_scores)) if row_scores else None
                best_updated = (
                    (getattr(canon, "approved_at", None) if bool(getattr(canon, "is_approved", False)) else None)
                    or getattr(canon, "updated_at", None)
                    if canon is not None
                    else None
                )
                best_note = ""
                is_flagged = bool(canon is not None and not bool(getattr(canon, "is_approved", False)))
            else:
                best_pct: float | None = None
                best_updated = None
                best_note = ""
                is_flagged = False
                for it in items_u:
                    canon = saved_by_item.get(int(it.id))
                    if canon is None:
                        continue
                    rows = _parse_saved_eval_rows(canon.payload_json)
                    if not rows:
                        continue
                    ridx = min(j, len(rows) - 1)
                    row = rows[ridx] if isinstance(rows[ridx], dict) else {}
                    pc = _eval_row_score_pct(row)
                    if pc is None:
                        continue
                    if best_pct is None or pc > best_pct:
                        best_pct = pc
                        best_updated = getattr(canon, "approved_at", None) or getattr(canon, "updated_at", None)
                        best_note = (row.get("notes") or "").strip()
            band = _pct_status_band(best_pct)
            h_fill = int(round(best_pct or 0)) if best_pct is not None else 0
            h_rest = max(0, 100 - h_fill) if best_pct is not None else 100
            bar_h = 44
            if best_pct is not None:
                fill_px = max(4, int(round(bar_h * h_fill / 100.0)))
                rest_px = max(4, bar_h - fill_px)
            else:
                fill_px = 4
                rest_px = bar_h - fill_px
            cells.append(
                {
                    "pct": best_pct,
                    "band": band,
                    "is_flagged": is_flagged,
                    "updated_at": best_updated,
                    "note": best_note,
                    "bar_fill": h_fill,
                    "bar_rest": h_rest,
                    "bar_fill_px": fill_px,
                    "bar_rest_px": rest_px,
                }
            )
        matrix_rows.append(
            {
                "unit_key": uk,
                "unit_label": ulab,
                "hierarchy_idx": ul_pos + 1,
                "row_sub": sub or "—",
                "has_lists": bool(items_u),
                "cells": cells,
            }
        )

    # قوائم لم تُعبأ بعد
    not_assessed: list[dict] = []
    for it in eval_items:
        uk = (it.unit_level_key or "").strip()
        canon = saved_by_item.get(int(it.id))
        rows = _parse_saved_eval_rows(getattr(canon, "payload_json", None) if canon else None)
        scored = any(_eval_row_score_pct(r) is not None for r in rows if isinstance(r, dict))
        if canon is None or not rows or not scored:
            not_assessed.append(
                {
                    "objective": ((it.text or "قائمة تقييم").strip()[:100]),
                    "entities": label_for_unit_level_key(uk) or uk or "—",
                    "activity": f"{_phase_label_ar(getattr(it, 'exercise_phase', None))} — بانتظار {'نتيجة معتمدة' if approved_only else 'حفظ نتيجة'}",
                    "open_href": url_for(
                        "views.analyst_evaluation_list_file_viewer",
                        unit_key=uk,
                        item_id=it.id,
                    ),
                }
            )

    # مناطق تحتاج تطويراً (أدنى النِسَب في المصفوفة)
    dev_list: list[tuple[float, dict]] = []
    sustain_list: list[tuple[float, dict]] = []
    for mr in matrix_rows:
        for j, cell in enumerate(mr.get("cells") or []):
            pct = cell.get("pct")
            if pct is None:
                continue
            col = matrix_columns[j] if j < len(matrix_columns) else {}
            row_info = {
                "objective": (col.get("text") or col.get("code") or "—")[:160],
                "entities": mr.get("unit_label") or "—",
                "activity": "نتيجة معتمدة من قوائم التقييم",
                "grade": grade_label_from_percent(pct),
                "pct": pct,
                "note": (cell.get("note") or "").strip(),
            }
            if cell.get("band") == "low":
                dev_list.append((float(pct), row_info))
            elif cell.get("band") == "high":
                sustain_list.append((float(pct), row_info))
    dev_list.sort(key=lambda x: x[0])
    sustain_list.sort(key=lambda x: x[0], reverse=True)
    development_areas = [d for _, d in dev_list[:12]]
    sustainability_areas = [d for _, d in sustain_list[:12]]
    development_notes = [d for d in development_areas if (d.get("note") or "").strip()]
    sustainability_notes = [d for d in sustainability_areas if (d.get("note") or "").strip()]

    # تحليل حسب مستوى الوحدة من النتائج المعتمدة فقط
    unit_eval_analysis: list[dict] = []
    for ul_pos, ul in enumerate(UNIT_LEVELS):
        uk = ul.get("key") or ""
        scores: list[dict] = []
        for it in by_unit.get(uk) or []:
            canon = saved_by_item.get(int(it.id))
            if canon is None:
                continue
            for idx, row in enumerate(_parse_saved_eval_rows(canon.payload_json)):
                if not isinstance(row, dict):
                    continue
                pct = _eval_row_score_pct(row)
                if pct is None:
                    continue
                element = (row.get("element") or "").strip()
                label = element or (it.text or f"بند {idx + 1}").strip()
                scores.append(
                    {
                        "pct": float(pct),
                        "label": label[:180],
                        "note": (row.get("notes") or "").strip(),
                        "grade": grade_label_from_percent(pct),
                        "evaluation_list": (it.text or "قائمة تقييم").strip()[:120],
                    }
                )
        scores.sort(key=lambda x: x["pct"])
        avg_pct = (sum(s["pct"] for s in scores) / len(scores)) if scores else None
        unit_eval_analysis.append(
            {
                "unit_key": uk,
                "unit_label": ul.get("label") or uk,
                "hierarchy_idx": ul_pos + 1,
                "approved_items_count": len([it for it in by_unit.get(uk) or [] if int(it.id) in saved_by_item]),
                "scored_rows_count": len(scores),
                "avg_pct": avg_pct,
                "avg_grade": grade_label_from_percent(avg_pct) if avg_pct is not None else "—",
                "development_points": [s for s in scores if s["pct"] < 50.0][:5],
                "sustainability_points": sorted(
                    [s for s in scores if s["pct"] >= 75.0],
                    key=lambda x: x["pct"],
                    reverse=True,
                )[:5],
            }
        )

    since_dt = getattr(ex, "planned_start", None) or getattr(ex, "created_at", None)

    return {
        "has_exercise": True,
        "exercise": ex,
        "matrix_columns": matrix_columns,
        "matrix_rows": matrix_rows,
        "not_assessed": not_assessed,
        "development_areas": development_areas,
        "sustainability_areas": sustainability_areas,
        "development_notes": development_notes,
        "sustainability_notes": sustainability_notes,
        "unit_eval_analysis": unit_eval_analysis,
        "since_dt": since_dt,
        "n_objectives": n_cols,
        "n_eval_lists": len(eval_items),
        "n_approved_eval_lists": len([r for r in saved_by_item.values() if bool(getattr(r, "is_approved", False))]),
        "n_saved_eval_lists": len(saved_by_item),
        "using_objectives": bool(objectives),
        "using_evaluation_lists": matrix_mode == "evaluation_lists",
    }


def _saved_eval_list_has_measurable_result(canon: EvaluationListSavedResult | None) -> bool:
    if canon is None:
        return False
    if _evaluation_saved_total_pct(canon) is not None:
        return True
    rows = _parse_saved_eval_rows(getattr(canon, "payload_json", None))
    return any(_eval_row_score_pct(r) is not None for r in rows if isinstance(r, dict))


def _chart_bar_height_px(value: float, values: list[float], *, max_px: int = 100, min_px: int = 10) -> int:
    if value <= 0 or not values:
        return min_px
    peak = max(values)
    if peak <= 0:
        return min_px
    return max(min_px, int(round(max_px * float(value) / peak)))


_AERC_UNIT_COUNT_COLORS: tuple[str, ...] = (
    "#4a7c59",
    "#6b5a48",
    "#2563eb",
    "#b45309",
    "#7c3aed",
    "#0d9488",
    "#be123c",
    "#ca8a04",
    "#64748b",
    "#c2410c",
)


def _unit_count_donut_style(bars: list[dict], total: int) -> str:
    """تدرج دائري لتوزيع عدد القوائم المحفوظة حسب الوحدة."""
    if total <= 0 or not bars:
        return "conic-gradient(#e8e0d8 0deg 360deg)"
    stops: list[str] = []
    acc = 0.0
    for i, bar in enumerate(bars):
        cnt = int(bar.get("count") or 0)
        share = (cnt / total) * 100.0
        bar["share_pct"] = round(share, 1)
        color = _AERC_UNIT_COUNT_COLORS[i % len(_AERC_UNIT_COUNT_COLORS)]
        bar["segment_color"] = color
        nxt = acc + share
        stops.append(f"{color} {acc:.2f}% {nxt:.2f}%")
        acc = nxt
    return f"conic-gradient(from 0.25turn, {', '.join(stops)})"


def _build_analyst_saved_results_charts(db, user: User) -> dict:
    """صفحة واحدة للمحللين: رسوم لقوائم التقييم المحفوظة فقط + نسب الوحدات والمراحل."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    eval_items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id)
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    saved_by_item = _evaluation_canonical_map_for_items(
        db,
        int(ex.id),
        [int(it.id) for it in eval_items if getattr(it, "id", None) is not None],
    )

    saved_entries: list[dict] = []
    for it in eval_items:
        iid = int(it.id)
        canon = saved_by_item.get(iid)
        if not _saved_eval_list_has_measurable_result(canon):
            continue
        pct = _evaluation_saved_total_pct(canon)
        if pct is None:
            continue
        uk = (it.unit_level_key or "").strip()
        phase_key = _normalized_exercise_phase(getattr(it, "exercise_phase", None))
        saved_entries.append(
            {
                "item_id": iid,
                "pct": float(pct),
                "unit_key": uk,
                "unit_label": label_for_unit_level_key(uk) or uk or "—",
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
            }
        )

    n_saved = len(saved_entries)
    all_pcts = [e["pct"] for e in saved_entries]
    overall_avg = (sum(all_pcts) / len(all_pcts)) if all_pcts else None

    unit_counts: dict[str, int] = {}
    unit_pcts: dict[str, list[float]] = {}
    for e in saved_entries:
        uk = e["unit_key"]
        unit_counts[uk] = unit_counts.get(uk, 0) + 1
        unit_pcts.setdefault(uk, []).append(e["pct"])

    unit_count_bars: list[dict] = []
    for ul in UNIT_LEVELS:
        uk = ul.get("key") or ""
        cnt = unit_counts.get(uk, 0)
        if cnt <= 0:
            continue
        unit_count_bars.append(
            {
                "unit_key": uk,
                "label": ul.get("label") or uk,
                "count": cnt,
            }
        )
    unit_count_donut_css = _unit_count_donut_style(unit_count_bars, n_saved)

    pct_values = [
        (sum(unit_pcts.get(ul.get("key") or "", [])) / len(unit_pcts[ul.get("key") or ""]))
        for ul in UNIT_LEVELS
        if unit_pcts.get(ul.get("key") or "")
    ]
    unit_pct_bars: list[dict] = []
    for ul in UNIT_LEVELS:
        uk = ul.get("key") or ""
        pcts_u = unit_pcts.get(uk) or []
        if not pcts_u:
            continue
        avg_u = sum(pcts_u) / len(pcts_u)
        unit_pct_bars.append(
            {
                "unit_key": uk,
                "label": ul.get("label") or uk,
                "avg_pct": _round_pct_display(avg_u),
                "saved_count": len(pcts_u),
                "grade": grade_label_from_percent(avg_u),
                "band": _pct_status_band(avg_u),
                "bar_w_pct": max(8, int(round(avg_u))),
                "color": _control_report_dot_color(avg_u),
            }
        )

    phase_summary = _phase_summary_from_eval_items(
        [it for it in eval_items if int(it.id) in saved_by_item and _saved_eval_list_has_measurable_result(saved_by_item.get(int(it.id)))],
        saved_by_item,
    )
    phase_bars = _distribution_from_phase_summary(phase_summary)
    for pb in phase_bars:
        pct_f = float(pb.get("pct") or 0)
        pb["bar_h_pct"] = max(12, int(round(pct_f)))
        pb["grade"] = grade_label_from_percent(pct_f)
        pb["band"] = _pct_status_band(pct_f)

    return {
        "has_exercise": True,
        "exercise": ex,
        "n_saved_lists": n_saved,
        "overall_avg_pct": _round_pct_display(overall_avg) if overall_avg is not None else None,
        "overall_grade": grade_label_from_percent(overall_avg) if overall_avg is not None else "—",
        "overall_band": _pct_status_band(overall_avg),
        "unit_count_bars": unit_count_bars,
        "unit_count_donut_css": unit_count_donut_css,
        "unit_pct_bars": unit_pct_bars,
        "phase_bars": phase_bars,
        "phase_exercise_pct": phase_summary.get("exercise_pct"),
        "phase_exercise_grade": phase_summary.get("exercise_grade"),
        "grade_legend": _control_report_grade_legend(),
    }


def _build_analyst_evaluation_criteria_distribution(db, user: User) -> dict:
    """توزيع نتائج مراحل التقييم اليدوية، مستقل عن قوائم التقييم."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    criteria_units = _sync_analyst_criteria_units_from_planner(db, ex)
    phase_items = (
        db.query(AnalystEvaluationCriteriaPhaseItem)
        .filter(AnalystEvaluationCriteriaPhaseItem.exercise_id == ex.id)
        .all()
    )
    marks_by_unit_phase: dict[tuple[int, str], list[float]] = {}
    for item in phase_items:
        if item.allocated_mark is None:
            continue
        marks_by_unit_phase.setdefault(
            (int(item.criteria_unit_id), item.phase_key or ""),
            [],
        ).append(float(item.allocated_mark))

    rows: list[dict] = []
    grand_total = 0.0
    for unit in criteria_units:
        phase_totals: dict[str, float | None] = {}
        phase_order = ANALYST_EVALUATION_CRITERIA_PHASE_ORDER or tuple(EXERCISE_PHASE_OPTIONS)
        for phase_key, _label in phase_order:
            marks = marks_by_unit_phase.get((unit.id, phase_key), [])
            phase_totals[phase_key] = sum(marks) if marks else None
        parts = [x for x in phase_totals.values() if x is not None]
        total_mark = sum(parts) if parts else None
        if total_mark is not None:
            grand_total += total_mark
        unit_level_key = _resolve_unit_level_key_for_criteria_label(unit.label or "")
        unit_label = label_for_unit_level_key(unit_level_key) or (unit.label or "—")
        rows.append(
            {
                "unit_id": unit.id,
                "unit_level_key": unit_level_key,
                "unit_label": unit_label,
                "phase_totals": phase_totals,
                "preparation_total": phase_totals.get("preparation"),
                "evaluation_tracks_total": phase_totals.get("evaluation_tracks"),
                "opening_total": phase_totals.get("opening"),
                "operations_total": phase_totals.get("main"),
                "total_mark": total_mark,
                "allocated_pct": None,
            }
        )

    if grand_total > 0:
        for row in rows:
            if row["total_mark"] is not None:
                row["allocated_pct"] = (float(row["total_mark"]) / grand_total) * 100.0

    present_keys = {(r.get("unit_level_key") or "").strip() for r in rows}
    available_unit_levels = [
        u for u in UNIT_LEVELS if (u.get("key") or "").strip() not in present_keys
    ]

    return {
        "has_exercise": True,
        "exercise": ex,
        "distribution_rows": rows,
        "grand_total": grand_total if grand_total > 0 else None,
        "criteria_phases": _analyst_criteria_phases_for_display(bool(rows)),
        "available_unit_levels": available_unit_levels,
        "planner_unit_levels": list(UNIT_LEVELS),
    }


def _build_analyst_final_evaluation_report(db, user: User) -> dict:
    """ملخص نهائي من مساحة المحكمين/قوائم التقييم، سواء كانت النتيجة محفوظة أو معتمدة."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None and is_system_admin(user):
        ex0 = db.query(Exercise).order_by(Exercise.id.desc()).first()
    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    eval_items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id)
        .order_by(
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    latest_by_item = _evaluation_canonical_map_for_items(
        db,
        int(ex.id),
        [int(it.id) for it in eval_items],
    )
    saved_source_rows = [
        row
        for row in latest_by_item.values()
        if (getattr(row, "payload_json", "") or "").strip()
    ]
    approved_source_count = len([row for row in saved_source_rows if bool(getattr(row, "is_approved", False))])
    saved_pending_source_count = len(saved_source_rows) - approved_source_count
    manual_max_by_item = _final_eval_manual_max_map(db, int(ex.id))
    phase_manual_max_map = _final_eval_phase_manual_max_map(db, int(ex.id))
    # «القصوى» الافتراضية لكل (وحدة، مرحلة) مأخوذة من جدول معايير التقييم
    # (مساحة المحللين / معايير التقييم) — تظهر في التقرير النهائي عند عدم
    # إدخال قيمة يدوية في AnalystFinalEvaluationPhaseAllocatedMax.
    criteria_phase_max_map = _analyst_criteria_phase_max_map(db, int(ex.id))
    included_unit_keys = planning_included_unit_keys()

    phase_acquired_totals: dict[tuple[str, str], dict] = {}
    unit_phase_slots: dict[tuple[str, str], dict] = {}
    list_rows_by_unit_phase: dict[tuple[str, str], list[dict]] = {}
    template_rows_cache: dict[str, list[dict]] = {}
    for item in eval_items:
        unit_key = (item.unit_level_key or "").strip()
        if unit_key and unit_key not in included_unit_keys:
            continue
        phase_key = _normalized_exercise_phase(getattr(item, "exercise_phase", None))
        if unit_key and phase_key:
            unit_phase_slots.setdefault(
                (unit_key, phase_key),
                {
                    "unit_key": unit_key,
                    "unit_label": label_for_unit_level_key(unit_key, db) or unit_key or "—",
                    "phase_key": phase_key,
                    "phase_label": _phase_label_ar(phase_key),
                    "acquired_mark": 0.0,
                },
            )
        list_label = (item.text or "").strip() or "—"
        saved = latest_by_item.get(int(item.id))
        item_id = int(item.id)
        manual_max = manual_max_by_item.get(item_id)
        allocated_max_mark: float | None = float(manual_max) if manual_max is not None else None
        has_allocated_max = manual_max is not None
        allocated_acquired_mark = 0.0
        allocated_pct: float | None = None
        allocated_grade = "—"
        evaluation_list_max_mark = 0.0
        evaluation_list_acquired_mark = 0.0
        evaluation_list_actual_pct: float | None = None
        evaluation_list_grade = "—"
        has_saved_payload = False
        acquired_mark = 0.0
        if saved is not None and (getattr(saved, "payload_json", "") or "").strip():
            has_saved_payload = True
            rows = _parse_saved_eval_rows(saved.payload_json)
            _payload_max, acquired_mark = _evaluation_payload_mark_totals(rows)
            relpath = (item.pdf_relpath or "").strip()
            if relpath not in template_rows_cache:
                template_rows_cache[relpath] = _evaluation_list_template_rows(item)
            footer = _evaluation_list_page_footer_stats(
                rows,
                template_rows_cache.get(relpath) or [],
                saved_total_pct=getattr(saved, "total_pct", None),
                saved_grade=getattr(saved, "grade_label", None),
            )
            evaluation_list_max_mark = float(footer["sum_max"] or 0.0)
            evaluation_list_acquired_mark = float(footer["sum_acquired"] or 0.0)
            evaluation_list_actual_pct = footer.get("actual_pct")
            evaluation_list_grade = footer.get("grade") or "—"
        if has_saved_payload:
            allocated_pct = evaluation_list_actual_pct
            if allocated_pct is None and evaluation_list_max_mark > 0:
                allocated_pct = _allocated_pct_from_marks(
                    evaluation_list_acquired_mark, evaluation_list_max_mark
                )
            allocated_grade = (
                evaluation_list_grade
                if evaluation_list_grade and evaluation_list_grade != "—"
                else (
                    grade_label_from_percent(allocated_pct)
                    if allocated_pct is not None
                    else "—"
                )
            )
        elif has_allocated_max and allocated_max_mark is not None and allocated_max_mark > 0:
            allocated_pct = _allocated_pct_from_marks(acquired_mark, allocated_max_mark)
            allocated_grade = (
                grade_label_from_percent(allocated_pct) if allocated_pct is not None else "—"
            )
        if (
            has_allocated_max
            and allocated_max_mark is not None
            and allocated_pct is not None
        ):
            computed = _allocated_acquired_from_pct_and_max(
                allocated_pct, allocated_max_mark
            )
            allocated_acquired_mark = float(computed) if computed is not None else 0.0
        if has_saved_payload and acquired_mark > 0 and unit_key and phase_key:
            block = phase_acquired_totals.setdefault(
                (unit_key, phase_key),
                {
                    "unit_key": unit_key,
                    "unit_label": label_for_unit_level_key(unit_key, db) or unit_key or "—",
                    "phase_key": phase_key,
                    "phase_label": _phase_label_ar(phase_key),
                    "acquired_mark": 0.0,
                },
            )
            block["acquired_mark"] += acquired_mark
            slot = unit_phase_slots.setdefault((unit_key, phase_key), {**block})
            slot["acquired_mark"] = block["acquired_mark"]
        list_rows_by_unit_phase.setdefault((unit_key, phase_key), []).append(
            {
                "unit_key": unit_key,
                "unit_label": label_for_unit_level_key(unit_key, db) or unit_key or "—",
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "list_label": list_label,
                "item_sort_order": int(item.sort_order or 0),
                "item_id": item_id,
                "allocated_max_mark": allocated_max_mark,
                "allocated_acquired_mark": allocated_acquired_mark,
                "allocated_pct": allocated_pct,
                "allocated_grade": allocated_grade,
                "evaluation_list_max_mark": evaluation_list_max_mark,
                "evaluation_list_acquired_mark": evaluation_list_acquired_mark,
                "evaluation_list_actual_pct": evaluation_list_actual_pct,
                "evaluation_list_grade": evaluation_list_grade,
                "has_allocated_max": has_allocated_max,
                "has_saved_payload": has_saved_payload,
            }
        )

    units_in_exercise: set[str] = set()
    for item in eval_items:
        uk = (item.unit_level_key or "").strip()
        if uk and uk in included_unit_keys:
            units_in_exercise.add(uk)
    for unit_key in units_in_exercise:
        unit_label = label_for_unit_level_key(unit_key, db) or unit_key or "—"
        for phase_key in exercise_phase_keys():
            unit_phase_slots.setdefault(
                (unit_key, phase_key),
                {
                    "unit_key": unit_key,
                    "unit_label": unit_label,
                    "phase_key": phase_key,
                    "phase_label": _phase_label_ar(phase_key),
                    "acquired_mark": 0.0,
                },
            )

    unit_order = {row["key"]: idx for idx, row in enumerate(UNIT_LEVELS)}
    phase_order = {key: idx for idx, key in enumerate(exercise_phase_keys())}
    unit_phase_pcts: dict[str, list[float]] = {}

    final_rows: list[dict] = []
    for slot_key in sorted(
        (k for k in unit_phase_slots.keys() if k[0] in included_unit_keys),
        key=lambda k: (
            unit_order.get(k[0], len(unit_order)),
            phase_order.get(k[1], len(phase_order)),
            unit_phase_slots[k].get("unit_label") or k[0],
        ),
    ):
        block = unit_phase_slots[slot_key]
        unit_key = block["unit_key"]
        phase_key = block["phase_key"]
        acquired_mark = float(
            phase_acquired_totals.get(slot_key, {}).get("acquired_mark")
            or block.get("acquired_mark")
            or 0.0
        )
        manual_max = phase_manual_max_map.get(slot_key)
        has_phase_manual_max = manual_max is not None
        if manual_max is not None:
            max_mark = float(manual_max)
        else:
            # عدم وجود إدخال يدوي ⇒ استخدم القيمة من معايير التقييم (إن وُجدت)
            max_mark = float(criteria_phase_max_map.get(slot_key) or 0.0)
        phase_pct: float | None = None
        if max_mark > 0 and acquired_mark > 0:
            phase_pct = (acquired_mark / max_mark) * 100.0
            unit_phase_pcts.setdefault(unit_key, []).append(phase_pct)
        elif max_mark > 0 and acquired_mark <= 0:
            phase_pct = 0.0
            unit_phase_pcts.setdefault(unit_key, []).append(phase_pct)
        final_rows.append(
            {
                **block,
                "acquired_mark": acquired_mark,
                "max_mark": max_mark,
                "manual_max_mark": float(manual_max) if manual_max is not None else None,
                "has_phase_manual_max": has_phase_manual_max,
                "phase_max_field": _report_phase_max_field_name(unit_key, phase_key),
                "phase_pct": phase_pct,
                "phase_grade": grade_label_from_percent(phase_pct) if phase_pct is not None else "—",
                "unit_total_pct": None,
                "unit_grade": "—",
            }
        )

    unit_pcts_computed: dict[str, float] = {}
    for unit_key, pcts in unit_phase_pcts.items():
        if pcts:
            unit_pcts_computed[unit_key] = sum(pcts) / len(pcts)
    for row in final_rows:
        uk = row.get("unit_key") or ""
        if uk in unit_pcts_computed:
            row["unit_total_pct"] = unit_pcts_computed[uk]
            row["unit_grade"] = grade_label_from_percent(unit_pcts_computed[uk])

    detail_rows: list[dict] = []
    for rows in list_rows_by_unit_phase.values():
        rows.sort(key=lambda x: (x.get("item_sort_order", 0), x.get("item_id", 0)))
        detail_rows.extend(rows)
    detail_rows.sort(
        key=lambda x: (
            unit_order.get(x["unit_key"], len(unit_order)),
            phase_order.get(x["phase_key"], len(phase_order)),
            x.get("item_sort_order", 0),
            x.get("item_id", 0),
        ),
    )

    unit_keys: list[str] = []
    for row in UNIT_LEVELS:
        key = (row.get("key") or "").strip()
        if key and key not in unit_keys:
            unit_keys.append(key)

    rows_by_unit: dict[str, list[dict]] = {}
    for row in final_rows:
        rows_by_unit.setdefault(row["unit_key"], []).append(row)
    details_by_unit_phase = list_rows_by_unit_phase

    report_units: list[dict] = []
    final_rows_all: list[dict] = []
    for idx, unit_key in enumerate(unit_keys):
        anchor = f"final-unit-{idx + 1}"
        unit_label = label_for_unit_level_key(unit_key, db) or unit_key or "—"
        unit_phase_rows = _ensure_unit_phase_rows_for_all_phases(
            unit_key, unit_label, rows_by_unit.get(unit_key) or []
        )
        rows_by_unit[unit_key] = unit_phase_rows
        phase_rows = unit_phase_rows
        preparation_detail_rows = _final_eval_detail_rows_for_unit_phase(
            details_by_unit_phase, unit_key, "preparation"
        )
        opening_detail_rows = _final_eval_detail_rows_for_unit_phase(
            details_by_unit_phase, unit_key, "opening"
        )
        main_detail_rows = _final_eval_detail_rows_for_unit_phase(
            details_by_unit_phase, unit_key, "main"
        )
        reorg_detail_rows = _final_eval_detail_rows_for_unit_phase(
            details_by_unit_phase, unit_key, "reorg"
        )
        preparation_criteria_items = _final_eval_criteria_items_for_unit_phase(
            db, ex, unit_key, "preparation"
        )
        opening_criteria_items = _final_eval_criteria_items_for_unit_phase(
            db, ex, unit_key, "opening"
        )
        main_criteria_items = _final_eval_criteria_items_for_unit_phase(
            db, ex, unit_key, "main"
        )
        reorg_criteria_items = _final_eval_criteria_items_for_unit_phase(
            db, ex, unit_key, "reorg"
        )
        preparation_criteria_items = _dedupe_criteria_items_by_label(preparation_criteria_items)
        opening_criteria_items = _dedupe_criteria_items_by_label(opening_criteria_items)
        main_criteria_items = _dedupe_criteria_items_by_label(main_criteria_items)
        reorg_criteria_items = _dedupe_criteria_items_by_label(reorg_criteria_items)
        preparation_criteria_items = _enrich_criteria_items_with_acquired(
            preparation_criteria_items, preparation_detail_rows
        )
        opening_criteria_items = _enrich_criteria_items_with_acquired(
            opening_criteria_items, opening_detail_rows
        )
        main_criteria_items = _enrich_criteria_items_with_acquired(
            main_criteria_items, main_detail_rows
        )
        reorg_criteria_items = _enrich_criteria_items_with_acquired(
            reorg_criteria_items, reorg_detail_rows
        )
        preparation_detail_rows = _filter_eval_detail_rows_excluding_criteria(
            preparation_detail_rows, preparation_criteria_items
        )
        opening_detail_rows = _filter_eval_detail_rows_excluding_criteria(
            opening_detail_rows, opening_criteria_items
        )
        main_detail_rows = _filter_eval_detail_rows_excluding_criteria(
            main_detail_rows, main_criteria_items
        )
        reorg_detail_rows = _filter_eval_detail_rows_excluding_criteria(
            reorg_detail_rows, reorg_criteria_items
        )
        evaluation_tracks_detail_rows_raw = [
            {**r, "phase_label": FINAL_EVALUATION_TRACK_PHASE_LABEL}
            for r in _final_eval_detail_rows_for_unit_phase(
                details_by_unit_phase, unit_key, FINAL_EVALUATION_TRACK_PHASE_KEY
            )
        ]
        evaluation_tracks_criteria_items = _final_eval_criteria_items_for_unit_phase(
            db, ex, unit_key, "evaluation_tracks"
        )
        evaluation_tracks_criteria_items = _dedupe_criteria_items_by_label(
            evaluation_tracks_criteria_items
        )
        evaluation_tracks_criteria_items = _enrich_criteria_items_with_acquired(
            evaluation_tracks_criteria_items, evaluation_tracks_detail_rows_raw
        )
        evaluation_tracks_detail_rows = _filter_eval_detail_rows_excluding_criteria(
            evaluation_tracks_detail_rows_raw, evaluation_tracks_criteria_items
        )
        preparation_criteria_total = _criteria_items_allocated_total(preparation_criteria_items)
        opening_criteria_total = _criteria_items_allocated_total(opening_criteria_items)
        main_criteria_total = _criteria_items_allocated_total(main_criteria_items)
        reorg_criteria_total = _criteria_items_allocated_total(reorg_criteria_items)
        preparation_allocated_acquired = _criteria_items_allocated_acquired_total(
            preparation_criteria_items
        )
        opening_allocated_acquired = _criteria_items_allocated_acquired_total(
            opening_criteria_items
        )
        main_allocated_acquired = _criteria_items_allocated_acquired_total(main_criteria_items)
        reorg_allocated_acquired = _criteria_items_allocated_acquired_total(reorg_criteria_items)
        preparation_allocated_pct = _criteria_items_eval_list_footer_pct(
            preparation_criteria_items
        )
        opening_allocated_pct = _criteria_items_eval_list_footer_pct(opening_criteria_items)
        main_allocated_pct = _criteria_items_eval_list_footer_pct(main_criteria_items)
        reorg_allocated_pct = _criteria_items_eval_list_footer_pct(reorg_criteria_items)

        preparation_allocated_grade = _criteria_items_eval_list_footer_grade(
            preparation_criteria_items, preparation_allocated_pct
        )
        opening_allocated_grade = _criteria_items_eval_list_footer_grade(
            opening_criteria_items, opening_allocated_pct
        )
        main_allocated_grade = _criteria_items_eval_list_footer_grade(
            main_criteria_items, main_allocated_pct
        )
        reorg_allocated_grade = _criteria_items_eval_list_footer_grade(
            reorg_criteria_items, reorg_allocated_pct
        )
        _apply_criteria_totals_to_phase_rows(
            unit_phase_rows,
            criteria_max_by_phase={
                "preparation": preparation_criteria_total,
                "opening": opening_criteria_total,
                "main": main_criteria_total,
                "reorg": reorg_criteria_total,
            },
            criteria_acquired_by_phase={
                "preparation": preparation_allocated_acquired,
                "opening": opening_allocated_acquired,
                "main": main_allocated_acquired,
                "reorg": reorg_allocated_acquired,
            },
            criteria_pct_by_phase={
                "preparation": preparation_allocated_pct,
                "opening": opening_allocated_pct,
                "main": main_allocated_pct,
                "reorg": reorg_allocated_pct,
            },
            criteria_grade_by_phase={
                "preparation": preparation_allocated_grade,
                "opening": opening_allocated_grade,
                "main": main_allocated_grade,
                "reorg": reorg_allocated_grade,
            },
        )
        _recompute_unit_total_pct_on_phase_rows(unit_phase_rows)
        for row in unit_phase_rows:
            row["unit_anchor"] = anchor
            final_rows_all.append(row)
        report_units.append(
            {
                "unit_key": unit_key,
                "unit_label": unit_label,
                "anchor": anchor,
                "show_evaluation_tracks": unit_key in FINAL_EVALUATION_TRACK_UNIT_KEYS,
                "phase_rows": phase_rows,
                "phase_summary": _final_report_phase_summary(phase_rows),
                "preparation_detail_rows": preparation_detail_rows,
                "preparation_criteria_items": preparation_criteria_items,
                "preparation_criteria_total": preparation_criteria_total,
                "preparation_criteria_eval_max_total": _criteria_items_eval_list_max_total(
                    preparation_criteria_items
                ),
                "preparation_criteria_allocated_acquired_total": preparation_allocated_acquired,
                "preparation_criteria_acquired_total": _criteria_items_eval_list_acquired_total(
                    preparation_criteria_items
                ),
                "preparation_criteria_allocated_pct": preparation_allocated_pct,
                "preparation_criteria_allocated_grade": preparation_allocated_grade,
                "opening_detail_rows": opening_detail_rows,
                "opening_criteria_items": opening_criteria_items,
                "opening_criteria_total": opening_criteria_total,
                "opening_criteria_eval_max_total": _criteria_items_eval_list_max_total(
                    opening_criteria_items
                ),
                "opening_criteria_allocated_acquired_total": opening_allocated_acquired,
                "opening_criteria_acquired_total": _criteria_items_eval_list_acquired_total(
                    opening_criteria_items
                ),
                "opening_criteria_allocated_pct": opening_allocated_pct,
                "opening_criteria_allocated_grade": opening_allocated_grade,
                "main_detail_rows": main_detail_rows,
                "main_criteria_items": main_criteria_items,
                "main_criteria_total": main_criteria_total,
                "main_criteria_eval_max_total": _criteria_items_eval_list_max_total(
                    main_criteria_items
                ),
                "main_criteria_allocated_acquired_total": main_allocated_acquired,
                "main_criteria_acquired_total": _criteria_items_eval_list_acquired_total(
                    main_criteria_items
                ),
                "main_criteria_allocated_pct": main_allocated_pct,
                "main_criteria_allocated_grade": main_allocated_grade,
                "reorg_detail_rows": reorg_detail_rows,
                "reorg_criteria_items": reorg_criteria_items,
                "reorg_criteria_total": reorg_criteria_total,
                "reorg_criteria_eval_max_total": _criteria_items_eval_list_max_total(
                    reorg_criteria_items
                ),
                "reorg_criteria_allocated_acquired_total": reorg_allocated_acquired,
                "reorg_criteria_acquired_total": _criteria_items_eval_list_acquired_total(
                    reorg_criteria_items
                ),
                "reorg_criteria_allocated_pct": reorg_allocated_pct,
                "reorg_criteria_allocated_grade": reorg_allocated_grade,
                "evaluation_tracks_detail_rows": evaluation_tracks_detail_rows,
                "evaluation_tracks_criteria_items": evaluation_tracks_criteria_items,
                "evaluation_tracks_criteria_total": _criteria_items_allocated_total(
                    evaluation_tracks_criteria_items
                ),
                "evaluation_tracks_criteria_eval_max_total": _criteria_items_eval_list_max_total(
                    evaluation_tracks_criteria_items
                ),
                "evaluation_tracks_criteria_allocated_acquired_total": _criteria_items_allocated_acquired_total(
                    evaluation_tracks_criteria_items
                ),
                "evaluation_tracks_criteria_acquired_total": _criteria_items_eval_list_acquired_total(
                    evaluation_tracks_criteria_items
                ),
                "evaluation_tracks_criteria_allocated_pct": _criteria_items_eval_list_footer_pct(
                    evaluation_tracks_criteria_items
                ),
                "evaluation_tracks_criteria_allocated_grade": _criteria_items_eval_list_footer_grade(
                    evaluation_tracks_criteria_items,
                    _criteria_items_eval_list_footer_pct(evaluation_tracks_criteria_items),
                ),
            }
        )

    final_rows_all.sort(
        key=lambda r: (
            unit_order.get(r.get("unit_key") or "", len(unit_order)),
            phase_order.get(r.get("phase_key") or "", len(phase_order)),
        )
    )
    for row_idx, row in enumerate(final_rows_all):
        row["show_unit_total"] = (
            row_idx == 0
            or (final_rows_all[row_idx - 1].get("unit_key") != row.get("unit_key"))
        )
        uk = row.get("unit_key") or ""
        row["unit_rowspan"] = sum(1 for r in final_rows_all if r.get("unit_key") == uk)

    return {
        "has_exercise": True,
        "exercise": ex,
        "final_rows": final_rows_all,
        "report_summary": _build_final_report_exercise_summary(final_rows_all),
        "report_units": report_units,
        "all_phase_rows": final_rows_all,
        "detail_rows": detail_rows,
        "preparation_detail_rows": [r for r in detail_rows if r["phase_key"] == "preparation"],
        "opening_detail_rows": [r for r in detail_rows if r["phase_key"] == "opening"],
        "main_detail_rows": [r for r in detail_rows if r["phase_key"] == "main"],
        "reorg_detail_rows": [r for r in detail_rows if r["phase_key"] == "reorg"],
        "n_eval_lists": len(eval_items),
        "n_saved_eval_lists": len(saved_source_rows),
        "n_approved_eval_lists": approved_source_count,
        "n_saved_pending_eval_lists": saved_pending_source_count,
    }


DEFAULT_ANALYST_EVALUATION_CRITERIA_UNITS: tuple[str, ...] = ()

ANALYST_EVALUATION_CRITERIA_PHASE_ORDER: tuple[tuple[str, str], ...] = ()
# عرض البيانات القديمة عند فراغ كتالوج مراحل التخطيط
_LEGACY_ANALYST_CRITERIA_PHASE_ORDER: tuple[tuple[str, str], ...] = (
    ("preparation", "التهيئة"),
    ("evaluation_tracks", "مسارات التقييم"),
    ("opening", "الافتتاح"),
    ("main", "العمليات"),
)
ANALYST_EVALUATION_CRITERIA_PHASES: dict[str, str] = dict(
    ANALYST_EVALUATION_CRITERIA_PHASE_ORDER or _LEGACY_ANALYST_CRITERIA_PHASE_ORDER
)


def _analyst_criteria_phases_for_display(has_rows: bool) -> list[tuple[str, str]]:
    if ANALYST_EVALUATION_CRITERIA_PHASE_ORDER:
        return list(ANALYST_EVALUATION_CRITERIA_PHASE_ORDER)
    if EXERCISE_PHASE_OPTIONS:
        return list(EXERCISE_PHASE_OPTIONS)
    return []


_CRITERIA_PHASE_LEGACY_ALIASES: tuple[tuple[str, str], ...] = (
    ("main", "battle_exposure"),
    ("reorg", "reorganization"),
    ("evaluation_tracks", "reorganization"),
)


def _analyst_criteria_valid_phase_keys() -> set[str]:
    return {pk for pk, _ in _analyst_criteria_phases_for_display(True)}


def _resolve_analyst_criteria_phase_key(raw: str) -> str | None:
    """مفتاح مرحلة معيّن للتخزين والمسار — متوافق مع كتالوج التخطيط والمفاتيح القديمة."""
    key = (raw or "").strip()
    if not key:
        return None
    valid = _analyst_criteria_valid_phase_keys()
    if key in valid:
        return key
    for legacy_key, catalog_key in _CRITERIA_PHASE_LEGACY_ALIASES:
        if key == legacy_key and catalog_key in valid:
            return catalog_key
        if key == catalog_key and legacy_key in valid:
            return legacy_key
    norm = normalize_exercise_phase(key)
    if norm and norm in valid:
        return norm
    for pk, _ in _CONTROL_REPORT_PHASE_FALLBACK:
        if pk == key and pk in valid:
            return pk
    legacy_label = ANALYST_EVALUATION_CRITERIA_PHASES.get(key)
    if legacy_label:
        for pk, lbl in _analyst_criteria_phases_for_display(True):
            if lbl == legacy_label:
                return pk
        return key if key in ANALYST_EVALUATION_CRITERIA_PHASES else None
    return None


def _analyst_criteria_phase_label(phase_key: str) -> str:
    """تسمية المرحلة كما في جدول معايير التقييم."""
    key = (phase_key or "").strip()
    resolved = _resolve_analyst_criteria_phase_key(key) or key
    for pk, lbl in _analyst_criteria_phases_for_display(True):
        if pk == resolved:
            return lbl
    label = exercise_phase_label(resolved)
    if label:
        return label
    return ANALYST_EVALUATION_CRITERIA_PHASES.get(resolved) or ANALYST_EVALUATION_CRITERIA_PHASES.get(key) or key


def _analyst_criteria_phase_db_keys(canonical_key: str) -> list[str]:
    """كل مفاتيح phase_key المحتملة في قاعدة البيانات لنفس المرحلة."""
    keys = [(canonical_key or "").strip()]
    for a, b in _CRITERIA_PHASE_LEGACY_ALIASES:
        if canonical_key == a and b not in keys:
            keys.append(b)
        elif canonical_key == b and a not in keys:
            keys.append(a)
    return [k for k in keys if k]


def _planner_unit_keys_for_exercise(db, ex: Exercise) -> list[str]:
    """مفاتيح مستويات الوحدة التي لها قوائم تقييم في التمرين (كما في التخطيط)."""
    raw_keys = [
        (row[0] or "").strip()
        for row in (
            db.query(EvaluationListPdfItem.unit_level_key)
            .filter(EvaluationListPdfItem.exercise_id == int(ex.id))
            .distinct()
            .all()
        )
        if (row[0] or "").strip()
    ]
    order = {u["key"]: idx for idx, u in enumerate(UNIT_LEVELS)}
    return sorted(set(raw_keys), key=lambda k: order.get(k, len(order)))


def _sync_analyst_criteria_units_from_planner(
    db,
    ex: Exercise,
) -> list[AnalystEvaluationCriteriaUnit]:
    """تحديث تسميات الوحدات من كتالوج مستوى الوحدة في التخطيط (إنشاء قائمة التقييم)."""
    existing = (
        db.query(AnalystEvaluationCriteriaUnit)
        .filter(AnalystEvaluationCriteriaUnit.exercise_id == ex.id)
        .order_by(AnalystEvaluationCriteriaUnit.sort_order, AnalystEvaluationCriteriaUnit.id)
        .all()
    )
    if not existing:
        seed_keys = _planner_unit_keys_for_exercise(db, ex)
        if not seed_keys:
            for label in DEFAULT_ANALYST_EVALUATION_CRITERIA_UNITS:
                key = _resolve_unit_level_key_for_criteria_label(label)
                if key and key not in seed_keys:
                    seed_keys.append(key)
        for idx, key in enumerate(seed_keys):
            label = label_for_unit_level_key(key) or key
            db.add(
                AnalystEvaluationCriteriaUnit(
                    exercise_id=ex.id,
                    sort_order=idx,
                    label=label[:300],
                )
            )
        db.commit()
    else:
        for row in existing:
            key = _resolve_unit_level_key_for_criteria_label(row.label or "")
            if key:
                catalog_label = label_for_unit_level_key(key)
                if catalog_label:
                    row.label = catalog_label[:300]
        db.commit()
    return (
        db.query(AnalystEvaluationCriteriaUnit)
        .filter(AnalystEvaluationCriteriaUnit.exercise_id == ex.id)
        .order_by(AnalystEvaluationCriteriaUnit.sort_order, AnalystEvaluationCriteriaUnit.id)
        .all()
    )


def _ensure_analyst_evaluation_criteria_units(
    db,
    ex: Exercise,
) -> list[AnalystEvaluationCriteriaUnit]:
    return _sync_analyst_criteria_units_from_planner(db, ex)


def _parse_pct_form_value(raw: str | None) -> float | None:
    s = (raw or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        v = float(s)
    except (TypeError, ValueError):
        return None
    return max(0.0, min(100.0, v))


def _save_analyst_evaluation_criteria_distribution(db, user: User, ex: Exercise) -> None:
    del user  # محفوظ للتوقيع المتسق مع بقية دوال الحفظ.
    existing = {
        int(row.id): row
        for row in (
            db.query(AnalystEvaluationCriteriaUnit)
            .filter(AnalystEvaluationCriteriaUnit.exercise_id == ex.id)
            .all()
        )
    }
    delete_ids = {
        int(x)
        for x in request.form.getlist("delete_unit_ids")
        if (x or "").strip().isdigit()
    }
    ordered_ids = [
        int(x)
        for x in request.form.getlist("unit_ids")
        if (x or "").strip().isdigit()
    ]
    for sort_order, uid in enumerate(ordered_ids):
        row = existing.get(uid)
        if row is None:
            continue
        if uid in delete_ids:
            db.query(AnalystEvaluationCriteriaPhaseItem).filter(
                AnalystEvaluationCriteriaPhaseItem.criteria_unit_id == uid
            ).delete(synchronize_session=False)
            db.delete(row)
            continue
        unit_key = (request.form.get(f"unit_level_key__{uid}") or "").strip()
        if unit_key:
            catalog_label = label_for_unit_level_key(unit_key)
            if catalog_label:
                row.label = catalog_label[:300]
        row.sort_order = sort_order
    new_key = (request.form.get("new_unit_level_key") or "").strip()
    if new_key:
        catalog_label = label_for_unit_level_key(new_key) or new_key
        exists = (
            db.query(AnalystEvaluationCriteriaUnit)
            .filter(AnalystEvaluationCriteriaUnit.exercise_id == ex.id)
            .all()
        )
        if not any(
            _resolve_unit_level_key_for_criteria_label(r.label or "") == new_key
            for r in exists
        ):
            db.add(
                AnalystEvaluationCriteriaUnit(
                    exercise_id=ex.id,
                    sort_order=len(ordered_ids),
                    label=catalog_label[:300],
                )
            )
    db.commit()


def _parse_mark_form_value(raw: str | None) -> float | None:
    s = (raw or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        v = float(s)
    except (TypeError, ValueError):
        return None
    return max(0.0, v)


def _norm_unit_label_for_match(label: str) -> str:
    import re

    s = (label or "").strip().casefold()
    s = re.sub(r"\s+", " ", s)
    for old, new in (("أ", "ا"), ("إ", "ا"), ("آ", "ا"), ("ى", "ي")):
        s = s.replace(old, new)
    return s.replace("ـ", "")


def _norm_eval_list_label_for_match(label: str) -> str:
    """تطبيع اسم قائمة التقييم لمقارنة التكرار بين المعايير وقوائم المحكمين."""
    import re

    s = (label or "").strip().casefold()
    if s.endswith(".xlsx"):
        s = s[:-5].strip()
    s = re.sub(r"\s+", " ", s)
    for old, new in (("أ", "ا"), ("إ", "ا"), ("آ", "ا"), ("ى", "ي")):
        s = s.replace(old, new)
    return s.replace("ـ", "")


def _dedupe_criteria_items_by_label(items: list[dict]) -> list[dict]:
    seen: set[str] = set()
    out: list[dict] = []
    for item in items or []:
        key = _norm_eval_list_label_for_match(item.get("criteria_text") or "")
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def _filter_eval_detail_rows_excluding_criteria(
    detail_rows: list[dict],
    criteria_items: list[dict],
) -> list[dict]:
    """إزالة صفوف قوائم المحكمين المكررة مع صفوف معايير التقييم."""
    if not criteria_items:
        return list(detail_rows or [])
    criteria_keys = {
        _norm_eval_list_label_for_match(item.get("criteria_text") or "")
        for item in criteria_items
    }
    criteria_keys.discard("")
    if not criteria_keys:
        return list(detail_rows or [])
    return [
        row
        for row in (detail_rows or [])
        if _norm_eval_list_label_for_match(row.get("list_label") or "") not in criteria_keys
    ]


def _resolve_unit_level_key_for_criteria_label(label: str) -> str:
    """ربط صف وحدة في معايير التقييم بمفتاح مستوى الوحدة في قوائم التخطيط."""
    key = normalize_unit_level_key(label)
    if key:
        return key
    norm = _norm_unit_label_for_match(label)
    if not norm:
        return ""
    for row in UNIT_LEVELS:
        if _norm_unit_label_for_match(row["label"]) == norm:
            return row["key"]
    best_key = ""
    best_len = 0
    for row in UNIT_LEVELS:
        unit_norm = _norm_unit_label_for_match(row["label"])
        if not unit_norm:
            continue
        if norm in unit_norm or unit_norm in norm:
            if len(unit_norm) > best_len:
                best_key = row["key"]
                best_len = len(unit_norm)
    return best_key


def _evaluation_list_phase_for_criteria_phase(criteria_phase_key: str) -> str:
    """ربط مرحلة معايير التقييم بمرحلة قوائم التقييم في التخطيط."""
    key = (criteria_phase_key or "").strip()
    if key == "evaluation_tracks":
        return FINAL_EVALUATION_TRACK_PHASE_KEY
    return _normalized_exercise_phase(key)


def _evaluation_list_phases_for_criteria_phase(criteria_phase_key: str) -> list[str]:
    """مراحل قوائم التقييم المحتملة (للمرادفات مثل المعركة التعرضية ومسارات التقييم)."""
    primary = _evaluation_list_phase_for_criteria_phase(criteria_phase_key)
    phases: list[str] = []
    if primary:
        phases.append(primary)
    key = (criteria_phase_key or "").strip()
    if key and key not in phases:
        phases.append(key)
    if primary == "main" or key in ("main", "battle_exposure"):
        for alias in ("main", "battle_exposure"):
            if alias not in phases:
                phases.append(alias)
    if primary in ("reorg", "reorganization", "evaluation_tracks") or key in (
        "reorg",
        "reorganization",
        "evaluation_tracks",
    ):
        for alias in ("reorg", "reorganization", "evaluation_tracks"):
            if alias not in phases:
                phases.append(alias)
    return phases


def _evaluation_list_titles_for_criteria_unit(
    db,
    ex: Exercise,
    unit: AnalystEvaluationCriteriaUnit,
    criteria_phase_key: str,
) -> list[str]:
    """عناوين القائمة الحالية من مساحة التخطيط (إنشاء قائمة التقييم) لهذه الوحدة والمرحلة."""
    from sqlalchemy import or_

    unit_label = (unit.label or "").strip()
    unit_level_key = _resolve_unit_level_key_for_criteria_label(unit_label)
    unit_label_norm = _norm_unit_label_for_match(unit_label)
    eval_phases = _evaluation_list_phases_for_criteria_phase(criteria_phase_key)

    q = db.query(EvaluationListPdfItem).filter(
        EvaluationListPdfItem.exercise_id == int(ex.id),
        EvaluationListPdfItem.exercise_phase.in_(eval_phases),
    )
    unit_filters = []
    if unit_level_key:
        unit_filters.append(EvaluationListPdfItem.unit_level_key == unit_level_key)
    if unit_label:
        unit_filters.append(EvaluationListPdfItem.unit_level_label == unit_label)
    if not unit_filters:
        return []
    rows = (
        q.filter(or_(*unit_filters))
        .order_by(
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    if not rows and (unit_label_norm or unit_level_key):
        fuzzy: list[EvaluationListPdfItem] = []
        for row in q.order_by(
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        ).all():
            row_key = (row.unit_level_key or "").strip()
            row_label_norm = _norm_unit_label_for_match(row.unit_level_label or "")
            if unit_level_key and row_key == unit_level_key:
                fuzzy.append(row)
            elif unit_label_norm and row_label_norm == unit_label_norm:
                fuzzy.append(row)
            elif unit_label_norm and row_label_norm and (
                unit_label_norm in row_label_norm or row_label_norm in unit_label_norm
            ):
                fuzzy.append(row)
        rows = fuzzy

    titles: list[str] = []
    seen: set[str] = set()
    for row in rows:
        title = (row.text or "").strip()
        if not title or title in seen:
            continue
        seen.add(title)
        titles.append(title[:1000])
    return titles


def _criteria_unit_for_unit_level(
    db, exercise_id: int, unit_level_key: str
) -> AnalystEvaluationCriteriaUnit | None:
    unit_level_key = (unit_level_key or "").strip()
    if not unit_level_key:
        return None
    for row in (
        db.query(AnalystEvaluationCriteriaUnit)
        .filter(AnalystEvaluationCriteriaUnit.exercise_id == int(exercise_id))
        .order_by(AnalystEvaluationCriteriaUnit.sort_order, AnalystEvaluationCriteriaUnit.id)
        .all()
    ):
        if _resolve_unit_level_key_for_criteria_label(row.label or "") == unit_level_key:
            return row
    return None


def _criteria_items_allocated_total(items: list[dict]) -> float:
    return sum(
        float(item["allocated_mark"])
        for item in (items or [])
        if item.get("allocated_mark") is not None
    )


def _final_eval_criteria_items_for_unit_phase(
    db, ex: Exercise, unit_key: str, final_phase_key: str
) -> list[dict]:
    """صفوف المعايير + العلامة المخصصة من معايير التقييم (عرض في التقييم النهائي)."""
    criteria_unit = _criteria_unit_for_unit_level(db, int(ex.id), unit_key)
    if criteria_unit is None:
        return []
    for criteria_phase in _FINAL_EVAL_TO_CRITERIA_PHASE.get(
        (final_phase_key or "").strip(), ((final_phase_key or "").strip(),)
    ):
        if not criteria_phase:
            continue
        items = _criteria_phase_items_for_unit(db, ex, criteria_unit, criteria_phase)
        if items:
            return _dedupe_criteria_items_by_label(items)
    return []


def _criteria_phase_items_for_unit(
    db,
    ex: Exercise,
    unit: AnalystEvaluationCriteriaUnit,
    phase_key: str,
) -> list[dict]:
    phase_db_keys = _analyst_criteria_phase_db_keys(phase_key)
    rows = (
        db.query(AnalystEvaluationCriteriaPhaseItem)
        .filter(
            AnalystEvaluationCriteriaPhaseItem.exercise_id == ex.id,
            AnalystEvaluationCriteriaPhaseItem.criteria_unit_id == unit.id,
            AnalystEvaluationCriteriaPhaseItem.phase_key.in_(phase_db_keys),
        )
        .order_by(
            AnalystEvaluationCriteriaPhaseItem.sort_order,
            AnalystEvaluationCriteriaPhaseItem.id,
        )
        .all()
    )
    saved: list[dict] = []
    for row in rows:
        mark = float(row.allocated_mark) if row.allocated_mark is not None else None
        saved.append(
            {
                "criteria_text": (row.criteria_text or "").strip(),
                "allocated_mark": mark,
            }
        )
    return _merge_criteria_items_with_evaluation_lists(
        db, ex, unit, phase_key, saved_items=saved
    )


def _merge_criteria_items_with_evaluation_lists(
    db,
    ex: Exercise,
    unit: AnalystEvaluationCriteriaUnit,
    criteria_phase_key: str,
    *,
    saved_items: list[dict],
) -> list[dict]:
    """دمج علامات المعايير المحفوظة مع عناوين قوائم التقييم من التخطيط."""
    list_titles = _evaluation_list_titles_for_criteria_unit(
        db, ex, unit, criteria_phase_key
    )
    if not list_titles:
        total_mark = sum(
            float(item["allocated_mark"] or 0)
            for item in saved_items
            if item.get("allocated_mark") is not None
        )
        out: list[dict] = []
        for item in saved_items:
            mark = item.get("allocated_mark")
            pct = (
                (float(mark) / total_mark * 100.0)
                if mark is not None and total_mark > 0
                else None
            )
            out.append(
                {
                    "criteria_text": item.get("criteria_text") or "",
                    "allocated_mark": mark,
                    "allocated_pct": pct,
                    "from_evaluation_list": False,
                }
            )
        return out

    marks_by_text: dict[str, float | None] = {}
    marks_by_index: list[float | None] = []
    for item in saved_items:
        text = (item.get("criteria_text") or "").strip()
        mark = item.get("allocated_mark")
        marks_by_index.append(mark)
        if text and text not in marks_by_text:
            marks_by_text[text] = mark

    merged: list[dict] = []
    for idx, title in enumerate(list_titles):
        mark = marks_by_text.get(title)
        if mark is None and idx < len(marks_by_index):
            mark = marks_by_index[idx]
        merged.append(
            {
                "criteria_text": title,
                "allocated_mark": mark,
                "from_evaluation_list": True,
            }
        )
    total_mark = sum(float(m or 0) for m in (r["allocated_mark"] for r in merged) if m is not None)
    for row in merged:
        mark = row.get("allocated_mark")
        row["allocated_pct"] = (
            (float(mark) / total_mark * 100.0)
            if mark is not None and total_mark > 0
            else None
        )
    return merged


def _save_criteria_phase_items_for_unit(
    db,
    ex: Exercise,
    unit: AnalystEvaluationCriteriaUnit,
    phase_key: str,
) -> None:
    phase_db_keys = _analyst_criteria_phase_db_keys(phase_key)
    db.query(AnalystEvaluationCriteriaPhaseItem).filter(
        AnalystEvaluationCriteriaPhaseItem.exercise_id == ex.id,
        AnalystEvaluationCriteriaPhaseItem.criteria_unit_id == unit.id,
        AnalystEvaluationCriteriaPhaseItem.phase_key.in_(phase_db_keys),
    ).delete(synchronize_session=False)
    storage_key = _resolve_analyst_criteria_phase_key(phase_key) or phase_key
    list_titles = _evaluation_list_titles_for_criteria_unit(db, ex, unit, phase_key)
    marks = request.form.getlist("allocated_mark")
    if list_titles:
        criteria_texts = list_titles
    else:
        criteria_texts = [
            (t or "").strip()[:1000] for t in request.form.getlist("criteria_text")
        ]
    n = max(len(criteria_texts), len(marks))
    for idx in range(n):
        text_value = (criteria_texts[idx] if idx < len(criteria_texts) else "").strip()[:1000]
        mark = _parse_mark_form_value(marks[idx] if idx < len(marks) else "")
        if not text_value and mark is None:
            continue
        db.add(
            AnalystEvaluationCriteriaPhaseItem(
                exercise_id=ex.id,
                criteria_unit_id=unit.id,
                phase_key=storage_key,
                sort_order=idx,
                criteria_text=text_value,
                allocated_mark=mark,
            )
        )
    db.commit()


def _ensure_judge_roster_synced(db, user: User, ex: Exercise | None) -> None:
    """إذا كانت بيانات قائمة المحكمين موجودة لكن الربط غير مُنشأ، أنشئه تلقائياً.

    هذا مهم في حال تم تحديث النظام بعد إدخال القوائم أو لم يتم إعادة حفظ قائمة المحكمين.
    """
    if ex is None or user is None:
        return
    if is_system_admin(user):
        return
    if not is_judge(user):
        return
    judge_id = int(getattr(user, "id", 0) or 0)
    if judge_id <= 0:
        return
    exists = (
        db.query(JudgeTraineeAssignment)
        .filter(
            JudgeTraineeAssignment.exercise_id == ex.id,
            JudgeTraineeAssignment.judge_user_id == judge_id,
        )
        .first()
    )
    if exists is None:
        _sync_judges_from_roster(db, ex)


@bp.route("/")
def home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login")
    return redirect("/dashboard")


@bp.route("/exercises")
def exercises_removed():
    """الصفحة أُلغيت حسب الطلب."""
    return redirect("/dashboard")


@bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        if get_current_user_optional():
            return redirect("/dashboard")
        return render_template("login.html", next_url=request.args.get("next", ""), error="")
    # POST
    from flask import g
    db = g.db
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    next_url = (request.form.get("next") or "").strip()
    u = (
        db.query(User)
        .filter(User.username == username, User.is_active == True)  # noqa: E712
        .first()
    )
    if not u or not verify_password(password, u.password_hash):
        return (
            render_template(
                "login.html",
                next_url=next_url,
                error="بيانات الدخول غير صحيحة",
            ),
            401,
        )
    u.last_login = datetime.utcnow()
    db.add(u)
    db.commit()
    session["user_id"] = u.id
    # للمحكمين: افتح مساحة المحكمين مباشرةً (ما لم يحدد النظام next)
    if (u.role_key or "") == RoleKey.JUDGE.value and not next_url:
        return redirect("/judge")
    if (u.role_key or "") == RoleKey.CHIEF_JUDGE.value and not next_url:
        return redirect("/judge")
    return redirect(next_url or "/dashboard")


@bp.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


def _dashboard_role_card_target(role_key: str, user: User) -> tuple[str, str, str]:
    """وجهة البطاقة في الصفحة الرئيسية: (المسار، aria، تلميح) مع احترام صلاحيات المستخدم."""
    if role_key == RoleKey.SYSTEM_ADMIN.value:
        if is_system_admin(user):
            return (
                "/admin/exercises/create",
                "فتح إنشاء تمرين جديد",
                "إبدأ",
            )
        return ("/library", "فتح المكتبة", "إبدأ")
    if role_key == RoleKey.ANALYST.value:
        return ("/analyst", "فتح مساحة المحللين", "إبدأ")
    if role_key == RoleKey.CONTROL.value:
        return ("/control", "فتح مساحة السيطرة", "إبدأ")
    if role_key == RoleKey.PLANNER.value:
        return ("/planner", "فتح مساحة التخطيط", "إبدأ")
    if role_key == RoleKey.JUDGE.value:
        return ("/judge", "فتح مساحة المحكمين", "إبدأ")
    if role_key == RoleKey.CHIEF_JUDGE.value:
        return (
            "/judge",
            "فتح مساحة المحكمين (مع صلاحيات كبير المحكمين)",
            "إبدأ",
        )
    if role_key == RoleKey.STANDARDS_LIBRARY.value:
        return ("/library", "فتح مكتبة المراجع والمعايير", "إبدأ")
    return ("/library", "فتح المكتبة", "إبدأ")


# ترتيب بطاقات الصفحة الرئيسية: المحكمين قبل المحللين ضمن الشبكة
_DASHBOARD_CARD_ORDER: tuple[str, ...] = (
    RoleKey.SYSTEM_ADMIN.value,
    RoleKey.PLANNER.value,
    RoleKey.CONTROL.value,
    RoleKey.JUDGE.value,
    RoleKey.ANALYST.value,
)
_DASHBOARD_CARD_ORDER_RANK: dict[str, int] = {rk: i for i, rk in enumerate(_DASHBOARD_CARD_ORDER)}


@bp.route("/dashboard")
def dashboard():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/dashboard")
    from flask import g
    db = g.db
    roles = db.query(RoleDef).order_by(RoleDef.id).all()
    rk = RoleKey.from_value(user.role_key)
    role_title = next(
        (r.title_ar for r in roles if r.role_key == user.role_key), rk.value
    )
    # لا بطاقة منفصلة لـ chief_judge: نفس نقطة الدخول «المحكمين» مع امتيازات الاعتماد الثاني من صلاحيات الدور.
    role_defs_home = [
        r
        for r in roles
        if r.role_key != RoleKey.STANDARDS_LIBRARY.value
        and r.role_key != RoleKey.CHIEF_JUDGE.value
    ]
    dashboard_cards: list[dict] = []
    for r in role_defs_home:
        href, aria, hint = _dashboard_role_card_target(r.role_key, user)
        dashboard_cards.append(
            {
                "role_key": r.role_key,
                "title_ar": r.title_ar,
                "duties_ar": r.duties_ar,
                "href": href,
                "aria_label": aria,
                "hint": hint,
            }
        )
    dashboard_cards.sort(
        key=lambda c: _DASHBOARD_CARD_ORDER_RANK.get(c["role_key"], 99)
    )
    return render_template(
        "dashboard.html",
        **_ctx(
            user,
            dashboard_cards=dashboard_cards,
            current_rk=rk,
            role_title=role_title,
            demo_password_note=DEMO_PASSWORD,
        ),
    )


# مساحة المحللين — عناصر الشريط (المعرّف، العنوان، أيقونة Font Awesome)
ANALYST_HUB_ITEMS: tuple[tuple[str, str, str], ...] = (
    ("evaluation-criteria", "معايير التقييم", "fa-list-check"),
    ("positives-negatives", "عرض الإيجابيات والسلبيات", "fa-plus-minus"),
    ("evaluation-results", "عرض نتائج التقييم", "fa-square-poll-vertical"),
    ("judges-eval-analysis", "تحليل وتقييم المحكمين", "fa-chart-column"),
    ("incomplete-tasks", "مهام غير مكتملة", "fa-clipboard-list"),
    ("after-action-review", "إنشاء مراجعة ما بعد العمل", "fa-people-arrows"),
    ("visual-documentation", "التوثيق المرئي", "fa-photo-film"),
    ("final-evaluation", "التقييم نهائي", "fa-file-signature"),
)
ANALYST_HUB_SLUGS: dict[str, str] = {s: t for s, t, _ in ANALYST_HUB_ITEMS}


@bp.route("/analyst")
def analyst_hub():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/analyst")
    if not can_access_analyst_hub(user):
        abort(403)
    hub_items = [{"slug": s, "title_ar": t, "icon": ic} for s, t, ic in ANALYST_HUB_ITEMS]
    return render_template(
        "analyst_hub.html",
        **_ctx(
            user,
            hub_items=hub_items,
        ),
    )


@bp.route("/analyst/<slug>", methods=["GET", "POST"])
def analyst_hub_section(slug: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/analyst/{slug}")
    if not can_access_analyst_hub(user):
        abort(403)
    slug_norm = (slug or "").strip().lower()
    if slug_norm == "visual-documentation":
        return redirect(url_for("views.visual_documentation", from_analyst=1))
    title = ANALYST_HUB_SLUGS.get(slug_norm)
    if not title:
        abort(404)
    def _actx(**extra):
        return _ctx(user, **_hub_back_ctx_for_request_path(), **extra)

    if slug_norm == "evaluation-criteria":
        from flask import g

        db = g.db
        dist = _build_analyst_evaluation_criteria_distribution(db, user)
        if not dist.get("has_exercise"):
            return render_template(
                "analyst_evaluation_criteria.html",
                **_actx( section_title=title, has_exercise=False),
            )
        if request.method == "POST":
            _save_analyst_evaluation_criteria_distribution(db, user, dist["exercise"])
            return redirect(url_for("views.analyst_hub_section", slug=slug_norm, ok=1))
        return render_template(
            "analyst_evaluation_criteria.html",
            **_actx(
                section_title=title,
                has_exercise=True,
                exercise=dist["exercise"],
                distribution_rows=dist["distribution_rows"],
                grand_total=dist["grand_total"],
                criteria_phases=dist.get("criteria_phases")
                or _analyst_criteria_phases_for_display(bool(dist.get("distribution_rows"))),
                available_unit_levels=dist.get("available_unit_levels") or [],
                ok_msg="تم الحفظ بنجاح." if request.args.get("ok") else "",
            ),
        )
    if slug_norm == "positives-negatives":
        from flask import g

        db = g.db
        pn = _build_control_positives_negatives(
            db,
            user,
            list_viewer="views.analyst_evaluation_list_file_viewer",
        )
        if not pn.get("has_exercise"):
            return render_template(
                "analyst_positives_negatives.html",
                **_actx(section_title=title, has_exercise=False),
            )
        return render_template(
            "analyst_positives_negatives.html",
            **_actx(section_title=title, **pn),
        )
    if slug_norm == "evaluation-results":
        from flask import g

        db = g.db
        dash = _build_analyst_saved_results_charts(db, user)
        if not dash.get("has_exercise"):
            return render_template(
                "analyst_evaluation_results_dashboard.html",
                **_actx(section_title=title, has_exercise=False),
            )
        return render_template(
            "analyst_evaluation_results_dashboard.html",
            **_actx(section_title=title, **dash),
        )
    if slug_norm == "judges-eval-analysis":
        from flask import g

        db = g.db
        ex = _admin_current_workspace_exercise(db, user)
        if ex is None:
            return render_template(
                "analyst_judges_evaluation_analysis.html",
                **_actx( section_title=title, has_exercise=False),
            )

        saved_rows = (
            db.query(EvaluationListSavedResult)
            .filter(EvaluationListSavedResult.exercise_id == ex.id)
            .order_by(EvaluationListSavedResult.updated_at.desc(), EvaluationListSavedResult.id.desc())
            .all()
        )

        eval_items = (
            db.query(EvaluationListPdfItem)
            .filter(EvaluationListPdfItem.exercise_id == ex.id)
            .order_by(
                _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
                _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
                EvaluationListPdfItem.sort_order,
                EvaluationListPdfItem.id,
            )
            .all()
        )

        # أحدث نتيجة لكل (محكم, ملف تقييم)
        latest_by_judge_item: dict[tuple[int, int], EvaluationListSavedResult] = {}
        for r in saved_rows:
            jid = getattr(r, "saved_by_id", None)
            if jid is None:
                continue
            key = (int(jid), int(r.evaluation_item_id))
            prev = latest_by_judge_item.get(key)
            if prev is None:
                latest_by_judge_item[key] = r
                continue
            pdt = getattr(prev, "updated_at", None)
            rdt = getattr(r, "updated_at", None)
            if (rdt and pdt and rdt > pdt) or (rdt and not pdt) or (rdt == pdt and r.id > prev.id):
                latest_by_judge_item[key] = r

        # إكمال قوائم التقييم لكل وحدة (معبّأة / غير معبّأة)
        saved_by_item_ids = {int(r.evaluation_item_id) for r in saved_rows if (r.payload_json or "").strip()}
        unit_completion: list[dict] = []
        by_unit_items: dict[str, list[EvaluationListPdfItem]] = {}
        for it in eval_items:
            uk = (it.unit_level_key or "").strip()
            if not uk:
                continue
            by_unit_items.setdefault(uk, []).append(it)
        for uk, items in by_unit_items.items():
            filled = []
            unfilled = []
            for it in items:
                x = {
                    "id": it.id,
                    "title": it.text or "تقييم",
                    "phase": _normalized_exercise_phase(getattr(it, "exercise_phase", None)),
                    "is_filled": int(it.id) in saved_by_item_ids,
                }
                (filled if x["is_filled"] else unfilled).append(x)
            unit_completion.append(
                {
                    "unit_key": uk,
                    "unit_label": label_for_unit_level_key(uk) or uk,
                    "total": len(items),
                    "filled_n": len(filled),
                    "unfilled_n": len(unfilled),
                    "filled": filled,
                    "unfilled": unfilled,
                }
            )
        unit_completion.sort(key=lambda x: (x["unfilled_n"], x["total"]), reverse=True)

        # إحصاءات المحكمين لكل وحدة + نتائج الملفات بشكل منفرد
        from app.models import User

        judge_ids = sorted({int(k[0]) for k in latest_by_judge_item.keys()})
        judge_users: dict[int, User] = {}
        if judge_ids:
            for u in db.query(User).filter(User.id.in_(judge_ids)).all():
                judge_users[int(u.id)] = u

        judge_stats: list[dict] = []
        for jid in judge_ids:
            u = judge_users.get(jid)
            display = (
                (getattr(u, "full_name", "") or "").strip()
                or (getattr(u, "username", "") or "").strip()
                or f"محكم #{jid}"
            )
            rows = [r for (jj, _iid), r in latest_by_judge_item.items() if jj == jid]
            by_unit: dict[str, list[EvaluationListSavedResult]] = {}
            for r in rows:
                uk = (r.unit_level_key or "").strip()
                by_unit.setdefault(uk or "—", []).append(r)
            unit_blocks = []
            for uk, rs in sorted(by_unit.items(), key=lambda x: x[0]):
                unit_blocks.append(
                    {
                        "unit_key": uk,
                        "unit_label": (label_for_unit_level_key(uk) or uk) if uk != "—" else "—",
                        "count": len(rs),
                        "results": [
                            {
                                "saved_id": rr.id,
                                "evaluation_item_id": rr.evaluation_item_id,
                                "exercise_phase": rr.exercise_phase,
                                "total_pct": rr.total_pct,
                                "grade_label": rr.grade_label,
                                "is_approved": bool(getattr(rr, "is_approved", False)),
                                "updated_at": rr.updated_at,
                            }
                            for rr in sorted(
                                rs,
                                key=lambda x: (x.updated_at or datetime.min, x.id),
                                reverse=True,
                            )
                        ],
                    }
                )
            judge_stats.append(
                {
                    "judge_id": jid,
                    "judge_name": display,
                    "total_results": len(rows),
                    "units": unit_blocks,
                }
            )

        # نتائج حسب الوحدة ثم حسب كل ملف تقييم منفرد (أحدث نتيجة لكل محكم)
        unit_item_results: list[dict] = []
        for ublk in unit_completion:
            uk = ublk["unit_key"]
            items = by_unit_items.get(uk) or []
            item_blocks = []
            for it in items:
                per_item = []
                for (jid, iid), r in latest_by_judge_item.items():
                    if int(iid) != int(it.id):
                        continue
                    u = judge_users.get(int(jid))
                    display = (
                        (getattr(u, "full_name", "") or "").strip()
                        or (getattr(u, "username", "") or "").strip()
                        or f"محكم #{jid}"
                    )
                    per_item.append(
                        {
                            "judge_id": jid,
                            "judge_name": display,
                            "saved_id": r.id,
                            "total_pct": r.total_pct,
                            "grade_label": r.grade_label,
                            "is_approved": bool(getattr(r, "is_approved", False)),
                            "updated_at": r.updated_at,
                        }
                    )
                per_item.sort(
                    key=lambda x: (x["is_approved"], x["updated_at"] or datetime.min),
                    reverse=True,
                )
                item_blocks.append(
                    {
                        "item_id": it.id,
                        "title": it.text or "تقييم",
                        "phase": _normalized_exercise_phase(getattr(it, "exercise_phase", None)),
                        "is_filled": int(it.id) in saved_by_item_ids,
                        "results": per_item,
                    }
                )
            unit_item_results.append(
                {
                    "unit_key": uk,
                    "unit_label": ublk["unit_label"],
                    "items": item_blocks,
                }
            )

        # نتيجة كل وحدة (المجموع العام): من أحدث نتيجة لكل (محكم, ملف) لتجنب التكرار
        by_unit_latest: dict[str, list[float]] = {}
        for (_jid, _iid), r in latest_by_judge_item.items():
            if r.total_pct is None:
                continue
            uk = (r.unit_level_key or "").strip()
            if not uk:
                continue
            try:
                by_unit_latest.setdefault(uk, []).append(float(r.total_pct))
            except Exception:
                continue

        unit_totals: list[dict] = []
        for uk, vals in by_unit_latest.items():
            if not vals:
                continue
            unit_totals.append(
                {
                    "unit_key": uk,
                    "unit_label": label_for_unit_level_key(uk) or uk,
                    "avg_pct": sum(vals) / len(vals),
                    "n": len(vals),
                }
            )
        unit_totals.sort(key=lambda x: x["avg_pct"], reverse=True)

        # نقاط القوة/الضعف على مستوى التمرين (حسب نتيجة الوحدات)
        strengths_units = unit_totals[:3]
        weaknesses_units = list(reversed(unit_totals[-3:])) if len(unit_totals) > 3 else []

        # تبويبات الوحدات: بيانات مختصرة للرسم + عرض كتابي + إكمال لكل محكم
        unit_tabs: list[dict] = []
        # خريطة سريعة: unit_key -> completion block
        completion_by_unit = {u["unit_key"]: u for u in unit_completion}
        # خريطة سريعة: unit_key -> unit_item_results block
        results_by_unit = {u["unit_key"]: u for u in unit_item_results}

        # تجهيز إكمال القوائم لكل محكم لكل وحدة (مخصّص/معبّأ/غير معبّأ)
        judge_name_by_id: dict[int, str] = {}
        for j in judge_stats:
            judge_name_by_id[int(j["judge_id"])] = j.get("judge_name") or f"محكم #{j['judge_id']}"

        unit_judge_completion: dict[str, list[dict]] = {}
        for ukey, items in by_unit_items.items():
            item_ids = {int(it.id) for it in items}
            rows = []
            for jid in judge_ids:
                filled_n = 0
                for iid in item_ids:
                    if (int(jid), int(iid)) in latest_by_judge_item:
                        filled_n += 1
                total_n = len(item_ids)
                rows.append(
                    {
                        "judge_id": jid,
                        "judge_name": judge_name_by_id.get(int(jid), f"محكم #{jid}"),
                        "total": total_n,
                        "filled_n": filled_n,
                        "unfilled_n": max(total_n - filled_n, 0),
                    }
                )
            rows.sort(key=lambda x: (x["unfilled_n"], x["total"]), reverse=True)
            unit_judge_completion[ukey] = rows

        for u in unit_completion:
            ukey = u["unit_key"]
            ures = results_by_unit.get(ukey) or {}
            items = ures.get("items") or []
            chart_items = []
            for it in items:
                vals = []
                for r in (it.get("results") or []):
                    v = r.get("total_pct")
                    if v is None:
                        continue
                    try:
                        vals.append(float(v))
                    except Exception:
                        continue
                avg = (sum(vals) / len(vals)) if vals else None
                chart_items.append(
                    {
                        "item_id": it.get("item_id"),
                        "title": it.get("title") or "تقييم",
                        "phase": _normalized_exercise_phase(it.get("phase")),
                        "is_filled": bool(it.get("is_filled")),
                        "avg_pct": avg,
                        "n": len(vals),
                    }
                )
            chart_items_scored = [x for x in chart_items if x["avg_pct"] is not None]
            chart_items_scored.sort(key=lambda x: float(x["avg_pct"]), reverse=True)
            unit_strengths = chart_items_scored[:3]
            unit_weaknesses = list(reversed(chart_items_scored[-3:])) if len(chart_items_scored) > 3 else []

            unit_vals = by_unit_latest.get(ukey) or []
            unit_total_pct = (sum(unit_vals) / len(unit_vals)) if unit_vals else None
            unit_total_grade = (
                grade_label_from_percent(unit_total_pct) if unit_total_pct is not None else "غير محسوب"
            )

            unit_tabs.append(
                {
                    "unit_key": ukey,
                    "unit_label": u["unit_label"],
                    "completion": u,
                    "unit_total_pct": unit_total_pct,
                    "unit_total_grade": unit_total_grade,
                    "chart_items": chart_items,
                    "strengths": unit_strengths,
                    "weaknesses": unit_weaknesses,
                    # لا نستخدم key اسمها items لأن Jinja تعتبرها dict.items()
                    "eval_lists": items,
                    "judge_completion": unit_judge_completion.get(ukey) or [],
                }
            )
        unit_tabs.sort(key=lambda x: x["unit_label"])

        # متوسط عام
        scored = [float(r.total_pct) for r in saved_rows if r.total_pct is not None]
        overall_avg = (sum(scored) / len(scored)) if scored else None
        overall_grade = grade_label_from_percent(overall_avg) if overall_avg is not None else "غير محسوب"

        # المحكم الأقدم: أول صف في قائمة المحكمين (حسب sort_order) ضمن التمرين الحالي
        oldest_judge = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        oldest_judge_unit_label = ""
        if oldest_judge is not None:
            uk_m = (getattr(oldest_judge, "unit_level_key", None) or "").strip()
            if uk_m:
                oldest_judge_unit_label = label_for_unit_level_key(uk_m) or uk_m
            else:
                oldest_judge_unit_label = (getattr(oldest_judge, "position_ar", None) or "").strip()

        saved_display = []
        for srow in saved_rows:
            uk_s = (srow.unit_level_key or "").strip()
            saved_display.append(
                {
                    "row": srow,
                    "unit_label_ar": (
                        label_for_unit_level_key(uk_s)
                        if uk_s
                        else ""
                    )
                    or (uk_s if uk_s else "—"),
                }
            )

        # توصية بسيطة
        if overall_avg is None:
            recommendation = "لا توجد نتائج محفوظة بعد. احفظ نتائج التقييم من صفحة قائمة التقييم أولاً."
        elif overall_avg >= 85:
            recommendation = "النتائج مرتفعة؛ يوصى بتثبيت الإجراءات الحالية والتركيز على التحسينات الدقيقة."
        elif overall_avg >= 70:
            recommendation = "النتائج جيدة؛ يوصى بوضع خطة تحسين مركّزة للعناصر الأقل أداءً."
        elif overall_avg >= 55:
            recommendation = "النتائج متوسطة؛ يوصى بإعادة التدريب على العناصر الأساسية وإعادة التقييم."
        else:
            recommendation = "النتائج منخفضة؛ يوصى بخطة تصحيح عاجلة وإعادة تنظيم التدريب قبل الاعتماد."

        return render_template(
            "analyst_judges_evaluation_analysis.html",
            **_actx(
                section_title=title,
                has_exercise=True,
                exercise=ex,
                saved_rows=saved_rows,
                saved_display=saved_display,
                unit_totals=unit_totals,
                strengths_units=strengths_units,
                weaknesses_units=weaknesses_units,
                overall_avg=overall_avg,
                overall_grade=overall_grade,
                oldest_judge=oldest_judge,
                oldest_judge_unit_label=oldest_judge_unit_label or "—",
                recommendation=recommendation,
                unit_completion=unit_completion,
                judge_stats=judge_stats,
                unit_item_results=unit_item_results,
                unit_tabs=unit_tabs,
            ),
        )
    if slug_norm == "final-evaluation":
        from flask import g

        db = g.db
        ex0 = _current_workspace_exercise(db, user)
        if ex0 is None and is_system_admin(user):
            ex0 = db.query(Exercise).order_by(Exercise.id.desc()).first()
        report = _build_analyst_final_evaluation_report(db, user)
        if not report.get("has_exercise"):
            return render_template(
                "analyst_final_evaluation.html",
                **_actx( section_title=title, has_exercise=False),
            )
        return render_template(
            "analyst_final_evaluation.html",
            **_actx(
                section_title=title,
                has_exercise=True,
                exercise=report["exercise"],
                final_rows=report["final_rows"],
                report_summary=report.get("report_summary"),
                all_phase_rows=report["all_phase_rows"],
                detail_rows=report["detail_rows"],
                preparation_detail_rows=report["preparation_detail_rows"],
                opening_detail_rows=report["opening_detail_rows"],
                main_detail_rows=report["main_detail_rows"],
                report_units=report["report_units"],
                n_eval_lists=report["n_eval_lists"],
                n_saved_eval_lists=report["n_saved_eval_lists"],
                n_approved_eval_lists=report["n_approved_eval_lists"],
                n_saved_pending_eval_lists=report["n_saved_pending_eval_lists"],
                final_eval_can_edit=True,
            ),
        )
    if slug_norm == "incomplete-tasks":
        return _render_incomplete_evaluations_page(
            user,
            section_title=title,
            section_icon="fa-clipboard-list",
            role="analyst",
        )
    return render_template(
        "analyst_section_placeholder.html",
        **_actx( section_title=title, section_slug=slug),
    )


@bp.route("/analyst/final-evaluation/<unit_key>", methods=["GET", "POST"])
def analyst_final_evaluation_unit_detail(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/analyst/final-evaluation/{unit_key}")
    if not can_access_analyst_hub(user):
        abort(403)
    from flask import g

    db = g.db
    unit_key_norm = (unit_key or "").strip()
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None and is_system_admin(user):
        ex0 = db.query(Exercise).order_by(Exercise.id.desc()).first()
    if request.method == "POST" and ex0 is not None:
        item_ids = [
            int(x)
            for x in request.form.getlist("evaluation_item_id")
            if (x or "").strip().isdigit()
        ]
        _save_final_eval_manual_maxes(
            db,
            exercise_id=int(ex0.id),
            unit_key=unit_key_norm,
            item_ids=item_ids,
        )
        return redirect(
            url_for(
                "views.analyst_final_evaluation_unit_detail",
                unit_key=unit_key_norm,
                saved=1,
            )
        )
    report = _build_analyst_final_evaluation_report(db, user)
    if not report.get("has_exercise"):
        abort(404)
    unit = next(
        (u for u in report.get("report_units", []) if (u.get("unit_key") or "").strip() == unit_key_norm),
        None,
    )
    if unit is None:
        abort(404)
    ok_msg = "تم حفظ علامات القصوى المخصصة." if request.args.get("saved") == "1" else None
    return render_template(
        "analyst_final_evaluation_unit.html",
        **_ctx(
            user,
            section_title=ANALYST_HUB_SLUGS.get("final-evaluation", "التقييم نهائي"),
            exercise=report["exercise"],
            unit=unit,
            n_approved_eval_lists=report["n_approved_eval_lists"],
            n_saved_pending_eval_lists=report["n_saved_pending_eval_lists"],
            final_eval_can_edit=True,
            ok_msg=ok_msg,
            **_subpage_close_ctx(
                url_for("views.analyst_hub_section", slug="final-evaluation")
            ),
        ),
    )


@bp.route("/api/analyst/final-evaluation/report-phase-max", methods=["POST"])
def api_analyst_final_eval_report_phase_max():
    """حفظ تلقائي لعلامة القصوى اليدوية (صف وحدة+مرحلة في التقرير النهائي)."""
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_access_analyst_hub(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    try:
        ex0 = _current_workspace_exercise(db, user)
        if ex0 is None and is_system_admin(user):
            ex0 = db.query(Exercise).order_by(Exercise.id.desc()).first()
        if ex0 is None:
            return jsonify({"ok": False, "error": "no_exercise"}), 400

        payload = request.get_json(silent=True) if request.is_json else {}
        if not isinstance(payload, dict):
            payload = {}
        field_name = (request.form.get("field_name") or payload.get("field_name") or "").strip()
        unit_key = (request.form.get("unit_key") or payload.get("unit_key") or "").strip()
        phase_key = request.form.get("phase_key") or payload.get("phase_key")
        if field_name:
            parsed = _parse_report_phase_max_field_name(field_name)
            if parsed:
                unit_key, phase_key = parsed
        mark_raw = request.form.get("max_mark")
        if mark_raw is None:
            mark_raw = payload.get("max_mark")
        if mark_raw is None and field_name:
            mark_raw = request.form.get(field_name)
        unit_key = (unit_key or "").strip()
        phase_key = _normalized_exercise_phase(str(phase_key or ""))
        if not unit_key or not phase_key:
            return jsonify({"ok": False, "error": "invalid_request"}), 400

        saved_mark = _upsert_final_eval_report_phase_max(
            db,
            exercise_id=int(ex0.id),
            unit_key=unit_key,
            phase_key=phase_key,
            mark_raw=None if mark_raw is None else str(mark_raw),
        )
        db.commit()
        metrics = _final_eval_phase_row_metrics(
            db,
            exercise_id=int(ex0.id),
            unit_key=unit_key,
            phase_key=phase_key,
            manual_max=saved_mark,
        )
        return jsonify(
            {
                "ok": True,
                "saved_mark": saved_mark,
                "acquired_mark": metrics["acquired_mark"],
                "phase_pct": metrics["phase_pct"],
                "phase_grade": metrics["phase_grade"],
            }
        )
    except Exception:
        db.rollback()
        return jsonify({"ok": False, "error": "server_error"}), 500


@bp.route("/api/analyst/final-evaluation/item-allocated-max", methods=["POST"])
def api_analyst_final_eval_item_allocated_max():
    """حفظ تلقائي لعلامة القصوى المخصصة لقائمة تقييم في تفاصيل الوحدة."""
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_access_analyst_hub(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None and is_system_admin(user):
        ex0 = db.query(Exercise).order_by(Exercise.id.desc()).first()
    if ex0 is None:
        return jsonify({"ok": False, "error": "no_exercise"}), 400

    try:
        payload = request.get_json(silent=True) if request.is_json else {}
        if not isinstance(payload, dict):
            payload = {}
        unit_key = (request.form.get("unit_key") or payload.get("unit_key") or "").strip()
        item_id_raw = request.form.get("item_id") or payload.get("item_id")
        mark_raw = request.form.get("max_mark")
        if mark_raw is None:
            mark_raw = payload.get("max_mark")
        if item_id_raw is None or not str(item_id_raw).strip().isdigit():
            return jsonify({"ok": False, "error": "invalid_item"}), 400
        item_id = int(item_id_raw)
        if mark_raw is None:
            mark_raw = request.form.get(f"allocated_max__{item_id}")

        saved_mark = _upsert_final_eval_item_allocated_max(
            db,
            exercise_id=int(ex0.id),
            unit_key=unit_key,
            item_id=item_id,
            mark_raw=None if mark_raw is None else str(mark_raw),
        )
        db.commit()
        return jsonify({"ok": True, "saved_mark": saved_mark, "item_id": item_id})
    except Exception:
        db.rollback()
        return jsonify({"ok": False, "error": "server_error"}), 500


@bp.route("/analyst/evaluation-criteria/<int:unit_id>/<phase_key>", methods=["GET", "POST"])
def analyst_evaluation_criteria_phase(unit_id: int, phase_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/analyst/evaluation-criteria/{unit_id}/{phase_key}")
    if not can_access_analyst_hub(user):
        abort(403)
    phase_key_raw = (phase_key or "").strip()
    phase_key = _resolve_analyst_criteria_phase_key(phase_key_raw)
    if not phase_key:
        abort(404)
    phase_label = _analyst_criteria_phase_label(phase_key)
    from flask import g

    db = g.db
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return render_template(
            "analyst_evaluation_criteria_phase.html",
            **_ctx(
                user,
                has_exercise=False,
                section_title="معايير التقييم",
                phase_label=phase_label,
                unit=None,
                items=[],
                total_mark=None,
                eval_list_driven=False,
                eval_list_count=0,
                criteria_unit_level_key="",
                **_subpage_close_ctx(
                    url_for("views.analyst_hub_section", slug="evaluation-criteria")
                ),
            ),
        )
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        abort(404)
    unit = db.get(AnalystEvaluationCriteriaUnit, unit_id)
    if unit is None or unit.exercise_id != ex.id:
        abort(404)
    if request.method == "POST":
        _save_criteria_phase_items_for_unit(db, ex, unit, phase_key)
        return redirect(
            url_for("views.analyst_hub_section", slug="evaluation-criteria", ok=1)
        )
    items = _criteria_phase_items_for_unit(db, ex, unit, phase_key)
    eval_list_titles = _evaluation_list_titles_for_criteria_unit(db, ex, unit, phase_key)
    eval_list_driven = bool(eval_list_titles)
    unit_level_key = _resolve_unit_level_key_for_criteria_label(unit.label or "")
    total_mark = sum(
        float(item["allocated_mark"] or 0)
        for item in items
        if item.get("allocated_mark") is not None
    )
    return render_template(
        "analyst_evaluation_criteria_phase.html",
        **_ctx(
            user,
            has_exercise=True,
            section_title="معايير التقييم",
            exercise=ex,
            unit=unit,
            phase_key=phase_key,
            phase_label=phase_label,
            items=items,
            eval_list_driven=eval_list_driven,
            eval_list_count=len(eval_list_titles),
            criteria_unit_level_key=unit_level_key,
            total_mark=total_mark if total_mark > 0 else None,
            ok_msg="تم حفظ جدول المرحلة." if request.args.get("ok") else "",
            **_subpage_close_ctx(
                url_for("views.analyst_hub_section", slug="evaluation-criteria")
            ),
        ),
    )


def _get_or_create_planner_bundle(
    db,
    exercise_id: int,
    phase: str,
    unit_key: str,
    unit_label: str,
) -> ExercisePlannerFlowBundle:
    phase_n = normalize_exercise_phase(phase)
    row = (
        db.query(ExercisePlannerFlowBundle)
        .filter(
            ExercisePlannerFlowBundle.exercise_id == exercise_id,
            ExercisePlannerFlowBundle.exercise_phase == phase_n,
            ExercisePlannerFlowBundle.unit_level_key == unit_key,
        )
        .first()
    )
    if row:
        return row
    row = ExercisePlannerFlowBundle(
        exercise_id=exercise_id,
        exercise_phase=phase_n,
        unit_level_key=unit_key,
        unit_level_label=(unit_label or "")[:200],
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _get_planner_bundle_if_exists(
    db,
    exercise_id: int,
    phase: str,
    unit_key: str,
) -> ExercisePlannerFlowBundle | None:
    """جلب الحزمة دون إنشاء (لصفحات العرض فقط)."""
    phase_n = normalize_exercise_phase(phase)
    return (
        db.query(ExercisePlannerFlowBundle)
        .filter(
            ExercisePlannerFlowBundle.exercise_id == exercise_id,
            ExercisePlannerFlowBundle.exercise_phase == phase_n,
            ExercisePlannerFlowBundle.unit_level_key == unit_key,
        )
        .first()
    )


_PLANNER_BUNDLE_MAX_ACTION_SLOTS = 200
_PLANNER_BUNDLE_MAX_EVENT_FLOW_SLOTS = 200


def _planner_upload_display_name(filename: str) -> str:
    base = Path(filename or "").name.strip()
    if not base:
        return ""
    return base[:500]


def _planner_bundle_event_flow_rows(
    db, bundle: ExercisePlannerFlowBundle
) -> list[ExercisePlannerFlowBundleEventFlow]:
    return (
        db.query(ExercisePlannerFlowBundleEventFlow)
        .filter(ExercisePlannerFlowBundleEventFlow.bundle_id == bundle.id)
        .order_by(ExercisePlannerFlowBundleEventFlow.slot_index)
        .all()
    )


def _migrate_legacy_bundle_event_flow(
    db, bundle: ExercisePlannerFlowBundle
) -> list[ExercisePlannerFlowBundleEventFlow]:
    """ينقل المسار القديم الوحيد على الحزمة إلى جدول ملفات المجرى إن لزم."""
    rows = _planner_bundle_event_flow_rows(db, bundle)
    legacy_rel = (bundle.event_flow_file_relpath or "").strip()
    if rows or not legacy_rel:
        return rows
    row = ExercisePlannerFlowBundleEventFlow(
        bundle_id=bundle.id,
        slot_index=1,
        title=(bundle.event_flow_title or "")[:500],
        file_relpath=legacy_rel.replace("\\", "/"),
    )
    db.add(row)
    db.flush()
    for slot_row in bundle.action_eval_slots:
        if slot_row.event_flow_item_id is None:
            slot_row.event_flow_item_id = row.id
    db.commit()
    return _planner_bundle_event_flow_rows(db, bundle)


def _sync_bundle_legacy_event_flow_columns(
    bundle: ExercisePlannerFlowBundle,
    event_rows: list[ExercisePlannerFlowBundleEventFlow],
) -> None:
    """يبقي أعمدة الحزمة القديمة متوافقة مع أول ملف مجرى (واجهة المحكم)."""
    primary = next(
        (r for r in event_rows if (r.file_relpath or "").strip()),
        None,
    )
    if primary is None:
        bundle.event_flow_file_relpath = ""
        bundle.event_flow_title = ""
        return
    bundle.event_flow_file_relpath = (primary.file_relpath or "").replace("\\", "/")
    bundle.event_flow_title = (primary.title or "")[:500]


def _sanitize_planner_bundle_orphan_event_flow_items(
    db, bundle: ExercisePlannerFlowBundle | None
) -> bool:
    if bundle is None:
        return False
    changed = False
    for row in _planner_bundle_event_flow_rows(db, bundle):
        rel = (row.file_relpath or "").strip()
        if rel and _planner_bundle_file_abspath(rel) is None:
            _unlink_planner_bundle_file(rel)
            row.file_relpath = ""
            changed = True
    if changed:
        bundle.linked_at = None
        bundle.updated_at = datetime.utcnow()
        _sync_bundle_legacy_event_flow_columns(
            bundle, _planner_bundle_event_flow_rows(db, bundle)
        )
        db.commit()
    return changed


def _planner_bundle_slot_view(
    *,
    slot_row: ExercisePlannerFlowBundleActionEval,
    bundle: ExercisePlannerFlowBundle,
    event_rows_by_id: dict[int, ExercisePlannerFlowBundleEventFlow],
) -> dict:
    rel = (slot_row.file_relpath or "").strip()
    has_file = bool(rel) and _planner_bundle_file_abspath(rel) is not None
    disp = _planner_blob_display_filename(
        stored_title=slot_row.title or "",
        relpath=rel,
        fallback=f"قائمة {slot_row.slot_index}",
    )
    ef_id = slot_row.event_flow_item_id
    ef_row = event_rows_by_id.get(int(ef_id)) if ef_id else None
    ef_name = ""
    if ef_row is not None:
        ef_name = _planner_blob_display_filename(
            stored_title=ef_row.title or "",
            relpath=ef_row.file_relpath or "",
            fallback=f"مجرى {ef_row.slot_index}",
        )
    linked = bool(bundle.linked_at) and has_file and ef_id is not None
    return {
        "id": int(slot_row.id),
        "slot_index": int(slot_row.slot_index),
        "title": disp,
        "has_file": has_file,
        "insert_label": "تم الإدراج" if has_file else "لم يُدرج",
        "link_label": "مرتبط" if linked else ("جاهز للربط" if has_file else "—"),
        "event_flow_item_id": int(ef_id) if ef_id else None,
        "event_flow_name": ef_name,
    }


def _planner_bundle_event_flow_view(
    *,
    row: ExercisePlannerFlowBundleEventFlow,
    bundle: ExercisePlannerFlowBundle,
) -> dict:
    rel = (row.file_relpath or "").strip()
    has_file = bool(rel) and _planner_bundle_file_abspath(rel) is not None
    disp = _planner_blob_display_filename(
        stored_title=row.title or "",
        relpath=rel,
        fallback=f"مجرى {row.slot_index}",
    )
    linked = bool(bundle.linked_at) and has_file
    return {
        "id": int(row.id),
        "slot_index": int(row.slot_index),
        "title": disp,
        "has_file": has_file,
        "insert_label": "تم الإدراج" if has_file else "لم يُدرج",
        "link_label": "مرتبط" if linked else ("جاهز للربط" if has_file else "—"),
    }


def _planner_bundle_sync_dilemma_count_from_slots(db, bundle: ExercisePlannerFlowBundle) -> None:
    """يُحدّث ``dilemma_count`` من عدد صفوف قوائم تقييم الإجراءات الفعلي."""
    n = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
        .count()
    )
    bundle.dilemma_count = int(n)


def _planner_bundle_judge_assignments(db, exercise_id: int, unit_level_key: str):
    return (
        db.query(JudgeTraineeAssignment)
        .filter(
            JudgeTraineeAssignment.exercise_id == exercise_id,
            JudgeTraineeAssignment.unit_level_key == unit_level_key,
        )
        .options(joinedload(JudgeTraineeAssignment.judge_user))
        .order_by(JudgeTraineeAssignment.id)
        .all()
    )


def _normalize_planner_flow_table_rows(raw_rows) -> list[dict]:
    if not isinstance(raw_rows, list):
        return []
    out: list[dict] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("kind") or "row").strip().lower()
        if kind not in ("event", "dilemma", "row"):
            kind = "row"
        if kind in ("event", "dilemma"):
            out.append({"kind": kind, "text": str(item.get("text") or "")[:4000]})
        else:
            out.append(
                {
                    "kind": "row",
                    "time": str(item.get("time") or "")[:500],
                    "description": str(item.get("description") or "")[:4000],
                    "assignee": str(item.get("assignee") or "")[:500],
                    "method": str(item.get("method") or "")[:500],
                    "reaction": str(item.get("reaction") or "")[:500],
                }
            )
    return out


def _parse_planner_flow_table_rows(payload_json: str | None) -> list[dict]:
    days, _active = _parse_planner_flow_table_days(payload_json)
    if not days:
        return []
    return list(days[0].get("rows") or [])


PLANNER_FLOW_DAY_ONE_ID = "day-1"
PLANNER_FLOW_DAY_ONE_LABEL = "اليوم/1"


def _default_planner_flow_table_days() -> tuple[list[dict], str]:
    return [
        {
            "id": PLANNER_FLOW_DAY_ONE_ID,
            "label": PLANNER_FLOW_DAY_ONE_LABEL,
            "note": "",
            "rows": [],
        }
    ], PLANNER_FLOW_DAY_ONE_ID


def _ensure_day_one_tab(days: list[dict]) -> list[dict]:
    """يضمن وجود تبويب اليوم/1 — يُعاد تلقائياً إن حُذف بالغلط."""
    out = [dict(d) for d in days if isinstance(d, dict)]
    idx = next(
        (i for i, d in enumerate(out) if d.get("id") == PLANNER_FLOW_DAY_ONE_ID),
        -1,
    )
    if idx < 0:
        idx = next(
            (
                i
                for i, d in enumerate(out)
                if (d.get("label") or "").strip() == PLANNER_FLOW_DAY_ONE_LABEL
            ),
            -1,
        )
    if idx < 0:
        return [
            {
                "id": PLANNER_FLOW_DAY_ONE_ID,
                "label": PLANNER_FLOW_DAY_ONE_LABEL,
                "note": "",
                "rows": [],
            },
            *out,
        ]
    day_one = out.pop(idx)
    if day_one.get("id") != PLANNER_FLOW_DAY_ONE_ID:
        day_one["id"] = PLANNER_FLOW_DAY_ONE_ID
    if not (day_one.get("label") or "").strip():
        day_one["label"] = PLANNER_FLOW_DAY_ONE_LABEL
    return [day_one, *out]


def _parse_planner_flow_table_days(
    payload_json: str | None,
) -> tuple[list[dict], str]:
    default_days, default_active = _default_planner_flow_table_days()
    if not (payload_json or "").strip():
        return default_days, default_active
    try:
        data = json.loads(payload_json)
    except json.JSONDecodeError:
        return default_days, default_active
    if isinstance(data, list):
        rows = _normalize_planner_flow_table_rows(data)
        return [{"id": "day-1", "label": "اليوم/1", "note": "", "rows": rows}], "day-1"
    if not isinstance(data, dict):
        return default_days, default_active
    raw_days = data.get("days")
    if not isinstance(raw_days, list) or not raw_days:
        return default_days, default_active
    out_days: list[dict] = []
    for idx, item in enumerate(raw_days):
        if not isinstance(item, dict):
            continue
        day_id = str(item.get("id") or "").strip() or f"day-{idx + 1}"
        label = str(item.get("label") or "").strip() or f"اليوم/{idx + 1}"
        rows = _normalize_planner_flow_table_rows(item.get("rows"))
        note = str(item.get("note") or "")[:4000]
        out_days.append(
            {"id": day_id[:64], "label": label[:200], "note": note, "rows": rows}
        )
    if not out_days:
        return default_days, default_active
    out_days = _ensure_day_one_tab(out_days)
    active = str(data.get("active_day_id") or "").strip()
    if not any(d["id"] == active for d in out_days):
        active = out_days[0]["id"]
    return out_days, active


def _normalize_planner_flow_table_document(payload) -> dict:
    default_days, default_active = _default_planner_flow_table_days()
    if isinstance(payload, dict) and isinstance(payload.get("days"), list):
        out_days: list[dict] = []
        for idx, item in enumerate(payload["days"]):
            if not isinstance(item, dict):
                continue
            day_id = str(item.get("id") or "").strip() or f"day-{idx + 1}"
            label = str(item.get("label") or "").strip() or f"اليوم/{idx + 1}"
            rows = _normalize_planner_flow_table_rows(item.get("rows"))
            note = str(item.get("note") or "")[:4000]
            out_days.append(
                {"id": day_id[:64], "label": label[:200], "note": note, "rows": rows}
            )
        if not out_days:
            out_days = list(default_days)
        active = str(payload.get("active_day_id") or "").strip()
        if not any(d["id"] == active for d in out_days):
            active = out_days[0]["id"]
        out_days = _ensure_day_one_tab(out_days)
        if not any(d["id"] == active for d in out_days):
            active = out_days[0]["id"]
        return {"version": 2, "active_day_id": active, "days": out_days}
    raw_rows = None
    if isinstance(payload, dict):
        raw_rows = payload.get("rows")
    elif isinstance(payload, list):
        raw_rows = payload
    rows = _normalize_planner_flow_table_rows(raw_rows)
    return {
        "version": 2,
        "active_day_id": default_active,
        "days": [{"id": "day-1", "label": "اليوم/1", "note": "", "rows": rows}],
    }


_UI_MSG_PLANNER_BUNDLE = {
    "bad_event_file": "يُقبل ملف PDF أو Word (.doc/.docx) فقط.",
    "bad_xlsx": "يُقبل ملف Excel (.xlsx) فقط.",
    "no_file": "لم يُحدد ملف.",
    "link_count": "أدرِج ملفاً واحداً على الأقل لقائمة تقييم الإجراءات (Excel) قبل الربط.",
    "link_master": "أدرج ملف مجرى الأحداث والمعاضل أولاً.",
    "link_master_disk": "مسار ملف المجرى مسجّل لكن الملف غير موجود على الخادم — أعد إدراج الملف.",
    "link_slots": "يجب إدراج ملف Excel لكل قائمة تقييم إجراءات.",
    "link_slots_disk": "يُشار إلى ملفات تقييم مسجّلة لكن أحدها غير موجود على الخادم — أعد إدراج ملفات Excel.",
    "link_ok": "تم تخصيص الربط بنجاح.",
    "upload_ok": "تم إدراج الملف.",
    "bulk_ok": "تم إدراج ملفات التقييم في الجدول — يمكنك الضغط على «تخصيص وربط» عند اكتمال الإدراج.",
    "bulk_event_ok": "تم إدراج ملف(ات) مجرى الأحداث في الجدول.",
    "bulk_event_limit": "تجاوز الحد الأقصى لعدد ملفات مجرى الأحداث في هذه الحزمة.",
    "delete_event_ok": "تم حذف ملف المجرى.",
    "delete_action_ok": "تم حذف قائمة تقييم الإجراءات.",
    "bulk_limit": "تجاوز الحد الأقصى لعدد قوائم تقييم الإجراءات في هذه الحزمة — احذف قوائم أو أنشئ حزمة أخرى.",
    "assign_ok": "تم ربط الحزمة بالمحكم.",
    "assign_clear_ok": "تمت إزالة ربط الحزمة عن المحكم.",
    "save_err": "تعذر حفظ الملف.",
    "import_flow_ok": "تم استيراد مجرى الأحداث والمعاضل من ملف Word.",
    "import_flow_bad_file": "يُقبل ملف Word (.docx) فقط.",
    "import_flow_empty": "لم يُعثر على جدول قابل للاستيراد في الملف.",
    "import_flow_fail": "تعذر استيراد ملف Word.",
}


def _build_planner_flow_bundle_page_context(
    db,
    user: User,
    ex: Exercise | None,
    *,
    readonly: bool,
    pf_workspace_endpoint: str,
    chief_hub_query_on_judge_links: bool,
    err: str,
    ok: str,
) -> dict:
    """سياق صفحة حزمة المجرى وتقييم الإجراءات (تخطيط أو عرض للإطلاع فقط)."""
    oversee = can_oversee_judge_planner_flow_materials(user)
    empty_base: dict = dict(
        has_exercise=False,
        bundle=None,
        slots=[],
        phase_key="",
        unit_key="",
        phase_options=EXERCISE_PHASE_OPTIONS,
        unit_levels=UNIT_LEVELS,
        judges=[],
        err_msg="",
        ok_msg="",
        phase_label_display="",
        unit_label_display="",
        planner_event_flow_file_ok=False,
        event_flow_rows=[],
        action_slot_rows=[],
        selected_event_flow_id=None,
        flow_table_rows=[],
        flow_table_days=[],
        flow_table_active_day_id="",
        flow_table_active_day_note="",
        readonly_mode=readonly,
        pf_workspace_endpoint=pf_workspace_endpoint,
        chief_hub_query_on_judge_links=chief_hub_query_on_judge_links,
        can_oversee_judge_planner_flow=oversee,
    )
    if ex is None:
        return {**empty_base, "exercise": None}
    phase_key = normalize_exercise_phase(request.args.get("phase") or "")
    if not phase_key and EXERCISE_PHASE_OPTIONS:
        phase_key = default_exercise_phase_key()
    err_msg = _UI_MSG_PLANNER_BUNDLE.get(err, "") if err else ""
    ok_msg = _UI_MSG_PLANNER_BUNDLE.get(ok, "") if ok else ""
    if readonly:
        err_msg = ""
        ok_msg = ""
    unit_param = (request.args.get("unit") or "").strip()
    unit_key = unit_param if unit_param else default_unit_level_key()
    unit = unit_level_row(unit_key)
    if unit is None or not phase_key or not EXERCISE_PHASE_OPTIONS:
        return {
            **empty_base,
            "has_exercise": True,
            "exercise": ex,
            "bundle": None,
            "slots": [],
            "judges": [],
            "phase_key": phase_key if phase_key else "",
            "unit_key": unit_key if unit else "",
            "phase_options": EXERCISE_PHASE_OPTIONS,
            "unit_levels": UNIT_LEVELS,
            "phase_label_display": _phase_label_ar(phase_key) if phase_key else "",
            "unit_label_display": unit["label"] if unit else "",
            "err_msg": err_msg,
            "ok_msg": ok_msg,
            "catalog_empty": not UNIT_LEVELS,
            "phases_catalog_empty": not EXERCISE_PHASE_OPTIONS,
            "event_flow_rows": [],
            "action_slot_rows": [],
            "selected_event_flow_id": None,
            "flow_table_rows": [],
            "flow_table_days": [],
            "flow_table_active_day_id": "",
            "planner_event_flow_file_ok": False,
        }
    bundle = (
        _get_planner_bundle_if_exists(db, ex.id, phase_key, unit_key)
        if readonly
        else _get_or_create_planner_bundle(
            db, ex.id, phase_key, unit_key, unit["label"]
        )
    )
    judges = _planner_bundle_judge_assignments(db, ex.id, unit_key)
    if bundle is None:
        return {
            **empty_base,
            "has_exercise": True,
            "exercise": ex,
            "bundle": None,
            "judges": judges,
            "slots": [],
            "phase_key": phase_key,
            "unit_key": unit_key,
            "phase_options": EXERCISE_PHASE_OPTIONS,
            "unit_levels": UNIT_LEVELS,
            "phase_label_display": _phase_label_ar(phase_key),
            "unit_label_display": unit["label"],
            "err_msg": err_msg,
            "ok_msg": ok_msg,
        }
    _sanitize_planner_bundle_orphan_event_flow(db, bundle)
    _sanitize_planner_bundle_orphan_event_flow_items(db, bundle)
    _migrate_legacy_bundle_event_flow(db, bundle)
    event_db_rows = _planner_bundle_event_flow_rows(db, bundle)
    _sync_bundle_legacy_event_flow_columns(bundle, event_db_rows)
    db.commit()
    slots = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
        .order_by(ExercisePlannerFlowBundleActionEval.slot_index)
        .all()
    )
    if int(bundle.dilemma_count or 0) != len(slots):
        bundle.dilemma_count = len(slots)
        bundle.updated_at = datetime.utcnow()
        db.commit()
    event_rows_by_id = {int(r.id): r for r in event_db_rows}
    event_flow_rows = [
        _planner_bundle_event_flow_view(row=r, bundle=bundle) for r in event_db_rows
    ]
    action_slot_rows = [
        _planner_bundle_slot_view(
            slot_row=s, bundle=bundle, event_rows_by_id=event_rows_by_id
        )
        for s in slots
    ]
    planner_ef_ok = any(r["has_file"] for r in event_flow_rows)
    sel_raw = (request.args.get("event_flow") or "").strip()
    selected_event_flow_id = int(sel_raw) if sel_raw.isdigit() else None
    if (
        selected_event_flow_id is not None
        and selected_event_flow_id not in event_rows_by_id
    ):
        selected_event_flow_id = None
    if selected_event_flow_id is None and event_flow_rows:
        selected_event_flow_id = event_flow_rows[0]["id"]
    flow_table_days, flow_table_active_day_id = _parse_planner_flow_table_days(
        getattr(bundle, "flow_table_json", None)
    )
    flow_table_rows = next(
        (
            d["rows"]
            for d in flow_table_days
            if d["id"] == flow_table_active_day_id
        ),
        flow_table_days[0]["rows"] if flow_table_days else [],
    )
    flow_table_active_day_note = next(
        (
            d.get("note") or ""
            for d in flow_table_days
            if d["id"] == flow_table_active_day_id
        ),
        (flow_table_days[0].get("note") or "") if flow_table_days else "",
    )
    return {
        **empty_base,
        "has_exercise": True,
        "exercise": ex,
        "bundle": bundle,
        "slots": slots,
        "phase_key": phase_key,
        "unit_key": unit_key,
        "phase_options": EXERCISE_PHASE_OPTIONS,
        "unit_levels": UNIT_LEVELS,
        "judges": judges,
        "err_msg": err_msg,
        "ok_msg": ok_msg,
        "phase_label_display": _phase_label_ar(phase_key),
        "unit_label_display": unit["label"],
        "planner_event_flow_file_ok": planner_ef_ok,
        "event_flow_rows": event_flow_rows,
        "action_slot_rows": action_slot_rows,
        "selected_event_flow_id": selected_event_flow_id,
        "flow_table_rows": flow_table_rows,
        "flow_table_days": flow_table_days,
        "flow_table_active_day_id": flow_table_active_day_id,
        "flow_table_active_day_note": flow_table_active_day_note,
    }


@bp.route("/planner/create-flow", methods=["GET"])
def planner_flow_bundle_workspace():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/planner/create-flow")
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    err = (request.args.get("err") or "").strip()
    ok = (request.args.get("ok") or "").strip()
    page_ctx = _build_planner_flow_bundle_page_context(
        db,
        user,
        ex,
        readonly=False,
        pf_workspace_endpoint="views.planner_flow_bundle_workspace",
        chief_hub_query_on_judge_links=False,
        err=err,
        ok=ok,
    )
    return render_template(
        "planner_flow_bundle.html",
        **_ctx(
            user,
            **page_ctx,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/admin/exercises/planner-flow-overview", methods=["GET"])
def admin_planner_flow_bundle_overview():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/planner-flow-overview")
    if not is_system_admin(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    page_ctx = _build_planner_flow_bundle_page_context(
        db,
        user,
        ex,
        readonly=True,
        pf_workspace_endpoint="views.admin_planner_flow_bundle_overview",
        chief_hub_query_on_judge_links=False,
        err="",
        ok="",
    )
    return render_template(
        "admin_planner_flow_bundle_overview.html",
        **_ctx(
            user,
            **page_ctx,
            hub_back_href=url_for("views.admin_exercise_judge_unit_roster"),
            hub_back_label="العودة إلى قائمة المحكمين",
        ),
    )


@bp.route("/chief-judge/planner-flow-bundle-overview", methods=["GET"])
def chief_judge_planner_flow_bundle_overview():
    """اعتماد المجرى وقوائم الإجراءات — نفس واجهة المحكم مع اختيار المحكم."""
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/chief-judge/planner-flow-bundle-overview")
    if not can_access_chief_judge_hub(user):
        abort(403)
    return redirect(
        url_for("views.judge_planner_flow_materials", from_chief_judge=1)
    )


def _redirect_planner_bundle_workspace(bundle: ExercisePlannerFlowBundle, *, err: str = "", ok: str = ""):
    kw = {"phase": bundle.exercise_phase, "unit": bundle.unit_level_key}
    if err:
        kw["err"] = err
    if ok:
        kw["ok"] = ok
    return redirect(url_for("views.planner_flow_bundle_workspace", **kw))


@bp.route("/planner/create-flow/<int:bundle_id>/upload-event-flow-bulk", methods=["POST"])
def planner_flow_bundle_upload_event_flow_bulk(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    raw_files = request.files.getlist("files")
    files = [f for f in raw_files if f and getattr(f, "filename", "").strip()]
    if not files:
        return _redirect_planner_bundle_workspace(bundle, err="no_file")
    rows = _planner_bundle_event_flow_rows(db, bundle)
    empty_count = sum(1 for r in rows if not (r.file_relpath or "").strip())
    need_new = max(0, len(files) - empty_count)
    if len(rows) + need_new > _PLANNER_BUNDLE_MAX_EVENT_FLOW_SLOTS:
        return _redirect_planner_bundle_workspace(bundle, err="bulk_event_limit")
    root = _planner_flow_bundle_root()

    def _assign_event_to_slot(
        slot_row: ExercisePlannerFlowBundleEventFlow, fstor
    ) -> str | None:
        try:
            data = fstor.read()
        except Exception:
            return "save_err"
        ext = _info_bank_event_flow_sniff_ext(data)
        if not ext:
            return "bad_event_file"
        sn = int(slot_row.slot_index)
        rel = f"{bundle.id}/event_flow_{sn}{ext}"
        dest = (root / rel).resolve()
        try:
            dest.relative_to(root)
        except ValueError:
            abort(400)
        dest.parent.mkdir(parents=True, exist_ok=True)
        try:
            dest.write_bytes(data)
        except OSError:
            return "save_err"
        old_rp = (slot_row.file_relpath or "").strip()
        old_nm = old_rp.replace("\\", "/")
        nw_nm = rel.replace("\\", "/")
        if old_nm and old_nm != nw_nm:
            _unlink_planner_bundle_file(old_rp)
        slot_row.file_relpath = nw_nm
        disp = _planner_upload_display_name(getattr(fstor, "filename", "") or "")
        slot_row.title = disp[:500] or slot_row.title
        return None

    try:
        fi = 0
        for row in rows:
            if fi >= len(files):
                break
            if not (row.file_relpath or "").strip():
                err_k = _assign_event_to_slot(row, files[fi])
                if err_k:
                    db.rollback()
                    return _redirect_planner_bundle_workspace(bundle, err=err_k)
                fi += 1
        mx = max((int(r.slot_index) for r in rows), default=0)
        while fi < len(files):
            mx += 1
            fn = getattr(files[fi], "filename", "") or ""
            title_base = _planner_upload_display_name(fn)
            new_row = ExercisePlannerFlowBundleEventFlow(
                bundle_id=bundle.id,
                slot_index=mx,
                title=title_base or f"مجرى الأحداث — {mx}",
            )
            db.add(new_row)
            db.flush()
            err_k = _assign_event_to_slot(new_row, files[fi])
            if err_k:
                db.rollback()
                return _redirect_planner_bundle_workspace(bundle, err=err_k)
            fi += 1
        bundle.linked_at = None
        bundle.updated_at = datetime.utcnow()
        event_rows = _planner_bundle_event_flow_rows(db, bundle)
        _sync_bundle_legacy_event_flow_columns(bundle, event_rows)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return _redirect_planner_bundle_workspace(bundle, ok="bulk_event_ok")


@bp.route(
    "/planner/create-flow/<int:bundle_id>/delete-event-flow/<int:item_id>",
    methods=["POST"],
)
def planner_flow_bundle_delete_event_flow(bundle_id: int, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    row = (
        db.query(ExercisePlannerFlowBundleEventFlow)
        .filter(
            ExercisePlannerFlowBundleEventFlow.bundle_id == bundle.id,
            ExercisePlannerFlowBundleEventFlow.id == int(item_id),
        )
        .first()
    )
    if row is None:
        abort(404)
    rel = (row.file_relpath or "").strip()
    if rel:
        _unlink_planner_bundle_file(rel)
    for slot_row in (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(
            ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id,
            ExercisePlannerFlowBundleActionEval.event_flow_item_id == row.id,
        )
        .all()
    ):
        slot_row.event_flow_item_id = None
    db.delete(row)
    bundle.linked_at = None
    bundle.updated_at = datetime.utcnow()
    event_rows = _planner_bundle_event_flow_rows(db, bundle)
    _sync_bundle_legacy_event_flow_columns(bundle, event_rows)
    db.commit()
    return _redirect_planner_bundle_workspace(bundle, ok="delete_event_ok")


@bp.route("/planner/create-flow/<int:bundle_id>/upload-actions-bulk", methods=["POST"])
def planner_flow_bundle_upload_actions_bulk(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    raw_files = request.files.getlist("files")
    files = [f for f in raw_files if f and getattr(f, "filename", "").strip()]
    if not files:
        return _redirect_planner_bundle_workspace(bundle, err="no_file")
    ef_raw = (request.form.get("event_flow_item_id") or "").strip()
    target_ef_id: int | None = int(ef_raw) if ef_raw.isdigit() else None
    if target_ef_id is not None:
        ef_row = (
            db.query(ExercisePlannerFlowBundleEventFlow)
            .filter(
                ExercisePlannerFlowBundleEventFlow.bundle_id == bundle.id,
                ExercisePlannerFlowBundleEventFlow.id == target_ef_id,
            )
            .first()
        )
        if ef_row is None:
            target_ef_id = None
    if target_ef_id is None:
        ef_first = (
            db.query(ExercisePlannerFlowBundleEventFlow)
            .filter(ExercisePlannerFlowBundleEventFlow.bundle_id == bundle.id)
            .order_by(ExercisePlannerFlowBundleEventFlow.slot_index)
            .first()
        )
        target_ef_id = int(ef_first.id) if ef_first else None

    rows = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
        .order_by(ExercisePlannerFlowBundleActionEval.slot_index)
        .all()
    )
    empty_count = sum(1 for r in rows if not (r.file_relpath or "").strip())
    need_new = max(0, len(files) - empty_count)
    if len(rows) + need_new > _PLANNER_BUNDLE_MAX_ACTION_SLOTS:
        return _redirect_planner_bundle_workspace(bundle, err="bulk_limit")

    root = _planner_flow_bundle_root()

    def _assign_xlsx_to_slot(slot_row: ExercisePlannerFlowBundleActionEval, fstor) -> str | None:
        """يكتب ملف xlsx على الصف؛ تعيد مفتاح رسالة خطأ أو None."""
        try:
            data = fstor.read()
        except Exception:
            return "save_err"
        if not _is_xlsx_bytes(data):
            return "bad_xlsx"
        sn = int(slot_row.slot_index)
        rel = f"{bundle.id}/action_{sn}.xlsx"
        dest = (root / rel).resolve()
        try:
            dest.relative_to(root)
        except ValueError:
            abort(400)
        dest.parent.mkdir(parents=True, exist_ok=True)
        try:
            dest.write_bytes(data)
        except OSError:
            return "save_err"
        old_rp = (slot_row.file_relpath or "").strip()
        old_nm = old_rp.replace("\\", "/")
        nw_nm = rel.replace("\\", "/")
        if old_nm and old_nm != nw_nm:
            _unlink_planner_bundle_file(old_rp)
        slot_row.file_relpath = nw_nm
        disp = _planner_upload_display_name(getattr(fstor, "filename", "") or "")
        slot_row.title = disp[:500] or slot_row.title
        if target_ef_id is not None:
            slot_row.event_flow_item_id = target_ef_id
        return None

    try:
        fi = 0
        for row in rows:
            if fi >= len(files):
                break
            if not (row.file_relpath or "").strip():
                err_k = _assign_xlsx_to_slot(row, files[fi])
                if err_k:
                    db.rollback()
                    return _redirect_planner_bundle_workspace(bundle, err=err_k)
                fi += 1

        mx = max((int(r.slot_index) for r in rows), default=0)
        while fi < len(files):
            mx += 1
            fn = getattr(files[fi], "filename", "") or ""
            title_base = _planner_upload_display_name(fn)
            new_row = ExercisePlannerFlowBundleActionEval(
                bundle_id=bundle.id,
                slot_index=mx,
                title=title_base or f"قائمة تقييم الإجراءات — {mx}",
                event_flow_item_id=target_ef_id,
            )
            db.add(new_row)
            db.flush()
            err_k = _assign_xlsx_to_slot(new_row, files[fi])
            if err_k:
                db.rollback()
                return _redirect_planner_bundle_workspace(bundle, err=err_k)
            fi += 1

        bundle.linked_at = None
        bundle.updated_at = datetime.utcnow()
        _planner_bundle_sync_dilemma_count_from_slots(db, bundle)
        db.commit()
    except Exception:
        db.rollback()
        raise

    return _redirect_planner_bundle_workspace(bundle, ok="bulk_ok")


@bp.route(
    "/planner/create-flow/<int:bundle_id>/delete-action/<int:slot_id>",
    methods=["POST"],
)
def planner_flow_bundle_delete_action(bundle_id: int, slot_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    row = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(
            ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id,
            ExercisePlannerFlowBundleActionEval.id == int(slot_id),
        )
        .first()
    )
    if row is None:
        abort(404)
    rel = (row.file_relpath or "").strip()
    if rel:
        _unlink_planner_bundle_file(rel)
    db.delete(row)
    bundle.linked_at = None
    bundle.updated_at = datetime.utcnow()
    _planner_bundle_sync_dilemma_count_from_slots(db, bundle)
    db.commit()
    return _redirect_planner_bundle_workspace(bundle, ok="delete_action_ok")


@bp.route("/planner/create-flow/<int:bundle_id>/upload-action/<int:slot>", methods=["POST"])
def planner_flow_bundle_upload_action(bundle_id: int, slot: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    row = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(
            ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id,
            ExercisePlannerFlowBundleActionEval.slot_index == int(slot),
        )
        .first()
    )
    if row is None:
        abort(404)
    f = request.files.get("file")
    if not f or not getattr(f, "filename", ""):
        return _redirect_planner_bundle_workspace(bundle, err="no_file")
    try:
        data = f.read()
    except Exception:
        return _redirect_planner_bundle_workspace(bundle, err="save_err")
    if not _is_xlsx_bytes(data):
        return _redirect_planner_bundle_workspace(bundle, err="bad_xlsx")
    root = _planner_flow_bundle_root()
    rel = f"{bundle.id}/action_{int(slot)}.xlsx"
    dest = (root / rel).resolve()
    try:
        dest.relative_to(root)
    except ValueError:
        abort(400)
    dest.parent.mkdir(parents=True, exist_ok=True)
    try:
        dest.write_bytes(data)
    except OSError:
        return _redirect_planner_bundle_workspace(bundle, err="save_err")
    old_rp = (row.file_relpath or "").strip()
    old_nm = old_rp.replace("\\", "/")
    nw_nm = rel.replace("\\", "/")
    if old_nm and old_nm != nw_nm:
        _unlink_planner_bundle_file(old_rp)
    row.file_relpath = nw_nm
    disp = _planner_upload_display_name(f.filename or "")
    row.title = disp[:500] or row.title
    bundle.linked_at = None
    bundle.updated_at = datetime.utcnow()
    db.commit()
    return _redirect_planner_bundle_workspace(bundle, ok="upload_ok")


@bp.route("/planner/create-flow/<int:bundle_id>/link", methods=["POST"])
def planner_flow_bundle_link(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    event_rows = _migrate_legacy_bundle_event_flow(db, bundle)
    event_with_file = [
        r
        for r in event_rows
        if (r.file_relpath or "").strip()
        and _planner_bundle_file_abspath(r.file_relpath) is not None
    ]
    if not event_with_file:
        return _redirect_planner_bundle_workspace(bundle, err="link_master")
    slots = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
        .order_by(ExercisePlannerFlowBundleActionEval.slot_index)
        .all()
    )
    if len(slots) < 1:
        return _redirect_planner_bundle_workspace(bundle, err="link_count")
    for s in slots:
        rel = (s.file_relpath or "").strip()
        if not rel:
            return _redirect_planner_bundle_workspace(bundle, err="link_slots")
        if _planner_bundle_file_abspath(rel) is None:
            return _redirect_planner_bundle_workspace(bundle, err="link_slots_disk")
        if s.event_flow_item_id is None and event_with_file:
            s.event_flow_item_id = int(event_with_file[0].id)
    _sync_bundle_legacy_event_flow_columns(bundle, event_rows)
    bundle.linked_at = datetime.utcnow()
    bundle.updated_at = datetime.utcnow()
    db.commit()
    return _redirect_planner_bundle_workspace(bundle, ok="link_ok")


@bp.route("/planner/create-flow/<int:bundle_id>/save-flow-table", methods=["POST"])
def planner_flow_bundle_save_flow_table(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_access_planner_hub(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        return jsonify({"ok": False, "error": "not_found"}), 404
    payload = request.get_json(silent=True)
    doc = _normalize_planner_flow_table_document(payload)
    bundle.flow_table_json = json.dumps(doc, ensure_ascii=False)
    bundle.updated_at = datetime.utcnow()
    db.commit()
    active_rows = next(
        (d["rows"] for d in doc["days"] if d["id"] == doc["active_day_id"]),
        [],
    )
    return jsonify(
        {
            "ok": True,
            "day_count": len(doc["days"]),
            "row_count": len(active_rows),
        }
    )


@bp.route("/planner/create-flow/<int:bundle_id>/import-flow-docx", methods=["POST"])
def planner_flow_bundle_import_flow_docx(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_access_planner_hub(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    from app.planner_flow_docx_import import parse_planner_flow_docx_bytes

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        return jsonify({"ok": False, "error": "not_found"}), 404

    up = request.files.get("file")
    if up is None or not (up.filename or "").strip():
        return jsonify({"ok": False, "error": "no_file"}), 400
    raw = up.read()
    if not raw:
        return jsonify({"ok": False, "error": "no_file"}), 400
    if not _is_docx_bytes(raw):
        return jsonify({"ok": False, "error": "bad_docx"}), 400

    parsed = parse_planner_flow_docx_bytes(raw)
    if not parsed.get("ok"):
        err = parsed.get("error") or "import_fail"
        return jsonify({"ok": False, "error": err, "warnings": parsed.get("warnings") or []}), 400

    day_id = (request.form.get("day_id") or "").strip()
    days, active_id = _parse_planner_flow_table_days(
        getattr(bundle, "flow_table_json", None)
    )
    if day_id and any(d["id"] == day_id for d in days):
        target_id = day_id
    else:
        target_id = active_id or (days[0]["id"] if days else "day-1")

    updated = False
    for d in days:
        if d["id"] == target_id:
            d["note"] = parsed.get("note") or ""
            d["rows"] = parsed.get("rows") or []
            updated = True
            break
    if not updated:
        days.append(
            {
                "id": target_id,
                "label": f"اليوم/{len(days) + 1}",
                "note": parsed.get("note") or "",
                "rows": parsed.get("rows") or [],
            }
        )

    days = _ensure_day_one_tab(days)
    doc = {"version": 2, "active_day_id": target_id, "days": days}
    bundle.flow_table_json = json.dumps(doc, ensure_ascii=False)
    bundle.updated_at = datetime.utcnow()
    db.commit()

    return jsonify(
        {
            "ok": True,
            "active_day_id": target_id,
            "days": days,
            "note": parsed.get("note") or "",
            "row_count": len(parsed.get("rows") or []),
            "warnings": parsed.get("warnings") or [],
        }
    )


@bp.route("/planner/create-flow/<int:bundle_id>/assign-judge", methods=["POST"])
def planner_flow_bundle_assign_judge(bundle_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    bundle = db.get(ExercisePlannerFlowBundle, bundle_id)
    if ex is None or bundle is None or bundle.exercise_id != ex.id:
        abort(404)
    raw_j = (request.form.get("judge_user_id") or "").strip()
    if not raw_j.isdigit():
        abort(400)
    judge_uid = int(raw_j)
    clear = (request.form.get("clear") or "").strip() == "1"
    ja = (
        db.query(JudgeTraineeAssignment)
        .filter(
            JudgeTraineeAssignment.exercise_id == ex.id,
            JudgeTraineeAssignment.judge_user_id == judge_uid,
        )
        .first()
    )
    if ja is None or ja.unit_level_key != bundle.unit_level_key:
        abort(400)
    ja.planner_flow_bundle_id = None if clear else bundle.id
    db.commit()
    return _redirect_planner_bundle_workspace(
        bundle,
        ok="assign_clear_ok" if clear else "assign_ok",
    )


def _planner_flow_oversee_judge_id_from_request() -> int | None:
    raw = (request.args.get("judge_user_id") or "").strip()
    if raw.isdigit():
        return int(raw)
    from flask import session

    sid = session.get("planner_flow_oversee_judge_id")
    if sid is not None:
        try:
            return int(sid)
        except (TypeError, ValueError):
            pass
    return None


def _planner_flow_is_readonly_oversee(
    user: User, *, oversee_judge_id: int | None = None
) -> bool:
    """عرض فقط عند إطلاع إدارة النظام/كبير المحكمين على حزمة محكم آخر (وليس حزمتهم)."""
    if not can_oversee_judge_planner_flow_materials(user):
        return False
    jid = (
        oversee_judge_id
        if oversee_judge_id is not None
        else _planner_flow_oversee_judge_id_from_request()
    )
    if jid is None:
        return True
    return int(jid) != int(getattr(user, "id", 0) or 0)


def _planner_flow_action_lists_editable(user: User) -> bool:
    """المحكم وكبير المحكمين (وإدارة النظام) يقيّمون ويحفظون ويعتمدون قوائم إجراءات الحزمة."""
    return bool(can_save_evaluation_results(user))


def _planner_flow_eval_list_viewer_ctx(user: User, saved) -> dict:
    """سياق واجهة تقييم قائمة إجراءات الحزمة — محكم أو كبير محكمين (اعتماد/إعادة)."""
    wf = dict(_eval_list_viewer_ctx(user, saved))
    wf["eval_chief_next_step_hint"] = True
    if _planner_flow_is_readonly_oversee(user):
        wf["eval_can_edit"] = False
        wf["show_eval_approve"] = False
        return wf
    if not _planner_flow_action_lists_editable(user):
        return wf
    return {
        **wf,
        "eval_can_edit": True,
        "show_eval_approve": bool(
            can_approve_evaluation_results(user)
            and eval_judge_can_approve(saved)
            and _evaluation_saved_allows_judge_approve(saved)
        ),
        "show_chief_approve": bool(
            can_chief_approve_evaluation_results(user) and eval_chief_can_approve(saved)
        ),
        "show_chief_reopen": bool(
            can_chief_reopen_evaluation_for_judge(user) and eval_chief_can_reopen(saved)
        ),
    }


def _planner_flow_materials_query_kwargs(viewer: User) -> dict[str, int]:
    if not can_oversee_judge_planner_flow_materials(viewer):
        return {}
    jid = _planner_flow_oversee_judge_id_from_request()
    if jid is not None:
        return {"judge_user_id": jid}
    return {}


_JUDGE_PF_TAB_QUERY = "tab"
_JUDGE_PF_TAB_EVENT = "pf-event"
_JUDGE_PF_TAB_EVAL_LISTS = "pf-eval"


def _judge_planner_flow_materials_active_tab() -> str:
    tab = (request.args.get(_JUDGE_PF_TAB_QUERY) or "").strip()
    if tab in (_JUDGE_PF_TAB_EVENT, _JUDGE_PF_TAB_EVAL_LISTS):
        return tab
    return _JUDGE_PF_TAB_EVENT


def _planner_flow_oversee_assignment_rows(db, ex: Exercise) -> list[dict]:
    """محكمون لديهم حزمة مجرى مربوطة — لاختيار الإطلاع."""
    rows = (
        db.query(JudgeTraineeAssignment)
        .filter(
            JudgeTraineeAssignment.exercise_id == ex.id,
            JudgeTraineeAssignment.planner_flow_bundle_id.isnot(None),
        )
        .options(joinedload(JudgeTraineeAssignment.judge_user))
        .options(joinedload(JudgeTraineeAssignment.planner_flow_bundle))
        .order_by(JudgeTraineeAssignment.unit_level_key, JudgeTraineeAssignment.id)
        .all()
    )
    out: list[dict] = []
    for ja in rows:
        ju = ja.judge_user
        bundle = ja.planner_flow_bundle
        jlabel = (
            (getattr(ju, "full_name", None) or "").strip()
            or (getattr(ju, "username", None) or "").strip()
            or f"محكم #{ja.judge_user_id}"
        )
        uk = (ja.unit_level_key or "").strip()
        link_kw: dict = {"judge_user_id": int(ja.judge_user_id)}
        if _request_from_chief_judge_hub():
            link_kw["from_chief_judge"] = 1
        bundle_label = "—"
        if bundle is not None:
            bundle_label = (
                (bundle.event_flow_title or "").strip()
                or label_for_unit_level_key(bundle.unit_level_key)
                or f"حزمة #{bundle.id}"
            )
        out.append(
            {
                "judge_user_id": int(ja.judge_user_id),
                "judge_label": jlabel,
                "trainee_name": (ja.trainee_name or "").strip() or "—",
                "unit_label": label_for_unit_level_key(uk) or uk or "—",
                "bundle_label": bundle_label,
                "bundle_linked": bool(
                    bundle and getattr(bundle, "linked_at", None)
                ),
                "view_href": url_for(
                    "views.judge_planner_flow_materials",
                    **link_kw,
                ),
            }
        )
    return out


def _judge_assigned_planner_bundle(
    db,
    user: User,
    ex: Exercise | None,
    *,
    judge_user_id: int | None = None,
) -> ExercisePlannerFlowBundle | None:
    if ex is None:
        return None
    if can_oversee_judge_planner_flow_materials(user):
        jid = (
            judge_user_id
            if judge_user_id is not None
            else _planner_flow_oversee_judge_id_from_request()
        )
        if jid is None:
            return None
        ja = (
            db.query(JudgeTraineeAssignment)
            .filter(
                JudgeTraineeAssignment.exercise_id == ex.id,
                JudgeTraineeAssignment.judge_user_id == int(jid),
            )
            .first()
        )
    else:
        ja = _judge_assignment_for_current_exercise(db, user, ex)
    if ja is None or not ja.planner_flow_bundle_id:
        return None
    b = db.get(ExercisePlannerFlowBundle, ja.planner_flow_bundle_id)
    if b is None or b.exercise_id != ex.id:
        return None
    return b


@bp.route("/judge/planner-flow-materials", methods=["GET"])
def judge_planner_flow_materials():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/judge/planner-flow-materials")
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    oversee = can_oversee_judge_planner_flow_materials(user)
    oversee_jid = _planner_flow_oversee_judge_id_from_request() if oversee else None
    if oversee and ex is not None and oversee_jid is None:
        return render_template(
            "judge_planner_flow_materials.html",
            **_ctx(
                user,
                has_exercise=True,
                exercise=ex,
                bundle=None,
                slots=[],
                planner_eval_rows=[],
                planner_flow_unit_label="",
                planner_event_flow_file_ok=False,
                planner_event_flow_display_name="",
                planner_flow_oversee_picker=True,
                planner_flow_oversee_rows=_planner_flow_oversee_assignment_rows(db, ex),
                planner_flow_view_only=True,
                planner_flow_viewing_judge_label="",
                **_hub_back_ctx_for_request_path(),
            ),
        )
    bundle = _judge_assigned_planner_bundle(
        db, user, ex, judge_user_id=oversee_jid
    )
    viewing_judge_label = ""
    if oversee and oversee_jid is not None:
        ja_view = (
            db.query(JudgeTraineeAssignment)
            .filter(
                JudgeTraineeAssignment.exercise_id == ex.id,
                JudgeTraineeAssignment.judge_user_id == int(oversee_jid),
            )
            .options(joinedload(JudgeTraineeAssignment.judge_user))
            .first()
            if ex is not None
            else None
        )
        if ja_view and ja_view.judge_user:
            ju = ja_view.judge_user
            viewing_judge_label = (
                (getattr(ju, "full_name", None) or "").strip()
                or (getattr(ju, "username", None) or "").strip()
                or f"محكم #{oversee_jid}"
            )
        else:
            viewing_judge_label = f"محكم #{oversee_jid}"
    pf_qs = _planner_flow_materials_query_kwargs(user)
    _sanitize_planner_bundle_orphan_event_flow(db, bundle)
    _sanitize_planner_bundle_orphan_event_flow_items(db, bundle)
    slots = []
    planner_eval_rows: list[dict] = []
    unit_label_pf = ""
    planner_event_flow_file_ok = False
    planner_event_flow_display_name = ""
    if bundle:
        _migrate_legacy_bundle_event_flow(db, bundle)
        event_db_rows = _planner_bundle_event_flow_rows(db, bundle)
        _sync_bundle_legacy_event_flow_columns(bundle, event_db_rows)
        db.commit()
        primary_ef = next((r for r in event_db_rows if (r.file_relpath or "").strip()), None)
        if primary_ef is None:
            ev_rel_j = (bundle.event_flow_file_relpath or "").strip()
        else:
            ev_rel_j = (primary_ef.file_relpath or "").strip()
        planner_event_flow_file_ok = bool(
            ev_rel_j and _planner_bundle_file_abspath(ev_rel_j) is not None
        )
        if primary_ef is not None:
            planner_event_flow_display_name = _planner_blob_display_filename(
                stored_title=primary_ef.title or "",
                relpath=ev_rel_j,
                fallback="ملف المجرى",
            )
        else:
            planner_event_flow_display_name = _planner_blob_display_filename(
                stored_title=bundle.event_flow_title or "",
                relpath=ev_rel_j,
                fallback="ملف المجرى",
            )
        slots = (
            db.query(ExercisePlannerFlowBundleActionEval)
            .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
            .order_by(ExercisePlannerFlowBundleActionEval.slot_index)
            .all()
        )
        unit_label_pf = label_for_unit_level_key(bundle.unit_level_key) or (
            bundle.unit_level_label or ""
        ).strip() or bundle.unit_level_key
        if ex is not None:
            for slot_row in slots:
                s_canon = _planner_bundle_eval_canonical_saved(db, ex.id, slot_row.id)
                has_xlsx = bool((slot_row.file_relpath or "").strip())
                title_s = _planner_blob_display_filename(
                    stored_title=slot_row.title or "",
                    relpath=slot_row.file_relpath or "",
                    fallback=f"قائمة {slot_row.slot_index}",
                )
                planner_eval_rows.append(
                    build_planner_flow_eval_row(
                        slot_index=int(slot_row.slot_index),
                        item_title=title_s,
                        saved=s_canon,
                        exercise=ex,
                        open_href=(
                            url_for(
                                "views.judge_planner_flow_materials_action_evaluate",
                                slot=slot_row.slot_index,
                                **pf_qs,
                            )
                            if has_xlsx
                            else ""
                        ),
                        dt_fallback=getattr(slot_row, "created_at", None),
                    )
                )
    if oversee and ex is not None and oversee_jid is not None and bundle is None:
        abort(404)
    from flask import session

    if oversee and oversee_jid is not None:
        session["planner_flow_oversee_judge_id"] = int(oversee_jid)
    else:
        session.pop("planner_flow_oversee_judge_id", None)
    hub_nav_href = url_for("views.judge_planner_flow_materials", **_role_hub_preserve_link_kwargs())
    if not oversee:
        hub_nav_href = ""
    return render_template(
        "judge_planner_flow_materials.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            bundle=bundle,
            slots=slots,
            planner_eval_rows=planner_eval_rows,
            planner_flow_unit_label=unit_label_pf,
            planner_event_flow_file_ok=planner_event_flow_file_ok,
            planner_event_flow_display_name=planner_event_flow_display_name,
            planner_event_flow_relpath=ev_rel_j if bundle else "",
            planner_flow_oversee_picker=False,
            planner_flow_oversee_rows=[],
            planner_flow_view_only=_planner_flow_is_readonly_oversee(user, oversee_judge_id=oversee_jid),
            planner_flow_viewing_judge_label=viewing_judge_label,
            planner_flow_url_kwargs=pf_qs,
            active_pf_tab=_judge_planner_flow_materials_active_tab(),
            hub_nav_href=hub_nav_href,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/planner-flow-materials/event-flow", methods=["GET"])
def judge_planner_flow_materials_event_flow():
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    oversee_jid = (
        _planner_flow_oversee_judge_id_from_request()
        if can_oversee_judge_planner_flow_materials(user)
        else None
    )
    bundle = _judge_assigned_planner_bundle(
        db, user, ex, judge_user_id=oversee_jid
    )
    if bundle is None:
        abort(404)
    _migrate_legacy_bundle_event_flow(db, bundle)
    event_rows = _planner_bundle_event_flow_rows(db, bundle)
    primary = next((r for r in event_rows if (r.file_relpath or "").strip()), None)
    rel = (
        (primary.file_relpath or "").strip()
        if primary is not None
        else (bundle.event_flow_file_relpath or "").strip()
    )
    if not rel:
        abort(404)
    path = _planner_bundle_file_abspath(rel)
    if path is None:
        abort(404)
    mt = _mimetype_info_bank_event_flow(path)
    return send_file(path, mimetype=mt, as_attachment=False)


@bp.route("/judge/planner-flow-materials/action/<int:slot>", methods=["GET"])
def judge_planner_flow_materials_action(slot: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    oversee_jid = (
        _planner_flow_oversee_judge_id_from_request()
        if can_oversee_judge_planner_flow_materials(user)
        else None
    )
    bundle = _judge_assigned_planner_bundle(
        db, user, ex, judge_user_id=oversee_jid
    )
    if bundle is None:
        abort(404)
    row = (
        db.query(ExercisePlannerFlowBundleActionEval)
        .filter(
            ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id,
            ExercisePlannerFlowBundleActionEval.slot_index == int(slot),
        )
        .first()
    )
    if row is None or not (row.file_relpath or "").strip():
        abort(404)
    path = _planner_bundle_file_abspath(row.file_relpath)
    if path is None:
        abort(404)
    return send_file(
        path,
        mimetype=_mimetype_for_eval_list_file(path),
        as_attachment=True,
        download_name=path.name or f"action_eval_{slot}.xlsx",
    )


@bp.route(
    "/judge/planner-flow-materials/action/<int:slot>/evaluate",
    methods=["GET"],
)
def judge_planner_flow_materials_action_evaluate(slot: int):
    """واجهة تقييم تفاعلية لقائمة إجراءات الحزمة (مثل صفحة قوائم التقييم العامة)."""
    user = get_current_user_optional()
    if not user:
        return redirect(
            f"/login?next=/judge/planner-flow-materials/action/{int(slot)}/evaluate"
        )
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    pair = _judge_planner_flow_action_bundle_row(db, user, ex, slot)
    if pair is None or ex is None:
        abort(404)
    bundle, action_row = pair
    unit_key = bundle.unit_level_key
    unit = unit_level_row(unit_key)
    unit_label_pf = (
        (unit.get("label") or "").strip()
        if unit
        else (bundle.unit_level_label or "").strip() or unit_key
    )
    path = _planner_bundle_file_abspath(action_row.file_relpath)
    pf_qs = _planner_flow_materials_query_kwargs(user)
    if path is None:
        return redirect(url_for("views.judge_planner_flow_materials", **pf_qs))
    ev = _evaluation_sheet_view_context(path)

    saved_payload = {}
    saved_updated_at = None
    saved_row_id = None
    canon = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)

    def _load_payload(sr: PlannerFlowBundleEvalSavedResult | None) -> dict:
        if not sr or not (sr.payload_json or "").strip():
            return {}
        try:
            p = json.loads(sr.payload_json)
        except Exception:
            return {}
        return p if isinstance(p, dict) else {}

    if canon is not None:
        saved_payload = _load_payload(canon)
        saved_updated_at = canon.updated_at
        saved_row_id = canon.id
    saved_payload = _saved_payload_aligned_with_eval_rows(saved_payload, ev.get("eval_rows"))
    wf = _planner_flow_eval_list_viewer_ctx(user, canon)

    item_title = _planner_blob_display_filename(
        stored_title=action_row.title or "",
        relpath=action_row.file_relpath or "",
        fallback=f"قائمة تقييم إجراءات — {slot}",
    ).strip()
    eval_save_url = url_for(
        "views.judge_planner_flow_materials_action_save_results",
        slot=int(slot),
        **pf_qs,
    )
    eval_approve_url = url_for(
        "views.judge_planner_flow_materials_action_approve",
        slot=int(slot),
        **pf_qs,
    )
    eval_chief_approve_url = url_for(
        "views.judge_planner_flow_materials_action_chief_approve",
        slot=int(slot),
        **pf_qs,
    )
    eval_chief_reopen_url = url_for(
        "views.judge_planner_flow_materials_action_chief_reopen",
        slot=int(slot),
        **pf_qs,
    )
    eval_close_href = url_for(
        "views.judge_planner_flow_materials",
        tab=_JUDGE_PF_TAB_EVAL_LISTS,
        **pf_qs,
    )

    shown_date = getattr(ex, "planned_start", None) or getattr(ex, "created_at", None)
    commander_name = "—"
    commander_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if commander_row is not None:
        commander_name = (commander_row.full_name or "").strip() or commander_name

    judge_name = (
        (getattr(user, "full_name", "") or "").strip()
        or (getattr(user, "username", "") or "").strip()
        or f"محكم #{getattr(user, 'id', '')}"
    )
    judge_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if judge_row is not None:
        judge_name = (judge_row.full_name or "").strip() or judge_name

    return render_template(
        "judge_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=action_row.id,
            evaluation_item_id=action_row.id,
            saved_row_id=saved_row_id,
            item_title=item_title,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            **wf,
            eval_close_href=eval_close_href,
            subpage_close_fallback=eval_close_href,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=ex,
                list_item_id=None,
                bundle_action_eval_id=int(action_row.id),
                eval_can_edit=bool(wf.get("eval_can_edit")),
            ),
            unit_label=unit_label_pf or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url=eval_save_url,
            eval_approve_url=eval_approve_url,
            eval_chief_approve_url=eval_chief_approve_url,
            eval_chief_reopen_url=eval_chief_reopen_url,
            eval_approve_incomplete=request.args.get("eval_approve_incomplete", type=int)
            == 1,
            eval_readonly_notice=request.args.get("eval_readonly", type=int) == 1,
            eval_action_failed=request.args.get("eval_action_failed", type=int) == 1,
            eval_saved_notice=request.args.get("eval_saved", type=int) == 1,
        ),
    )


@bp.route(
    "/judge/planner-flow-materials/action/<int:slot>/save-results",
    methods=["GET", "POST"],
)
def judge_planner_flow_materials_action_save_results(slot: int):
    user = get_current_user_optional()
    pf_qs = _planner_flow_materials_query_kwargs(user)
    eval_url = url_for(
        "views.judge_planner_flow_materials_action_evaluate",
        slot=int(slot),
        **pf_qs,
    )
    if request.method == "GET":
        if not user:
            return redirect(
                f"/login?next=/judge/planner-flow-materials/action/{int(slot)}/evaluate"
            )
        return redirect(eval_url)
    if not user:
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    pair = _judge_planner_flow_action_bundle_row(db, user, ex, slot)
    if pair is None or ex is None:
        abort(404)
    bundle, action_row = pair
    if _planner_flow_is_readonly_oversee(user):
        saved_ro = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)
        if (
            saved_ro is not None
            and can_chief_reopen_evaluation_for_judge(user)
            and eval_chief_can_reopen(saved_ro)
        ):
            apply_chief_reopen(saved_ro)
            from app.notifications_service import (
                notify_evaluation_reopened_by_chief_judge,
            )

            unit_key = bundle.unit_level_key
            unit_label = label_for_unit_level_key(unit_key) or unit_key
            item_title = _planner_blob_display_filename(
                stored_title=action_row.title or "",
                relpath=action_row.file_relpath or "",
                fallback=f"قائمة {action_row.slot_index}",
            )
            notify_evaluation_reopened_by_chief_judge(
                db,
                exercise_id=int(ex.id),
                unit_key=unit_key,
                unit_label=unit_label,
                item_title=item_title,
                item_id=int(action_row.id),
                saved_by_user_id=getattr(saved_ro, "saved_by_id", None),
                exclude_user_id=getattr(user, "id", None),
            )
            db.commit()
            return redirect(
                url_for(
                    "views.judge_planner_flow_materials",
                    tab=_JUDGE_PF_TAB_EVAL_LISTS,
                    pf_reopened=1,
                    **pf_qs,
                )
            )
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                eval_readonly=1,
                **pf_qs,
            )
        )
    raw = (request.form.get("payload_json") or "").strip()
    if not raw:
        return redirect(eval_url)
    if len(raw) > 250_000:
        abort(400)
    _planner_bundle_eval_commit_payload_save(
        db,
        user=user,
        action_row=action_row,
        bundle=bundle,
        current_exercise=ex,
        raw=raw,
    )
    return redirect(
        url_for(
            "views.judge_planner_flow_materials_action_evaluate",
            slot=int(slot),
            eval_saved=1,
            **pf_qs,
        )
    )


@bp.route(
    "/judge/planner-flow-materials/action/<int:slot>/approve",
    methods=["POST"],
)
def judge_planner_flow_materials_action_approve(slot: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if _planner_flow_is_readonly_oversee(user):
        abort(403)
    if not can_approve_evaluation_results(user):
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    pair = _judge_planner_flow_action_bundle_row(db, user, ex, slot)
    if pair is None or ex is None:
        abort(404)
    _bundle, action_row = pair

    pf_qs = _planner_flow_materials_query_kwargs(user)
    saved = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)
    if saved is None or not (saved.payload_json or "").strip():
        abort(400)
    if not eval_judge_can_approve(saved):
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                **pf_qs,
            )
        )
    rows = _parse_saved_eval_rows(saved.payload_json)
    if _evaluation_payload_has_empty_acquired_for_approve(rows):
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                eval_approve_incomplete=1,
                **pf_qs,
            )
        )
    if not _evaluation_saved_allows_judge_approve(saved):
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                eval_approve_grade_blocked=1,
                **pf_qs,
            )
        )
    apply_judge_approve(saved, getattr(user, "id", None))
    db.commit()
    return redirect(
        url_for(
            "views.judge_planner_flow_materials_action_evaluate",
            slot=int(slot),
            **pf_qs,
        )
    )


@bp.route(
    "/judge/planner-flow-materials/action/<int:slot>/chief-approve",
    methods=["POST"],
)
def judge_planner_flow_materials_action_chief_approve(slot: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_chief_approve_evaluation_results(user):
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    pair = _judge_planner_flow_action_bundle_row(db, user, ex, slot)
    if pair is None or ex is None:
        abort(404)
    _bundle, action_row = pair
    pf_qs = _planner_flow_materials_query_kwargs(user)
    saved = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)
    if saved is None or not eval_chief_can_approve(saved):
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                eval_action_failed=1,
                **pf_qs,
            )
        )
    apply_chief_approve(saved, getattr(user, "id", None))
    db.commit()
    return redirect(
        url_for(
            "views.judge_planner_flow_materials",
            tab=_JUDGE_PF_TAB_EVAL_LISTS,
            pf_chief_approved=1,
            **pf_qs,
        )
    )


@bp.route(
    "/judge/planner-flow-materials/action/<int:slot>/chief-reopen",
    methods=["POST"],
)
def judge_planner_flow_materials_action_chief_reopen(slot: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_chief_reopen_evaluation_for_judge(user):
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    pair = _judge_planner_flow_action_bundle_row(db, user, ex, slot)
    if pair is None or ex is None:
        abort(404)
    bundle, action_row = pair
    pf_qs = _planner_flow_materials_query_kwargs(user)
    saved = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)
    if saved is None or not eval_chief_can_reopen(saved):
        return redirect(
            url_for(
                "views.judge_planner_flow_materials_action_evaluate",
                slot=int(slot),
                eval_action_failed=1,
                **pf_qs,
            )
        )
    apply_chief_reopen(saved)
    from app.notifications_service import notify_evaluation_reopened_by_chief_judge

    unit_key = bundle.unit_level_key
    unit_label = label_for_unit_level_key(unit_key) or unit_key
    item_title = _planner_blob_display_filename(
        stored_title=action_row.title or "",
        relpath=action_row.file_relpath or "",
        fallback=f"قائمة {action_row.slot_index}",
    )
    notify_evaluation_reopened_by_chief_judge(
        db,
        exercise_id=int(ex.id),
        unit_key=unit_key,
        unit_label=unit_label,
        item_title=item_title,
        item_id=int(action_row.id),
        saved_by_user_id=getattr(saved, "saved_by_id", None),
        exclude_user_id=getattr(user, "id", None),
    )
    db.commit()
    return redirect(
        url_for(
            "views.judge_planner_flow_materials",
            tab=_JUDGE_PF_TAB_EVAL_LISTS,
            pf_reopened=1,
            **pf_qs,
        )
    )


# مساحة التخطيط — عناصر الشريط (المعرّف، العنوان، أيقونة Font Awesome)
PLANNER_HUB_ITEMS: tuple[tuple[str, str, str], ...] = (
    ("new-flow", "المجرى وتقييم الإجراءات", "fa-diagram-project"),
    ("new-evaluation-list", "إنشاء قائمة تقييم", "fa-file-circle-plus"),
    ("evaluation-lists", "قوائم التقييم — إدخال النتائج", "fa-file-excel"),
    ("incomplete-tasks", "موقف المهام غير المكتملة", "fa-hourglass-half"),
    ("battle-overview", "الصورة العامة للمعركة", "fa-map"),
    ("assign-task", "إسناد مهمة جديدة", "fa-user-plus"),
)
PLANNER_HUB_SLUGS: dict[str, str] = {s: t for s, t, _ in PLANNER_HUB_ITEMS}


@bp.route("/planner")
def planner_hub():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/planner")
    if not can_access_planner_hub(user):
        abort(403)
    hub_items = [{"slug": s, "title_ar": t, "icon": ic} for s, t, ic in PLANNER_HUB_ITEMS]
    return render_template(
        "planner_hub.html",
        **_ctx(
            user,
            hub_items=hub_items,
        ),
    )


@bp.route("/planner/<slug>")
def planner_hub_section(slug: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/planner/{slug}")
    if not can_access_planner_hub(user):
        abort(403)
    slug_norm = (slug or "").strip().lower()
    if slug_norm == "new-flow":
        return redirect(url_for("views.planner_flow_bundle_workspace"))
    if slug_norm == "new-evaluation-list":
        return redirect(url_for("views.admin_evaluation_lists_home"))
    if slug_norm == "evaluation-lists":
        return redirect(url_for("views.planner_evaluation_lists_home"))
    if slug_norm == "visual-documentation":
        return redirect(url_for("views.visual_documentation", from_planner=1))
    title = PLANNER_HUB_SLUGS.get(slug_norm)
    if not title:
        abort(404)
    if slug_norm == "incomplete-tasks":
        return _render_incomplete_evaluations_page(
            user,
            section_title=title,
            section_icon="fa-hourglass-half",
            role="planner",
        )
    return render_template(
        "planner_section_placeholder.html",
        **_ctx(
            user,
            section_title=title,
            section_slug=slug,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/planner/evaluation-lists", methods=["GET"])
def planner_evaluation_lists_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/planner/evaluation-lists")
    if not can_access_planner_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    units = list(UNIT_LEVELS)
    unit_rows = evaluation_unit_home_rows(db, ex, units)
    return render_template(
        "judge_evaluation_lists_home.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            unit_levels=units,
            unit_rows=unit_rows,
            unit_totals=evaluation_unit_home_totals(unit_rows),
            unit_list_endpoint="views.planner_evaluation_lists",
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/planner/evaluation-lists/<unit_key>", methods=["GET"])
def planner_evaluation_lists(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/planner/evaluation-lists/{unit_key}")
    if not can_access_planner_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    if ex is None:
        return redirect("/planner/evaluation-lists")
    items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id, EvaluationListPdfItem.unit_level_key == unit_key)
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    item_ids = [int(it.id) for it in items]
    canonical_by_item = _evaluation_canonical_map_for_items(db, ex.id, item_ids)

    evaluation_lists_rows: list[dict] = []
    for it in items:
        s = canonical_by_item.get(int(it.id))
        evaluation_lists_rows.append(
            build_evaluation_list_row(
                item=it,
                saved=s,
                exercise=ex,
                open_href=url_for(
                    "views.planner_evaluation_list_file_viewer",
                    unit_key=unit_key,
                    item_id=it.id,
                ),
            )
        )
    return render_template(
        "judge_evaluation_lists.html",
        **_ctx(
            user,
            exercise=ex,
            unit=unit,
            unit_key=unit_key,
            items=items,
            evaluation_lists_rows=evaluation_lists_rows,
            eval_lists_parent_href=url_for("views.planner_evaluation_lists_home"),
            subpage_close_fallback=url_for("views.planner_evaluation_lists_home"),
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/planner/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def planner_evaluation_list_file_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/planner/evaluation-lists/{unit_key}/view/{item_id}")
    if not can_access_planner_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or current_exercise is None or row.exercise_id != current_exercise.id:
        abort(404)
    list_url = url_for("views.planner_evaluation_lists", unit_key=unit_key)
    if not (row.pdf_relpath or "").strip():
        return redirect(list_url)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        return redirect(list_url)
    ev = _evaluation_sheet_view_context(fspath)

    saved_payload = {}
    saved_updated_at = None
    saved_is_approved = False
    saved_approved_at = None
    saved_row_id = None
    canon = _evaluation_canonical_saved_row(db, current_exercise.id, row.id)

    def _load_payload(sr: EvaluationListSavedResult | None) -> dict:
        if not sr or not (sr.payload_json or "").strip():
            return {}
        try:
            p = json.loads(sr.payload_json)
        except Exception:
            return {}
        return p if isinstance(p, dict) else {}

    if canon is not None:
        saved_payload = _load_payload(canon)
        saved_updated_at = canon.updated_at
        saved_is_approved = bool(getattr(canon, "is_approved", False))
        saved_approved_at = getattr(canon, "approved_at", None)
        saved_row_id = canon.id
    saved_payload = _saved_payload_aligned_with_eval_rows(saved_payload, ev.get("eval_rows"))

    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = getattr(current_exercise, "planned_start", None) or getattr(current_exercise, "created_at", None)

    commander_name = "—"
    commander_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if commander_row is not None:
        commander_name = (commander_row.full_name or "").strip() or commander_name

    judge_name = "—"
    judge_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if judge_row is not None:
        judge_name = (judge_row.full_name or "").strip() or judge_name

    eval_save_url = url_for("views.planner_evaluation_list_save_results", unit_key=unit_key, item_id=item_id)
    eval_approve_url = url_for("views.planner_evaluation_list_approve", unit_key=unit_key, item_id=item_id)
    wf = _eval_list_viewer_ctx(user, canon)
    crit_edit = bool(
        not saved_is_approved and can_save_evaluation_results(user)
    )
    wf["eval_can_edit"] = crit_edit

    return render_template(
        "judge_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            **wf,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=current_exercise,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=crit_edit,
            ),
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url=eval_save_url,
            eval_approve_url=eval_approve_url,
            eval_approve_incomplete=request.args.get("eval_approve_incomplete", type=int) == 1,
            subpage_close_fallback=url_for("views.planner_evaluation_lists", unit_key=unit_key),
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route(
    "/planner/evaluation-lists/<unit_key>/view/<int:item_id>/save-results",
    methods=["POST"],
)
def planner_evaluation_list_save_results(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)

    raw = (request.form.get("payload_json") or "").strip()
    if not raw:
        return redirect(url_for("views.planner_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    if len(raw) > 250_000:
        abort(400)
    _evaluation_commit_payload_save(db, user=user, item=item, current_exercise=current_exercise, raw=raw)
    return redirect(
        url_for(
            "views.planner_evaluation_list_file_viewer",
            unit_key=unit_key,
            item_id=item_id,
            eval_saved=1,
        )
    )


@bp.route(
    "/planner/evaluation-lists/<unit_key>/view/<int:item_id>/approve",
    methods=["POST"],
)
def planner_evaluation_list_approve(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_planner_hub(user):
        abort(403)
    if not can_approve_evaluation_results(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)

    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is None or not (saved.payload_json or "").strip():
        abort(400)
    if bool(getattr(saved, "is_approved", False)):
        return redirect(url_for("views.planner_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    rows = _parse_saved_eval_rows(saved.payload_json)
    if _evaluation_payload_has_empty_acquired_for_approve(rows):
        return redirect(
            url_for(
                "views.planner_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_incomplete=1,
            )
        )
    if not _evaluation_saved_allows_judge_approve(saved):
        return redirect(
            url_for(
                "views.planner_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_grade_blocked=1,
            )
        )
    saved.is_approved = True
    saved.approved_by_id = getattr(user, "id", None)
    saved.approved_at = datetime.utcnow()
    db.commit()
    return redirect(url_for("views.planner_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))


# مساحة المحكمين — عناصر الشريط (المعرّف، العنوان، أيقونة Font Awesome)
JUDGE_HUB_ITEMS: tuple[tuple[str, str, str], ...] = (
    ("dilemmas", "قوائم تقييم الإجراءات", "fa-file-pdf"),
    ("evaluation-lists", "قوائم التقييم", "fa-file-excel"),
    ("planner-flow-materials", "المجرى وتقييم الإجراءات", "fa-diagram-project"),
    ("visual-documentation", "التوثيق المرئي", "fa-photo-film"),
    ("incomplete-tasks", "مهام غير مكتملة", "fa-clipboard-list"),
    ("battle-overview", "الصورة العامة للمعركة", "fa-map"),
)
JUDGE_HUB_SLUGS: dict[str, str] = {s: t for s, t, _ in JUDGE_HUB_ITEMS}

_JUDGE_HUB_ITEMS_NO_PLANNER_FLOW_MATERIALS: tuple[tuple[str, str, str], ...] = tuple(
    x for x in JUDGE_HUB_ITEMS if x[0] != "planner-flow-materials"
)


def _judge_hub_menu_items(user: User) -> tuple[tuple[str, str, str], ...]:
    """عناصر أوامر مساحة المحكمين حسب الدور."""
    _hub_by_slug = {x[0]: x for x in JUDGE_HUB_ITEMS}
    _judge_individual_slugs = (
        "evaluation-lists",
        "planner-flow-materials",
        "incomplete-tasks",
    )
    if is_system_admin(user):
        return CHIEF_JUDGE_ONLY_HUB_ITEMS + _JUDGE_HUB_ITEMS_NO_PLANNER_FLOW_MATERIALS
    if can_access_chief_judge_hub(user):
        return _chief_judge_hub_items()
    return tuple(_hub_by_slug[s] for s in _judge_individual_slugs if s in _hub_by_slug)


@bp.route("/judge")
def judge_hub():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/judge")
    if not can_access_judge_hub(user):
        abort(403)
    # المحكم الفردي: قائمة مخفّضة؛ كبير المحكمين (أو المسؤول بصلاحيته): أوامر خاصة + كامل أوامر المحكمين.
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    _ensure_judge_roster_synced(db, user, ex)

    items_src = _judge_hub_menu_items(user)
    hub_items = [{"slug": s, "title_ar": t, "icon": ic} for s, t, ic in items_src]
    return render_template(
        "judge_hub.html",
        **_ctx(
            user,
            hub_items=hub_items,
            judge_hub_show_chief_subtitle=bool(can_access_chief_judge_hub(user)),
            judge_hub_hide_close=bool(
                is_judge(user)
                and not can_access_chief_judge_hub(user)
                and not is_system_admin(user)
            ),
        ),
    )


@bp.route("/judge/<slug>")
def judge_hub_section(slug: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/judge/{slug}")
    if not can_access_judge_hub(user):
        abort(403)
    slug_norm = (slug or "").strip().lower()
    if slug_norm == "evaluation-lists-chief":
        if not can_access_chief_judge_hub(user):
            abort(403)
        return redirect(url_for("views.chief_judge_evaluation_lists_home"))
    if slug_norm == "planner-flow-bundle-overview":
        if not can_access_chief_judge_hub(user):
            abort(403)
        return redirect(
            url_for("views.judge_planner_flow_materials", from_chief_judge=1)
        )
    title = JUDGE_HUB_SLUGS.get(slug_norm)
    if not title:
        abort(404)
    if slug_norm == "evaluation-lists":
        return redirect("/judge/evaluation-lists")
    if slug_norm == "planner-flow-materials":
        return redirect(url_for("views.judge_planner_flow_materials", from_judge=1))
    if slug_norm == "dilemmas":
        return redirect("/judge/dilemmas")
    if slug_norm == "visual-documentation":
        return redirect(url_for("views.visual_documentation", from_judge=1))
    if slug_norm == "incomplete-tasks":
        return _render_incomplete_evaluations_page(
            user,
            section_title=title,
            section_icon="fa-clipboard-list",
            role="judge",
        )
    return render_template(
        "judge_section_placeholder.html",
        **_ctx(
            user,
            section_title=title,
            section_slug=slug_norm,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/incomplete-tasks/update", methods=["POST"])
def judge_incomplete_tasks_update():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/judge/incomplete-tasks")
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    if ex is None:
        return redirect("/judge/incomplete-tasks")

    from app.models import JudgeIncompleteTaskStatus, JudgeTaskStatusKey

    unit_key = (request.form.get("unit_key") or "").strip()
    phase = _normalized_exercise_phase(request.form.get("phase"))
    pair_index_raw = (request.form.get("pair_index") or "").strip()
    status_key = (request.form.get("status_key") or "").strip().lower()
    if status_key not in (
        JudgeTaskStatusKey.LATE.value,
        JudgeTaskStatusKey.ONTIME.value,
        JudgeTaskStatusKey.DONE.value,
    ):
        status_key = JudgeTaskStatusKey.ONTIME.value
    try:
        pair_index = int(pair_index_raw)
    except Exception:
        pair_index = 0
    dilemma_id_raw = (request.form.get("dilemma_id") or "").strip()
    evaluation_item_id_raw = (request.form.get("evaluation_item_id") or "").strip()
    dilemma_id = int(dilemma_id_raw) if dilemma_id_raw.isdigit() else None
    evaluation_item_id = int(evaluation_item_id_raw) if evaluation_item_id_raw.isdigit() else None

    judge_id = int(getattr(user, "id", 0) or 0)
    row = (
        db.query(JudgeIncompleteTaskStatus)
        .filter(
            JudgeIncompleteTaskStatus.exercise_id == ex.id,
            JudgeIncompleteTaskStatus.judge_id == judge_id,
            JudgeIncompleteTaskStatus.unit_level_key == unit_key,
            JudgeIncompleteTaskStatus.exercise_phase == phase,
            JudgeIncompleteTaskStatus.pair_index == pair_index,
        )
        .first()
    )
    if row is None:
        row = JudgeIncompleteTaskStatus(
            exercise_id=ex.id,
            judge_id=judge_id,
            unit_level_key=unit_key,
            exercise_phase=phase,
            pair_index=pair_index,
        )
        db.add(row)
    row.dilemma_id = dilemma_id
    row.evaluation_item_id = evaluation_item_id
    row.status_key = status_key
    db.commit()
    return redirect("/judge/incomplete-tasks")


def _notifications_scope_exercise(db, user: User) -> Exercise | None:
    """التمرين المرتبط بسجل الإشعارات (نفس منطق غرف المحادثة)."""
    if is_system_admin(user):
        return _admin_current_workspace_exercise(db, user)
    return _current_workspace_exercise(db, user)


@bp.route("/notifications", methods=["GET"])
def notifications_log():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/notifications")
    if not can_view_notifications_log(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _notifications_scope_exercise(db, user)
    rows: list[ExerciseNotification] = []
    if ex is not None:
        rows = (
            db.query(ExerciseNotification)
            .filter(
                ExerciseNotification.user_id == int(user.id),
                ExerciseNotification.exercise_id == int(ex.id),
            )
            .order_by(desc(ExerciseNotification.created_at), desc(ExerciseNotification.id))
            .limit(500)
            .all()
        )
    return render_template(
        "notifications_log.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            notifications=rows,
            **_hub_back_ctx_for_request_path(),
            hub_from_form_param=_role_hub_from_form_param(),
        ),
    )


@bp.route("/notifications/<int:nid>/read", methods=["POST"])
def notification_mark_read(nid: int):
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/notifications")
    if not can_view_notifications_log(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _notifications_scope_exercise(db, user)
    if ex is None:
        return redirect(url_for("views.notifications_log", **_role_hub_preserve_link_kwargs()))
    row = db.get(ExerciseNotification, nid)
    if (
        row
        and int(row.user_id) == int(user.id)
        and int(row.exercise_id) == int(ex.id)
    ):
        row.is_read = True
        db.add(row)
        db.commit()
    hub_kwargs = _role_hub_preserve_link_kwargs()
    return redirect(
        url_for("views.notifications_log", **hub_kwargs) + f"#notif-{int(nid)}"
    )


@bp.route("/notifications/read-all", methods=["POST"])
def notifications_mark_all_read():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/notifications")
    if not can_view_notifications_log(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _notifications_scope_exercise(db, user)
    if ex is not None:
        for row in (
            db.query(ExerciseNotification)
            .filter(
                ExerciseNotification.user_id == int(user.id),
                ExerciseNotification.exercise_id == int(ex.id),
                ExerciseNotification.is_read == False,
            )
            .all()
        ):
            row.is_read = True
            db.add(row)
        db.commit()
    return redirect(url_for("views.notifications_log", **_role_hub_preserve_link_kwargs()))


@bp.route("/api/system/heartbeat", methods=["GET"])
def api_system_heartbeat():
    """نقطة نهاية «نبضة الحياة» للتحديث التلقائي بين الخادم والعملاء.

    تُرجع توقيع نسخة مركّب يتغيّر عند أي تغيير في الجداول الرئيسية
    (الإشعارات، قوائم التقييم، نتائج التقييم، تقييمات المجرى/الإجراءات،
    رسائل المحادثة، توثيق مرئي، معاضل). يستدعيها العميل بشكل دوري
    لاكتشاف الحاجة لإعادة تحميل الصفحة تلقائياً.
    """
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    from flask import g

    db = g.db
    parts: list[str] = []

    def _add_max(query):
        try:
            ts = query.scalar()
        except Exception:
            ts = None
        if ts is not None:
            try:
                parts.append(str(int(ts.timestamp())))
            except Exception:
                parts.append(str(ts))

    # نطاق التمرين الحالي (لكل من المسؤول وغيره).
    ex = None
    try:
        if is_system_admin(user):
            ex = _admin_current_workspace_exercise(db, user)
        else:
            ex = _current_workspace_exercise(db, user)
    except Exception:
        ex = None
    ex_id = int(ex.id) if ex is not None else 0
    parts.append(f"ex={ex_id}")

    # سجل الإشعارات الخاص بالمستخدم.
    _add_max(
        db.query(func.max(ExerciseNotification.created_at)).filter(
            ExerciseNotification.user_id == int(user.id)
        )
    )
    unread = (
        db.query(func.count(ExerciseNotification.id))
        .filter(
            ExerciseNotification.user_id == int(user.id),
            ExerciseNotification.is_read == False,
        )
        .scalar()
        or 0
    )
    parts.append(f"unread={int(unread)}")

    if ex_id:
        # قوائم التقييم.
        _add_max(
            db.query(func.max(EvaluationListPdfItem.created_at)).filter(
                EvaluationListPdfItem.exercise_id == ex_id
            )
        )
        # نتائج التقييم (تشمل الحفظ والاعتماد والإعادة).
        _add_max(
            db.query(func.max(EvaluationListSavedResult.updated_at)).filter(
                EvaluationListSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(EvaluationListSavedResult.approved_at)).filter(
                EvaluationListSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(EvaluationListSavedResult.chief_approved_at)).filter(
                EvaluationListSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(EvaluationListSavedResult.control_approved_at)).filter(
                EvaluationListSavedResult.exercise_id == ex_id
            )
        )
        # نتائج تقييم إجراءات حزمة المجرى (تقييم المجرى/الإجراءات).
        _add_max(
            db.query(func.max(PlannerFlowBundleEvalSavedResult.updated_at)).filter(
                PlannerFlowBundleEvalSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(PlannerFlowBundleEvalSavedResult.approved_at)).filter(
                PlannerFlowBundleEvalSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(PlannerFlowBundleEvalSavedResult.chief_approved_at)).filter(
                PlannerFlowBundleEvalSavedResult.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(PlannerFlowBundleEvalSavedResult.control_approved_at)).filter(
                PlannerFlowBundleEvalSavedResult.exercise_id == ex_id
            )
        )
        pf_saved_n = (
            db.query(func.count(PlannerFlowBundleEvalSavedResult.id))
            .filter(PlannerFlowBundleEvalSavedResult.exercise_id == ex_id)
            .scalar()
            or 0
        )
        parts.append(f"pf_saved={int(pf_saved_n)}")
        # حزم المجرى وقوائم الإجراءات (ربط/إنشاء جديد).
        _add_max(
            db.query(func.max(ExercisePlannerFlowBundle.updated_at)).filter(
                ExercisePlannerFlowBundle.exercise_id == ex_id
            )
        )
        _add_max(
            db.query(func.max(ExercisePlannerFlowBundleActionEval.created_at))
            .join(
                ExercisePlannerFlowBundle,
                ExercisePlannerFlowBundleActionEval.bundle_id
                == ExercisePlannerFlowBundle.id,
            )
            .filter(ExercisePlannerFlowBundle.exercise_id == ex_id)
        )
        _add_max(
            db.query(func.max(ExercisePlannerFlowBundleEventFlow.created_at))
            .join(
                ExercisePlannerFlowBundle,
                ExercisePlannerFlowBundleEventFlow.bundle_id
                == ExercisePlannerFlowBundle.id,
            )
            .filter(ExercisePlannerFlowBundle.exercise_id == ex_id)
        )
        # توثيق بنود التقييم (صور/فيديو).
        _add_max(
            db.query(func.max(EvaluationCriterionMedia.created_at)).filter(
                EvaluationCriterionMedia.exercise_id == ex_id
            )
        )
        # رسائل المحادثة لكل غرف هذا التمرين.
        _add_max(
            db.query(func.max(ChatMessage.created_at))
            .join(ChatRoom, ChatMessage.room_id == ChatRoom.id)
            .filter(ChatRoom.exercise_id == ex_id)
        )
        # توثيق مرئي.
        _add_max(
            db.query(func.max(VisualDocument.created_at)).filter(
                VisualDocument.exercise_id == ex_id
            )
        )
        # معاضل/تقييمات.
        _add_max(
            db.query(func.max(DilemmaItem.created_at)).filter(
                DilemmaItem.exercise_id == ex_id
            )
        )

    version = "|".join(parts)
    resp = jsonify(
        {
            "ok": True,
            "version": version,
            "unread_notifications": int(unread),
            "exercise_id": ex_id,
            "server_time": datetime.utcnow().isoformat(),
        }
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@bp.route("/api/notifications/summary", methods=["GET"])
def api_notifications_summary():
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_view_notifications_log(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    ex = _notifications_scope_exercise(db, user)
    if ex is None:
        return jsonify({"ok": True, "unread_count": 0, "latest": []})
    unread = (
        db.query(func.count(ExerciseNotification.id))
        .filter(
            ExerciseNotification.user_id == int(user.id),
            ExerciseNotification.exercise_id == int(ex.id),
            ExerciseNotification.is_read == False,
        )
        .scalar()
        or 0
    )
    latest_rows = (
        db.query(ExerciseNotification)
        .filter(
            ExerciseNotification.user_id == int(user.id),
            ExerciseNotification.exercise_id == int(ex.id),
        )
        .order_by(desc(ExerciseNotification.created_at), desc(ExerciseNotification.id))
        .limit(8)
        .all()
    )
    latest = [
        {
            "id": r.id,
            "title": r.title,
            "body": (r.body or "")[:500],
            "type": r.type,
            "priority": r.priority or "normal",
            "is_read": bool(r.is_read),
            "action_url": r.action_url or "",
            "created_at": r.created_at.isoformat() if r.created_at else "",
        }
        for r in latest_rows
    ]
    return jsonify({"ok": True, "unread_count": int(unread), "latest": latest})


@bp.route("/api/notifications/<int:nid>/read", methods=["POST"])
def api_notification_mark_read(nid: int):
    """تعليم إشعار كمقروء (للوحة التنبيه المنبثقة دون إعادة تحميل الصفحة)."""
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not can_view_notifications_log(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    ex = _notifications_scope_exercise(db, user)
    if ex is None:
        return jsonify({"ok": False, "error": "no_exercise"}), 400
    row = db.get(ExerciseNotification, nid)
    if (
        not row
        or int(row.user_id) != int(user.id)
        or int(row.exercise_id) != int(ex.id)
    ):
        return jsonify({"ok": False, "error": "not_found"}), 404
    if not row.is_read:
        row.is_read = True
        db.add(row)
        db.commit()
    unread = (
        db.query(func.count(ExerciseNotification.id))
        .filter(
            ExerciseNotification.user_id == int(user.id),
            ExerciseNotification.exercise_id == int(ex.id),
            ExerciseNotification.is_read == False,
        )
        .scalar()
        or 0
    )
    return jsonify({"ok": True, "id": int(nid), "unread_count": int(unread)})


def _visual_scope_exercise(db, user: User) -> Exercise | None:
    """نطاق التمرين للتوثيق المرئي (نفس منطق المحادثة/الإشعارات)."""
    if is_system_admin(user):
        return _admin_current_workspace_exercise(db, user)
    return _current_workspace_exercise(db, user)


def _visual_doc_disk_path(relpath: str) -> Path | None:
    if not relpath or ".." in relpath.replace("\\", "/"):
        return None
    root = VISUAL_DOC_DIR.resolve()
    p = (root / relpath).resolve()
    if str(p).startswith(str(root)):
        return p
    return None


def _visual_doc_file_size_kb(row: VisualDocument) -> int | None:
    p = _visual_doc_disk_path((row.file_relpath or "").strip())
    if p is None or not p.is_file():
        return None
    return max(1, int(p.stat().st_size / 1024))


def _visual_doc_rows_for_template(docs: list) -> list[dict]:
    out: list[dict] = []
    for row in docs:
        name = (row.description or "").strip()
        out.append(
            {
                "row": row,
                "display_name": name or "—",
                "size_kb": _visual_doc_file_size_kb(row),
            }
        )
    return out


def _visual_doc_can_access_row(db, user, ex, row: VisualDocument) -> bool:
    if ex is None or int(row.exercise_id) != int(ex.id):
        return False
    if is_system_admin(user):
        return True
    if is_judge(user):
        a = _judge_assignment_for_current_exercise(db, user, ex)
        assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
        if assigned_uk and (row.unit_level_key or "").strip() != assigned_uk:
            return False
    return True


_VISUAL_ALLOWED_SUFFIX = {
    ".jpg",
    ".jpeg",
    ".png",
    ".mp4",
    ".mp3",
    ".wav",
    ".m4a",
    ".ogg",
    ".webm",
}
_VISUAL_MAX_UPLOAD_BYTES = 80 * 1024 * 1024  # 80MB

_VISUAL_UPLOAD_ERROR_MESSAGES: dict[str, str] = {
    "no_file": "لم يُستلم ملف. أعد التسجيل أو اختر ملفاً ثم اضغط رفع.",
    "bad_type": "نوع الملف غير مدعوم.",
    "too_large": "حجم الملف يتجاوز الحد المسموح (80 ميجابايت).",
    "no_unit": "حدد مستوى الوحدة أولاً.",
}


def _visual_upload_error_redirect(unit_key: str, code: str):
    return redirect(
        url_for(
            "views.visual_documentation",
            **_visual_doc_redirect_kwargs(unit_key=unit_key, upload_error=code),
        )
    )


def _visual_infer_upload_suffix(raw_name: str, mime: str) -> str:
    suf = Path(secure_filename(raw_name or "")).suffix.lower()
    if suf in _VISUAL_ALLOWED_SUFFIX:
        return suf
    if mime.startswith("video/"):
        return ".webm" if "webm" in mime else ".mp4"
    if mime.startswith("audio/"):
        if "ogg" in mime:
            return ".ogg"
        return ".webm"
    if mime.startswith("image/"):
        if "png" in mime:
            return ".png"
        return ".jpg"
    return ""


def _visual_file_type_for_suffix(suf: str, mime: str) -> str:
    if mime.startswith("video/"):
        return "video"
    if mime.startswith("audio/"):
        return "audio"
    if suf == ".mp4":
        return "video"
    if suf in (".mp3", ".wav", ".m4a", ".ogg"):
        return "audio"
    if suf == ".webm":
        return "audio"
    return "image"


@bp.route("/visual-documentation", methods=["GET"])
def visual_documentation():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/visual-documentation")
    if not can_use_visual_documentation(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _visual_scope_exercise(db, user)
    if ex is None:
        return render_template(
            "visual_documentation.html",
            **_ctx(
                user,
                has_exercise=False,
                exercise=None,
                unit_levels=[],
                selected_unit_key="",
                doc_rows=[],
                **_hub_back_ctx_for_request_path(),
                hub_from_form_param=_role_hub_from_form_param(),
            ),
        )

    # نطاق الوحدة للمحكم غير إدارة النظام
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    unit_key = (request.args.get("unit_key") or "").strip()
    if assigned_uk and not is_system_admin(user):
        unit_key = assigned_uk
    if not unit_key:
        unit_key = assigned_uk or default_unit_level_key()

    q = db.query(VisualDocument).filter(VisualDocument.exercise_id == int(ex.id))
    if unit_key:
        q = q.filter(VisualDocument.unit_level_key == unit_key)
    if assigned_uk and not is_system_admin(user):
        q = q.filter(VisualDocument.unit_level_key == assigned_uk)
    docs = q.order_by(desc(VisualDocument.created_at), desc(VisualDocument.id)).limit(400).all()

    unit_levels = [u for u in UNIT_LEVELS if not assigned_uk or u.get("key") == assigned_uk]
    return render_template(
        "visual_documentation.html",
        **_ctx(
            user,
            has_exercise=True,
            exercise=ex,
            unit_levels=unit_levels,
            selected_unit_key=unit_key,
            doc_rows=_visual_doc_rows_for_template(docs),
            assigned_unit_key=assigned_uk,
            upload_error_code=(request.args.get("upload_error") or "").strip(),
            upload_error_message=_VISUAL_UPLOAD_ERROR_MESSAGES.get(
                (request.args.get("upload_error") or "").strip(), ""
            ),
            upload_ok=request.args.get("upload_ok", type=int) == 1,
            **_hub_back_ctx_for_request_path(),
            hub_from_form_param=_role_hub_from_form_param(),
        ),
    )


@bp.route("/visual-documentation/upload", methods=["POST"])
def visual_documentation_upload():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/visual-documentation")
    if not can_use_visual_documentation(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _visual_scope_exercise(db, user)
    if ex is None:
        return redirect(url_for("views.visual_documentation", **_visual_doc_redirect_kwargs()))

    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    unit_key = normalize_unit_level_key(request.form.get("unit_key") or "")
    if assigned_uk and not is_system_admin(user):
        unit_key = assigned_uk
    if not unit_key:
        unit_key = assigned_uk
    if not unit_key:
        return _visual_upload_error_redirect("", "no_unit")

    f = request.files.get("media_file")
    data = f.read() if f else b""
    if not data:
        return _visual_upload_error_redirect(unit_key, "no_file")

    uploaded_mime = (getattr(f, "mimetype", "") or "").lower()
    raw_name = (getattr(f, "filename", "") or "").strip()
    suf = _visual_infer_upload_suffix(raw_name, uploaded_mime)
    if not suf or suf not in _VISUAL_ALLOWED_SUFFIX:
        return _visual_upload_error_redirect(unit_key, "bad_type")
    if len(data) > _VISUAL_MAX_UPLOAD_BYTES:
        return _visual_upload_error_redirect(unit_key, "too_large")

    ft = _visual_file_type_for_suffix(suf, uploaded_mime)

    desc_txt = (request.form.get("display_name") or "").strip()[:200]
    loc = ""
    dilemma_id = None
    event_id_raw = (request.form.get("event_id") or "").strip()
    event_id = int(event_id_raw) if event_id_raw.isdigit() else None

    VISUAL_DOC_DIR.mkdir(parents=True, exist_ok=True)
    sub = VISUAL_DOC_DIR / str(int(ex.id)) / (unit_key or "misc")
    sub.mkdir(parents=True, exist_ok=True)
    store_name = f"{uuid.uuid4().hex}{suf}"
    disk_path = (sub / store_name).resolve()
    disk_path.write_bytes(data)
    rel = f"{int(ex.id)}/{(unit_key or 'misc')}/{store_name}".replace("\\", "/")

    row = VisualDocument(
        exercise_id=int(ex.id),
        event_id=event_id,
        dilemma_id=dilemma_id,
        unit_level_key=(unit_key or "").strip(),
        uploaded_by_id=int(user.id),
        file_type=ft,
        file_relpath=rel,
        description=desc_txt,
        location_label=loc,
    )
    db.add(row)
    db.commit()

    # إشعار للمستخدمين المعنيين
    try:
        from app.notifications_service import notify_visual_document_added

        unit_label = label_for_unit_level_key(unit_key) or unit_key
        notify_visual_document_added(
            db,
            exercise_id=int(ex.id),
            unit_key=unit_key,
            unit_label=unit_label,
            file_type=ft,
            action_url=url_for("views.visual_documentation", unit_key=unit_key),
        )
        db.commit()
    except Exception:
        db.rollback()

    return redirect(
        url_for(
            "views.visual_documentation",
            **_visual_doc_redirect_kwargs(unit_key=unit_key, upload_ok=1),
        )
    )


@bp.route("/visual-documents/<int:doc_id>/delete", methods=["POST"])
def visual_document_delete(doc_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/visual-documentation")
    if not can_use_visual_documentation(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(VisualDocument, doc_id)
    if row is None:
        abort(404)
    ex = _visual_scope_exercise(db, user)
    if not _visual_doc_can_access_row(db, user, ex, row):
        abort(403)

    unit_key = normalize_unit_level_key(request.form.get("unit_key") or row.unit_level_key or "")
    p = _visual_doc_disk_path((row.file_relpath or "").strip())
    if p is not None and p.is_file():
        try:
            p.unlink()
        except OSError:
            pass
    db.delete(row)
    db.commit()
    return redirect(url_for("views.visual_documentation", **_visual_doc_redirect_kwargs(unit_key=unit_key)))


@bp.route("/visual-documents/<int:doc_id>/file", methods=["GET"])
def visual_document_file(doc_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/visual-documents/{doc_id}/file")
    if not can_use_visual_documentation(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(VisualDocument, doc_id)
    if row is None:
        abort(404)
    ex = _visual_scope_exercise(db, user)
    if ex is None or int(row.exercise_id) != int(ex.id):
        abort(404)
    if not is_system_admin(user) and is_judge(user):
        a = _judge_assignment_for_current_exercise(db, user, ex)
        assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
        if assigned_uk and (row.unit_level_key or "").strip() != assigned_uk:
            abort(403)

    p = _visual_doc_disk_path((row.file_relpath or "").strip())
    if p is None or not p.is_file():
        abort(404)
    mime = mimetypes.guess_type(p.name)[0] or "application/octet-stream"
    return send_file(p, mimetype=mime)


# مساحة السيطرة — عناصر الشريط (المعرّف، العنوان، أيقونة Font Awesome)
CONTROL_HUB_ITEMS: tuple[tuple[str, str, str], ...] = (
    ("evaluation-lists-status", "موقف قوائم التقييم", "fa-file-excel"),
    ("top-positives-negatives", "عرض أبرز الإيجابيات والسلبيات", "fa-star"),
    ("evaluation-results", "عرض نتائج التقييم", "fa-square-poll-vertical"),
    ("incomplete-tasks-status", "موقف المهام غير المكتملة", "fa-hourglass-half"),
    ("visual-doc-status", "موقف التوثيق المرئي", "fa-photo-film"),
    ("battle-overview", "الصورة العامة للمعركة", "fa-map"),
    ("assign-task", "إسناد مهمة جديدة", "fa-user-plus"),
)
CONTROL_HUB_SLUGS: dict[str, str] = {s: t for s, t, _ in CONTROL_HUB_ITEMS}
CONTROL_HUB_ICONS: dict[str, str] = {s: ic for s, t, ic in CONTROL_HUB_ITEMS}


def _control_section_icon(slug: str) -> str:
    return CONTROL_HUB_ICONS.get((slug or "").strip().lower(), "fa-eye")


def _request_from_control_hub() -> bool:
    raw = (
        request.args.get("from")
        or request.args.get("from_control")
        or request.form.get("from_control")
        or ""
    ).strip().lower()
    return raw in ("1", "true", "yes", "control")


def _control_hub_back_ctx_always() -> dict:
    return {
        "hub_back_href": url_for("views.control_hub"),
        "hub_back_label": "العودة إلى مساحة السيطرة",
    }


def _control_hub_back_ctx() -> dict:
    return _resolve_role_hub_back_ctx()


def _control_hub_link_kwargs() -> dict:
    return _role_hub_preserve_link_kwargs()


def _request_from_judge_hub() -> bool:
    raw = (
        request.args.get("from")
        or request.args.get("from_judge")
        or request.form.get("from_judge")
        or ""
    ).strip().lower()
    return raw in ("1", "true", "yes", "judge")


def _request_from_analyst_hub() -> bool:
    raw = (
        request.args.get("from")
        or request.args.get("from_analyst")
        or request.form.get("from_analyst")
        or ""
    ).strip().lower()
    return raw in ("1", "true", "yes", "analyst")


def _request_from_planner_hub() -> bool:
    raw = (
        request.args.get("from")
        or request.args.get("from_planner")
        or request.form.get("from_planner")
        or ""
    ).strip().lower()
    return raw in ("1", "true", "yes", "planner")


def _request_from_chief_judge_hub() -> bool:
    raw = (
        request.args.get("from")
        or request.args.get("from_chief_judge")
        or request.form.get("from_chief_judge")
        or ""
    ).strip().lower()
    return raw in ("1", "true", "yes", "chief_judge", "chief-judge", "chiefjudge")


def _judge_hub_back_ctx_always() -> dict:
    return {
        "hub_back_href": url_for("views.judge_hub"),
        "hub_back_label": "العودة إلى مساحة المحكمين",
    }


def _analyst_hub_back_ctx_always() -> dict:
    return {
        "hub_back_href": url_for("views.analyst_hub"),
        "hub_back_label": "العودة إلى مساحة المحللين",
    }


def _planner_hub_back_ctx_always() -> dict:
    return {
        "hub_back_href": url_for("views.planner_hub"),
        "hub_back_label": "العودة إلى مساحة التخطيط",
    }


def _chief_judge_hub_back_ctx_always() -> dict:
    """مسارات الاعتماد الثاني ما زالت تحت /chief-judge؛ العودة إلى مركز المحكمين الموحد."""
    return {
        "hub_back_href": url_for("views.judge_hub"),
        "hub_back_label": "العودة إلى مساحة المحكمين",
    }


def _resolve_role_hub_back_ctx() -> dict:
    """سياق العودة عند الدخول من مساحة دور عبر معامل from_* (صفحات مشتركة)."""
    if _request_from_control_hub():
        return _control_hub_back_ctx_always()
    if _request_from_chief_judge_hub():
        return _chief_judge_hub_back_ctx_always()
    if _request_from_judge_hub():
        return _judge_hub_back_ctx_always()
    if _request_from_planner_hub():
        return _planner_hub_back_ctx_always()
    if _request_from_analyst_hub():
        return _analyst_hub_back_ctx_always()
    return {}


def _hub_back_ctx_for_request_path() -> dict:
    """سياق العودة حسب مسار الطلب الحالي أو معامل from_*."""
    p = (request.path or "").rstrip("/").lower() or "/"
    hub_dashboard = {
        "hub_back_href": "/dashboard",
        "hub_back_label": "العودة إلى لوحة المستخدم",
    }
    if p in ("/control", "/judge", "/analyst", "/planner"):
        return hub_dashboard
    if p.startswith("/control"):
        return _control_hub_back_ctx_always()
    if p.startswith("/chief-judge"):
        return _chief_judge_hub_back_ctx_always()
    if p.startswith("/judge"):
        return _judge_hub_back_ctx_always()
    if p.startswith("/analyst"):
        return _analyst_hub_back_ctx_always()
    if p.startswith("/planner"):
        return _planner_hub_back_ctx_always()
    return _resolve_role_hub_back_ctx()


def _subpage_close_ctx(parent_href: str | None = None, **extra) -> dict:
    """زر الإغلاق: parent_href إن وُجد، وإلا hub_back حسب الدور/المسار."""
    ctx = dict(_hub_back_ctx_for_request_path())
    if parent_href:
        ctx["subpage_close_fallback"] = parent_href
    ctx.update(extra)
    return ctx


def _role_hub_preserve_link_kwargs() -> dict:
    kw: dict = {}
    if _request_from_control_hub():
        kw["from_control"] = 1
    if _request_from_chief_judge_hub():
        kw["from_chief_judge"] = 1
    if _request_from_judge_hub():
        kw["from_judge"] = 1
    if _request_from_planner_hub():
        kw["from_planner"] = 1
    if _request_from_analyst_hub():
        kw["from_analyst"] = 1
    return kw


def _role_hub_from_form_param() -> str | None:
    if _request_from_control_hub():
        return "from_control"
    if _request_from_chief_judge_hub():
        return "from_chief_judge"
    if _request_from_judge_hub():
        return "from_judge"
    if _request_from_planner_hub():
        return "from_planner"
    if _request_from_analyst_hub():
        return "from_analyst"
    return None


def _visual_doc_redirect_kwargs(**extra) -> dict:
    kw = dict(extra)
    kw.update(_role_hub_preserve_link_kwargs())
    return kw


_CONTROL_EVAL_LIST_TYPE_JUDGE = "judge_eval"
_CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION = "planner_flow_action"

_CONTROL_EVAL_LIST_TYPE_LABELS: dict[str, str] = {
    _CONTROL_EVAL_LIST_TYPE_JUDGE: "قائمة التقييم",
    _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION: "المجرى وتقييم الإجراءات — قوائم تقييم الإجراءات",
}

_CONTROL_EVAL_LIST_TYPE_SORT: dict[str, int] = {
    _CONTROL_EVAL_LIST_TYPE_JUDGE: 0,
    _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION: 1,
}


def _control_eval_list_type_label(kind: str) -> str:
    return _CONTROL_EVAL_LIST_TYPE_LABELS.get((kind or "").strip(), "—")


def _control_eval_status_row_sort_key(row: dict) -> tuple:
    kind = (row.get("list_type_kind") or "").strip()
    return (
        _CONTROL_EVAL_LIST_TYPE_SORT.get(kind, 9),
        (row.get("item_title") or ""),
        int(row.get("sort_tiebreaker") or 0),
    )


def _control_append_planner_flow_status_rows(
    db,
    *,
    exercise_id: int,
    exercise,
    by_phase: dict[str, list[dict]],
) -> None:
    """إضافة صفوف قوائم تقييم الإجراءات (المجرى) إلى موقف قوائم التقييم."""
    bundles = (
        db.query(ExercisePlannerFlowBundle)
        .filter(ExercisePlannerFlowBundle.exercise_id == int(exercise_id))
        .order_by(
            ExercisePlannerFlowBundle.unit_level_key,
            ExercisePlannerFlowBundle.exercise_phase,
            ExercisePlannerFlowBundle.id,
        )
        .all()
    )
    if not bundles:
        return

    for bundle in bundles:
        uk = (bundle.unit_level_key or "").strip()
        phase_key = _normalized_exercise_phase(getattr(bundle, "exercise_phase", None))
        if not uk:
            continue

        action_rows = (
            db.query(ExercisePlannerFlowBundleActionEval)
            .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
            .order_by(
                ExercisePlannerFlowBundleActionEval.slot_index,
                ExercisePlannerFlowBundleActionEval.id,
            )
            .all()
        )
        for action_row in action_rows:
            rel = (action_row.file_relpath or "").strip()
            if not rel:
                continue
            title = _planner_blob_display_filename(
                stored_title=action_row.title or "",
                relpath=rel,
                fallback=f"قائمة تقييم إجراءات {int(action_row.slot_index)}",
            )
            canon = _planner_bundle_eval_canonical_saved(
                db, int(exercise_id), int(action_row.id)
            )
            total_pct = _evaluation_saved_total_pct(canon)
            shim = type(
                "_PfActionShim",
                (),
                {
                    "id": int(action_row.id),
                    "text": title,
                    "created_at": getattr(action_row, "created_at", None),
                },
            )()
            by_phase.setdefault(phase_key, []).append(
                {
                    **build_evaluation_list_row(
                        item=shim,
                        saved=canon,
                        exercise=exercise,
                        open_href=url_for(
                            "views.control_planner_flow_action_view",
                            unit_key=uk,
                            action_eval_id=int(action_row.id),
                        ),
                    ),
                    "list_type_kind": _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION,
                    "list_type_label": _control_eval_list_type_label(
                        _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION
                    ),
                    "phase_key": phase_key,
                    "phase_label": _phase_label_ar(phase_key),
                    "workflow_label": eval_workflow_label_ar(canon),
                    "unit_key": uk,
                    "unit_label": label_for_unit_level_key(uk) or uk or "—",
                    "total_pct": total_pct,
                    "sort_tiebreaker": int(action_row.id),
                }
            )


def _build_control_evaluation_lists_status(db, user: User) -> dict:
    """موقف قوائم التقييم مجمّع حسب مراحل التمرين (تبويب لكل مرحلة)."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id)
        .order_by(
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    item_ids = [int(it.id) for it in items if getattr(it, "id", None) is not None]
    canonical_by_item = (
        _evaluation_canonical_map_for_items(db, int(ex.id), item_ids) if item_ids else {}
    )

    def _roster_person_label(row: ExerciseRosterRow | None) -> str:
        if row is None:
            return "—"
        display = (getattr(row, "full_name", None) or "").strip()
        rank = (getattr(row, "rank_ar", None) or "").strip()
        if rank and display:
            return f"{rank} {display}"
        if display:
            return display
        mil = (getattr(row, "military_number", None) or "").strip()
        return mil or "—"

    judge_name_by_unit: dict[str, str] = {}
    trainee_name_by_unit: dict[str, str] = {}
    for rr in (
        db.query(ExerciseRosterRow)
        .filter(ExerciseRosterRow.exercise_id == ex.id)
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .all()
    ):
        uk_r = (getattr(rr, "unit_level_key", None) or "").strip()
        if not uk_r:
            continue
        kind = (getattr(rr, "roster_kind", None) or "").strip()
        if kind == ExerciseRosterKind.JUDGE.value and uk_r not in judge_name_by_unit:
            judge_name_by_unit[uk_r] = _roster_person_label(rr)
        elif kind == ExerciseRosterKind.TRAINEE.value and uk_r not in trainee_name_by_unit:
            trainee_name_by_unit[uk_r] = _roster_person_label(rr)

    phase_order = {key: idx for idx, key in enumerate(exercise_phase_keys())}
    by_phase: dict[str, list[dict]] = {}

    for it in items:
        uk = (it.unit_level_key or "").strip()
        phase_key = _normalized_exercise_phase(getattr(it, "exercise_phase", None))
        saved = canonical_by_item.get(int(it.id))
        total_pct = _evaluation_saved_total_pct(saved)
        by_phase.setdefault(phase_key, []).append(
            {
                **build_evaluation_list_row(
                    item=it,
                    saved=saved,
                    exercise=ex,
                    open_href=url_for(
                        "views.control_evaluation_list_file_viewer",
                        unit_key=uk,
                        item_id=int(it.id),
                    ),
                ),
                "list_type_kind": _CONTROL_EVAL_LIST_TYPE_JUDGE,
                "list_type_label": _control_eval_list_type_label(
                    _CONTROL_EVAL_LIST_TYPE_JUDGE
                ),
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "workflow_label": eval_workflow_label_ar(saved),
                "unit_key": uk,
                "unit_label": label_for_unit_level_key(uk) or uk or "—",
                "total_pct": total_pct,
                "sort_tiebreaker": int(it.id),
            }
        )

    _control_append_planner_flow_status_rows(
        db, exercise_id=int(ex.id), exercise=ex, by_phase=by_phase
    )

    unit_order = {row.get("key"): idx for idx, row in enumerate(UNIT_LEVELS)}
    phase_keys_seen = set(by_phase.keys())
    ordered_phase_keys: list[str] = list(exercise_phase_keys())
    for pk in sorted(phase_keys_seen - set(ordered_phase_keys), key=lambda k: phase_order.get(k, 99)):
        ordered_phase_keys.append(pk)

    phase_tabs: list[dict] = []
    for pk in ordered_phase_keys:
        phase_rows = by_phase.get(pk, [])
        by_unit: dict[str, list[dict]] = {}
        for row in phase_rows:
            uk = (row.get("unit_key") or "").strip()
            by_unit.setdefault(uk, []).append(row)
        unit_tabs: list[dict] = []
        for uk in sorted(
            by_unit.keys(),
            key=lambda k: (unit_order.get(k, len(unit_order)), label_for_unit_level_key(k) or k),
        ):
            unit_rows = sorted(by_unit[uk], key=_control_eval_status_row_sort_key)
            n_assigned = len(unit_rows)
            n_done = sum(1 for r in unit_rows if r.get("status_done"))
            unit_tabs.append(
                {
                    "unit_key": uk,
                    "unit_label": label_for_unit_level_key(uk) or uk or "—",
                    "rows": unit_rows,
                    "total_count": n_assigned,
                    "judge_name": judge_name_by_unit.get(uk, "—"),
                    "trainee_name": trainee_name_by_unit.get(uk, "—"),
                    "n_assigned": n_assigned,
                    "n_done": n_done,
                    "n_not_done": n_assigned - n_done,
                }
            )
        phase_tabs.append(
            {
                "phase_key": pk,
                "phase_label": _phase_label_ar(pk),
                "unit_tabs": unit_tabs,
                "total_count": len(phase_rows),
            }
        )

    return {
        "has_exercise": True,
        "phase_tabs": phase_tabs,
    }


def _incomplete_eval_started_at_sort_key(row: dict) -> tuple:
    started = row.get("started_at")
    if started is None:
        return (1, datetime.max, int(row.get("sort_tiebreaker") or 0))
    return (0, started, int(row.get("sort_tiebreaker") or 0))


def _collect_all_eval_status_rows_flat(
    db,
    *,
    exercise: Exercise,
    unit_filter: str | None = None,
    eval_open_endpoint: str = "views.control_evaluation_list_file_viewer",
    planner_open_endpoint: str = "views.control_planner_flow_action_view",
    planner_open_uses_slot: bool = False,
) -> list[dict]:
    """كل قوائم التقييم + إجراءات المجرى كصفوف مسطحة (للتصفية والفرز)."""
    ex_id = int(exercise.id)
    items_q = db.query(EvaluationListPdfItem).filter(EvaluationListPdfItem.exercise_id == ex_id)
    if unit_filter:
        items_q = items_q.filter(EvaluationListPdfItem.unit_level_key == unit_filter)
    items = (
        items_q.order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    item_ids = [int(it.id) for it in items if getattr(it, "id", None) is not None]
    canonical_by_item = (
        _evaluation_canonical_map_for_items(db, ex_id, item_ids) if item_ids else {}
    )

    rows: list[dict] = []
    for it in items:
        uk = (it.unit_level_key or "").strip()
        phase_key = _normalized_exercise_phase(getattr(it, "exercise_phase", None))
        saved = canonical_by_item.get(int(it.id))
        total_pct = _evaluation_saved_total_pct(saved)
        started_at = getattr(it, "created_at", None)
        rows.append(
            {
                **build_evaluation_list_row(
                    item=it,
                    saved=saved,
                    exercise=exercise,
                    open_href=url_for(eval_open_endpoint, unit_key=uk, item_id=int(it.id)),
                ),
                "list_type_kind": _CONTROL_EVAL_LIST_TYPE_JUDGE,
                "list_type_label": _control_eval_list_type_label(_CONTROL_EVAL_LIST_TYPE_JUDGE),
                "phase_key": phase_key,
                "phase_label": _phase_label_ar(phase_key),
                "workflow_label": eval_workflow_label_ar(saved),
                "unit_key": uk,
                "unit_label": label_for_unit_level_key(uk) or uk or "—",
                "total_pct": total_pct,
                "sort_tiebreaker": int(it.id),
                "started_at": started_at,
            }
        )

    bundles_q = db.query(ExercisePlannerFlowBundle).filter(
        ExercisePlannerFlowBundle.exercise_id == ex_id
    )
    if unit_filter:
        bundles_q = bundles_q.filter(ExercisePlannerFlowBundle.unit_level_key == unit_filter)
    bundles = (
        bundles_q.order_by(
            ExercisePlannerFlowBundle.unit_level_key,
            ExercisePlannerFlowBundle.exercise_phase,
            ExercisePlannerFlowBundle.id,
        )
        .all()
    )
    for bundle in bundles:
        uk = (bundle.unit_level_key or "").strip()
        phase_key = _normalized_exercise_phase(getattr(bundle, "exercise_phase", None))
        if not uk:
            continue
        action_rows = (
            db.query(ExercisePlannerFlowBundleActionEval)
            .filter(ExercisePlannerFlowBundleActionEval.bundle_id == bundle.id)
            .order_by(
                ExercisePlannerFlowBundleActionEval.slot_index,
                ExercisePlannerFlowBundleActionEval.id,
            )
            .all()
        )
        for action_row in action_rows:
            if not (action_row.file_relpath or "").strip():
                continue
            title = _planner_blob_display_filename(
                stored_title=action_row.title or "",
                relpath=action_row.file_relpath or "",
                fallback=f"قائمة تقييم إجراءات {int(action_row.slot_index)}",
            )
            canon = _planner_bundle_eval_canonical_saved(db, ex_id, int(action_row.id))
            total_pct = _evaluation_saved_total_pct(canon)
            if planner_open_uses_slot:
                open_href = url_for(
                    planner_open_endpoint,
                    slot=int(action_row.slot_index),
                )
            else:
                open_href = url_for(
                    planner_open_endpoint,
                    unit_key=uk,
                    action_eval_id=int(action_row.id),
                )
            shim = type(
                "_PfActionShim",
                (),
                {
                    "id": int(action_row.id),
                    "text": title,
                    "created_at": getattr(action_row, "created_at", None),
                },
            )()
            rows.append(
                {
                    **build_evaluation_list_row(
                        item=shim,
                        saved=canon,
                        exercise=exercise,
                        open_href=open_href,
                    ),
                    "list_type_kind": _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION,
                    "list_type_label": _control_eval_list_type_label(
                        _CONTROL_EVAL_LIST_TYPE_PLANNER_ACTION
                    ),
                    "phase_key": phase_key,
                    "phase_label": _phase_label_ar(phase_key),
                    "workflow_label": eval_workflow_label_ar(canon),
                    "unit_key": uk,
                    "unit_label": label_for_unit_level_key(uk) or uk or "—",
                    "total_pct": total_pct,
                    "sort_tiebreaker": int(action_row.id),
                    "started_at": getattr(action_row, "created_at", None),
                }
            )
    return rows


def _build_incomplete_evaluations_report(db, user: User, *, role: str) -> dict:
    """قوائم التقييم غير المنجزة فقط — مرتبة حسب وقت بدء المهمة."""
    if role == "judge":
        ex0 = _current_workspace_exercise(db, user)
        unit_filter = None
        if (
            ex0 is not None
            and not is_system_admin(user)
            and not can_access_chief_judge_hub(user)
        ):
            a = _judge_assignment_for_current_exercise(db, user, ex0)
            unit_filter = (getattr(a, "unit_level_key", "") or "").strip() if a else None
        eval_ep = "views.judge_evaluation_list_file_viewer"
        pf_ep = "views.judge_planner_flow_materials_action_evaluate"
        pf_slot = True
    elif role == "planner":
        ex0 = _admin_current_workspace_exercise(db, user)
        unit_filter = None
        eval_ep = "views.planner_evaluation_list_file_viewer"
        pf_ep = "views.control_planner_flow_action_view"
        pf_slot = False
    elif role == "analyst":
        ex0 = _admin_current_workspace_exercise(db, user)
        unit_filter = None
        eval_ep = "views.analyst_evaluation_list_file_viewer"
        pf_ep = "views.control_planner_flow_action_view"
        pf_slot = False
    else:
        ex0 = _current_workspace_exercise(db, user)
        unit_filter = None
        eval_ep = "views.control_evaluation_list_file_viewer"
        pf_ep = "views.control_planner_flow_action_view"
        pf_slot = False

    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    all_rows = _collect_all_eval_status_rows_flat(
        db,
        exercise=ex,
        unit_filter=unit_filter or None,
        eval_open_endpoint=eval_ep,
        planner_open_endpoint=pf_ep,
        planner_open_uses_slot=pf_slot,
    )
    incomplete_rows = [
        r for r in all_rows if not r.get("status_done")
    ]
    incomplete_rows.sort(key=_incomplete_eval_started_at_sort_key)
    return {
        "has_exercise": True,
        "incomplete_rows": incomplete_rows,
        "incomplete_count": len(incomplete_rows),
    }


def _render_incomplete_evaluations_page(
    user: User,
    *,
    section_title: str,
    section_icon: str = "fa-hourglass-half",
    role: str,
    extra_ctx: dict | None = None,
):
    from flask import g

    report = _build_incomplete_evaluations_report(g.db, user, role=role)
    ctx = {
        "section_title": section_title,
        "section_icon": section_icon,
        "role": role,
    }
    if extra_ctx:
        ctx.update(extra_ctx)
    if not report.get("has_exercise"):
        ctx["has_exercise"] = False
        ctx["incomplete_rows"] = []
        ctx["incomplete_count"] = 0
    else:
        ctx.update(report)
    if role == "control":
        ctx.update(_control_hub_back_ctx_always())
    else:
        ctx.update(_hub_back_ctx_for_request_path())
    return render_template("incomplete_evaluations.html", **_ctx(user, **ctx))


def _build_control_positives_negatives(
    db,
    user: User,
    *,
    list_viewer: str = "views.control_evaluation_list_file_viewer",
) -> dict:
    """إيجابيات وسلبيات من الملاحظات الكتابية في قوائم التقييم بعد حفظ واعتماد المحكم."""
    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {"has_exercise": False}
    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    if ex is None:
        return {"has_exercise": False}

    items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id)
        .order_by(
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    item_ids = [int(it.id) for it in items if getattr(it, "id", None) is not None]
    canonical_by_item = (
        _evaluation_canonical_map_for_items(db, int(ex.id), item_ids) if item_ids else {}
    )

    positive_notes: list[dict] = []
    negative_notes: list[dict] = []
    lists_with_notes: set[int] = set()
    n_approved = 0

    for it in items:
        saved = canonical_by_item.get(int(it.id))
        if saved is None or not eval_judge_approved(saved):
            continue
        if not (getattr(saved, "payload_json", "") or "").strip():
            continue
        n_approved += 1
        uk = (it.unit_level_key or "").strip()
        unit_label = label_for_unit_level_key(uk) or uk or "—"
        list_title = (it.text or "قائمة تقييم").strip()
        phase_label = _phase_label_ar(getattr(it, "exercise_phase", None))
        open_href = url_for(
            list_viewer,
            unit_key=uk,
            item_id=int(it.id),
        )
        list_has_note = False
        for row in _parse_saved_eval_rows(saved.payload_json):
            if not isinstance(row, dict):
                continue
            note = (row.get("notes") or "").strip()
            if not note:
                continue
            list_has_note = True
            pct = _eval_row_score_pct(row)
            element = (row.get("element") or "").strip() or "—"
            entry = {
                "list_title": list_title,
                "unit_label": unit_label,
                "phase_label": phase_label,
                "element": element[:200],
                "note": note,
                "pct": pct,
                "grade": grade_label_from_percent(pct) if pct is not None else "—",
                "open_href": open_href,
                "approved_at": getattr(saved, "approved_at", None),
            }
            band = _pct_status_band(pct)
            if band == "high":
                positive_notes.append(entry)
            elif band == "low":
                negative_notes.append(entry)
        if list_has_note:
            lists_with_notes.add(int(it.id))

    positive_notes.sort(
        key=lambda x: (
            -(float(x["pct"]) if x.get("pct") is not None else -1.0),
            x["unit_label"],
            x["list_title"],
        )
    )
    negative_notes.sort(
        key=lambda x: (
            (float(x["pct"]) if x.get("pct") is not None else 101.0),
            x["unit_label"],
            x["list_title"],
        )
    )

    return {
        "has_exercise": True,
        "exercise": ex,
        "positive_notes": positive_notes,
        "negative_notes": negative_notes,
        "n_positive": len(positive_notes),
        "n_negative": len(negative_notes),
        "n_lists_with_notes": len(lists_with_notes),
        "n_eval_lists": len(items),
        "n_approved_eval_lists": n_approved,
    }


def _control_exercise_performance_report(db, user: User) -> dict:
    """تقرير السيطرة الشامل وفق النموذج التشغيلي المعتمد في الصورة المرجعية.

    ملاحظة: التصميم يُحافظ عليه كما هو. رسم المجموعات ومؤشر المتوسط العام
    يعتمدان على النتائج المحفوظة؛ بقية بعض المؤشرات ما زالت من المعتمد.
    """

    def _ar_month_name(m: int) -> str:
        return {
            1: "يناير",
            2: "فبراير",
            3: "مارس",
            4: "أبريل",
            5: "مايو",
            6: "يونيو",
            7: "يوليو",
            8: "أغسطس",
            9: "سبتمبر",
            10: "أكتوبر",
            11: "نوفمبر",
            12: "ديسمبر",
        }.get(int(m or 0), "")

    def _fmt_m_ss(seconds: float | None) -> str:
        if seconds is None:
            return "—"
        s = int(round(max(0.0, float(seconds))))
        mm = s // 60
        ss = s % 60
        return f"{mm}:{ss:02d}"

    def _avg(xs: list[float]) -> float | None:
        xs2 = [float(x) for x in xs if x is not None]
        return (sum(xs2) / len(xs2)) if xs2 else None

    def _radar_points(values_pct: list[float | None]) -> str:
        # نقاط سداسية (مثل SVG في القالب): أعلى ثم يمين-أعلى ثم يمين-أسفل ثم أسفل ثم يسار-أسفل ثم يسار-أعلى.
        cx, cy = 200.0, 127.5
        outer = [(200.0, 35.0), (285.0, 80.0), (285.0, 170.0), (200.0, 220.0), (115.0, 170.0), (115.0, 80.0)]
        pts: list[str] = []
        for i in range(6):
            v = values_pct[i] if i < len(values_pct) else None
            p = max(0.0, min(100.0, float(v))) / 100.0 if v is not None else 0.0
            ox, oy = outer[i]
            x = cx + (ox - cx) * p
            y = cy + (oy - cy) * p
            pts.append(f"{x:.0f},{y:.0f}")
        if pts:
            pts.append(pts[0])
        return " ".join(pts)

    ex0 = _current_workspace_exercise(db, user)
    if ex0 is None:
        return {
            "exercise": None,
            "crumb": "",
            "criteria_count": 0,
            "report_title": "تقرير شامل لأداء التمرين",
            "exercise_label": "—",
            "exercise_code": "—",
            "meta": [],
            "kpis": [],
            "axis_scores": [],
            "group_scores": [],
            "timeline": [],
            "distribution": [],
            "phase_donut_css": "conic-gradient(var(--tint-200, #e8e0d8) 0deg 360deg)",
            "phase_summary": {"phase_summaries": [], "exercise_pct": None, "exercise_grade": "—", "phase_count": 0},
            "unit_detail_rows": [],
            "unit_detail_phase_headers": [],
            "unit_detail_phase_max_dots": [],
            "unit_detail_list_number_row": [],
            "grade_legend": _control_report_grade_legend(),
            "detail_source_legend": _control_report_detail_source_legend(),
            "n_saved_eval": 0,
            "n_eval_lists_total": 0,
            "radar_series": [],
        }

    ex = db.query(Exercise).filter(Exercise.id == ex0.id).first()
    eval_items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex0.id)
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            _unit_level_order_expr(EvaluationListPdfItem.unit_level_key),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )

    item_ids = [int(it.id) for it in eval_items if getattr(it, "id", None) is not None]
    canon_by_item = _evaluation_canonical_map_for_items(db, ex0.id, item_ids) if item_ids else {}
    saved_by_item = {
        iid: sr
        for iid, sr in canon_by_item.items()
        if (getattr(sr, "payload_json", None) or "").strip()
    }
    approved_by_item = {iid: sr for iid, sr in canon_by_item.items() if bool(getattr(sr, "is_approved", False))}

    n_eval_lists = len(eval_items)
    n_saved = len(saved_by_item)
    n_approved = len(approved_by_item)

    # متوسط زمن التنفيذ لكل تقييم (تقريب من إنشاء السجل حتى آخر تحديث/اعتماد)
    durations: list[float] = []
    for sr in approved_by_item.values():
        t0 = getattr(sr, "created_at", None)
        t1 = getattr(sr, "approved_at", None) or getattr(sr, "updated_at", None)
        if not t0 or not t1:
            continue
        try:
            sec = (t1 - t0).total_seconds()
        except Exception:
            continue
        if 0 <= sec <= 60 * 60 * 24:
            durations.append(float(sec))
    avg_duration = _avg(durations)

    # إجمالي المعايير/البنود المحسوبة + توزيعها
    scored_row_pcts: list[float] = []
    for sr in approved_by_item.values():
        for r in _parse_saved_eval_rows(getattr(sr, "payload_json", None)):
            if not isinstance(r, dict):
                continue
            pc = _eval_row_score_pct(r)
            if pc is None:
                continue
            scored_row_pcts.append(float(pc))
    criteria_count = len(scored_row_pcts)

    phase_summary = _phase_summary_for_control_report(
        db, ex0.id, eval_items, saved_by_item
    )
    distribution = _distribution_from_phase_summary(phase_summary)
    phase_donut_css = _distribution_phase_donut_css(distribution)

    # حالة الإنجاز: نسبة المعتمد مقابل الإجمالي (وأي عنصر غير معتمد يعتبر "قيد التقييم")
    pending_pct = int(round(((n_eval_lists - n_approved) / n_eval_lists) * 100.0)) if n_eval_lists else 0
    done_pct = max(0, 100 - pending_pct) if n_eval_lists else 0

    # متوسط عام من نتائج العناصر المعتمدة (إن وجدت)
    approved_totals: list[float] = []
    for sr in approved_by_item.values():
        v = getattr(sr, "total_pct", None)
        if v is None:
            # fallback: متوسط البنود داخل نفس العنصر
            rows = [float(x) for x in (_eval_row_score_pct(r) for r in _parse_saved_eval_rows(sr.payload_json)) if x is not None]
            v = (sum(rows) / len(rows)) if rows else None
        if v is not None:
            approved_totals.append(float(v))
    overall_avg = _avg(approved_totals) or (_avg(scored_row_pcts) if scored_row_pcts else None)
    saved_list_pcts: list[float] = []
    for sr in saved_by_item.values():
        sv = _evaluation_saved_total_pct(sr)
        if sv is not None:
            saved_list_pcts.append(float(sv))
    saved_overall_avg = _avg(saved_list_pcts)
    overall_avg_i = (
        int(round(saved_overall_avg))
        if saved_overall_avg is not None
        else (int(round(overall_avg)) if overall_avg is not None else 0)
    )

    # متوسط كل وحدة (مجموعات) من عناصر التقييم المحفوظة (لا يشترط الاعتماد)
    by_unit_vals: dict[str, list[float]] = {}
    for it in eval_items:
        sr = saved_by_item.get(int(getattr(it, "id", 0) or 0))
        if sr is None:
            continue
        uk = (
            (getattr(it, "unit_level_key", "") or getattr(sr, "unit_level_key", "") or "")
            .strip()
        )
        if not uk:
            continue
        v = _evaluation_saved_total_pct(sr)
        if v is None:
            continue
        by_unit_vals.setdefault(uk, []).append(float(v))

    for action_row in (
        db.query(ExercisePlannerFlowBundleActionEval)
        .join(ExercisePlannerFlowBundle)
        .filter(ExercisePlannerFlowBundle.exercise_id == ex0.id)
        .all()
    ):
        bundle = action_row.bundle
        if bundle is None:
            continue
        canon = _planner_bundle_eval_canonical_saved(db, ex0.id, int(action_row.id))
        if canon is None or not (getattr(canon, "payload_json", None) or "").strip():
            continue
        uk = (getattr(bundle, "unit_level_key", "") or getattr(canon, "unit_level_key", "") or "").strip()
        if not uk:
            continue
        v = _evaluation_saved_total_pct(canon)
        if v is None:
            continue
        by_unit_vals.setdefault(uk, []).append(float(v))

    # ألوان هادئة موحّدة مع تقرير السيطرة (تُطبَّق أيضاً عبر CSS للأعمدة)
    palette = [
        "#6b8cae",
        "#7a9a7e",
        "#9a8bb8",
        "#b89a6e",
        "#5f9a9f",
        "#a67f72",
        "#8a8f7a",
        "#7d8fa8",
    ]
    unit_order = {row["key"]: idx for idx, row in enumerate(UNIT_LEVELS)}
    unit_avg_rows: list[dict] = []
    seen_unit_keys: set[str] = set()

    def _append_unit_avg_row(unit_key: str, sort_idx: int) -> None:
        uk = (unit_key or "").strip()
        if not uk or uk in seen_unit_keys:
            return
        vals = by_unit_vals.get(uk) or []
        avg_u = _avg(vals)
        if avg_u is None:
            return
        seen_unit_keys.add(uk)
        unit_avg_rows.append(
            {
                "unit_key": uk,
                "label": label_for_unit_level_key(uk, db) or uk,
                "value": int(round(avg_u)),
                "raw": float(avg_u),
                "color": palette[len(unit_avg_rows) % len(palette)],
                "_sort": sort_idx,
            }
        )

    for idx, ul in enumerate(UNIT_LEVELS):
        _append_unit_avg_row(ul.get("key") or "", idx)
    for uk in sorted(by_unit_vals.keys()):
        _append_unit_avg_row(uk, unit_order.get(uk, 9000))
    unit_avg_rows.sort(key=lambda r: (int(r.get("_sort", 9000)), r["label"]))
    for row in unit_avg_rows:
        row.pop("_sort", None)
    sorted_by_perf = sorted(
        unit_avg_rows, key=lambda r: float(r.get("raw", 0.0)), reverse=True
    )
    top_unit = sorted_by_perf[0] if sorted_by_perf else None
    bottom_unit = sorted_by_perf[-1] if sorted_by_perf else None
    # ترتيب الرسم = تسلسل قوائم الوحدات (كتالوج UNIT_LEVELS) من الأعلى إلى الأسفل
    group_scores = []
    for seq, r in enumerate(unit_avg_rows, start=1):
        val = int(r["value"])
        band = _control_report_grade_band(float(val))
        group_scores.append(
            {
                "seq": seq,
                "label": r["label"],
                "value": val,
                "grade_band": band,
                "color": _CONTROL_REPORT_GRADE_COLORS[band],
            }
        )

    # محور/جدول: نستخدم مصفوفة الأهداف إن وجدت لتوليد 8 محاور بنفس عناوين الصورة.
    axis_labels = [
        ("القيادة والسيطرة", "#64b5f6"),
        ("التخطيط", "#7bd86f"),
        ("جمع المعلومات", "#78d06a"),
        ("العمليات", "#8b5cf6"),
        ("الاستطلاع", "#f59e0b"),
        ("الدعم اللوجستي", "#d7b735"),
        ("الاتصالات", "#fb923c"),
        ("الأمن والدفاع", "#ef4444"),
    ]
    dash = _build_analyst_evaluation_results_dashboard(db, user, approved_only=True, matrix_mode="objectives")
    matrix_rows = dash.get("matrix_rows") or []

    def _axis_value_for_unit(unit_key: str, axis_idx: int) -> float | None:
        for mr in matrix_rows:
            if (mr.get("unit_key") or "").strip() != (unit_key or "").strip():
                continue
            cells = mr.get("cells") or []
            if axis_idx < len(cells):
                v = cells[axis_idx].get("pct")
                return float(v) if v is not None else None
        return None

    axis_scores: list[dict] = []
    for i, (lbl, col) in enumerate(axis_labels):
        vals_i: list[float] = []
        for mr in matrix_rows:
            cells = mr.get("cells") or []
            if i < len(cells):
                v = cells[i].get("pct")
                if v is not None:
                    vals_i.append(float(v))
        vavg = _avg(vals_i)
        if vavg is None:
            vavg = float(overall_avg_i)
        axis_scores.append({"label": lbl, "value": int(round(vavg)), "color": col})

    table_headers = [
        "الوحدة / المحور",
        "القيادة والسيطرة",
        "جمع المعلومات",
        "العمليات",
        "المعلومات",
        "الاستطلاع",
        "الدعم اللوجستي",
        "الاتصالات",
        "المتوسط",
    ]

    # اختيار 6 وحدات للجدول (نفس الظاهر في عمود المجموعات، مع fallback لأول وحدات الكتالوج)
    units_for_table = []
    for r in unit_avg_rows[:6]:
        units_for_table.append({"unit_key": r["unit_key"], "label": r["label"]})
    if not units_for_table:
        for ul in UNIT_LEVELS[:6]:
            units_for_table.append({"unit_key": (ul.get("key") or ""), "label": (ul.get("label") or ul.get("key") or "—")})

    table_rows: list[list] = []
    # ملاحظة: الأعمدة الوسطية في الجدول تشمل "المعلومات" كعنوان مستقل في الصورة؛ نربطه بمحور التخطيط (index=1) إذا لم تتوفر أهداف كافية.
    for u in units_for_table:
        uk = (u.get("unit_key") or "").strip()
        if not uk:
            continue
        axis_vals = []
        # الأعمدة: القيادة(0), جمع المعلومات(2), العمليات(3), المعلومات(2 fallback), الاستطلاع(4), اللوجستي(5), الاتصالات(6)
        mapping = [0, 2, 3, 2, 4, 5, 6]
        for ax in mapping:
            v = _axis_value_for_unit(uk, ax)
            axis_vals.append(int(round(v)) if v is not None else 0)
        avg_row = int(round(sum(axis_vals) / len(axis_vals))) if axis_vals else 0
        table_rows.append([u["label"], *axis_vals, avg_row])

    # تقدم زمني: نسبة المعتمد تراكميًا حسب آخر 8 أيام بيانات
    end_dates: list[datetime] = []
    for sr in canon_by_item.values():
        t = getattr(sr, "approved_at", None) or getattr(sr, "updated_at", None) or getattr(sr, "created_at", None)
        if isinstance(t, datetime):
            end_dates.append(t)
    if end_dates:
        max_dt = max(end_dates)
        # اجمع آخر 8 أيام مميزة
        seen = set()
        cur = datetime(max_dt.year, max_dt.month, max_dt.day, 0, 0, 0)
        while len(seen) < 8:
            key = (cur.year, cur.month, cur.day)
            seen.add(key)
            cur = cur.replace()  # no-op for clarity
            cur = cur.fromtimestamp((cur.timestamp() - 86400))
            if len(seen) >= 8:
                break
        # إعادة بناء قائمة مرتبة تصاعديًا
        days = sorted(seen)
        day_dts = [datetime(y, m, d, 23, 59, 59) for (y, m, d) in days][-8:]
    else:
        day_dts = []

    def _approved_count_upto(dt: datetime) -> int:
        n = 0
        for sr in approved_by_item.values():
            t = getattr(sr, "approved_at", None) or getattr(sr, "updated_at", None)
            if not isinstance(t, datetime):
                continue
            if t <= dt:
                n += 1
        return n

    timeline = []
    for dt in day_dts:
        c = _approved_count_upto(dt)
        v = int(round((c / n_eval_lists) * 100.0)) if n_eval_lists else 0
        label = f"{dt.day} {_ar_month_name(dt.month)}"
        timeline.append({"label": label, "value": v})

    # رادار: أعلى 4 وحدات (6 محاور) — نأخذ أول 6 محاور من axis_scores.
    radar_palette = ["#3b82f6", "#84cc16", "#f59e0b", "#8b5cf6"]
    radar_series = []
    for i, u in enumerate(unit_avg_rows[:4]):
        uk = u.get("unit_key") or ""
        vals6 = []
        for ax in range(6):
            v = _axis_value_for_unit(uk, ax)
            vals6.append(v if v is not None else float(overall_avg_i))
        radar_series.append({"label": u.get("label") or uk, "color": radar_palette[i % len(radar_palette)], "points": _radar_points(vals6)})

    # بيانات Meta أعلى التقرير
    n_specialists = int(
        db.query(func.count(ExerciseRosterRow.id))
        .filter(
            ExerciseRosterRow.exercise_id == ex0.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
        )
        .scalar()
        or 0
    )
    ex_dt = getattr(ex, "planned_start", None) or getattr(ex, "created_at", None)
    dt_label = ex_dt.strftime("%d-%m-%Y") if isinstance(ex_dt, datetime) else "—"

    crumb = f"التمرين الحالي / {(getattr(ex, 'title', None) or 'تمرين')}"
    exercise_label = (getattr(ex, "title", "") or "تمرين").strip()
    exercise_code = (getattr(ex, "code", "") or "—").strip()

    meta = [
        f"قوائم التقييم: {n_eval_lists}",
        f"المحفوظة: {n_saved}",
        f"المختصين: {n_specialists}",
        f"تاريخ التمرين: {dt_label}",
    ]

    kpis = [
        {"label": "متوسط الزمن", "value": _fmt_m_ss(avg_duration), "hint": "لكل تقييم", "icon": "fa-clock", "tone": "blue"},
        {"label": "قيد التقييم", "value": f"{pending_pct}%", "hint": "من إجمالي التقييم", "icon": "fa-hexagon-nodes", "tone": "violet"},
        {"label": "نسبة الإستكمال", "value": f"{done_pct}%", "hint": "من إجمالي التقييم", "icon": "fa-circle-check", "tone": "purple"},
        {"label": "أقل مجموعة", "value": f"{int(bottom_unit['value']) if bottom_unit else 0}%", "hint": (bottom_unit["label"] if bottom_unit else "—"), "icon": "fa-arrow-down", "tone": "red"},
        {"label": "أعلى مجموعة", "value": f"{int(top_unit['value']) if top_unit else 0}%", "hint": (top_unit["label"] if top_unit else "—"), "icon": "fa-trophy", "tone": "cyan"},
        {"label": "المتوسط العام", "value": f"{overall_avg_i}%", "hint": "أداء التمرين", "icon": "fa-chart-line", "tone": "green"},
        {"label": "معايير التقييم", "value": str(criteria_count), "hint": "إجمالي المعايير", "icon": "fa-list", "tone": "indigo"},
    ]

    unit_detail_rows, unit_detail_phase_columns = _control_build_unit_detail_rows(
        db, ex0.id, eval_items, saved_by_item
    )
    unit_detail_phase_max_dots = _control_phase_max_dot_counts(unit_detail_rows)
    unit_detail_list_number_row = _control_build_list_number_row(unit_detail_phase_max_dots)

    return {
        "exercise": ex,
        "crumb": crumb,
        "criteria_count": criteria_count,
        "report_title": "تقرير شامل لأداء التمرين",
        "exercise_label": exercise_label,
        "exercise_code": exercise_code,
        "meta": meta,
        "kpis": kpis,
        "axis_scores": axis_scores,
        "group_scores": group_scores,
        "timeline": timeline,
        "distribution": distribution,
        "phase_donut_css": phase_donut_css,
        "phase_summary": phase_summary,
        "table_rows": table_rows,
        "table_headers": table_headers,
        "unit_detail_rows": unit_detail_rows,
        "unit_detail_phase_headers": [lbl for _, lbl in unit_detail_phase_columns],
        "unit_detail_phase_max_dots": unit_detail_phase_max_dots,
        "unit_detail_list_number_row": unit_detail_list_number_row,
        "grade_legend": _control_report_grade_legend(),
        "detail_source_legend": _control_report_detail_source_legend(),
        "n_saved_eval": n_saved,
        "n_eval_lists_total": n_eval_lists,
        "radar_series": radar_series,
    }


def _control_report_detail_source_legend() -> list[dict]:
    return [
        {
            "kind": "judge_eval",
            "label": "قوائم التقييم (المحكم)",
            "hint": "دائرة ملونة",
        },
        {
            "kind": "planner_flow",
            "label": "المجرى وتقييم الإجراءات — قائمة التقييم",
            "hint": "مربع ملون",
        },
    ]


@bp.route("/control/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def control_evaluation_list_file_viewer(unit_key: str, item_id: int):
    """عرض قائمة تقييم للسيطرة (قراءة فقط) من تقرير النتائج."""
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/control/evaluation-lists/{unit_key}/view/{item_id}")
    if not can_access_control_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    ex = _current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key:
        abort(404)
    if ex is not None and row.exercise_id not in (None, ex.id):
        abort(404)
    if not (row.pdf_relpath or "").strip():
        abort(404)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        abort(404)

    ev = _evaluation_sheet_view_context(fspath)
    saved_payload = {}
    saved_updated_at = None
    saved_row_id = None
    saved_is_approved = False
    saved_approved_at = None
    if ex is not None:
        saved_row = _evaluation_canonical_map_for_items(db, ex.id, [int(row.id)]).get(int(row.id))
        if saved_row and (saved_row.payload_json or "").strip():
            try:
                saved_payload = json.loads(saved_row.payload_json)
            except Exception:
                saved_payload = {}
            saved_updated_at = saved_row.updated_at
            saved_row_id = saved_row.id
            saved_is_approved = bool(getattr(saved_row, "is_approved", False))
            saved_approved_at = getattr(saved_row, "approved_at", None)

    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = None
    commander_name = "—"
    judge_name = "—"
    if ex is not None:
        shown_date = getattr(ex, "planned_start", None) or getattr(ex, "created_at", None)
        commander_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if commander_row is not None:
            commander_name = (commander_row.full_name or "").strip() or commander_name
        judge_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if judge_row is not None:
            judge_name = (judge_row.full_name or "").strip() or judge_name

    return render_template(
        "admin_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            can_edit=False,
            close_href=url_for("views.control_hub_section", slug="evaluation-lists-status"),
            subpage_close_fallback=url_for("views.control_hub_section", slug="evaluation-lists-status"),
            close_label="العودة إلى موقف القوائم",
            **_control_hub_back_ctx_always(),
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            saved_is_approved=saved_is_approved,
            saved_approved_at=saved_approved_at,
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name,
            judge_name=judge_name,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=ex,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=False,
            ),
        ),
    )


@bp.route(
    "/control/planner-flow-action/<unit_key>/view/<int:action_eval_id>",
    methods=["GET"],
)
def control_planner_flow_action_view(unit_key: str, action_eval_id: int):
    """عرض قائمة تقييم إجراءات (المجرى) للسيطرة — قراءة فقط من تقرير النتائج."""
    user = get_current_user_optional()
    if not user:
        return redirect(
            f"/login?next=/control/planner-flow-action/{unit_key}/view/{int(action_eval_id)}"
        )
    if not can_access_control_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    if ex is None:
        abort(404)
    action_row = db.get(ExercisePlannerFlowBundleActionEval, int(action_eval_id))
    if action_row is None:
        abort(404)
    bundle = db.get(ExercisePlannerFlowBundle, int(action_row.bundle_id))
    if (
        bundle is None
        or bundle.exercise_id != ex.id
        or (bundle.unit_level_key or "").strip() != unit_key
    ):
        abort(404)
    path = _planner_bundle_file_abspath(action_row.file_relpath)
    if path is None:
        abort(404)
    ev = _evaluation_sheet_view_context(path)
    canon = _planner_bundle_eval_canonical_saved(db, ex.id, action_row.id)
    saved_payload: dict = {}
    saved_updated_at = None
    saved_row_id = None
    if canon is not None and (canon.payload_json or "").strip():
        try:
            saved_payload = json.loads(canon.payload_json)
        except Exception:
            saved_payload = {}
        if not isinstance(saved_payload, dict):
            saved_payload = {}
        saved_updated_at = canon.updated_at
        saved_row_id = canon.id
    unit_label_pf = label_for_unit_level_key(unit_key) or (
        bundle.unit_level_label or ""
    ).strip() or unit_key
    item_title = _planner_blob_display_filename(
        stored_title=action_row.title or "",
        relpath=action_row.file_relpath or "",
        fallback=f"قائمة تقييم إجراءات — {int(action_row.slot_index)}",
    ).strip()
    shown_date = getattr(ex, "planned_start", None) or getattr(ex, "created_at", None)
    commander_name = "—"
    commander_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if commander_row is not None:
        commander_name = (commander_row.full_name or "").strip() or commander_name
    judge_name = "—"
    judge_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if judge_row is not None:
        judge_name = (judge_row.full_name or "").strip() or judge_name
    return render_template(
        "judge_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=int(action_row.id),
            evaluation_item_id=int(action_row.id),
            saved_row_id=saved_row_id,
            item_title=item_title,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            eval_can_edit=False,
            eval_can_approve=False,
            eval_can_chief_approve=False,
            eval_can_chief_reopen=False,
            eval_can_control_approve=False,
            eval_close_href=url_for("views.control_hub_section", slug="evaluation-results"),
            eval_close_label="العودة إلى التقرير الشامل",
            subpage_close_fallback=url_for("views.control_hub_section", slug="evaluation-results"),
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=ex,
                list_item_id=None,
                bundle_action_eval_id=int(action_row.id),
                eval_can_edit=False,
            ),
            unit_label=unit_label_pf or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url="",
            eval_approve_url="",
            eval_chief_approve_url="",
            eval_chief_reopen_url="",
            eval_approve_incomplete=False,
            **_control_hub_back_ctx_always(),
        ),
    )


@bp.route("/control")
def control_hub():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/control")
    if not can_access_control_hub(user):
        abort(403)
    hub_items = [{"slug": s, "title_ar": t, "icon": ic} for s, t, ic in CONTROL_HUB_ITEMS]
    return render_template(
        "control_hub.html",
        **_ctx(
            user,
            hub_items=hub_items,
        ),
    )


@bp.route("/control/<slug>")
def control_hub_section(slug: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/control/{slug}")
    if not can_access_control_hub(user):
        abort(403)
    slug_norm = (slug or "").strip().lower()
    if slug_norm in ("visual-doc-status", "visual-documentation"):
        return redirect(url_for("views.visual_documentation", from_control=1))
    title = CONTROL_HUB_SLUGS.get(slug_norm)
    if not title:
        abort(404)
    if slug_norm == "evaluation-results":
        from flask import g

        return render_template(
            "control_evaluation_results_report.html",
            **_ctx(
                user,
                section_title=title,
                section_icon=_control_section_icon(slug_norm),
                report=_control_exercise_performance_report(g.db, user),
                **_control_hub_back_ctx_always(),
            ),
        )
    if slug_norm == "evaluation-lists-status":
        from flask import g

        status = _build_control_evaluation_lists_status(g.db, user)
        if not status.get("has_exercise"):
            return render_template(
                "control_evaluation_lists_status.html",
                **_ctx(
                    user,
                    section_title=title,
                    section_icon=_control_section_icon(slug_norm),
                    has_exercise=False,
                    **_control_hub_back_ctx_always(),
                ),
            )
        return render_template(
            "control_evaluation_lists_status.html",
            **_ctx(
                user,
                section_title=title,
                section_icon=_control_section_icon(slug_norm),
                **_control_hub_back_ctx_always(),
                **status,
            ),
        )
    if slug_norm == "incomplete-tasks-status":
        return _render_incomplete_evaluations_page(
            user,
            section_title=title,
            section_icon=_control_section_icon(slug_norm),
            role="control",
        )
    if slug_norm == "top-positives-negatives":
        from flask import g

        pn = _build_control_positives_negatives(g.db, user)
        if not pn.get("has_exercise"):
            return render_template(
                "control_positives_negatives.html",
                **_ctx(
                    user,
                    section_title=title,
                    section_icon=_control_section_icon(slug_norm),
                    has_exercise=False,
                    **_control_hub_back_ctx_always(),
                ),
            )
        return render_template(
            "control_positives_negatives.html",
            **_ctx(
                user,
                section_title=title,
                section_icon=_control_section_icon(slug_norm),
                **_control_hub_back_ctx_always(),
                **pn,
            ),
        )
    return render_template(
        "control_section_placeholder.html",
        **_ctx(
            user,
            section_title=title,
            section_slug=slug,
            section_icon=_control_section_icon(slug_norm),
            **_control_hub_back_ctx_always(),
        ),
    )


@bp.route("/admin")
def admin_root_redirect():
    """لوحة /admin أُلغيت؛ إعادة التوجيه للصفحة الرئيسية أو تسجيل الدخول."""
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/dashboard")
    if not is_system_admin(user):
        abort(403)
    return redirect("/dashboard")


@bp.route("/admin/exercises/create", methods=["GET", "POST"])
def admin_exercise_create():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/create")
    if not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db

    def _render_create_page(*, error: str = "", success: str = ""):
        ex_cur = _workspace_exercise_for_admin_form(db, user)
        if request.method == "GET":
            form_prefill = (
                _prefill_create_form_from_exercise(ex_cur)
                if ex_cur is not None
                else _empty_create_form_prefill()
            )
        else:
            form_prefill = _prefill_create_form_from_request()
        from flask import make_response

        resp = make_response(
            render_template(
                "admin_exercise_create.html",
                **_ctx(
                    user,
                    error=error,
                    success=success,
                    export_dir=str(export_directory()),
                    form_prefill=form_prefill,
                    has_current_exercise=ex_cur is not None,
                    workspace_exercise=ex_cur,
                    form_build_tag="20260602-create-v4",
                    **_admin_exercise_form_ctx(),
                ),
            )
        )
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        return resp

    if request.method == "GET":
        qerr = (request.args.get("err") or "").strip()
        qok = (request.args.get("ok") or "").strip()
        return _render_create_page(error=qerr, success=qok)

    def _pick(field: str, allowed: list[str]) -> str | None:
        v = (request.form.get(field) or "").strip()
        return v if v in allowed else None

    def _text(field: str, max_len: int) -> str | None:
        v = (request.form.get(field) or "").strip()
        return v[:max_len] if v else None

    def _parse_dt_local(field: str):
        raw = (request.form.get(field) or "").strip()
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw)
        except Exception:
            return None

    title = _text("exercise_name", 500)
    et = _pick("exercise_type", ex_opts.EXERCISE_TYPES)
    el = _pick("exercise_level", ex_opts.EXERCISE_LEVELS)
    mission = _pick("mission", ex_opts.MISSIONS)
    unit = _text("trained_unit", 400)
    loc = _text("location_label", 400)
    planned_start = _parse_dt_local("planned_start")
    planned_end = _parse_dt_local("planned_end")

    if planned_start and planned_end and planned_end < planned_start:
        return _render_create_page(error="تاريخ/وقت النهاية يجب أن يكون بعد البداية."), 400

    missing: list[str] = []
    if not title:
        missing.append("اسم التمرين")
    if not et:
        missing.append("نوع التمرين")
    if not el:
        missing.append("مستوى التمرين")
    if not mission:
        missing.append("المهمة")
    if not unit:
        missing.append("اسم الوحدة المتدربة")
    if not loc:
        missing.append("مكان التمرين")
    if missing:
        return (
            _render_create_page(
                error="يرجى تعبئة الحقول التالية: " + "، ".join(missing)
            ),
            400,
        )

    ex_cur = _workspace_exercise_for_admin_form(db, user)
    if ex_cur is not None:
        ex_cur.title = title
        ex_cur.exercise_type = et
        ex_cur.exercise_level = el
        ex_cur.mission_label = mission
        ex_cur.trained_unit = unit
        ex_cur.location_label = loc
        ex_cur.planned_start = planned_start
        ex_cur.planned_end = planned_end
        db.commit()
        write_exercise_json_file(db, ex_cur.id)
        return redirect(
            "/admin/exercises/create?ok="
            + quote("تم حفظ بيانات التمرين الحالي (المعاضل والتقييمات والرسائل لم تُمس).", safe="")
        )

    purge_all_exercises_and_dilemmas(db)
    ex = Exercise(
        code=f"EX-{uuid.uuid4().hex[:8].upper()}",
        title=title,
        description="",
        exercise_type=et,
        exercise_level=el,
        mission_label=mission,
        trained_unit=unit,
        location_label=loc,
        status=ExerciseStatus.DRAFT.value,
        owner_id=user.id,
        planned_start=planned_start,
        planned_end=planned_end,
    )
    db.add(ex)
    db.commit()
    db.refresh(ex)
    write_exercise_json_file(db, ex.id)

    return redirect(
        "/admin/exercises/create?ok="
        + quote("تم إنشاء التمرين وحفظ بياناته.", safe="")
    )


@bp.route("/admin/exercises/import-full-json", methods=["POST"])
def admin_exercise_import_full_json():
    """مسح التمارين السابقة واستيراد تمرين كامل من ملف JSON في مجلد التصدير."""
    user = get_current_user_optional()
    if not user:
        if _wants_import_json_response():
            return jsonify({"ok": False, "error": "يجب تسجيل الدخول"}), 401
        return redirect("/login?next=/admin/exercises/create")
    if not can_manage_users(user):
        if _wants_import_json_response():
            return jsonify({"ok": False, "error": "غير مسموح"}), 403
        abort(403)
    from flask import g

    db = g.db
    up = request.files.get("json_file")
    if not up or not getattr(up, "filename", ""):
        return _import_full_json_error("يرجى اختيار ملف JSON من مجلد التمارين.")
    try:
        raw = up.read()
    except Exception:
        raw = b""
    if not raw:
        return _import_full_json_error("الملف فارغ.")
    try:
        data = json.loads(raw.decode("utf-8"))
    except Exception:
        return _import_full_json_error("ملف JSON غير صالح أو تالف.")
    if not isinstance(data, dict) or not isinstance(data.get("exercise"), dict):
        return _import_full_json_error(
            "الملف لا يحتوي على كائن «exercise» كما في ملفات التصدير."
        )

    cur = _admin_current_workspace_exercise(db, user)
    if cur is not None:
        pwd = (request.form.get("system_admin_password") or "").strip()
        if not pwd:
            return _import_full_json_error(
                "يجب إدخال كلمة مرور إدارة النظام لتغيير التمرين الحالي."
            )
        if not verify_password(pwd, user.password_hash):
            return _import_full_json_error("كلمة مرور إدارة النظام غير صحيحة.")

    try:
        purge_all_exercises_and_dilemmas(db)
        eid = import_exercise_bundle_from_dict(db, data, user.id)
        if eid is None:
            db.rollback()
            return _import_full_json_error(
                "تعذر استيراد البيانات من الملف (بنية غير متوقعة)."
            )
        db.commit()
    except Exception:
        db.rollback()
        return _import_full_json_error("حدث خطأ أثناء استيراد التمرين.")
    return _import_full_json_ok(eid)


@bp.route("/admin/exercises/finish", methods=["GET", "POST"])
def admin_exercise_finish():
    """إنهاء التمرين الحالي: أرشفة JSON كامل ثم مسح بيانات التمرين (يبقى بنك المعلومات)."""
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/create")
    if not can_manage_users(user):
        abort(403)
    from flask import g
    from urllib.parse import quote

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    if ex is None:
        return redirect(
            "/admin/exercises/create?err="
            + quote("لا يوجد تمرين حالي لإنهائه.", safe="")
        )
    pwd = (request.form.get("system_admin_password") or "").strip()
    if not pwd:
        return redirect(
            "/admin/exercises/create?err="
            + quote("يجب إدخال كلمة مرور إدارة النظام لإنهاء التمرين.", safe="")
        )
    if not verify_password(pwd, user.password_hash):
        return redirect(
            "/admin/exercises/create?err="
            + quote("كلمة مرور إدارة النظام غير صحيحة.", safe="")
        )
    save_archive = (request.form.get("save_archive") or "1").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )
    try:
        if save_archive:
            path = archive_and_clear_current_exercise(db, ex.id, finished_by_id=user.id)
            if path is None:
                db.rollback()
                return redirect(
                    "/admin/exercises/create?err="
                    + quote("تعذر أرشفة التمرين الحالي.", safe="")
                )
            db.commit()
            msg = f"تم مسح التمرين من النظام وحفظ نسخة أرشيف في: {path}"
        else:
            if not wipe_exercise_from_system(db, ex.id):
                db.rollback()
                return redirect(
                    "/admin/exercises/create?err="
                    + quote("تعذر مسح التمرين الحالي.", safe="")
                )
            db.commit()
            msg = "تم مسح جميع بيانات التمرين والنتائج والتقييمات من النظام (بدون أرشفة). بنك المعلومات لم يُمس."
    except Exception:
        db.rollback()
        return redirect(
            "/admin/exercises/create?err="
            + quote("حدث خطأ أثناء مسح التمرين.", safe="")
        )
    return redirect("/admin/exercises/create?ok=" + quote(msg, safe=""))


@bp.route("/admin/exercises/open-export-dir", methods=["POST"])
def admin_open_export_dir():
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "يجب تسجيل الدخول"}), 401
    if not can_manage_users(user):
        return jsonify({"ok": False, "error": "غير مسموح"}), 403
    ok, err = open_export_directory_in_os()
    if ok:
        return jsonify({"ok": True, "path": str(export_directory())})
    return jsonify({"ok": False, "error": err or "تعذر فتح المجلد"}), 500


@bp.route("/admin/exercises/import-json-prefill", methods=["POST"])
def admin_exercise_import_json_prefill():
    """رفع ملف JSON (تصدير تمرين) واستخراج حقول تطابق نموذج الإنشاء."""
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "يجب تسجيل الدخول"}), 401
    if not can_manage_users(user):
        return jsonify({"ok": False, "error": "غير مسموح"}), 403
    f = request.files.get("json_file")
    if not f or not getattr(f, "filename", None):
        return jsonify({"ok": False, "error": "لم يُرفع ملف."}), 400
    raw = f.read()
    if not raw or not raw.strip():
        return jsonify({"ok": False, "error": "الملف فارغ."}), 400
    try:
        data = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError):
        return jsonify({"ok": False, "error": "صيغة JSON غير صالحة."}), 400
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": "ملف غير صالح."}), 400
    fields, warnings = extract_create_form_prefill_from_export_json(data)
    select_keys = (
        "exercise_name",
        "exercise_type",
        "exercise_level",
        "mission",
        "trained_unit",
        "location_label",
    )
    has_select = any(fields.get(k) for k in select_keys)
    has_dt = bool(fields.get("planned_start") or fields.get("planned_end"))
    if not has_select and not has_dt:
        msg = "لم تُطابق أي قيمة من الملف القوائم الحالية في النظام."
        if warnings:
            msg = warnings[0]
        return jsonify({"ok": False, "error": msg, "warnings": warnings}), 400
    out_fields = {k: v for k, v in fields.items() if v is not None and v != ""}
    return jsonify({"ok": True, "fields": out_fields, "warnings": warnings})


def _admin_current_workspace_exercise(db, user: User) -> Exercise | None:
    """التمرين الحالي لمسؤول النظام — آخر تمرين في النظام (عادة واحد فقط)."""
    return (
        db.query(Exercise)
        .options(
            joinedload(Exercise.objectives),
            joinedload(Exercise.roster_rows),
        )
        .order_by(Exercise.id.desc())
        .first()
    )


def _current_workspace_exercise(db, user: User) -> Exercise | None:
    """التمرين الحالي — آخر تمرين في قاعدة البيانات (عادة واحد فقط)."""
    return db.query(Exercise).order_by(Exercise.id.desc()).first()


def _sync_judges_from_roster(db, ex: Exercise) -> None:
    """مزامنة حسابات المحكمين وتخصيصاتهم اعتماداً على قائمة المحكمين + قائمة المتدربين للتمرين الحالي.

    القاعدة المطلوبة:
    - username = الرقم العسكري
    - password = الرقم العسكري
    - ربط المحكم بوحدة/متدرب عبر unit_level_key (نفس مفتاح المعاضل والتقييم)
    """
    if ex is None:
        return

    trainees = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
        )
        .all()
    )
    trainee_by_unit: dict[str, ExerciseRosterRow] = {}
    for tr in trainees:
        uk = (tr.unit_level_key or "").strip()
        if uk and uk not in trainee_by_unit:
            trainee_by_unit[uk] = tr

    judges = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
        )
        .all()
    )
    for jr in judges:
        mil = (jr.military_number or "").strip()
        if not mil:
            continue
        uk = (jr.unit_level_key or "").strip()
        # إن لم تتوفر ملفات معاضل/تقييم لهذا المفتاح، لا ننشئ تخصيصاً "فارغاً" لأن المحكم لن يرى شيئاً.
        has_any = (
            db.query(DilemmaItem)
            .filter(DilemmaItem.exercise_id == ex.id, DilemmaItem.unit_level_key == uk)
            .first()
            is not None
        ) or (
            db.query(EvaluationListPdfItem)
            .filter(EvaluationListPdfItem.exercise_id == ex.id, EvaluationListPdfItem.unit_level_key == uk)
            .first()
            is not None
        )
        if not uk or not has_any:
            # نترك الحساب يُنشأ/يُحدّث، لكن بدون تخصيص وحدة حتى يتم تصحيح القائمة.
            uk = ""

        u = db.query(User).filter(User.username == mil).first()
        if u is None:
            u = User(
                username=mil,
                full_name=(jr.full_name or "").strip(),
                role_key=RoleKey.JUDGE.value,
                password_hash=hash_password(mil),
                is_active=True,
            )
            db.add(u)
            db.commit()
            db.refresh(u)
        else:
            # لا نغيّر الأدوار الأخرى
            if (u.role_key or "") != RoleKey.JUDGE.value:
                continue
            u.full_name = (jr.full_name or "").strip() or u.full_name
            u.is_active = True
            # ضمان أن كلمة المرور = الرقم العسكري (حسب الطلب)
            u.password_hash = hash_password(mil)
            db.add(u)
            db.commit()

        # Upsert تخصيص المحكم لهذا التمرين
        db.execute(
            delete(JudgeTraineeAssignment).where(
                JudgeTraineeAssignment.exercise_id == ex.id,
                JudgeTraineeAssignment.judge_user_id == u.id,
            )
        )
        tr = trainee_by_unit.get(uk) if uk else None
        db.add(
            JudgeTraineeAssignment(
                exercise_id=ex.id,
                judge_user_id=u.id,
                unit_level_key=uk,
                trainee_name=(tr.full_name or "").strip() if tr else "",
                trainee_military_number=(tr.military_number or "").strip() if tr else "",
            )
        )
        db.commit()


@bp.route("/admin/exercises/objectives", methods=["GET", "POST"])
def admin_exercise_objectives():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/objectives")
    if not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db

    def _render(*, error: str, ok_msg: str, exercise: Exercise | None):
        return render_template(
            "admin_exercise_objectives.html",
            **_ctx(user, error=error, ok_msg=ok_msg, current_exercise=exercise),
        )

    if request.method == "GET":
        ex = _admin_current_workspace_exercise(db, user)
        return _render(error="", ok_msg="", exercise=ex)

    ex = _admin_current_workspace_exercise(db, user)
    if not ex:
        return (
            _render(
                error="لا يوجد تمرين حالي. أنشئ تمريناً جديداً أو افتح تمريناً من ملف JSON من صفحة «إنشاء تمرين جديد».",
                ok_msg="",
                exercise=None,
            ),
            400,
        )

    if (request.form.get("objectives_action") or "").strip() == "clear_all":
        db.execute(
            delete(ExerciseObjective).where(ExerciseObjective.exercise_id == ex.id)
        )
        db.commit()
        ex_show = (
            db.query(Exercise)
            .options(joinedload(Exercise.objectives))
            .filter(Exercise.id == ex.id)
            .one()
        )
        write_exercise_json_file(db, ex_show.id)
        return _render(
            error="",
            ok_msg="تم حذف جميع الأهداف التدريبية وتحديث ملف JSON في مجلد التصدير.",
            exercise=ex_show,
        )

    source = (request.form.get("objectives_source") or "manual").strip()
    if source == "excel":
        objectives = _parse_objectives_file_storage(request.files.get("objectives_file"))
    else:
        objectives = []
        for s in request.form.getlist("objective_items"):
            t = (s or "").strip()[:2000]
            if t:
                objectives.append(t)
        objectives = objectives[:200]

    if not objectives:
        return (
            _render(
                error="لم يُرسل أي هدف. اختر «إدخال يدوي» وأدخل النصوص، أو «إدراج ملف خارجي» وارفع ملفاً (Excel أو PDF أو CSV أو XML أو نص).",
                ok_msg="",
                exercise=ex,
            ),
            400,
        )

    db.execute(
        delete(ExerciseObjective).where(ExerciseObjective.exercise_id == ex.id)
    )
    for i, txt in enumerate(objectives):
        db.add(ExerciseObjective(exercise_id=ex.id, sort_order=i, text=txt))
    db.commit()
    ex_show = (
        db.query(Exercise)
        .options(joinedload(Exercise.objectives))
        .filter(Exercise.id == ex.id)
        .one()
    )
    write_exercise_json_file(db, ex_show.id)
    n = len(objectives)
    return _render(
        error="",
        ok_msg=f"تم حفظ {n} هدفاً (استبدال كامل للقائمة). تم تحديث ملف JSON في مجلد التصدير.",
        exercise=ex_show,
    )


def _zip_roster_rows_from_manual_form() -> list[tuple[str, str, str, str]]:
    ms = request.form.getlist("roster_military")
    rs = request.form.getlist("roster_rank")
    ns = request.form.getlist("roster_full_name")
    uls = request.form.getlist("roster_unit_level")
    n = max(len(ms), len(rs), len(ns), len(uls))
    out: list[tuple[str, str, str, str]] = []
    for i in range(n):
        a = (ms[i] if i < len(ms) else "").strip()[:128]
        b = (rs[i] if i < len(rs) else "").strip()[:256]
        c = (ns[i] if i < len(ns) else "").strip()[:256]
        uk = normalize_unit_level_key(uls[i] if i < len(uls) else "")
        if a or b or c or uk:
            out.append((a, b, c, uk))
    return out[:500]


def _exercise_roster_page(roster_kind: str):
    user = get_current_user_optional()
    if not user:
        nxt = (
            "/admin/exercises/trainee-unit-roster"
            if roster_kind == ExerciseRosterKind.TRAINEE.value
            else "/admin/exercises/judge-unit-roster"
        )
        return redirect(f"/login?next={nxt}")
    if not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db

    meta = {
        ExerciseRosterKind.TRAINEE.value: {
            "page_title": "قائمة الوحدة المتدربة",
            "h1": "قائمة الوحدة المتدربة",
            "icon": "fa-users",
            "form_action": "/admin/exercises/trainee-unit-roster",
            "add_label": "إضافة سطر",
            "save_label": "حفظ القائمة",
            "clear_label": "حذف جميع الأسطر",
            "file_panel_hint": "CSV أو TXT أو Excel (.xlsx): الرقم العسكري، الرتبة، الاسم، ثم عمود مستوى الوحدة (المفتاح بالإنجليزية أو التسمية العربية كما في قوائم المعاضل).",
        },
        ExerciseRosterKind.JUDGE.value: {
            "page_title": "قائمة المحكمين",
            "h1": "قائمة المحكمين",
            "icon": "fa-gavel",
            "form_action": "/admin/exercises/judge-unit-roster",
            "add_label": "إضافة سطر",
            "save_label": "حفظ القائمة",
            "clear_label": "حذف جميع الأسطر",
            "file_panel_hint": "كقائمة المتدربين: العمود الرابع هو مستوى الوحدة (نفس قائمة المعاضل والتقييم).",
        },
    }[roster_kind]

    def _render(*, error: str, ok_msg: str, exercise: Exercise | None):
        rows_f = []
        if exercise and exercise.roster_rows:
            rows_f = sorted(
                [r for r in exercise.roster_rows if r.roster_kind == roster_kind],
                key=lambda r: (r.sort_order, r.id),
            )
        return render_template(
            "admin_exercise_unit_roster.html",
            **_ctx(
                user,
                error=error,
                ok_msg=ok_msg,
                current_exercise=exercise,
                roster_kind=roster_kind,
                roster_meta=meta,
                roster_rows=rows_f,
                unit_levels=UNIT_LEVELS,
            ),
        )

    if request.method == "GET":
        ex = _admin_current_workspace_exercise(db, user)
        return _render(error="", ok_msg="", exercise=ex)

    ex = _admin_current_workspace_exercise(db, user)
    if not ex:
        return (
            _render(
                error="لا يوجد تمرين حالي. أنشئ تمريناً أو افتحه من ملف JSON.",
                ok_msg="",
                exercise=None,
            ),
            400,
        )

    if (request.form.get("roster_action") or "").strip() == "clear_all":
        db.execute(
            delete(ExerciseRosterRow).where(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == roster_kind,
            )
        )
        db.commit()
        ex_show = (
            db.query(Exercise)
            .options(joinedload(Exercise.roster_rows))
            .filter(Exercise.id == ex.id)
            .one()
        )
        write_exercise_json_file(db, ex_show.id)
        return _render(
            error="",
            ok_msg="تم حذف جميع الأسطر لهذه القائمة وتحديث ملف التصدير.",
            exercise=ex_show,
        )

    source = (request.form.get("roster_source") or "manual").strip()
    if source == "excel":
        tuples = parse_roster_rows_from_upload(request.files.get("roster_file"))
    else:
        tuples = _zip_roster_rows_from_manual_form()

    if not tuples:
        return (
            _render(
                error="لم يُرسل أي صف. اختر «إدخال يدوي» واملأ الحقول، أو «إدراج ملف خارجي» وارفع ملفاً صالحاً.",
                ok_msg="",
                exercise=ex,
            ),
            400,
        )

    db.execute(
        delete(ExerciseRosterRow).where(
            ExerciseRosterRow.exercise_id == ex.id,
            ExerciseRosterRow.roster_kind == roster_kind,
        )
    )
    for i, (mil, rk, nm, cell4) in enumerate(tuples):
        uk, pos_ar = coerce_roster_import_position_cell(cell4)
        db.add(
            ExerciseRosterRow(
                exercise_id=ex.id,
                roster_kind=roster_kind,
                sort_order=i,
                military_number=mil,
                rank_ar=rk,
                full_name=nm,
                unit_level_key=uk,
                position_ar=pos_ar,
            )
        )
    db.commit()
    # بعد حفظ قائمة المحكمين: أنشئ/حدّث حساباتهم وتخصيصاتهم تلقائياً
    if roster_kind == ExerciseRosterKind.JUDGE.value:
        _sync_judges_from_roster(db, ex)
    ex_show = (
        db.query(Exercise)
        .options(joinedload(Exercise.roster_rows))
        .filter(Exercise.id == ex.id)
        .one()
    )
    write_exercise_json_file(db, ex_show.id)
    return _render(
        error="",
        ok_msg=f"تم حفظ {len(tuples)} صفاً. تم تحديث ملف JSON في مجلد التصدير.",
        exercise=ex_show,
    )


@bp.route("/admin/exercises/trainee-unit-roster", methods=["GET", "POST"])
def admin_exercise_trainee_unit_roster():
    return _exercise_roster_page(ExerciseRosterKind.TRAINEE.value)


@bp.route("/admin/exercises/judge-unit-roster", methods=["GET", "POST"])
def admin_exercise_judge_unit_roster():
    return _exercise_roster_page(ExerciseRosterKind.JUDGE.value)


@bp.route("/admin/exercises/exports", methods=["GET", "POST"])
def admin_exercise_exports():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/exports")
    if not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db
    error = ""
    ok_msg = ""
    if request.method == "POST":
        up = request.files.get("json_file")
        if not up or not getattr(up, "filename", ""):
            error = "يرجى اختيار ملف JSON."
        else:
            try:
                raw = up.read()
            except Exception:
                raw = b""
            eid = read_exercise_id_from_json_bytes(raw)
            if eid is None:
                error = "الملف لا يحتوي على معرف تمرين صالح (exercise.id)."
            elif not db.get(Exercise, eid):
                error = "التمرين المذكور في الملف غير موجود في قاعدة البيانات."
            else:
                return redirect(f"/exercises/{eid}")
    files = list_export_json_files()
    return render_template(
        "admin_exercise_exports.html",
        **_ctx(
            user,
            error=error,
            ok_msg=ok_msg,
            export_files=files,
            export_dir=str(export_directory()),
        ),
    )


@bp.route("/admin/exercises/open-export")
def admin_exercise_open_export():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/exercises/exports")
    if not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db
    raw = unquote((request.args.get("f") or "").strip())
    if not raw or raw != Path(raw).name:
        abort(400)
    if not raw.lower().endswith(".json"):
        abort(400)
    base = export_directory().resolve()
    path = (base / raw).resolve()
    try:
        path.relative_to(base)
    except ValueError:
        abort(400)
    if not path.is_file():
        abort(404)
    eid = read_exercise_id_from_json_path(path)
    if eid is None:
        abort(400)
    if not db.get(Exercise, eid):
        abort(404)
    return redirect(f"/exercises/{eid}")


@bp.route("/admin/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def admin_evaluation_list_file_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or (current_exercise is not None and row.exercise_id != current_exercise.id):
        abort(404)
    list_url = url_for("views.admin_evaluation_lists", unit_key=unit_key)
    if not (row.pdf_relpath or "").strip():
        return redirect(list_url)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        return redirect(list_url)
    ev = _evaluation_sheet_view_context(fspath)

    saved_payload = {}
    saved_updated_at = None
    saved_row_id = None
    saved_is_approved = False
    saved_approved_at = None
    saved_by_id = None
    saved_row = None

    # معلومات إضافية أعلى مربع قائمة التقييم (تعريف افتراضي لتجنب NameError)
    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = getattr(current_exercise, "planned_start", None) if current_exercise is not None else None
    if shown_date is None and current_exercise is not None:
        shown_date = getattr(current_exercise, "created_at", None)
    commander_name = "—"
    judge_name = "—"
    if current_exercise is not None:
        commander_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == current_exercise.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if commander_row is not None:
            commander_name = (commander_row.full_name or "").strip() or commander_name
        judge_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == current_exercise.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if judge_row is not None:
            judge_name = (judge_row.full_name or "").strip() or judge_name

    if current_exercise is not None:
        saved_id_raw = (request.args.get("saved_id") or "").strip()
        saved_id = None
        if saved_id_raw.isdigit():
            try:
                saved_id = int(saved_id_raw)
            except Exception:
                saved_id = None
        saved_row = (
            db.query(EvaluationListSavedResult)
            .filter(
                EvaluationListSavedResult.evaluation_item_id == row.id,
                EvaluationListSavedResult.exercise_id == current_exercise.id,
            )
            .order_by(EvaluationListSavedResult.updated_at.desc(), EvaluationListSavedResult.id.desc())
            .first()
        )
        if saved_id is not None:
            picked = (
                db.query(EvaluationListSavedResult)
                .filter(
                    EvaluationListSavedResult.id == saved_id,
                    EvaluationListSavedResult.evaluation_item_id == row.id,
                    EvaluationListSavedResult.exercise_id == current_exercise.id,
                )
                .first()
            )
            if picked is not None:
                saved_row = picked
        if saved_row and (saved_row.payload_json or "").strip():
            try:
                saved_payload = json.loads(saved_row.payload_json)
            except Exception:
                saved_payload = {}
            saved_updated_at = saved_row.updated_at
            saved_row_id = saved_row.id
            saved_is_approved = bool(getattr(saved_row, "is_approved", False))
            saved_approved_at = getattr(saved_row, "approved_at", None)
            saved_by_id = getattr(saved_row, "saved_by_id", None)
    saved_payload = _saved_payload_aligned_with_eval_rows(saved_payload, ev.get("eval_rows"))
    admin_crit = bool(not saved_is_approved and can_save_evaluation_results(user))
    canon_admin = saved_row if current_exercise is not None else None
    wf_admin = _eval_list_viewer_ctx(user, canon_admin)
    wf_admin["eval_can_edit"] = admin_crit
    return render_template(
        "admin_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            can_edit=not saved_is_approved,
            close_href=url_for("views.admin_evaluation_lists", unit_key=unit_key),
            subpage_close_fallback=url_for("views.admin_evaluation_lists", unit_key=unit_key),
            close_label="إغلاق والعودة",
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            saved_by_id=saved_by_id,
            eval_save_url=url_for("views.admin_evaluation_list_save_results", unit_key=unit_key, item_id=item_id),
            eval_approve_url=url_for("views.admin_evaluation_list_approve", unit_key=unit_key, item_id=item_id),
            **wf_admin,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=current_exercise,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=admin_crit,
            ),
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_approve_incomplete=request.args.get("eval_approve_incomplete", type=int) == 1,
        ),
    )


@bp.route("/analyst/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def analyst_evaluation_list_file_viewer(unit_key: str, item_id: int):
    """عرض ملف التقييم للمحلل (قراءة فقط)."""
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/analyst/evaluation-lists/{unit_key}/view/{item_id}")
    # نسمح للمحلل (وأيضاً إدارة النظام) بالعرض
    if not (can_access_analyst_hub(user) or is_system_admin(user)):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or (
        current_exercise is not None and row.exercise_id != current_exercise.id
    ):
        abort(404)
    if not (row.pdf_relpath or "").strip():
        abort(404)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        abort(404)

    ev = _evaluation_sheet_view_context(fspath)

    # نعرض أحدث نتيجة محفوظة إن وجدت (اختيار saved_id اختياري)
    saved_payload = {}
    saved_updated_at = None
    saved_row_id = None
    saved_is_approved = False
    saved_approved_at = None
    saved_by_id = None
    if current_exercise is not None:
        saved_id_raw = (request.args.get("saved_id") or "").strip()
        saved_id = int(saved_id_raw) if saved_id_raw.isdigit() else None
        saved_row = (
            db.query(EvaluationListSavedResult)
            .filter(
                EvaluationListSavedResult.evaluation_item_id == row.id,
                EvaluationListSavedResult.exercise_id == current_exercise.id,
            )
            .order_by(EvaluationListSavedResult.updated_at.desc(), EvaluationListSavedResult.id.desc())
            .first()
        )
        if saved_id is not None:
            picked = (
                db.query(EvaluationListSavedResult)
                .filter(
                    EvaluationListSavedResult.id == saved_id,
                    EvaluationListSavedResult.evaluation_item_id == row.id,
                    EvaluationListSavedResult.exercise_id == current_exercise.id,
                )
                .first()
            )
            if picked is not None:
                saved_row = picked
        if saved_row and (saved_row.payload_json or "").strip():
            try:
                saved_payload = json.loads(saved_row.payload_json)
            except Exception:
                saved_payload = {}
            saved_updated_at = saved_row.updated_at
            saved_row_id = saved_row.id
            saved_is_approved = bool(getattr(saved_row, "is_approved", False))
            saved_approved_at = getattr(saved_row, "approved_at", None)
            saved_by_id = getattr(saved_row, "saved_by_id", None)

    # معلومات إضافية أعلى مربع قائمة التقييم
    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = None
    commander_name = "—"
    judge_name = "—"
    if current_exercise is not None:
        shown_date = getattr(current_exercise, "planned_start", None) or getattr(current_exercise, "created_at", None)
        commander_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == current_exercise.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if commander_row is not None:
            commander_name = (commander_row.full_name or "").strip() or commander_name
        judge_row = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == current_exercise.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
                ExerciseRosterRow.unit_level_key == unit_key,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .first()
        )
        if judge_row is not None:
            judge_name = (judge_row.full_name or "").strip() or judge_name

    return render_template(
        "admin_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            can_edit=False,
            close_href="/analyst/judges-eval-analysis",
            subpage_close_fallback="/analyst/judges-eval-analysis",
            close_label="إغلاق والعودة للتحليل",
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            saved_is_approved=saved_is_approved,
            saved_approved_at=saved_approved_at,
            saved_by_id=saved_by_id,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=current_exercise,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=False,
            ),
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url="#",
            eval_approve_url="#",
            show_eval_approve=False,
            eval_can_edit=False,
        ),
    )


@bp.route("/admin/evaluation-lists/<unit_key>/item/<int:item_id>/file", methods=["GET"])
def admin_evaluation_list_file(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or (current_exercise is not None and row.exercise_id != current_exercise.id):
        abort(404)
    rel = (row.pdf_relpath or "").strip()
    if not rel:
        abort(404)
    path = _evaluation_list_file_abspath(rel)
    if path is None:
        abort(404)
    mt = _mimetype_for_eval_list_file(path)
    return send_file(
        path,
        mimetype=mt,
        as_attachment=True,
        download_name=path.name,
    )


@bp.route(
    "/admin/evaluation-lists/<unit_key>/item/<int:item_id>/delete",
    methods=["POST"],
)
def admin_evaluation_list_delete_item(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not row
        or row.unit_level_key != unit_key
        or (current_exercise is not None and row.exercise_id != current_exercise.id)
    ):
        abort(404)
    if row.pdf_relpath:
        _unlink_evaluation_list_stored_file(row.pdf_relpath)
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_evaluation_lists", unit_key=unit_key))


@bp.route(
    "/admin/evaluation-lists/<unit_key>/item/<int:item_id>/phase",
    methods=["POST"],
)
def admin_evaluation_list_set_phase(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not row
        or row.unit_level_key != unit_key
        or (current_exercise is not None and row.exercise_id != current_exercise.id)
    ):
        abort(404)
    phase = _normalized_exercise_phase(request.form.get("exercise_phase"))
    if not phase and EXERCISE_PHASE_OPTIONS:
        return redirect(url_for("views.admin_evaluation_lists", unit_key=unit_key))
    if phase:
        _sync_evaluation_list_item_phase(db, row, phase)
    db.commit()
    return redirect(url_for("views.admin_evaluation_lists", unit_key=unit_key))


@bp.route(
    "/admin/evaluation-lists/<unit_key>/view/<int:item_id>/save-results",
    methods=["POST"],
)
def admin_evaluation_list_save_results(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)

    raw = (request.form.get("payload_json") or "").strip()
    if not raw:
        return redirect(url_for("views.admin_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    if len(raw) > 250_000:
        abort(400)
    _evaluation_commit_payload_save(db, user=user, item=item, current_exercise=current_exercise, raw=raw)
    return redirect(url_for("views.admin_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))


@bp.route(
    "/admin/evaluation-lists/<unit_key>/view/<int:item_id>/approve",
    methods=["POST"],
)
def admin_evaluation_list_approve(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user or not is_system_admin(user):
        abort(403)
    if not can_approve_evaluation_results(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)

    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is None or not (saved.payload_json or "").strip():
        abort(400)
    if bool(getattr(saved, "is_approved", False)):
        return redirect(url_for("views.admin_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    rows = _parse_saved_eval_rows(saved.payload_json)
    if _evaluation_payload_has_empty_acquired_for_approve(rows):
        return redirect(
            url_for(
                "views.admin_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_incomplete=1,
            )
        )
    if not _evaluation_saved_allows_judge_approve(saved):
        return redirect(
            url_for(
                "views.admin_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_grade_blocked=1,
            )
        )
    saved.is_approved = True
    saved.approved_by_id = getattr(user, "id", None)
    saved.approved_at = datetime.utcnow()
    db.commit()
    return redirect(url_for("views.admin_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))


@bp.route("/judge/evaluation-lists", methods=["GET"])
def judge_evaluation_lists_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/judge/evaluation-lists")
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    _ensure_judge_roster_synced(db, user, ex)
    # إن كان المحكم مخصصاً لوحدة واحدة، افتحها مباشرة
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if assigned_uk and unit_level_row(assigned_uk) and not is_system_admin(user):
        return redirect(url_for("views.judge_evaluation_lists", unit_key=assigned_uk))
    units = [u for u in UNIT_LEVELS if not assigned_uk or u.get("key") == assigned_uk]
    unit_rows = evaluation_unit_home_rows(db, ex, units)
    return render_template(
        "judge_evaluation_lists_home.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            unit_levels=units,
            unit_rows=unit_rows,
            unit_totals=evaluation_unit_home_totals(unit_rows),
            unit_list_endpoint="views.judge_evaluation_lists",
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/dilemmas", methods=["GET"])
def judge_dilemmas_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/judge/dilemmas")
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    _ensure_judge_roster_synced(db, user, ex)
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if assigned_uk and unit_level_row(assigned_uk) and not is_system_admin(user):
        return redirect(url_for("views.judge_dilemmas", unit_key=assigned_uk))
    return render_template(
        "judge_dilemmas_home.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            unit_levels=[u for u in UNIT_LEVELS if not assigned_uk or u.get("key") == assigned_uk],
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/dilemmas/<unit_key>", methods=["GET"])
def judge_dilemmas(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/judge/dilemmas/{unit_key}")
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    if ex is None:
        return redirect("/judge/dilemmas")
    _enforce_judge_unit_scope(db, user, ex, unit_key)
    items = (
        db.query(DilemmaItem)
        .filter(DilemmaItem.exercise_id == ex.id, DilemmaItem.unit_level_key == unit_key)
        .order_by(
            _exercise_phase_order_expr(DilemmaItem.exercise_phase),
            DilemmaItem.sort_order,
            DilemmaItem.id,
        )
        .all()
    )
    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if assigned_uk and not is_system_admin(user):
        dilemmas_parent_href = url_for("views.judge_hub")
    else:
        dilemmas_parent_href = url_for("views.judge_dilemmas_home")
    return render_template(
        "judge_dilemmas.html",
        **_ctx(
            user,
            exercise=ex,
            unit=unit,
            unit_key=unit_key,
            items=items,
            subpage_close_fallback=dilemmas_parent_href,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/dilemmas/<unit_key>/view/<int:item_id>", methods=["GET"])
def judge_dilemma_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/judge/dilemmas/{unit_key}/view/{item_id}")
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or current_exercise is None or row.exercise_id != current_exercise.id:
        abort(404)
    _enforce_judge_unit_scope(db, user, current_exercise, unit_key)
    if _dilemma_pdf_abspath(row.pdf_relpath) is None:
        return redirect(url_for("views.judge_dilemmas", unit_key=unit_key))
    pdf_url = url_for("views.judge_dilemma_pdf_file", unit_key=unit_key, item_id=item_id)
    return render_template(
        "judge_dilemma_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "معضلة",
            pdf_url=pdf_url,
            subpage_close_fallback=url_for("views.judge_dilemmas", unit_key=unit_key),
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/dilemmas/<unit_key>/item/<int:item_id>/pdf", methods=["GET"])
def judge_dilemma_pdf_file(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    ex = _current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or ex is None or row.exercise_id != ex.id:
        abort(404)
    _enforce_judge_unit_scope(db, user, ex, unit_key)
    rel = (row.pdf_relpath or "").strip()
    path = _dilemma_pdf_abspath(rel)
    if path is None:
        abort(404)
    return send_file(path, mimetype="application/pdf", as_attachment=False)


@bp.route("/judge/evaluation-lists/<unit_key>", methods=["GET"])
def judge_evaluation_lists(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/judge/evaluation-lists/{unit_key}")
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    if ex is None:
        return redirect("/judge/evaluation-lists")
    _enforce_judge_unit_scope(db, user, ex, unit_key)
    items = (
        db.query(EvaluationListPdfItem)
        .filter(EvaluationListPdfItem.exercise_id == ex.id, EvaluationListPdfItem.unit_level_key == unit_key)
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )

    item_ids = [int(it.id) for it in items]
    canonical_by_item = _evaluation_canonical_map_for_items(db, ex.id, item_ids)

    a = _judge_assignment_for_current_exercise(db, user, ex)
    assigned_uk = (getattr(a, "unit_level_key", "") or "").strip() if a else ""
    if assigned_uk and not is_system_admin(user):
        eval_lists_parent_href = url_for("views.judge_hub")
    else:
        eval_lists_parent_href = url_for("views.judge_evaluation_lists_home")

    evaluation_lists_rows: list[dict] = []
    for it in items:
        s = canonical_by_item.get(int(it.id))
        evaluation_lists_rows.append(
            build_evaluation_list_row(
                item=it,
                saved=s,
                exercise=ex,
                open_href=url_for(
                    "views.judge_evaluation_list_file_viewer",
                    unit_key=unit_key,
                    item_id=it.id,
                ),
            )
        )
    return render_template(
        "judge_evaluation_lists.html",
        **_ctx(
            user,
            exercise=ex,
            unit=unit,
            unit_key=unit_key,
            items=items,
            evaluation_lists_rows=evaluation_lists_rows,
            eval_lists_parent_href=eval_lists_parent_href,
            subpage_close_fallback=eval_lists_parent_href,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/judge/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def judge_evaluation_list_file_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/judge/evaluation-lists/{unit_key}/view/{item_id}")
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or current_exercise is None or row.exercise_id != current_exercise.id:
        abort(404)
    _enforce_judge_unit_scope(db, user, current_exercise, unit_key)
    list_url = url_for("views.judge_evaluation_lists", unit_key=unit_key)
    if not (row.pdf_relpath or "").strip():
        return redirect(list_url)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        return redirect(list_url)
    ev = _evaluation_sheet_view_context(fspath)

    saved_payload = {}
    saved_updated_at = None
    saved_row_id = None
    canon = _evaluation_canonical_saved_row(db, current_exercise.id, row.id)

    def _load_payload(sr: EvaluationListSavedResult | None) -> dict:
        if not sr or not (sr.payload_json or "").strip():
            return {}
        try:
            p = json.loads(sr.payload_json)
        except Exception:
            return {}
        return p if isinstance(p, dict) else {}

    if canon is not None:
        saved_payload = _load_payload(canon)
        saved_updated_at = canon.updated_at
        saved_row_id = canon.id
    saved_payload = _saved_payload_aligned_with_eval_rows(saved_payload, ev.get("eval_rows"))

    eval_save_url = url_for("views.judge_evaluation_list_save_results", unit_key=unit_key, item_id=item_id)
    eval_approve_url = url_for("views.judge_evaluation_list_approve", unit_key=unit_key, item_id=item_id)
    wf = _eval_list_viewer_ctx(user, canon)

    # معلومات إضافية أعلى مربع قائمة التقييم
    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = getattr(current_exercise, "planned_start", None) or getattr(current_exercise, "created_at", None)

    commander_name = "—"
    commander_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if commander_row is not None:
        commander_name = (commander_row.full_name or "").strip() or commander_name

    judge_name = (
        (getattr(user, "full_name", "") or "").strip()
        or (getattr(user, "username", "") or "").strip()
        or f"محكم #{getattr(user, 'id', '')}"
    )
    judge_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if judge_row is not None:
        judge_name = (judge_row.full_name or "").strip() or judge_name

    return render_template(
        "judge_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            **wf,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=current_exercise,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=bool(wf.get("eval_can_edit")),
            ),
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url=eval_save_url,
            eval_approve_url=eval_approve_url,
            eval_chief_approve_url="",
            eval_chief_reopen_url="",
            eval_approve_incomplete=request.args.get("eval_approve_incomplete", type=int) == 1,
            eval_saved_notice=request.args.get("eval_saved", type=int) == 1,
            subpage_close_fallback=list_url,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route(
    "/judge/evaluation-lists/<unit_key>/view/<int:item_id>/save-results",
    methods=["POST"],
)
def judge_evaluation_list_save_results(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)
    _enforce_judge_unit_scope(db, user, current_exercise, unit_key)

    raw = (request.form.get("payload_json") or "").strip()
    if not raw:
        return redirect(url_for("views.judge_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    if len(raw) > 250_000:
        abort(400)
    _evaluation_commit_payload_save(db, user=user, item=item, current_exercise=current_exercise, raw=raw)
    return redirect(
        url_for(
            "views.judge_evaluation_list_file_viewer",
            unit_key=unit_key,
            item_id=item_id,
            eval_saved=1,
        )
    )


@bp.route(
    "/judge/evaluation-lists/<unit_key>/view/<int:item_id>/approve",
    methods=["POST"],
)
def judge_evaluation_list_approve(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_approve_evaluation_results(user):
        abort(403)
    if not can_access_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)
    _enforce_judge_unit_scope(db, user, current_exercise, unit_key)

    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is None or not (saved.payload_json or "").strip():
        abort(400)
    if not eval_judge_can_approve(saved):
        return redirect(url_for("views.judge_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))
    rows = _parse_saved_eval_rows(saved.payload_json)
    if _evaluation_payload_has_empty_acquired_for_approve(rows):
        return redirect(
            url_for(
                "views.judge_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_incomplete=1,
            )
        )
    if not _evaluation_saved_allows_judge_approve(saved):
        return redirect(
            url_for(
                "views.judge_evaluation_list_file_viewer",
                unit_key=unit_key,
                item_id=item_id,
                eval_approve_grade_blocked=1,
            )
        )
    apply_judge_approve(saved, getattr(user, "id", None))
    db.commit()
    return redirect(url_for("views.judge_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id))


@bp.route("/eval-criterion-media/<int:media_id>/stream", methods=["GET"])
def eval_criterion_media_stream(media_id: int):
    """تسليم ملف توثيق (صورة أو فيديو) لمستخدم مسموح."""
    user = get_current_user_optional()
    if not user:
        abort(403)
    from flask import g

    db = g.db
    m = db.get(EvaluationCriterionMedia, media_id)
    if m is None or not (m.file_relpath or "").strip():
        abort(404)
    if not _eval_crit_user_can_stream_media(db, user, m):
        abort(403)
    abs_p = criterion_media_absolute_path((m.file_relpath or "").strip())
    if abs_p is None or not abs_p.is_file():
        abort(404)
    dl_name = Path(m.file_relpath or "").name
    mime = ((m.mime_type or "").strip() or mimetypes.guess_type(dl_name)[0] or "").strip()
    return send_file(abs_p, mimetype=mime or None)


@bp.route("/eval-criterion-media/upload", methods=["POST"])
def eval_criterion_media_upload():
    user = get_current_user_optional()
    if not user:
        abort(403)
    from flask import g

    db = g.db
    li_raw = (request.form.get("evaluation_list_item_id") or "").strip()
    ba_raw = (request.form.get("bundle_action_eval_id") or "").strip()
    li_id = int(li_raw) if li_raw.isdigit() else None
    ba_id = int(ba_raw) if ba_raw.isdigit() else None
    if li_id is not None and li_id <= 0:
        li_id = None
    if ba_id is not None and ba_id <= 0:
        ba_id = None
    if li_id is None and ba_id is None:
        return jsonify(ok=False, error="scope_missing"), 400
    if li_id is not None and ba_id is not None:
        return jsonify(ok=False, error="scope_conflict"), 400
    ri_raw = (request.form.get("row_index") or "").strip()
    row_index = int(ri_raw) if ri_raw.isdigit() else -1
    if row_index < 0:
        return jsonify(ok=False, error="row"), 400
    media_kind = (request.form.get("media_kind") or "photo").strip().lower()
    if media_kind not in ("photo", "video"):
        return jsonify(ok=False, error="kind"), 400
    uf = request.files.get("file")
    if uf is None or not (uf.filename or "").strip():
        return jsonify(ok=False, error="file"), 400
    blob = uf.read()
    if not blob:
        return jsonify(ok=False, error="empty"), 400
    ct_in = uf.mimetype or mimetypes.guess_type((uf.filename or "").strip())[0] or ""

    exercise_id = 0
    unit_key = ""
    try:
        if li_id is not None:
            item = db.get(EvaluationListPdfItem, li_id)
            if item is None:
                return jsonify(ok=False, error="item"), 404
            exercise_id = int(item.exercise_id or 0)
            unit_key = (item.unit_level_key or "").strip()
            ba_id = None
        elif ba_id is not None:
            ar = db.get(ExercisePlannerFlowBundleActionEval, int(ba_id))
            if ar is None:
                return jsonify(ok=False, error="action"), 404
            bd = db.get(ExercisePlannerFlowBundle, int(ar.bundle_id))
            if bd is None:
                return jsonify(ok=False, error="bundle"), 404
            exercise_id = int(bd.exercise_id)
            unit_key = (bd.unit_level_key or "").strip()
            li_id = None

        canon = _eval_row_canonical_saved_for_crit_upload(
            db,
            exercise_id=int(exercise_id),
            list_item_id=li_id,
            bundle_action_eval_id=ba_id,
        )
        if not _eval_crit_user_can_upload_media(
            db,
            user,
            exercise_id=int(exercise_id),
            unit_level_key=unit_key,
            list_item_id=li_id,
            bundle_action_eval_id=ba_id,
            canonical_saved=canon,
        ):
            abort(403)
        rec = persist_criterion_medium(
            db,
            exercise_id=int(exercise_id),
            unit_level_key=unit_key,
            list_item_id=li_id,
            bundle_action_eval_id=ba_id,
            row_index=int(row_index),
            media_kind=media_kind,
            mime_type_in=ct_in,
            bin_data=blob,
            uploaded_by_id=getattr(user, "id", None),
        )
        db.commit()
        stream_u = url_for("views.eval_criterion_media_stream", media_id=int(rec.id))
        del_u = url_for("views.eval_criterion_media_delete", media_id=int(rec.id))
        return jsonify(
            ok=True,
            id=int(rec.id),
            media_kind=rec.media_kind,
            stream_url=stream_u,
            delete_url=del_u,
        )
    except ValueError as err:
        db.rollback()
        tag = "".join(str(x) for x in err.args)
        if "mime" in tag:
            return jsonify(ok=False, error="mime"), 415
        if "size" in tag:
            return jsonify(ok=False, error="size"), 413
        return jsonify(ok=False, error="reject"), 400


@bp.route("/eval-criterion-media/<int:media_id>/delete", methods=["POST"])
def eval_criterion_media_delete(media_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    from flask import g

    db = g.db
    m = db.get(EvaluationCriterionMedia, media_id)
    if m is None:
        abort(404)
    li_raw = getattr(m, "evaluation_list_item_id", None)
    ba_raw = getattr(m, "bundle_action_eval_id", None)
    li_pk = int(li_raw) if li_raw is not None else None
    ba_pk = int(ba_raw) if ba_raw is not None else None
    canon = _eval_row_canonical_saved_for_crit_upload(
        db,
        exercise_id=int(m.exercise_id),
        list_item_id=li_pk,
        bundle_action_eval_id=ba_pk,
    )
    if not _eval_crit_user_can_upload_media(
        db,
        user,
        exercise_id=int(m.exercise_id),
        unit_level_key=(getattr(m, "unit_level_key", "") or "").strip(),
        list_item_id=li_pk,
        bundle_action_eval_id=ba_pk,
        canonical_saved=canon,
    ):
        abort(403)
    rel = (m.file_relpath or "").strip()
    abs_p = criterion_media_absolute_path(rel) if rel else None
    db.delete(m)
    db.commit()
    if abs_p is not None and abs_p.is_file():
        try:
            abs_p.unlink()
        except OSError:
            pass
    return jsonify(ok=True)


# ——— مساحة كبير المحكمين ———
CHIEF_JUDGE_ONLY_HUB_ITEMS: tuple[tuple[str, str, str], ...] = (
    ("evaluation-lists-chief", "اعتماد قوائم التقييم (كبير المحكمين)", "fa-stamp"),
    (
        "planner-flow-bundle-overview",
        "المجرى وتقييم الإجراءات (اعتماد)",
        "fa-diagram-project",
    ),
)


def _chief_judge_hub_items() -> tuple[tuple[str, str, str], ...]:
    """امتيازات كبير المحكمين الخاصة + جميع أوامر مساحة المحكمين."""
    return CHIEF_JUDGE_ONLY_HUB_ITEMS + JUDGE_HUB_ITEMS


CHIEF_JUDGE_HUB_SLUGS: dict[str, str] = {s: t for s, t, _ in _chief_judge_hub_items()}


@bp.route("/chief-judge")
def chief_judge_hub():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/chief-judge")
    if not can_access_chief_judge_hub(user):
        abort(403)
    return redirect(url_for("views.judge_hub"))


@bp.route("/chief-judge/<slug>")
def chief_judge_hub_section(slug: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/chief-judge/{slug}")
    if not can_access_chief_judge_hub(user):
        abort(403)
    slug_norm = (slug or "").strip().lower()
    if slug_norm not in CHIEF_JUDGE_HUB_SLUGS:
        abort(404)
    return redirect(url_for("views.judge_hub_section", slug=slug_norm))


@bp.route("/chief-judge/evaluation-lists", methods=["GET"])
def chief_judge_evaluation_lists_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/chief-judge/evaluation-lists")
    if not can_access_chief_judge_hub(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    units = list(UNIT_LEVELS)
    unit_rows = evaluation_unit_home_rows(db, ex, units)
    return render_template(
        "judge_evaluation_lists_home.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            unit_levels=units,
            unit_rows=unit_rows,
            unit_totals=evaluation_unit_home_totals(unit_rows),
            unit_list_endpoint="views.chief_judge_evaluation_lists",
            page_title="قوائم التقييم — اعتماد كبير المحكمين",
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/chief-judge/evaluation-lists/<unit_key>", methods=["GET"])
def chief_judge_evaluation_lists(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/chief-judge/evaluation-lists/{unit_key}")
    if not can_access_chief_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    ex = _current_workspace_exercise(db, user)
    if ex is None:
        return redirect(url_for("views.chief_judge_evaluation_lists_home"))
    items = (
        db.query(EvaluationListPdfItem)
        .filter(
            EvaluationListPdfItem.exercise_id == ex.id,
            EvaluationListPdfItem.unit_level_key == unit_key,
        )
        .order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        )
        .all()
    )
    item_ids = [int(it.id) for it in items]
    canonical_by_item = _evaluation_canonical_map_for_items(db, ex.id, item_ids)
    evaluation_lists_rows: list[dict] = []
    for it in items:
        s = canonical_by_item.get(int(it.id))
        evaluation_lists_rows.append(
            build_evaluation_list_row(
                item=it,
                saved=s,
                exercise=ex,
                open_href=url_for(
                    "views.chief_judge_evaluation_list_file_viewer",
                    unit_key=unit_key,
                    item_id=it.id,
                ),
                chief_workflow_label=eval_workflow_label_ar(s),
            )
        )
    return render_template(
        "chief_judge_evaluation_lists.html",
        **_ctx(
            user,
            exercise=ex,
            unit=unit,
            unit_key=unit_key,
            evaluation_lists_rows=evaluation_lists_rows,
            eval_lists_parent_href=url_for("views.chief_judge_evaluation_lists_home"),
            subpage_close_fallback=url_for("views.chief_judge_evaluation_lists_home"),
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/chief-judge/evaluation-lists/<unit_key>/view/<int:item_id>", methods=["GET"])
def chief_judge_evaluation_list_file_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/chief-judge/evaluation-lists/{unit_key}/view/{item_id}")
    if not can_access_chief_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or current_exercise is None or row.exercise_id != current_exercise.id:
        abort(404)
    list_url = url_for("views.chief_judge_evaluation_lists", unit_key=unit_key)
    if not (row.pdf_relpath or "").strip():
        return redirect(list_url)
    fspath = _evaluation_list_file_abspath(row.pdf_relpath)
    if fspath is None:
        return redirect(list_url)
    ev = _evaluation_sheet_view_context(fspath)
    canon = _evaluation_canonical_saved_row(db, current_exercise.id, row.id)

    def _load_payload(sr: EvaluationListSavedResult | None) -> dict:
        if not sr or not (sr.payload_json or "").strip():
            return {}
        try:
            p = json.loads(sr.payload_json)
        except Exception:
            return {}
        return p if isinstance(p, dict) else {}

    saved_payload = _load_payload(canon)
    saved_updated_at = getattr(canon, "updated_at", None) if canon else None
    saved_row_id = getattr(canon, "id", None) if canon else None
    wf = _eval_list_viewer_ctx(user, canon)
    wf = {
        **wf,
        "eval_can_edit": False,
        "show_eval_approve": False,
    }
    unit_label = (unit.get("label") or "").strip() if isinstance(unit, dict) else ""
    shown_date = getattr(current_exercise, "planned_start", None) or getattr(current_exercise, "created_at", None)
    commander_name = "—"
    commander_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if commander_row is not None:
        commander_name = (commander_row.full_name or "").strip() or commander_name
    judge_name = "—"
    judge_row = (
        db.query(ExerciseRosterRow)
        .filter(
            ExerciseRosterRow.exercise_id == current_exercise.id,
            ExerciseRosterRow.roster_kind == ExerciseRosterKind.JUDGE.value,
            ExerciseRosterRow.unit_level_key == unit_key,
        )
        .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
        .first()
    )
    if judge_row is not None:
        judge_name = (judge_row.full_name or "").strip() or judge_name

    return render_template(
        "judge_evaluation_list_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "تقييم",
            evaluation_item_id=row.id,
            saved_row_id=saved_row_id,
            saved_payload=saved_payload,
            saved_updated_at=saved_updated_at,
            **wf,
            **ev,
            **_eval_crit_media_sheet_ctx(
                db,
                user,
                exercise=current_exercise,
                list_item_id=int(row.id),
                bundle_action_eval_id=None,
                eval_can_edit=False,
            ),
            unit_label=unit_label or "—",
            shown_date=shown_date,
            commander_name=commander_name or "—",
            judge_name=judge_name or "—",
            has_saved_rows=bool(saved_payload and (saved_payload.get("rows") or [])),
            eval_save_url="",
            eval_approve_url="",
            eval_chief_approve_url=url_for(
                "views.chief_judge_evaluation_list_chief_approve",
                unit_key=unit_key,
                item_id=item_id,
            ),
            eval_chief_reopen_url=url_for(
                "views.chief_judge_evaluation_list_chief_reopen",
                unit_key=unit_key,
                item_id=item_id,
            ),
            eval_close_href=list_url,
            subpage_close_fallback=list_url,
            eval_approve_incomplete=False,
            viewer_readonly_chief=True,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route(
    "/chief-judge/evaluation-lists/<unit_key>/view/<int:item_id>/chief-approve",
    methods=["POST"],
)
def chief_judge_evaluation_list_chief_approve(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_chief_approve_evaluation_results(user):
        abort(403)
    if not can_access_chief_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)
    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is None or not eval_chief_can_approve(saved):
        abort(400)
    apply_chief_approve(saved, getattr(user, "id", None))
    db.commit()
    return redirect(
        url_for("views.chief_judge_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id)
    )


@bp.route(
    "/chief-judge/evaluation-lists/<unit_key>/view/<int:item_id>/chief-reopen",
    methods=["POST"],
)
def chief_judge_evaluation_list_chief_reopen(unit_key: str, item_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_chief_reopen_evaluation_for_judge(user):
        abort(403)
    if not can_access_chief_judge_hub(user):
        abort(403)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    item = db.get(EvaluationListPdfItem, item_id)
    current_exercise = _current_workspace_exercise(db, user)
    if (
        not item
        or item.unit_level_key != unit_key
        or current_exercise is None
        or item.exercise_id != current_exercise.id
    ):
        abort(404)
    saved = _evaluation_canonical_saved_row(db, current_exercise.id, item.id)
    if saved is None or not eval_chief_can_reopen(saved):
        abort(400)
    apply_chief_reopen(saved)
    from app.notifications_service import notify_evaluation_reopened_by_chief_judge

    unit_label = label_for_unit_level_key(unit_key) or unit.get("label") or unit_key
    item_title = (getattr(item, "text", None) or "قائمة التقييم").strip()
    notify_evaluation_reopened_by_chief_judge(
        db,
        exercise_id=int(current_exercise.id),
        unit_key=unit_key,
        unit_label=unit_label,
        item_title=item_title,
        item_id=int(item.id),
        saved_by_user_id=getattr(saved, "saved_by_id", None),
        exclude_user_id=getattr(user, "id", None),
    )
    db.commit()
    return redirect(
        url_for("views.chief_judge_evaluation_list_file_viewer", unit_key=unit_key, item_id=item_id)
    )


@bp.route("/admin/evaluation-lists/<unit_key>/clear", methods=["POST"])
def admin_evaluation_list_clear(unit_key: str):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    current_exercise = _admin_current_workspace_exercise(db, user)
    q = db.query(EvaluationListPdfItem).filter(EvaluationListPdfItem.unit_level_key == unit_key)
    if current_exercise is not None:
        q = q.filter(EvaluationListPdfItem.exercise_id == current_exercise.id)
    else:
        q = q.filter(EvaluationListPdfItem.exercise_id == -1)
    for row in q.all():
        if row.pdf_relpath:
            _unlink_evaluation_list_stored_file(row.pdf_relpath)
    q.delete(synchronize_session=False)
    db.commit()
    return redirect(url_for("views.admin_evaluation_lists", unit_key=unit_key))


def _render_admin_evaluation_lists(
    db,
    user: User,
    *,
    unit_key: str,
    unit: dict[str, str] | None,
):
    unit_key = (unit_key or "").strip()
    current_exercise = _admin_current_workspace_exercise(db, user)

    def _display_name_for_upload(filename: str) -> str:
        base = Path(filename or "").name.strip()
        if not base:
            return "تقييم"
        return base[:2000]

    error = ""
    ok_msg = ""
    if request.method == "POST":
        if unit is None:
            error = (
                "لا توجد مستويات وحدة في كتالوج التخطيط. "
                "أضف مستويات الوحدة إلى الكتالوج أولاً (بنك المعلومات له كتالوج مستقل)."
            )
        phase = _normalized_exercise_phase(request.form.get("exercise_phase"))
        if unit is not None and not phase and EXERCISE_PHASE_OPTIONS:
            error = "اختر مرحلة التمرين قبل إضافة الملفات."
        if unit is not None and not EXERCISE_PHASE_OPTIONS:
            error = (
                "لا توجد مراحل تمرين في كتالوج التخطيط. "
                "أضف المراحل أولاً (بنك المعلومات له كتالوج مستقل)."
            )
        if unit is not None and current_exercise is None:
            error = "لا يوجد تمرين حالي. أنشئ تمريناً جديداً قبل إدراج قوائم التقييم."
            files = []
        elif unit is not None and phase:
            files = request.files.getlist("evaluation_lists_file")
        else:
            files = []
        valid_files: list[tuple[bytes, str]] = []
        for f in files:
            if not f or not getattr(f, "filename", ""):
                continue
            try:
                data = f.read()
            except Exception:
                error = "تعذر قراءة أحد الملفات."
                break
            fn = (getattr(f, "filename", "") or "").strip()
            if not fn.lower().endswith(".xlsx"):
                error = "يُقبل فقط ملف Excel بصيغة .xlsx."
                break
            if not _is_xlsx_bytes(data):
                error = "الملف ليس مصنفاً Excel صالحاً (.xlsx)."
                break
            if len(data) > 30 * 1024 * 1024:
                error = "الملف كبير جداً (الحد 30 ميغابايت لكل ملف)."
                break
            valid_files.append((data, _display_name_for_upload(f.filename)))
        if unit is not None and phase and not error and not valid_files:
            error = "اختر ملفاً بصيغة .xlsx (يمكن اختيار عدة ملفات دفعة واحدة)."
        if unit is not None and phase and not error and valid_files:
            stored_hashes = _hashes_of_unit_pdfs(
                db,
                EvaluationListPdfItem,
                unit_key,
                _evaluation_list_file_abspath,
                current_exercise.id if current_exercise else None,
                exercise_phase=phase,
            )
            batch_hashes: set[str] = set()
            to_add: list[tuple[bytes, str]] = []
            skipped_labels: list[str] = []
            for data, label in valid_files:
                h = hashlib.sha256(data).hexdigest()
                if h in stored_hashes or h in batch_hashes:
                    skipped_labels.append(label)
                    continue
                to_add.append((data, label))
                batch_hashes.add(h)
                stored_hashes.add(h)

            if not to_add and skipped_labels:
                error = (
                    "لا يمكن الإضافة: الملف (أو الملفات) مطابق لملف موجود مسبقاً في القائمة الحالية "
                    "(No Duplicate)."
                )
            elif to_add:
                EVALUATION_LIST_XLSX_DIR.mkdir(parents=True, exist_ok=True)
                udir = EVALUATION_LIST_XLSX_DIR / unit_key
                udir.mkdir(parents=True, exist_ok=True)
                mx = (
                    db.query(func.max(EvaluationListPdfItem.sort_order))
                    .filter(
                        EvaluationListPdfItem.unit_level_key == unit_key,
                        EvaluationListPdfItem.exercise_id == current_exercise.id,
                    )
                    .scalar()
                )
                start = (int(mx) if mx is not None else -1) + 1
                n = 0
                for i, (data, label) in enumerate(to_add):
                    rel_name = f"{unit_key}/{uuid.uuid4().hex}.xlsx"
                    full = (EVALUATION_LIST_XLSX_DIR / rel_name).resolve()
                    try:
                        full.parent.mkdir(parents=True, exist_ok=True)
                        full.write_bytes(data)
                    except OSError:
                        error = "تعذر حفظ الملفات على الخادم."
                        db.rollback()
                        break
                    db.add(
                        EvaluationListPdfItem(
                            exercise_id=current_exercise.id if current_exercise else None,
                            exercise_phase=phase,
                            unit_level_key=unit_key,
                            unit_level_label=unit["label"],
                            sort_order=start + i,
                            text=label,
                            pdf_relpath=rel_name.replace("\\", "/"),
                        )
                    )
                    n += 1
                if not error:
                    db.commit()
                    if n > 0 and current_exercise is not None:
                        from app.notifications_service import notify_evaluation_lists_added

                        notify_evaluation_lists_added(
                            db,
                            exercise_id=int(current_exercise.id),
                            unit_key=unit_key,
                            unit_label=unit["label"],
                            n_files=n,
                        )
                        db.commit()
                    if skipped_labels:
                        preview = "، ".join(skipped_labels[:8])
                        if len(skipped_labels) > 8:
                            preview += " …"
                        ok_msg = (
                            f"تمت إضافة {n} ملفاً إلى قوائم التقييم لهذا المستوى. "
                            f"تُرك دون إضافة — مكرر (No Duplicate): {preview}"
                        )
                    else:
                        ok_msg = f"تمت إضافة {n} ملفاً إلى قوائم التقييم لهذا المستوى."

    existing: list = []
    if unit is not None:
        existing_q = db.query(EvaluationListPdfItem).filter(
            EvaluationListPdfItem.unit_level_key == unit_key
        )
        if current_exercise is not None:
            existing_q = existing_q.filter(EvaluationListPdfItem.exercise_id == current_exercise.id)
        else:
            existing_q = existing_q.filter(EvaluationListPdfItem.exercise_id == -1)
        existing = existing_q.order_by(
            _exercise_phase_order_expr(EvaluationListPdfItem.exercise_phase),
            EvaluationListPdfItem.sort_order,
            EvaluationListPdfItem.id,
        ).all()
    unit_label = unit["label"] if unit else "—"
    return render_template(
        "admin_evaluation_lists.html",
        **_ctx(
            user,
            unit_levels=UNIT_LEVELS,
            selected_unit_key=unit_key,
            selected_unit_label=unit_label,
            exercise_phase_options=EXERCISE_PHASE_OPTIONS,
            upload_phase_default=DEFAULT_EXERCISE_PHASE,
            items=existing,
            error=error,
            ok_msg=ok_msg,
            catalog_empty=not UNIT_LEVELS,
            phases_catalog_empty=not EXERCISE_PHASE_OPTIONS,
        ),
    )


@bp.route("/admin/evaluation-lists/<unit_key>", methods=["GET", "POST"])
def admin_evaluation_lists(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/admin/evaluation-lists/{unit_key}")
    _require_planner_hub_catalog_access(user)
    from flask import g

    unit = _require_unit_level_row(unit_key)
    return _render_admin_evaluation_lists(g.db, user, unit_key=unit_key, unit=unit)


@bp.route("/admin/evaluation-lists")
@bp.route("/admin/evaluation-lists/", methods=["GET", "POST"])
def admin_evaluation_lists_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/evaluation-lists")
    _require_planner_hub_catalog_access(user)
    from flask import g

    first = default_unit_level_key()
    if request.method == "GET" and first:
        return redirect(url_for("views.admin_evaluation_lists", unit_key=first))
    return _render_admin_evaluation_lists(g.db, user, unit_key="", unit=None)


@bp.route("/admin/battle-organization", methods=["GET"])
def admin_battle_organization():
    user = get_current_user_optional()
    if not user or not is_system_admin(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    personnel_by_unit: dict[str, dict[str, str]] = {}
    if ex:
        for row in (
            db.query(ExerciseBattleUnitPersonnel)
            .filter(ExerciseBattleUnitPersonnel.exercise_id == ex.id)
            .all()
        ):
            personnel_by_unit[row.unit_id] = {
                "trainee_name": row.trainee_name or "",
                "trainee_military_number": row.trainee_military_number or "",
                "rank_ar": row.rank_ar or "",
                "position_ar": row.position_ar or "",
                "judge_trainee_name": row.judge_trainee_name or "",
                "judge_military_number": row.judge_military_number or "",
                "judge_rank_ar": row.judge_rank_ar or "",
                "judge_position_ar": row.judge_position_ar or "",
            }
    ok_msg = (request.args.get("ok") or "").strip()
    err = (request.args.get("err") or "").strip()
    err_msgs = {
        "no_exercise": "لا يوجد تمرين حالي. أنشئ تمريناً أو اختر مساحة عمل قبل الحفظ.",
        "exercise": "معرّف التمرين غير مطابق للتمرين الحالي.",
        "unit": "معرّف الوحدة غير صالح.",
    }
    return render_template(
        "admin_battle_organization.html",
        **_ctx(
            user,
            battle_tree=BATTLE_ORG_DEMO_ROOT,
            current_exercise=ex,
            personnel_by_unit=personnel_by_unit,
            battle_ok_msg="تم حفظ البيانات." if ok_msg == "1" else "",
            battle_error_msg=err_msgs.get(err, "") if err else "",
        ),
    )


@bp.route("/admin/battle-organization/save", methods=["POST"])
def admin_battle_organization_save():
    user = get_current_user_optional()
    if not user or not is_system_admin(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    if not ex:
        return redirect("/admin/battle-organization?err=no_exercise")
    try:
        exercise_id = int(request.form.get("exercise_id") or "0")
    except ValueError:
        exercise_id = 0
    if exercise_id != ex.id:
        return redirect("/admin/battle-organization?err=exercise")
    unit_id = (request.form.get("unit_id") or "").strip()
    if not unit_id or len(unit_id) > 64:
        return redirect("/admin/battle-organization?err=unit")
    trainee_name = (request.form.get("trainee_name") or "").strip()[:256]
    trainee_military_number = (
        (request.form.get("trainee_military_number") or "").strip()[:128]
    )
    rank_ar = (request.form.get("rank_ar") or "").strip()[:256]
    position_ar = (request.form.get("position_ar") or "").strip()[:512]
    judge_trainee_name = (request.form.get("judge_trainee_name") or "").strip()[:256]
    judge_military_number = (
        (request.form.get("judge_military_number") or "").strip()[:128]
    )
    judge_rank_ar = (request.form.get("judge_rank_ar") or "").strip()[:256]
    judge_position_ar = (request.form.get("judge_position_ar") or "").strip()[:512]
    row = (
        db.query(ExerciseBattleUnitPersonnel)
        .filter(
            ExerciseBattleUnitPersonnel.exercise_id == ex.id,
            ExerciseBattleUnitPersonnel.unit_id == unit_id,
        )
        .first()
    )
    if row is None:
        row = ExerciseBattleUnitPersonnel(
            exercise_id=ex.id,
            unit_id=unit_id,
        )
        db.add(row)
    row.trainee_name = trainee_name
    row.trainee_military_number = trainee_military_number
    row.rank_ar = rank_ar
    row.position_ar = position_ar
    row.judge_trainee_name = judge_trainee_name
    row.judge_military_number = judge_military_number
    row.judge_rank_ar = judge_rank_ar
    row.judge_position_ar = judge_position_ar
    db.commit()
    return redirect("/admin/battle-organization?ok=1")


@bp.route("/admin/dilemmas")
@bp.route("/admin/dilemmas/", methods=["GET", "POST"])
def admin_dilemmas_home():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/dilemmas")
    _require_planner_hub_catalog_access(user)
    from flask import g

    first = default_unit_level_key()
    if request.method == "GET" and first:
        return redirect(url_for("views.admin_dilemmas", unit_key=first))
    return _render_admin_dilemmas(g.db, user, unit_key="", unit=None)


def _render_admin_dilemmas(
    db,
    user: User,
    *,
    unit_key: str,
    unit: dict[str, str] | None,
):
    unit_key = (unit_key or "").strip()
    current_exercise = _admin_current_workspace_exercise(db, user)

    def _display_name_for_upload(filename: str) -> str:
        base = Path(filename or "").name.strip()
        if not base:
            return "معضلة"
        return base[:2000]

    error = ""
    ok_msg = ""
    if request.method == "POST":
        if unit is None:
            error = (
                "لا توجد مستويات وحدة في كتالوج التخطيط. "
                "أضف مستويات الوحدة إلى الكتالوج أولاً (بنك المعلومات له كتالوج مستقل)."
            )
        phase = _normalized_exercise_phase(request.form.get("exercise_phase"))
        if unit is not None and not phase and EXERCISE_PHASE_OPTIONS:
            error = "اختر مرحلة التمرين قبل إضافة الملفات."
        if unit is not None and not EXERCISE_PHASE_OPTIONS:
            error = (
                "لا توجد مراحل تمرين في كتالوج التخطيط. "
                "أضف المراحل أولاً (بنك المعلومات له كتالوج مستقل)."
            )
        if unit is not None and current_exercise is None:
            error = "لا يوجد تمرين حالي. أنشئ تمريناً جديداً قبل إدراج قوائم المعاضل."
            files = []
        elif unit is not None and phase:
            files = request.files.getlist("dilemmas_file")
        else:
            files = []
        valid_files: list[tuple[bytes, str]] = []
        for f in files:
            if not f or not getattr(f, "filename", ""):
                continue
            try:
                data = f.read()
            except Exception:
                error = "تعذر قراءة أحد الملفات."
                break
            if not _is_pdf_bytes(data):
                error = "يُقبل فقط PDF (تعذر التحقق من الملف كـ PDF)."
                break
            if len(data) > 30 * 1024 * 1024:
                error = "الملف كبير جداً (الحد 30 ميغابايت لكل ملف)."
                break
            valid_files.append((data, _display_name_for_upload(f.filename)))
        if unit is not None and phase and not error and not valid_files:
            error = "اختر ملفاً بصيغة PDF (يمكن اختيار عدة ملفات دفعة واحدة)."
        if unit is not None and phase and not error and valid_files:
            stored_hashes = _hashes_of_unit_pdfs(
                db,
                DilemmaItem,
                unit_key,
                _dilemma_pdf_abspath,
                current_exercise.id if current_exercise else None,
                exercise_phase=phase,
            )
            batch_hashes: set[str] = set()
            to_add: list[tuple[bytes, str]] = []
            skipped_labels: list[str] = []
            for data, label in valid_files:
                h = hashlib.sha256(data).hexdigest()
                if h in stored_hashes or h in batch_hashes:
                    skipped_labels.append(label)
                    continue
                to_add.append((data, label))
                batch_hashes.add(h)
                stored_hashes.add(h)

            if not to_add and skipped_labels:
                error = (
                    "لا يمكن الإضافة: الملف (أو الملفات) مطابق لملف موجود مسبقاً في القائمة الحالية "
                    "(No Duplicate)."
                )
            elif to_add:
                DILEMMA_PDF_DIR.mkdir(parents=True, exist_ok=True)
                udir = DILEMMA_PDF_DIR / unit_key
                udir.mkdir(parents=True, exist_ok=True)
                mx = (
                    db.query(func.max(DilemmaItem.sort_order))
                    .filter(
                        DilemmaItem.unit_level_key == unit_key,
                        DilemmaItem.exercise_id == current_exercise.id,
                    )
                    .scalar()
                )
                start = (int(mx) if mx is not None else -1) + 1
                n = 0
                for i, (data, label) in enumerate(to_add):
                    rel_name = f"{unit_key}/{uuid.uuid4().hex}.pdf"
                    full = (DILEMMA_PDF_DIR / rel_name).resolve()
                    try:
                        full.parent.mkdir(parents=True, exist_ok=True)
                        full.write_bytes(data)
                    except OSError:
                        error = "تعذر حفظ الملفات على الخادم."
                        db.rollback()
                        break
                    db.add(
                        DilemmaItem(
                            exercise_id=current_exercise.id if current_exercise else None,
                            exercise_phase=phase,
                            unit_level_key=unit_key,
                            unit_level_label=unit["label"],
                            sort_order=start + i,
                            text=label,
                            pdf_relpath=rel_name.replace("\\", "/"),
                        )
                    )
                    n += 1
                if not error:
                    db.commit()
                    if n > 0 and current_exercise is not None:
                        from app.notifications_service import notify_dilemma_files_added

                        notify_dilemma_files_added(
                            db,
                            exercise_id=int(current_exercise.id),
                            unit_key=unit_key,
                            unit_label=unit["label"],
                            n_files=n,
                        )
                        db.commit()
                    if skipped_labels:
                        preview = "، ".join(skipped_labels[:8])
                        if len(skipped_labels) > 8:
                            preview += " …"
                        ok_msg = (
                            f"تمت إضافة {n} ملفاً معضلة لهذا المستوى. "
                            f"تُرك دون إضافة — مكرر (No Duplicate): {preview}"
                        )
                    else:
                        ok_msg = f"تمت إضافة {n} ملفاً معضلة لهذا المستوى."

    existing: list = []
    if unit is not None:
        existing_q = db.query(DilemmaItem).filter(DilemmaItem.unit_level_key == unit_key)
        if current_exercise is not None:
            existing_q = existing_q.filter(DilemmaItem.exercise_id == current_exercise.id)
        else:
            existing_q = existing_q.filter(DilemmaItem.exercise_id == -1)
        existing = existing_q.order_by(
            _exercise_phase_order_expr(DilemmaItem.exercise_phase),
            DilemmaItem.sort_order,
            DilemmaItem.id,
        ).all()
    unit_label = unit["label"] if unit else "—"
    return render_template(
        "admin_dilemmas.html",
        **_ctx(
            user,
            unit_levels=UNIT_LEVELS,
            selected_unit_key=unit_key,
            selected_unit_label=unit_label,
            exercise_phase_options=EXERCISE_PHASE_OPTIONS,
            upload_phase_default=DEFAULT_EXERCISE_PHASE,
            items=existing,
            error=error,
            ok_msg=ok_msg,
            catalog_empty=not UNIT_LEVELS,
            phases_catalog_empty=not EXERCISE_PHASE_OPTIONS,
        ),
    )


@bp.route("/admin/dilemmas/<unit_key>", methods=["GET", "POST"])
def admin_dilemmas(unit_key: str):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/admin/dilemmas/{unit_key}")
    _require_planner_hub_catalog_access(user)
    from flask import g

    unit = _require_unit_level_row(unit_key)
    return _render_admin_dilemmas(g.db, user, unit_key=unit_key, unit=unit)


@bp.route("/admin/dilemmas/<unit_key>/view/<int:item_id>", methods=["GET"])
def admin_dilemma_viewer(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or (current_exercise is not None and row.exercise_id != current_exercise.id):
        abort(404)
    list_url = url_for("views.admin_dilemmas", unit_key=unit_key)
    if not (row.pdf_relpath or "").strip():
        return redirect(list_url)
    if _dilemma_pdf_abspath(row.pdf_relpath) is None:
        return redirect(list_url)
    pdf_url = url_for(
        "views.admin_dilemma_pdf_file", unit_key=unit_key, item_id=item_id
    )
    return render_template(
        "admin_dilemma_viewer.html",
        **_ctx(
            user,
            unit_key=unit_key,
            item_id=item_id,
            item_title=row.text or "معضلة",
            pdf_url=pdf_url,
            subpage_close_fallback=list_url,
            **_hub_back_ctx_for_request_path(),
        ),
    )


@bp.route("/admin/dilemmas/<unit_key>/item/<int:item_id>/pdf", methods=["GET"])
def admin_dilemma_pdf_file(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if not row or row.unit_level_key != unit_key or (current_exercise is not None and row.exercise_id != current_exercise.id):
        abort(404)
    rel = (row.pdf_relpath or "").strip()
    if not rel:
        abort(404)
    path = _dilemma_pdf_abspath(rel)
    if path is None:
        abort(404)
    return send_file(path, mimetype="application/pdf", as_attachment=False)


@bp.route("/admin/dilemmas/<unit_key>/item/<int:item_id>/delete", methods=["POST"])
def admin_dilemmas_delete_item(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not row
        or row.unit_level_key != unit_key
        or (current_exercise is not None and row.exercise_id != current_exercise.id)
    ):
        abort(404)
    if row.pdf_relpath:
        _unlink_dilemma_stored_pdf(row.pdf_relpath)
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_dilemmas", unit_key=unit_key))


@bp.route(
    "/admin/dilemmas/<unit_key>/item/<int:item_id>/phase",
    methods=["POST"],
)
def admin_dilemmas_set_phase(unit_key: str, item_id: int):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    row = db.get(DilemmaItem, item_id)
    current_exercise = _admin_current_workspace_exercise(db, user)
    if (
        not row
        or row.unit_level_key != unit_key
        or (current_exercise is not None and row.exercise_id != current_exercise.id)
    ):
        abort(404)
    row.exercise_phase = _normalized_exercise_phase(request.form.get("exercise_phase"))
    db.commit()
    return redirect(url_for("views.admin_dilemmas", unit_key=unit_key))


@bp.route("/admin/dilemmas/<unit_key>/clear", methods=["POST"])
def admin_dilemmas_clear(unit_key: str):
    user = get_current_user_optional()
    _require_planner_hub_catalog_access(user)
    unit = _require_unit_level_row(unit_key)
    from flask import g

    db = g.db
    current_exercise = _admin_current_workspace_exercise(db, user)
    q = db.query(DilemmaItem).filter(DilemmaItem.unit_level_key == unit_key)
    if current_exercise is not None:
        q = q.filter(DilemmaItem.exercise_id == current_exercise.id)
    else:
        q = q.filter(DilemmaItem.exercise_id == -1)
    for row in q.all():
        if row.pdf_relpath:
            _unlink_dilemma_stored_pdf(row.pdf_relpath)
    q.delete(synchronize_session=False)
    db.commit()
    return redirect(url_for("views.admin_dilemmas", unit_key=unit_key))


_INFO_BANK_MAX_PDF = 50 * 1024 * 1024
_INFO_BANK_MAX_XLSX = 30 * 1024 * 1024


def _info_bank_next_sort_order(db, model, phase: str, unit: str) -> int:
    mx = (
        db.query(func.max(model.sort_order))
        .filter(
            model.training_phase_key == phase,
            model.unit_level_key == unit,
        )
        .scalar()
    )
    return (int(mx) if mx is not None else -1) + 1


def _ensure_information_bank_catalog_rows(db) -> None:
    """يجب أن تكون صفوف الكتالوج الافتراضي مطابقة لـ ``TRAINING_PHASES`` وقوالب مستويات الوحدات.

    تنشئ أي مفاتيح ناقصة عند أول تشغيل. مراحل التمرين: تُحدَّث التسمية من الكتالوج البرمجي.
    مستويات الوحدات: تُحدَّث الترتيب ومجموعة اللواء فقط — تسمية المستخدم من زر «تعديل» لا تُستبدل.
    """
    changed = False
    for idx, row in enumerate(TRAINING_PHASES):
        r = db.get(InformationBankTrainingPhase, row["key"])
        if r is None:
            db.add(
                InformationBankTrainingPhase(
                    key=row["key"],
                    label=row["label"],
                    sort_order=idx,
                    is_system=True,
                )
            )
            changed = True
        elif r.label != row["label"] or r.sort_order != idx:
            r.label = row["label"]
            r.sort_order = idx
            r.is_system = True
            changed = True
    from app.ibank_ui import ibank_brigade_groups_for_page

    for bg in ibank_brigade_groups_for_page():
        bg_key = bg["key"]
        for idx, row in enumerate(INFO_BANK_UNIT_LEVEL_TEMPLATES):
            catalog_key = unit_catalog_key_for_brigade(bg_key, row["key"])
            if not catalog_key:
                continue
            r = db.get(InformationBankUnitLevel, catalog_key)
            if r is None:
                db.add(
                    InformationBankUnitLevel(
                        key=catalog_key,
                        label=row["label"],
                        brigade_group=bg_key,
                        sort_order=idx,
                        is_system=True,
                    )
                )
                changed = True
            else:
                if getattr(r, "brigade_group", None) != bg_key:
                    r.brigade_group = bg_key
                    changed = True
                # لا نُعيد فرض التسمية من القالب — يُحفظ تعديل المستخدم عبر زر «تعديل»
                if r.sort_order != idx:
                    r.sort_order = idx
                    changed = True
                if not r.is_system:
                    r.is_system = True
                    changed = True
    legacy_rows = (
        db.query(InformationBankUnitLevel)
        .filter(
            (InformationBankUnitLevel.brigade_group == "")
            | (InformationBankUnitLevel.brigade_group.is_(None))
        )
        .all()
    )
    for r in legacy_rows:
        r.brigade_group = "1"
        changed = True
    if changed:
        db.commit()


def _information_bank_training_phases(db) -> list[dict[str, str | bool]]:
    _ensure_information_bank_catalog_rows(db)
    rows = (
        db.query(InformationBankTrainingPhase)
        .order_by(
            InformationBankTrainingPhase.sort_order,
            InformationBankTrainingPhase.created_at,
            InformationBankTrainingPhase.key,
        )
        .all()
    )
    return [
        {
            "key": r.key,
            "label": r.label,
            "included_in_exercise": bool(getattr(r, "included_in_exercise", False)),
        }
        for r in rows
        if (r.key or "").strip()
    ]


def _information_bank_unit_levels(db, brigade_group: str | None = None) -> list[dict[str, str | bool]]:
    _ensure_information_bank_catalog_rows(db)
    q = db.query(InformationBankUnitLevel)
    if brigade_group is not None:
        q = q.filter(InformationBankUnitLevel.brigade_group == (brigade_group or "").strip())
    rows = q.order_by(
        InformationBankUnitLevel.sort_order,
        InformationBankUnitLevel.created_at,
        InformationBankUnitLevel.key,
    ).all()
    return [
        {
            "key": r.key,
            "label": r.label,
            "brigade_group": getattr(r, "brigade_group", "") or "1",
            "included_in_exercise": bool(getattr(r, "included_in_exercise", False)),
        }
        for r in rows
        if (r.key or "").strip()
    ]


def _information_bank_brigade_units_map(db) -> dict[str, list[dict[str, str | bool]]]:
    from app.ibank_ui import ibank_brigade_groups_for_page

    return {
        bg["key"]: _information_bank_unit_levels(db, bg["key"])
        for bg in ibank_brigade_groups_for_page()
    }


def _information_bank_gate_ok() -> bool:
    return information_bank_gate_ok(session)


def _verify_system_admin_password(db, password: str) -> bool:
    """كلمة مرور أي حساب نشط بدور إدارة النظام."""
    pwd = (password or "").strip()
    if not pwd:
        return False
    admins = (
        db.query(User)
        .filter(User.role_key == RoleKey.SYSTEM_ADMIN.value, User.is_active.is_(True))
        .all()
    )
    for adm in admins:
        if verify_password(pwd, adm.password_hash):
            return True
    return False


def _require_information_bank_gate():
    """إعادة توجيه لبوابة كلمة مرور إدارة النظام إن لم تُفتح الجلسة."""
    if _information_bank_gate_ok():
        return None
    nxt = (request.full_path or request.path or "/admin/information-bank").strip()
    return redirect(url_for("views.admin_information_bank_gate", next=nxt))


def _included_unit_keys_from_form() -> set[str]:
    raw = request.form.getlist("included_unit_keys")
    return {(x or "").strip() for x in raw if (x or "").strip()}


def _included_phase_keys_from_form() -> set[str]:
    raw = request.form.getlist("included_phase_keys")
    return {(x or "").strip() for x in raw if (x or "").strip()}


def _ibank_included_save_http_response(*, tab: str, ok_msg: str = "", err_msg: str = ""):
    """حفظ التحديدات دون إعادة تحميل كاملة عند طلب AJAX من واجهة بنك المعلومات."""
    if is_ibank_included_save_request():
        if err_msg:
            return jsonify({"ok": False, "error": err_msg}), 400
        return jsonify({"ok": True, "message": ok_msg})
    if err_msg:
        return redirect(
            url_for("views.admin_information_bank", tab=tab, err=err_msg)
        )
    return redirect(url_for("views.admin_information_bank", tab=tab, ok=ok_msg))


def _ibank_included_save_auth_or_response(tab: str):
    """تحقق الجلسة والصلاحية؛ يُرجع استجابة JSON أو إعادة توجيه عند فشل حفظ AJAX."""
    user = get_current_user_optional()
    if not user:
        if is_ibank_included_save_request():
            return jsonify({"ok": False, "error": "يجب تسجيل الدخول أولاً."}), 401
        return redirect("/login?next=/admin/information-bank")
    if not can_manage_information_bank(user):
        if is_ibank_included_save_request():
            return jsonify({"ok": False, "error": "ليس لديك صلاحية تعديل بنك المعلومات."}), 403
        abort(403)
    return None


def _information_bank_training_phase_label(db, key: str | None) -> str:
    k = (key or "").strip()
    for row in _information_bank_training_phases(db):
        if row["key"] == k:
            return row["label"]
    return training_phase_label(k)


def _information_bank_unit_label(db, key: str | None) -> str:
    k = (key or "").strip()
    for row in _information_bank_unit_levels(db):
        if row["key"] == k:
            return row["label"]
    return info_bank_unit_label(k)


def _is_valid_information_bank_phase(db, key: str | None) -> bool:
    k = (key or "").strip()
    return any(row["key"] == k for row in _information_bank_training_phases(db))


def _is_valid_information_bank_unit(db, key: str | None) -> bool:
    k = (key or "").strip()
    return any(row["key"] == k for row in _information_bank_unit_levels(db))


def _custom_catalog_key(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


@bp.route("/admin/information-bank/exit", methods=["GET"])
def admin_information_bank_exit():
    """إنهاء جلسة بنك المعلومات والخروج دون طلب كلمة مرور."""
    user = get_current_user_optional()
    if not user:
        return redirect("/login")
    clear_information_bank_gate(session)
    dest = (request.args.get("next") or "/dashboard").strip()
    if not dest.startswith("/") or dest.startswith("//"):
        dest = "/dashboard"
    if is_information_bank_path(dest):
        dest = "/dashboard"
    return redirect(dest)


@bp.route("/admin/information-bank/gate", methods=["GET", "POST"])
def admin_information_bank_gate():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/information-bank/gate")
    if not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    nxt = (request.args.get("next") or request.form.get("next") or "/admin/information-bank").strip()
    if not nxt.startswith("/"):
        nxt = "/admin/information-bank"
    err = ""
    if request.method == "POST":
        pwd = (request.form.get("password") or "").strip()
        if _verify_system_admin_password(db, pwd):
            session[INFO_BANK_GATE_SESSION_KEY] = True
            session.modified = True
            return redirect(nxt)
        err = "كلمة المرور غير صحيحة. أدخل كلمة مرور حساب إدارة النظام."
    if _information_bank_gate_ok():
        return redirect(nxt)
    return render_template(
        "admin_information_bank_gate.html",
        **_ctx(user, next_url=nxt, gate_error=err),
    )


@bp.route("/admin/information-bank/phases/included", methods=["POST"])
def admin_information_bank_phases_included_save():
    auth_resp = _ibank_included_save_auth_or_response("phases")
    if auth_resp is not None:
        return auth_resp
    from flask import g

    db = g.db
    try:
        _ensure_information_bank_catalog_rows(db)
        new_keys = _included_phase_keys_from_form()
        rows = db.query(InformationBankTrainingPhase).all()
        prev_keys = {
            r.key for r in rows if bool(getattr(r, "included_in_exercise", False))
        }
        removed = prev_keys - new_keys
        if removed:
            uncheck_pwd = (request.form.get("uncheck_password") or "").strip()
            if not _verify_system_admin_password(db, uncheck_pwd):
                return _ibank_included_save_http_response(
                    tab="phases",
                    err_msg="إلغاء الإدراج في التمرين يتطلب كلمة مرور إدارة النظام الصحيحة.",
                )
        for row in rows:
            row.included_in_exercise = (row.key or "").strip() in new_keys
        db.commit()
        from app.planning_catalog_sync import sync_planning_catalogs_from_db

        sync_planning_catalogs_from_db(db)
        return _ibank_included_save_http_response(
            tab="phases",
            ok_msg="تم حفظ تحديدات وإلغاءات مراحل التمرين — تُطبَّق على قوائم التخطيط والمحكمين.",
        )
    except Exception:
        db.rollback()
        logging.getLogger(__name__).exception("phases included save failed")
        return _ibank_included_save_http_response(
            tab="phases",
            err_msg="حدث خطأ أثناء الحفظ. أعد المحاولة.",
        )


@bp.route("/admin/information-bank/units/included", methods=["POST"])
def admin_information_bank_units_included_save():
    return_tab = (request.form.get("return_tab") or "units-bg-1").strip()
    from app.ibank_ui import ibank_brigade_groups_for_page

    if return_tab not in {bg["tab"] for bg in ibank_brigade_groups_for_page()}:
        return_tab = "units-bg-1"
    auth_resp = _ibank_included_save_auth_or_response(return_tab)
    if auth_resp is not None:
        return auth_resp
    from flask import g

    db = g.db
    try:
        _ensure_information_bank_catalog_rows(db)
        new_keys = _included_unit_keys_from_form()
        rows = db.query(InformationBankUnitLevel).all()
        prev_keys = {
            r.key for r in rows if bool(getattr(r, "included_in_exercise", False))
        }
        removed = prev_keys - new_keys
        if removed:
            uncheck_pwd = (request.form.get("uncheck_password") or "").strip()
            if not _verify_system_admin_password(db, uncheck_pwd):
                return _ibank_included_save_http_response(
                    tab=return_tab,
                    err_msg="إلغاء الإدراج في التمرين يتطلب كلمة مرور إدارة النظام الصحيحة.",
                )
        from app.ibank_ui import unit_level_row_is_removed_brigade

        for row in rows:
            key = (row.key or "").strip()
            if unit_level_row_is_removed_brigade(
                key=key, brigade_group=getattr(row, "brigade_group", None)
            ):
                row.included_in_exercise = False
                continue
            row.included_in_exercise = key in new_keys
        db.commit()
        from app.planning_catalog_sync import sync_planning_catalogs_from_db

        sync_planning_catalogs_from_db(db)
        return _ibank_included_save_http_response(
            tab=return_tab,
            ok_msg="تم حفظ تحديدات وإلغاءات مستويات الوحدات — تُطبَّق على قوائم التخطيط والمحكمين.",
        )
    except Exception:
        db.rollback()
        logging.getLogger(__name__).exception("units included save failed")
        return _ibank_included_save_http_response(
            tab=return_tab,
            err_msg="حدث خطأ أثناء الحفظ. أعد المحاولة.",
        )


def _reload_information_bank_catalog_module():
    """إعادة تحميل كتالوج بنك المعلومات (debugpy لا يعيد تحميل Python تلقائياً)."""
    import importlib
    from app import information_bank_catalog as ibc

    return importlib.reload(ibc)


def _ibank_ui_brigade_groups() -> list[dict[str, str]]:
    """مجموعات الألوية الظاهرة في الواجهة — دائماً من أحدث كتالوج على القرص."""
    ibc = _reload_information_bank_catalog_module()
    return ibc.info_bank_brigade_groups_for_ui()


@bp.route("/admin/information-bank", methods=["GET"])
def admin_information_bank():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/admin/information-bank")
    if not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    training_phases = _information_bank_training_phases(db)
    info_bank_brigade_units = _information_bank_brigade_units_map(db)
    phase_notes = {r.phase_key: r.notes for r in db.query(InformationBankPhaseNote).all()}
    unit_notes = {r.unit_level_key: r.notes for r in db.query(InformationBankUnitNote).all()}
    from app.info_bank_tree import (
        build_tree_payload,
        ensure_all_information_bank_trees,
        exercise_judge_names_by_unit,
    )

    tree_event_flow: list = []
    tree_action_eval: list = []
    tree_dilemma_eval: list = []
    ibank_unit_labels = {
        (u.get("key") or "").strip(): (u.get("label") or "").strip()
        for u in UNIT_LEVELS
        if (u.get("key") or "").strip()
    }
    current_exercise = _admin_current_workspace_exercise(db, user)
    ibank_judge_names_by_unit = exercise_judge_names_by_unit(
        db, int(current_exercise.id) if current_exercise else None
    )
    try:
        ensure_all_information_bank_trees(db)
        tree_event_flow = build_tree_payload(db, "event_flow", unit_label_by_key=ibank_unit_labels)
        tree_action_eval = build_tree_payload(
            db,
            "action_eval",
            unit_label_by_key=ibank_unit_labels,
            judge_name_by_unit=ibank_judge_names_by_unit,
        )
        tree_dilemma_eval = build_tree_payload(
            db, "dilemma_eval", unit_label_by_key=ibank_unit_labels
        )
    except Exception as exc:
        db.rollback()
        current_app.logger.exception("information bank tree build failed: %s", exc)
    err = (request.args.get("err") or "").strip()[:2000]
    ok = (request.args.get("ok") or "").strip()[:500]
    from app.ibank_ui import ibank_brigade_groups_for_page, is_removed_brigade_tab

    ui_brigade_groups = ibank_brigade_groups_for_page()
    brigade_tabs = {bg["tab"] for bg in ui_brigade_groups}
    active_tab = (request.args.get("tab") or "phases").strip()
    if is_removed_brigade_tab(active_tab):
        active_tab = "units-bg-1"
    allowed_tabs = {"phases", "event-flow", "action-eval", "dilemma-eval"} | brigade_tabs
    if active_tab not in allowed_tabs:
        active_tab = "phases"
    from flask import make_response

    resp = make_response(
        render_template(
            "admin_information_bank.html",
            **_ctx(
                user,
                training_phases=training_phases,
                info_bank_brigade_groups=ui_brigade_groups,
                info_bank_brigade_units=info_bank_brigade_units,
                phase_notes=phase_notes,
                unit_notes=unit_notes,
                tree_event_flow=tree_event_flow,
                tree_action_eval=tree_action_eval,
                tree_dilemma_eval=tree_dilemma_eval,
                training_phase_label=lambda key: _information_bank_training_phase_label(
                    db, key
                ),
                info_bank_unit_label=lambda key: _information_bank_unit_label(db, key),
                error=err,
                ok_msg=ok,
                active_tab=active_tab,
                information_bank_can_manage=can_manage_information_bank(user),
                ibank_tree_unit_levels=list(UNIT_LEVELS),
                ibank_judge_names_by_unit=ibank_judge_names_by_unit,
            ),
        )
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["X-Ibank-Ui-Build"] = "action-eval-unit-v9"
    return resp


@bp.route("/admin/information-bank/phase-notes", methods=["POST"])
def admin_information_bank_phase_notes():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    for p in TRAINING_PHASES:
        key = p["key"]
        raw = request.form.get(f"note_{key}") or ""
        row = db.get(InformationBankPhaseNote, key)
        if row is None:
            row = InformationBankPhaseNote(phase_key=key, notes=raw)
            db.add(row)
        else:
            row.notes = raw
    db.commit()
    return redirect(url_for("views.admin_information_bank", ok="تم حفظ بيانات مراحل التمرين."))


@bp.route("/admin/information-bank/unit-notes", methods=["POST"])
def admin_information_bank_unit_notes():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    for u in INFO_BANK_UNIT_LEVELS:
        key = u["key"]
        raw = request.form.get(f"unote_{key}") or ""
        row = db.get(InformationBankUnitNote, key)
        if row is None:
            row = InformationBankUnitNote(unit_level_key=key, notes=raw)
            db.add(row)
        else:
            row.notes = raw
    db.commit()
    return redirect(url_for("views.admin_information_bank", ok="تم حفظ بيانات مستويات الوحدات."))


@bp.route("/admin/information-bank/phases/add", methods=["POST"])
def admin_information_bank_phase_add():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    label = (request.form.get("phase_label") or "").strip()[:300]
    if not label:
        return redirect(url_for("views.admin_information_bank", tab="phases", err="أدخل اسم المرحلة."))
    db = g.db
    _ensure_information_bank_catalog_rows(db)
    mx = db.query(func.max(InformationBankTrainingPhase.sort_order)).scalar()
    next_order = (int(mx) if mx is not None else -1) + 1
    phase_key = _custom_catalog_key("phase")
    db.add(
        InformationBankTrainingPhase(
            key=phase_key,
            label=label,
            sort_order=next_order,
            is_system=False,
        )
    )
    db.commit()
    from app.info_bank_tree import INFO_BANK_TREE_KINDS, _unit_rows, ensure_information_bank_tree, get_or_create_folder

    for k in INFO_BANK_TREE_KINDS:
        ensure_information_bank_tree(db, k)
        phase_node = get_or_create_folder(
            db,
            kind=k,
            parent_id=None,
            name=label,
            is_system=False,
            catalog_phase_key=phase_key,
        )
        for un in _unit_rows(db):
            get_or_create_folder(
                db,
                kind=k,
                parent_id=int(phase_node.id),
                name=(un.label or un.key)[:500],
                is_system=bool(un.is_system),
                catalog_unit_key=un.key,
            )
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab="phases", ok="تمت إضافة مرحلة التمرين."))


@bp.route("/admin/information-bank/phases/delete", methods=["POST"])
def admin_information_bank_phase_delete():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    key = (request.form.get("phase_key") or "").strip()
    db = g.db
    row = db.get(InformationBankTrainingPhase, key)
    if row is None:
        return redirect(url_for("views.admin_information_bank", tab="phases", err="اختر مرحلة صالحة للحذف."))
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab="phases", ok="تم حذف مرحلة التمرين."))


@bp.route("/admin/information-bank/units/add", methods=["POST"])
def admin_information_bank_unit_add():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    label = (request.form.get("unit_label") or "").strip()[:300]
    tab = (request.form.get("brigade_tab") or "units-bg-1").strip()
    from app.ibank_ui import ibank_brigade_groups_for_page, is_removed_brigade_tab

    if is_removed_brigade_tab(tab) or tab not in {bg["tab"] for bg in ibank_brigade_groups_for_page()}:
        tab = "units-bg-1"
    bg_key = brigade_group_for_tab(tab)
    if not label:
        return redirect(url_for("views.admin_information_bank", tab=tab, err="أدخل اسم مستوى الوحدة."))
    db = g.db
    _ensure_information_bank_catalog_rows(db)
    mx = (
        db.query(func.max(InformationBankUnitLevel.sort_order))
        .filter(InformationBankUnitLevel.brigade_group == bg_key)
        .scalar()
    )
    next_order = (int(mx) if mx is not None else -1) + 1
    unit_key = _custom_catalog_key(f"unit_bg{bg_key}")
    db.add(
        InformationBankUnitLevel(
            key=unit_key,
            label=label,
            brigade_group=bg_key,
            sort_order=next_order,
            is_system=False,
        )
    )
    db.commit()
    from app.info_bank_tree import INFO_BANK_TREE_KINDS, ensure_information_bank_tree, get_or_create_folder

    for k in INFO_BANK_TREE_KINDS:
        if k == "action_eval":
            continue
        ensure_information_bank_tree(db, k)
        phase_nodes = (
            db.query(InformationBankTreeNode)
            .filter(
                InformationBankTreeNode.kind == k,
                InformationBankTreeNode.parent_id.is_(None),
                InformationBankTreeNode.is_folder.is_(True),
            )
            .all()
        )
        for ph in phase_nodes:
            get_or_create_folder(
                db,
                kind=k,
                parent_id=int(ph.id),
                name=label,
                is_system=False,
                catalog_unit_key=unit_key,
            )
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab=tab, ok="تمت إضافة مستوى الوحدة."))


@bp.route("/admin/information-bank/units/edit", methods=["POST"])
def admin_information_bank_unit_edit():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    key = (request.form.get("unit_key") or "").strip()
    label = (request.form.get("unit_label") or "").strip()[:300]
    tab = (request.form.get("brigade_tab") or "units-bg-1").strip()
    from app.ibank_ui import ibank_brigade_groups_for_page, is_removed_brigade_tab

    if is_removed_brigade_tab(tab) or tab not in {bg["tab"] for bg in ibank_brigade_groups_for_page()}:
        tab = "units-bg-1"
    ajax = (request.headers.get("X-Requested-With") or "").strip() == "XMLHttpRequest"

    def _edit_response(*, ok: bool, err_msg: str = "", **extra):
        if ajax:
            if ok:
                return jsonify(ok=True, **extra)
            return jsonify(ok=False, error=err_msg), 400
        if ok:
            return redirect(
                url_for("views.admin_information_bank", tab=tab, ok="تم تعديل مستوى الوحدة.")
            )
        return redirect(url_for("views.admin_information_bank", tab=tab, err=err_msg))

    if not key or not label:
        return _edit_response(ok=False, err_msg="أدخل اسماً صالحاً للوحدة.")
    db = g.db
    row = db.get(InformationBankUnitLevel, key)
    if row is None:
        return _edit_response(ok=False, err_msg="مستوى الوحدة غير موجود.")
    row.label = label
    for node in (
        db.query(InformationBankTreeNode)
        .filter(
            InformationBankTreeNode.catalog_unit_key == key,
            InformationBankTreeNode.is_folder.is_(True),
        )
        .all()
    ):
        node.name = label[:500]
    db.commit()
    return _edit_response(ok=True, label=label, unit_key=key, tab=tab)


@bp.route("/admin/information-bank/units/delete", methods=["POST"])
def admin_information_bank_unit_delete():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    key = (request.form.get("unit_key") or "").strip()
    tab = (request.form.get("brigade_tab") or "units-bg-1").strip()
    from app.ibank_ui import ibank_brigade_groups_for_page, is_removed_brigade_tab

    if is_removed_brigade_tab(tab) or tab not in {bg["tab"] for bg in ibank_brigade_groups_for_page()}:
        tab = "units-bg-1"
    db = g.db
    row = db.get(InformationBankUnitLevel, key)
    if row is None:
        return redirect(url_for("views.admin_information_bank", tab=tab, err="اختر مستوى وحدة صالحاً للحذف."))
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab=tab, ok="تم حذف مستوى الوحدة."))


@bp.route("/admin/information-bank/tree/folder", methods=["POST"])
def admin_information_bank_tree_folder_add():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    from app.info_bank_tree import add_custom_folder, ensure_information_bank_tree, kind_tab

    db = g.db
    kind = (request.form.get("kind") or "").strip()
    if kind not in ("event_flow", "action_eval", "dilemma_eval"):
        abort(400)
    parent_raw = (request.form.get("parent_id") or "").strip()
    parent_id = int(parent_raw) if parent_raw.isdigit() else None
    name = (request.form.get("folder_name") or "").strip()
    tab = kind_tab(kind)
    if not name:
        return redirect(url_for("views.admin_information_bank", tab=tab, err="أدخل اسم المجلد."))
    ensure_information_bank_tree(db, kind)
    try:
        add_custom_folder(db, kind=kind, parent_id=parent_id, name=name)
        db.commit()
    except ValueError as exc:
        db.rollback()
        return redirect(url_for("views.admin_information_bank", tab=tab, err=str(exc) or "تعذر إنشاء المجلد."))
    return redirect(url_for("views.admin_information_bank", tab=tab, ok="تم إنشاء المجلد."))


@bp.route("/admin/information-bank/tree/upload", methods=["POST"])
def admin_information_bank_tree_upload():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    from app.info_bank_tree import (
        _is_phase_root_folder,
        _unit_key_for_node,
        ensure_information_bank_tree,
        get_node,
        kind_tab,
        upload_files_to_parent,
        upload_includes_subdirectory_paths,
    )

    db = g.db
    kind = (request.form.get("kind") or "").strip()
    tab = kind_tab(kind)
    if kind not in ("event_flow", "action_eval", "dilemma_eval"):
        abort(400)
    parent_raw = (request.form.get("parent_id") or "").strip()
    if not parent_raw.isdigit():
        return redirect(
            url_for("views.admin_information_bank", tab=tab, err="حدّد مجلداً مستهدفاً في الشجرة (زر تحديد).")
        )
    parent_id = int(parent_raw)
    ensure_information_bank_tree(db, kind)
    parent = get_node(db, parent_id, kind)
    if parent is None or not parent.is_folder:
        return redirect(url_for("views.admin_information_bank", tab=tab, err="المجلد المستهدف غير صالح."))
    files = [x for x in request.files.getlist("files") if x and getattr(x, "filename", "").strip()]
    if kind == "action_eval":
        has_subdirs = upload_includes_subdirectory_paths(files)
        if _is_phase_root_folder(parent):
            if not has_subdirs:
                return redirect(
                    url_for(
                        "views.admin_information_bank",
                        tab=tab,
                        err="حدّد مرحلة التمرين ثم «إرفاق مجلد» (مجلد كامل وليس ملفات متفرقة).",
                    )
                )
        elif not _unit_key_for_node(db, parent) and not has_subdirs:
            return redirect(
                url_for(
                    "views.admin_information_bank",
                    tab=tab,
                    err="اختر مستوى الوحدة من القائمة على المجلد المرفق، أو ارفق مجلداً فرعياً.",
                )
            )
    if not files:
        return redirect(url_for("views.admin_information_bank", tab=tab, err="اختر ملفاً أو مجلداً للإدراج."))
    added, errors = upload_files_to_parent(db, kind=kind, parent_id=parent_id, file_storages=files)
    if added:
        db.commit()
    else:
        db.rollback()
    err_q = " ".join(errors)[:2000] if errors else ""
    if not added:
        return redirect(url_for("views.admin_information_bank", tab=tab, err=err_q or "لم تُضف أي ملف."))
    ok_msg = f"تم إدراج {added} ملف(ات)."
    if err_q:
        return redirect(
            url_for("views.admin_information_bank", tab=tab, ok=ok_msg, err=f"تجاهل بعض الملفات: {err_q}")
        )
    return redirect(url_for("views.admin_information_bank", tab=tab, ok=ok_msg))


@bp.route("/admin/information-bank/tree/<int:node_id>/delete", methods=["POST"])
def admin_information_bank_tree_delete(node_id: int):
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    from app.info_bank_tree import delete_node, kind_tab

    db = g.db
    row = db.get(InformationBankTreeNode, node_id)
    if row is None:
        abort(404)
    tab = kind_tab(row.kind)
    delete_node(db, row)
    db.commit()
    if (
        request.accept_mimetypes.best_match(["application/json", "text/html"])
        == "application/json"
        or (request.headers.get("X-Requested-With") or "").strip() == "XMLHttpRequest"
    ):
        return jsonify(ok=True, node_id=int(node_id), tab=tab)
    return redirect(url_for("views.admin_information_bank", tab=tab, ok="تم الحذف."))


@bp.route("/admin/information-bank/tree/<int:node_id>/unit-level", methods=["POST"])
def admin_information_bank_tree_unit_level(node_id: int):
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    from app.info_bank_tree import kind_tab, set_folder_unit_level

    db = g.db
    kind = (request.form.get("kind") or "").strip()
    if kind != "action_eval":
        abort(400)
    tab = kind_tab(kind)
    unit_key = (request.form.get("unit_key") or "").strip()
    row = db.get(InformationBankTreeNode, node_id)
    is_folder = bool(row and row.is_folder)
    try:
        set_folder_unit_level(db, kind=kind, node_id=node_id, unit_key=unit_key)
        db.commit()
    except ValueError as exc:
        db.rollback()
        return redirect(
            url_for("views.admin_information_bank", tab=tab, err=str(exc) or "تعذّر التعيين.")
        )
    ok_msg = (
        "تم تعيين مستوى الوحدة وتطبيقه على المجلد وجميع محتوياته."
        if is_folder
        else "تم تعيين مستوى الوحدة للملف."
    )
    return redirect(url_for("views.admin_information_bank", tab=tab, ok=ok_msg))


@bp.route("/admin/information-bank/tree/move", methods=["POST"])
def admin_information_bank_tree_move():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        return jsonify(ok=False, error="غير مسموح."), 403
    from flask import g

    from app.info_bank_tree import move_tree_node

    data = request.get_json(force=True, silent=True) or {}
    kind = (data.get("kind") or "").strip()
    if kind not in ("event_flow", "action_eval", "dilemma_eval"):
        return jsonify(ok=False, error="نوع المرفقات غير صالح."), 400
    try:
        nid = int(data.get("node_id"))
    except (TypeError, ValueError):
        return jsonify(ok=False, error="بيانات غير صالحة."), 400
    parent_raw = data.get("parent_id")
    try:
        pid = int(parent_raw) if parent_raw is not None and parent_raw != "" else 0
    except (TypeError, ValueError):
        return jsonify(ok=False, error="بيانات غير صالحة."), 400
    parent_id = None if pid < 1 else pid
    db = g.db
    try:
        move_tree_node(db, kind=kind, node_id=nid, parent_id=parent_id)
        db.commit()
    except ValueError as exc:
        db.rollback()
        return jsonify(ok=False, error=str(exc) or "تعذّر النقل."), 400
    return jsonify(ok=True)


@bp.route("/admin/information-bank/tree/<int:node_id>/file", methods=["GET"])
def admin_information_bank_tree_file(node_id: int):
    user = get_current_user_optional()
    if not user or not can_view_information_bank(user):
        abort(403)
    from flask import g

    from app.info_bank_tree import node_file_abspath

    db = g.db
    row = db.get(InformationBankTreeNode, node_id)
    if row is None or row.is_folder or not (row.file_relpath or "").strip():
        abort(404)
    path = node_file_abspath(row.kind, row.file_relpath)
    if path is None:
        abort(404)
    low = path.name.lower()
    if low.endswith(".xlsx"):
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        mt = _mimetype_info_bank_event_flow(path)
    return send_file(path, mimetype=mt, as_attachment=False, download_name=row.name or path.name)


@bp.route("/admin/information-bank/event-flow/upload", methods=["POST"])
def admin_information_bank_event_flow_upload():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    phase = (request.form.get("training_phase_key") or "").strip()
    unit = (request.form.get("unit_level_key") or "").strip()
    if not _is_valid_information_bank_phase(db, phase) or not _is_valid_information_bank_unit(db, unit):
        return redirect(url_for("views.admin_information_bank", tab="event-flow", err="مرحلة أو مستوى وحدة غير صالح."))
    files = [x for x in request.files.getlist("file") if x and getattr(x, "filename", "").strip()]
    if not files:
        return redirect(url_for("views.admin_information_bank", tab="event-flow", err="اختر ملفاً واحداً على الأقل (PDF أو Word)."))
    INFO_BANK_DIR.mkdir(parents=True, exist_ok=True)
    sub = INFO_BANK_DIR / "event_flow"
    sub.mkdir(parents=True, exist_ok=True)
    sort_next = _info_bank_next_sort_order(db, InfoBankEventFlowPdf, phase, unit)
    added = 0
    errors: list[str] = []
    for f in files:
        fn = (getattr(f, "filename", "") or "").strip()
        try:
            data = f.read()
        except Exception:
            errors.append(f"{fn or 'ملف'}: تعذّر القراءة.")
            continue
        ext = _info_bank_event_flow_sniff_ext(data)
        if ext is None:
            errors.append(f"{fn}: يُقبل فقط PDF أو Word صالحاً.")
            continue
        if len(data) > _INFO_BANK_MAX_PDF:
            errors.append(f"{fn}: الملف كبير جداً (الحد 50 ميغابايت).")
            continue
        title = (Path(fn).stem or "مجرى أحداث").strip()[:500]
        rel_name = f"event_flow/{uuid.uuid4().hex}{ext}"
        full = (INFO_BANK_DIR / rel_name).resolve()
        full.write_bytes(data)
        row = InfoBankEventFlowPdf(
            training_phase_key=phase,
            unit_level_key=unit,
            title=title,
            file_relpath=rel_name.replace("\\", "/"),
            sort_order=sort_next,
        )
        sort_next += 1
        db.add(row)
        added += 1
    if added:
        db.commit()
    else:
        db.rollback()
    err_q = " ".join(errors)[:2000] if errors else ""
    if not added:
        return redirect(url_for("views.admin_information_bank", tab="event-flow", err=err_q or "لم تُضف أي ملف."))
    ok_msg = f"تمت إضافة {added} ملف(ات) لمجرى الأحداث والمعاضل."
    if err_q:
        return redirect(url_for("views.admin_information_bank", tab="event-flow", ok=ok_msg, err=f"تجاهل أو فشل بعض الملفات: {err_q}"))
    return redirect(url_for("views.admin_information_bank", tab="event-flow", ok=ok_msg))


@bp.route("/admin/information-bank/event-flow/<int:item_id>/delete", methods=["POST"])
def admin_information_bank_event_flow_delete(item_id: int):
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankEventFlowPdf, item_id)
    if not row:
        abort(404)
    if row.file_relpath:
        _unlink_info_bank_file("event_flow", row.file_relpath)
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab="event-flow", ok="تم حذف ملف مجرى الأحداث."))


@bp.route("/admin/information-bank/action-eval/upload", methods=["POST"])
def admin_information_bank_action_eval_upload():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    phase = (request.form.get("training_phase_key") or "").strip()
    unit = (request.form.get("unit_level_key") or "").strip()
    if not _is_valid_information_bank_phase(db, phase) or not _is_valid_information_bank_unit(db, unit):
        return redirect(url_for("views.admin_information_bank", tab="action-eval", err="مرحلة أو مستوى وحدة غير صالح."))
    files = [x for x in request.files.getlist("file") if x and getattr(x, "filename", "").strip()]
    if not files:
        return redirect(url_for("views.admin_information_bank", tab="action-eval", err="اختر ملفاً واحداً على الأقل (.xlsx)."))
    sort_next = _info_bank_next_sort_order(db, InfoBankActionEvalXlsx, phase, unit)
    added = 0
    errors: list[str] = []
    for f in files:
        fn = (getattr(f, "filename", "") or "").strip()
        try:
            data = f.read()
        except Exception:
            errors.append(f"{fn or 'ملف'}: تعذّر القراءة.")
            continue
        if not fn.lower().endswith(".xlsx") or not _is_xlsx_bytes(data):
            errors.append(f"{fn}: يُقبل فقط ملف .xlsx صالحاً.")
            continue
        if len(data) > _INFO_BANK_MAX_XLSX:
            errors.append(f"{fn}: ملف Excel كبير جداً (الحد 30 ميغابايت).")
            continue
        title = (Path(fn).stem or "قائمة تقييم إجراءات").strip()[:500]
        rel_name = f"action_eval/{uuid.uuid4().hex}.xlsx"
        full = (INFO_BANK_DIR / rel_name).resolve()
        full.parent.mkdir(parents=True, exist_ok=True)
        full.write_bytes(data)
        row = InfoBankActionEvalXlsx(
            training_phase_key=phase,
            unit_level_key=unit,
            title=title,
            file_relpath=rel_name.replace("\\", "/"),
            sort_order=sort_next,
        )
        sort_next += 1
        db.add(row)
        added += 1
    if added:
        db.commit()
    else:
        db.rollback()
    err_q = " ".join(errors)[:2000] if errors else ""
    if not added:
        return redirect(url_for("views.admin_information_bank", tab="action-eval", err=err_q or "لم تُضف أي ملف."))
    ok_msg = f"تمت إضافة {added} ملف(ات) لقوائم تقييم الإجراءات."
    if err_q:
        return redirect(url_for("views.admin_information_bank", tab="action-eval", ok=ok_msg, err=f"تجاهل أو فشل بعض الملفات: {err_q}"))
    return redirect(url_for("views.admin_information_bank", tab="action-eval", ok=ok_msg))


@bp.route("/admin/information-bank/action-eval/<int:item_id>/delete", methods=["POST"])
def admin_information_bank_action_eval_delete(item_id: int):
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankActionEvalXlsx, item_id)
    if not row:
        abort(404)
    if row.file_relpath:
        _unlink_info_bank_file("action_eval", row.file_relpath)
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab="action-eval", ok="تم حذف قائمة تقييم الإجراءات."))


@bp.route("/admin/information-bank/dilemma-eval/upload", methods=["POST"])
def admin_information_bank_dilemma_eval_upload():
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    phase = (request.form.get("training_phase_key") or "").strip()
    unit = (request.form.get("unit_level_key") or "").strip()
    if not _is_valid_information_bank_phase(db, phase) or not _is_valid_information_bank_unit(db, unit):
        return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", err="مرحلة أو مستوى وحدة غير صالح."))
    files = [x for x in request.files.getlist("file") if x and getattr(x, "filename", "").strip()]
    if not files:
        return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", err="اختر ملفاً واحداً على الأقل (.xlsx)."))
    sort_next = _info_bank_next_sort_order(db, InfoBankDilemmaEvalXlsx, phase, unit)
    added = 0
    errors: list[str] = []
    for f in files:
        fn = (getattr(f, "filename", "") or "").strip()
        try:
            data = f.read()
        except Exception:
            errors.append(f"{fn or 'ملف'}: تعذّر القراءة.")
            continue
        if not fn.lower().endswith(".xlsx") or not _is_xlsx_bytes(data):
            errors.append(f"{fn}: يُقبل فقط ملف .xlsx صالحاً.")
            continue
        if len(data) > _INFO_BANK_MAX_XLSX:
            errors.append(f"{fn}: ملف Excel كبير جداً (الحد 30 ميغابايت).")
            continue
        title = (Path(fn).stem or "قائمة تقييم معضلة").strip()[:500]
        rel_name = f"dilemma_eval/{uuid.uuid4().hex}.xlsx"
        full = (INFO_BANK_DIR / rel_name).resolve()
        full.parent.mkdir(parents=True, exist_ok=True)
        full.write_bytes(data)
        row = InfoBankDilemmaEvalXlsx(
            training_phase_key=phase,
            unit_level_key=unit,
            title=title,
            file_relpath=rel_name.replace("\\", "/"),
            sort_order=sort_next,
        )
        sort_next += 1
        db.add(row)
        added += 1
    if added:
        db.commit()
    else:
        db.rollback()
    err_q = " ".join(errors)[:2000] if errors else ""
    if not added:
        return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", err=err_q or "لم تُضف أي ملف."))
    ok_msg = f"تمت إضافة {added} ملف(ات) لتقييم المعاضل."
    if err_q:
        return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", ok=ok_msg, err=f"تجاهل أو فشل بعض الملفات: {err_q}"))
    return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", ok=ok_msg))


@bp.route("/admin/information-bank/dilemma-eval/<int:item_id>/delete", methods=["POST"])
def admin_information_bank_dilemma_eval_delete(item_id: int):
    user = get_current_user_optional()
    if not user or not can_manage_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankDilemmaEvalXlsx, item_id)
    if not row:
        abort(404)
    if row.file_relpath:
        _unlink_info_bank_file("dilemma_eval", row.file_relpath)
    db.delete(row)
    db.commit()
    return redirect(url_for("views.admin_information_bank", tab="dilemma-eval", ok="تم حذف قائمة تقييم المعاضل."))


@bp.route("/admin/information-bank/file/event-flow/<int:item_id>", methods=["GET"])
def admin_information_bank_event_flow_file(item_id: int):
    user = get_current_user_optional()
    if not user or not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankEventFlowPdf, item_id)
    if not row or not (row.file_relpath or "").strip():
        abort(404)
    path = _info_bank_file_abspath("event_flow", row.file_relpath)
    if path is None:
        abort(404)
    mt = _mimetype_info_bank_event_flow(path)
    return send_file(path, mimetype=mt, as_attachment=False)


@bp.route("/admin/information-bank/file/action-eval/<int:item_id>", methods=["GET"])
def admin_information_bank_action_eval_file(item_id: int):
    user = get_current_user_optional()
    if not user or not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankActionEvalXlsx, item_id)
    if not row or not (row.file_relpath or "").strip():
        abort(404)
    path = _info_bank_file_abspath("action_eval", row.file_relpath)
    if path is None:
        abort(404)
    return send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=False,
    )


@bp.route("/admin/information-bank/file/dilemma-eval/<int:item_id>", methods=["GET"])
def admin_information_bank_dilemma_eval_file(item_id: int):
    user = get_current_user_optional()
    if not user or not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    row = db.get(InfoBankDilemmaEvalXlsx, item_id)
    if not row or not (row.file_relpath or "").strip():
        abort(404)
    path = _info_bank_file_abspath("dilemma_eval", row.file_relpath)
    if path is None:
        abort(404)
    return send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=False,
    )


@bp.route("/admin/information-bank/manifest.json", methods=["GET"])
def admin_information_bank_manifest_json():
    """قائمة موحّدة للمرفقات (لصفحات أخرى تستهلك السحب والإفلات لاحقاً)."""
    user = get_current_user_optional()
    if not user or not can_view_information_bank(user):
        abort(403)
    from flask import g

    db = g.db
    items: list[dict] = []
    for row in db.query(InfoBankEventFlowPdf).order_by(InfoBankEventFlowPdf.id).all():
        if not (row.file_relpath or "").strip():
            continue
        p = _info_bank_file_abspath("event_flow", row.file_relpath)
        if p is None:
            continue
        items.append(
            {
                "kind": "event_flow_document",
                "id": row.id,
                "title": row.title or "",
                "file_ext": p.suffix.lower().lstrip("."),
                "training_phase_key": row.training_phase_key,
                "training_phase_label": training_phase_label(row.training_phase_key),
                "unit_level_key": row.unit_level_key,
                "unit_label": info_bank_unit_label(row.unit_level_key),
                "url": url_for("views.admin_information_bank_event_flow_file", item_id=row.id),
            }
        )
    for row in db.query(InfoBankActionEvalXlsx).order_by(InfoBankActionEvalXlsx.id).all():
        if not (row.file_relpath or "").strip():
            continue
        if _info_bank_file_abspath("action_eval", row.file_relpath) is None:
            continue
        items.append(
            {
                "kind": "action_eval_xlsx",
                "id": row.id,
                "title": row.title or "",
                "training_phase_key": row.training_phase_key,
                "training_phase_label": training_phase_label(row.training_phase_key),
                "unit_level_key": row.unit_level_key,
                "unit_label": info_bank_unit_label(row.unit_level_key),
                "url": url_for("views.admin_information_bank_action_eval_file", item_id=row.id),
            }
        )
    for row in db.query(InfoBankDilemmaEvalXlsx).order_by(InfoBankDilemmaEvalXlsx.id).all():
        if not (row.file_relpath or "").strip():
            continue
        if _info_bank_file_abspath("dilemma_eval", row.file_relpath) is None:
            continue
        items.append(
            {
                "kind": "dilemma_eval_xlsx",
                "id": row.id,
                "title": row.title or "",
                "training_phase_key": row.training_phase_key,
                "training_phase_label": training_phase_label(row.training_phase_key),
                "unit_level_key": row.unit_level_key,
                "unit_label": info_bank_unit_label(row.unit_level_key),
                "url": url_for("views.admin_information_bank_dilemma_eval_file", item_id=row.id),
            }
        )
    return jsonify({"version": 1, "items": items})


@bp.route("/exercises/<int:eid>")
def exercise_detail(eid):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/exercises/{eid}")
    from flask import g
    db = g.db
    ex = (
        db.query(Exercise)
        .options(
            joinedload(Exercise.objectives),
        )
        .filter(Exercise.id == eid)
        .first()
    )
    if not ex:
        abort(404)
    return render_template(
        "exercise_detail.html",
        **_ctx(
            user,
            ex=ex,
            can_plan=can_plan_exercises(user),
            can_ref=can_edit_references(user),
            **_subpage_close_ctx("/dashboard"),
        ),
    )


def _library_redirect(*, tab: str, ok: str = "", err: str = ""):
    kw: dict = {"tab": tab}
    if ok:
        kw["ok"] = ok
    if err:
        kw["err"] = err
    return redirect(url_for("views.library", **kw))


@bp.route("/library", methods=["GET"])
def library():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/library")
    from flask import g

    from app.library_tree import (
        LIBRARY_TAB_SPECS,
        LIBRARY_TREE_KINDS,
        build_tree_payload,
        library_active_tab_from_request,
        library_kind_title,
    )

    db = g.db
    active_tab = library_active_tab_from_request(request.args.get("tab"))
    library_trees = {
        kind: build_tree_payload(db, kind) for kind in LIBRARY_TREE_KINDS
    }
    ok_raw = (request.args.get("ok") or "").strip()
    err_raw = (request.args.get("err") or "").strip()
    return render_template(
        "library.html",
        **_ctx(
            user,
            library_tabs=LIBRARY_TAB_SPECS,
            library_trees=library_trees,
            active_tab=active_tab,
            library_can_manage=can_edit_references(user),
            ok_msg=ok_raw,
            error=err_raw,
            library_kind_title=library_kind_title,
            **_subpage_close_ctx("/dashboard"),
        ),
    )


@bp.route("/library/tree/folder", methods=["POST"])
def library_tree_folder_add():
    user = get_current_user_optional()
    if not user or not can_edit_references(user):
        abort(403)
    from flask import g

    from app.library_tree import (
        LIBRARY_TREE_KINDS,
        add_custom_folder,
        ensure_library_tree,
        library_kind_tab,
    )

    db = g.db
    kind = (request.form.get("kind") or "").strip()
    if kind not in LIBRARY_TREE_KINDS:
        abort(400)
    tab = library_kind_tab(kind)
    parent_raw = (request.form.get("parent_id") or "").strip()
    parent_id = int(parent_raw) if parent_raw.isdigit() else None
    name = (request.form.get("folder_name") or "").strip()
    if not name:
        return _library_redirect(tab=tab, err="أدخل اسم المجلد.")
    ensure_library_tree(db, kind)
    try:
        add_custom_folder(db, kind=kind, parent_id=parent_id, name=name)
        db.commit()
    except ValueError as exc:
        db.rollback()
        return _library_redirect(tab=tab, err=str(exc) or "تعذر إنشاء المجلد.")
    return _library_redirect(tab=tab, ok="تم إنشاء المجلد.")


@bp.route("/library/tree/upload", methods=["POST"])
def library_tree_upload():
    user = get_current_user_optional()
    if not user or not can_edit_references(user):
        abort(403)
    from flask import g

    from app.library_tree import (
        LIBRARY_TREE_KINDS,
        ensure_library_tree,
        get_node,
        library_kind_tab,
        upload_files_to_tree,
    )

    db = g.db
    kind = (request.form.get("kind") or "").strip()
    tab = library_kind_tab(kind)
    if kind not in LIBRARY_TREE_KINDS:
        abort(400)
    from app.config import LIBRARY_DIR

    LIBRARY_DIR.mkdir(parents=True, exist_ok=True)
    parent_raw = (request.form.get("parent_id") or "").strip()
    parent_id = int(parent_raw) if parent_raw.isdigit() else None
    ensure_library_tree(db, kind)
    if parent_id is not None:
        parent = get_node(db, parent_id, kind)
        if parent is None or not parent.is_folder:
            return _library_redirect(tab=tab, err="المجلد المستهدف غير صالح.")
    files = [x for x in request.files.getlist("files") if x and getattr(x, "filename", "").strip()]
    if not files:
        return _library_redirect(tab=tab, err="اختر ملفاً أو مجلداً للإدراج.")
    added, errors = upload_files_to_tree(
        db, kind=kind, parent_id=parent_id, file_storages=files
    )
    if added:
        db.commit()
    else:
        db.rollback()
    err_q = " ".join(errors)[:2000] if errors else ""
    if not added:
        return _library_redirect(tab=tab, err=err_q or "لم تُضف أي ملف.")
    ok_msg = f"تم إدراج {added} ملف(ات) مع الحفاظ على أسماء المسارات."
    if err_q:
        return _library_redirect(tab=tab, ok=ok_msg, err=f"تجاهل بعض الملفات: {err_q}")
    return _library_redirect(tab=tab, ok=ok_msg)


@bp.route("/library/tree/<int:node_id>/delete", methods=["POST"])
def library_tree_delete(node_id: int):
    user = get_current_user_optional()
    if not user or not can_edit_references(user):
        abort(403)
    from flask import g

    from app.library_tree import (
        delete_library_node,
        is_library_tree_kind,
        library_kind_tab,
    )

    db = g.db
    row = db.get(InformationBankTreeNode, node_id)
    if row is None or not is_library_tree_kind(row.kind):
        abort(404)
    tab = library_kind_tab(row.kind)
    delete_library_node(db, row)
    db.commit()
    if (
        request.accept_mimetypes.best_match(["application/json", "text/html"])
        == "application/json"
        or (request.headers.get("X-Requested-With") or "").strip() == "XMLHttpRequest"
    ):
        return jsonify(ok=True, node_id=int(node_id), tab=tab)
    return _library_redirect(tab=tab, ok="تم الحذف.")


@bp.route("/library/tree/move", methods=["POST"])
def library_tree_move():
    user = get_current_user_optional()
    if not user or not can_edit_references(user):
        return jsonify(ok=False, error="غير مسموح."), 403
    from flask import g

    from app.library_tree import LIBRARY_TREE_KINDS, is_library_tree_kind, move_tree_node

    data = request.get_json(force=True, silent=True) or {}
    kind = (data.get("kind") or "").strip()
    if kind not in LIBRARY_TREE_KINDS:
        return jsonify(ok=False, error="نوع غير صالح."), 400
    try:
        nid = int(data.get("node_id"))
    except (TypeError, ValueError):
        return jsonify(ok=False, error="بيانات غير صالحة."), 400
    parent_raw = data.get("parent_id")
    try:
        pid = int(parent_raw) if parent_raw is not None and parent_raw != "" else 0
    except (TypeError, ValueError):
        return jsonify(ok=False, error="بيانات غير صالحة."), 400
    parent_id = None if pid < 1 else pid
    db = g.db
    row = db.get(InformationBankTreeNode, nid)
    if row is None or not is_library_tree_kind(row.kind):
        return jsonify(ok=False, error="العنصر غير موجود."), 404
    try:
        move_tree_node(db, kind=kind, node_id=nid, parent_id=parent_id)
        db.commit()
    except ValueError as exc:
        db.rollback()
        return jsonify(ok=False, error=str(exc) or "تعذّر النقل."), 400
    return jsonify(ok=True)


@bp.route("/library/tree/<int:node_id>/file", methods=["GET"])
def library_tree_file(node_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    from flask import g

    from app.library_tree import is_library_tree_kind, node_file_abspath

    row = g.db.get(InformationBankTreeNode, node_id)
    if row is None or row.is_folder or not is_library_tree_kind(row.kind):
        abort(404)
    if not (row.file_relpath or "").strip():
        abort(404)
    path = node_file_abspath(row.kind, row.file_relpath)
    if path is None:
        abort(404)
    low = path.name.lower()
    if low.endswith(".xlsx"):
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif low.endswith(".docx"):
        mt = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif low.endswith(".doc"):
        mt = "application/msword"
    else:
        mt = _mimetype_info_bank_event_flow(path)
    return send_file(path, mimetype=mt, as_attachment=False, download_name=path.name)


@bp.route("/admin/users", methods=["GET", "POST"])
def admin_users():
    user = get_current_user_optional()
    if not user or not can_manage_users(user):
        abort(403)
    from flask import g
    db = g.db
    if request.method == "POST":
        rk = (request.form.get("role_key") or "judge").strip()
        if not any(rk == m.value for m in RoleKey):
            rk = RoleKey.JUDGE.value
        if rk == RoleKey.STANDARDS_LIBRARY.value:
            rk = RoleKey.JUDGE.value
        username = (request.form.get("username") or "").strip() or f"u{user.id}-new"
        # الافتراضي: demo123، أما المحكم فمطلوب (username=password=الرقم العسكري)
        temp_pwd = (request.form.get("temp_password") or "").strip()
        if not temp_pwd:
            temp_pwd = username if rk == RoleKey.JUDGE.value else "demo123"
        u = User(
            username=username,
            full_name=(request.form.get("full_name") or "").strip(),
            role_key=rk,
            password_hash=hash_password(temp_pwd),
        )
        db.add(u)
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            return redirect("/admin/users?err=duplicate_username")
        # ربط المحكم بمتدرب (اختياري) ضمن التمرين الحالي
        if rk == RoleKey.JUDGE.value:
            ex = _admin_current_workspace_exercise(db, user)
            trainee_row_id_raw = (request.form.get("trainee_row_id") or "").strip()
            if ex is not None and trainee_row_id_raw.isdigit():
                tr = db.get(ExerciseRosterRow, int(trainee_row_id_raw))
                if (
                    tr is not None
                    and tr.exercise_id == ex.id
                    and (tr.roster_kind or "") == ExerciseRosterKind.TRAINEE.value
                    and (tr.unit_level_key or "").strip()
                ):
                    db.add(
                        JudgeTraineeAssignment(
                            exercise_id=ex.id,
                            judge_user_id=u.id,
                            unit_level_key=(tr.unit_level_key or "").strip(),
                            trainee_name=(tr.full_name or "").strip(),
                            trainee_military_number=(tr.military_number or "").strip(),
                        )
                    )
                    db.commit()
        return redirect("/admin/users")
    us = db.query(User).order_by(User.id).all()
    rdefs = {r.role_key: r for r in db.query(RoleDef).all()}
    role_choices = [
        {
            "value": m.value,
            "label": rdefs[m.value].title_ar if m.value in rdefs else m.value,
        }
        for m in RoleKey
        if m != RoleKey.STANDARDS_LIBRARY
    ]
    ex = _admin_current_workspace_exercise(db, user)
    trainee_choices: list[dict] = []
    if ex is not None:
        trs = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .all()
        )
        for tr in trs:
            uk = (tr.unit_level_key or "").strip()
            trainee_choices.append(
                {
                    "id": int(tr.id),
                    "label": f"{(tr.full_name or '').strip() or 'متدرب'} — {(tr.military_number or '').strip() or '—'} — {label_for_unit_level_key(uk) or uk or '—'}",
                }
            )
    err_key = (request.args.get("err") or "").strip()
    user_errors = {
        "duplicate_username": "اسم المستخدم مستخدم مسبقاً. اختر اسماً آخر.",
    }
    return render_template(
        "admin_users.html",
        **_ctx(
            user,
            users=us,
            rdefs=rdefs,
            role_choices=role_choices,
            has_exercise=ex is not None,
            trainee_choices=trainee_choices,
            user_error=user_errors.get(err_key, ""),
        ),
    )


@bp.route("/admin/users/<int:uid>/edit", methods=["GET", "POST"])
def admin_user_edit(uid: int):
    user = get_current_user_optional()
    if not user or not can_manage_users(user):
        abort(403)
    from flask import g

    db = g.db
    target = db.query(User).filter(User.id == uid).first()
    if not target:
        abort(404)

    rdefs = {r.role_key: r for r in db.query(RoleDef).all()}
    role_choices = [
        {
            "value": m.value,
            "label": rdefs[m.value].title_ar if m.value in rdefs else m.value,
        }
        for m in RoleKey
        if m != RoleKey.STANDARDS_LIBRARY
    ]

    ex = _admin_current_workspace_exercise(db, user)
    existing_assignment = None
    trainee_choices: list[dict] = []
    if ex is not None:
        existing_assignment = (
            db.query(JudgeTraineeAssignment)
            .filter(
                JudgeTraineeAssignment.exercise_id == ex.id,
                JudgeTraineeAssignment.judge_user_id == target.id,
            )
            .first()
        )
        trs = (
            db.query(ExerciseRosterRow)
            .filter(
                ExerciseRosterRow.exercise_id == ex.id,
                ExerciseRosterRow.roster_kind == ExerciseRosterKind.TRAINEE.value,
            )
            .order_by(ExerciseRosterRow.sort_order, ExerciseRosterRow.id)
            .all()
        )
        for tr in trs:
            uk = (tr.unit_level_key or "").strip()
            trainee_choices.append(
                {
                    "id": int(tr.id),
                    "label": f"{(tr.full_name or '').strip() or 'متدرب'} — {(tr.military_number or '').strip() or '—'} — {label_for_unit_level_key(uk) or uk or '—'}",
                    "unit_key": uk,
                }
            )

    if request.method == "GET":
        return render_template(
            "admin_user_edit.html",
            **_ctx(
                user,
                target=target,
                role_choices=role_choices,
                error="",
                has_exercise=ex is not None,
                trainee_choices=trainee_choices,
                assignment=existing_assignment,
            ),
        )

    # POST: update
    full_name = (request.form.get("full_name") or "").strip()
    rk = (request.form.get("role_key") or target.role_key or RoleKey.JUDGE.value).strip()
    if not any(rk == m.value for m in RoleKey):
        rk = RoleKey.JUDGE.value
    if rk == RoleKey.STANDARDS_LIBRARY.value:
        rk = RoleKey.JUDGE.value

    is_active = (request.form.get("is_active") or "").strip() in ("1", "true", "on", "yes")
    temp_pwd = (request.form.get("temp_password") or "").strip()

    target.full_name = full_name
    target.role_key = rk
    target.is_active = is_active
    if temp_pwd:
        target.password_hash = hash_password(temp_pwd)
    db.add(target)
    db.commit()
    # تحديث تخصيص المحكم (ضمن التمرين الحالي) إن كانت الصلاحية محكم
    if ex is not None and rk == RoleKey.JUDGE.value:
        trainee_row_id_raw = (request.form.get("trainee_row_id") or "").strip()
        # حذف أي تخصيص سابق
        db.execute(
            delete(JudgeTraineeAssignment).where(
                JudgeTraineeAssignment.exercise_id == ex.id,
                JudgeTraineeAssignment.judge_user_id == target.id,
            )
        )
        db.commit()
        if trainee_row_id_raw.isdigit():
            tr = db.get(ExerciseRosterRow, int(trainee_row_id_raw))
            if (
                tr is not None
                and tr.exercise_id == ex.id
                and (tr.roster_kind or "") == ExerciseRosterKind.TRAINEE.value
                and (tr.unit_level_key or "").strip()
            ):
                db.add(
                    JudgeTraineeAssignment(
                        exercise_id=ex.id,
                        judge_user_id=target.id,
                        unit_level_key=(tr.unit_level_key or "").strip(),
                        trainee_name=(tr.full_name or "").strip(),
                        trainee_military_number=(tr.military_number or "").strip(),
                    )
                )
                db.commit()
    return redirect("/admin/users")


@bp.route("/admin/users/<int:uid>/delete", methods=["POST"])
def admin_user_delete(uid: int):
    user = get_current_user_optional()
    if not user or not can_manage_users(user):
        abort(403)
    if getattr(user, "id", None) == uid:
        abort(400)
    from flask import g

    db = g.db
    target = db.query(User).filter(User.id == uid).first()
    if not target:
        abort(404)
    # "حذف" آمن: تعطيل الحساب
    target.is_active = False
    db.add(target)
    db.commit()
    return redirect("/admin/users")


@bp.route("/api/admin/generate-username", methods=["POST"])
def api_admin_generate_username():
    user = get_current_user_optional()
    if not user or not can_manage_users(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    rk = (request.form.get("role_key") or "").strip()
    if not any(rk == m.value for m in RoleKey):
        rk = ""

    prefix_map = {
        RoleKey.SYSTEM_ADMIN.value: "admin",
        RoleKey.ANALYST.value: "analyst",
        RoleKey.PLANNER.value: "planner",
        RoleKey.JUDGE.value: "judge",
        RoleKey.CONTROL.value: "control",
    }
    prefix = prefix_map.get(rk, "user")

    # توليد اسم فريد: prefix + رقم متسلسل
    for i in range(1, 5000):
        cand = f"{prefix}{i:03d}"
        if not db.query(User).filter(User.username == cand).first():
            return jsonify({"ok": True, "username": cand})

    # احتياط: في حال امتلاء النطاق
    cand = f"{prefix}-{uuid.uuid4().hex[:6]}"
    return jsonify({"ok": True, "username": cand})


@bp.route("/api/admin/generate-usernames-batch", methods=["POST"])
def api_admin_generate_usernames_batch():
    user = get_current_user_optional()
    if not user or not can_manage_users(user):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    from flask import g

    db = g.db
    prefix_map = {
        RoleKey.SYSTEM_ADMIN.value: "admin",
        RoleKey.ANALYST.value: "analyst",
        RoleKey.PLANNER.value: "planner",
        RoleKey.JUDGE.value: "judge",
        RoleKey.CONTROL.value: "control",
    }

    f = request.files.get("names_file")
    if not f:
        return jsonify({"ok": False, "error": "missing_file"}), 400
    try:
        data = f.read()
    except Exception:
        data = b""
    if not data:
        return jsonify({"ok": False, "error": "empty_file"}), 400

    def _norm(s: str) -> str:
        return (
            (s or "")
            .replace("\u200f", "")
            .replace("\u200e", "")
            .replace("ـ", "")
            .strip()
            .casefold()
        )

    # خريطة أدوار عربية/مفاتيح إلى RoleKey.value
    rdefs = {r.role_key: r for r in db.query(RoleDef).all()}
    title_to_role = {_norm(r.title_ar): r.role_key for r in rdefs.values() if r.title_ar}
    # مرادفات شائعة
    title_to_role.update(
        {
            _norm("إدارة النظام"): RoleKey.SYSTEM_ADMIN.value,
            _norm("المحللين"): RoleKey.ANALYST.value,
            _norm("المحلّلون"): RoleKey.ANALYST.value,
            _norm("المحللون"): RoleKey.ANALYST.value,
            _norm("التخطيط"): RoleKey.PLANNER.value,
            _norm("المحكمين"): RoleKey.JUDGE.value,
            _norm("المحكّمون"): RoleKey.JUDGE.value,
            _norm("المحكمون"): RoleKey.JUDGE.value,
            _norm("السيطرة"): RoleKey.CONTROL.value,
        }
    )

    def _resolve_role(raw: str) -> str:
        r = (raw or "").strip()
        if not r:
            return ""
        if any(r == m.value for m in RoleKey):
            return r
        k = _norm(r)
        return title_to_role.get(k, "")

    def _parse_records_from_bytes(filename_lower: str, blob: bytes) -> list[tuple[str, str]]:
        # returns (name, role_raw)
        if filename_lower.endswith(".xlsx"):
            try:
                from io import BytesIO
                from openpyxl import load_workbook

                wb = load_workbook(BytesIO(blob), data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return []
                # محاولة إيجاد رأس أعمدة: الاسم + الدور/الوظيفة
                name_idx = role_idx = None
                for ridx, row in enumerate(rows[:10]):
                    vals = [str(c).strip() if c is not None else "" for c in row]
                    for i, v in enumerate(vals):
                        nv = _norm(v)
                        if name_idx is None and ("الاسم" in nv or "name" in nv):
                            name_idx = i
                        if role_idx is None and ("الدور" in nv or "الوظيفة" in nv or "role" in nv):
                            role_idx = i
                    if name_idx is not None:
                        start = ridx + 1
                        break
                else:
                    start = 0
                out: list[tuple[str, str]] = []
                for row in rows[start:]:
                    vals = [str(c).strip() if c is not None else "" for c in row]
                    name = vals[name_idx] if name_idx is not None and name_idx < len(vals) else ""
                    role = vals[role_idx] if role_idx is not None and role_idx < len(vals) else ""
                    name = name.strip()
                    role = role.strip()
                    if not name:
                        continue
                    out.append((name, role))
                return out
            except Exception:
                return []

        text = blob.decode("utf-8", errors="ignore")
        out: list[tuple[str, str]] = []
        for line in text.splitlines():
            s = line.strip()
            if not s:
                continue
            # CSV/TSV/pipe: الاسم,الدور
            for sep in ("\t", "|", ";", ","):
                if sep in s:
                    parts = [p.strip() for p in s.split(sep) if p.strip()]
                    if len(parts) >= 2:
                        out.append((parts[0], parts[1]))
                    else:
                        out.append((parts[0], ""))
                    break
            else:
                # دعم "الاسم - الدور" أو "الاسم — الدور"
                if "—" in s:
                    parts = [p.strip() for p in s.split("—") if p.strip()]
                    out.append((parts[0], parts[1] if len(parts) > 1 else ""))
                elif "-" in s:
                    parts = [p.strip() for p in s.split("-") if p.strip()]
                    out.append((parts[0], parts[1] if len(parts) > 1 else ""))
                else:
                    out.append((s, ""))
        return out

    filename_lower = (getattr(f, "filename", "") or "").lower()
    records = _parse_records_from_bytes(filename_lower, data)

    # إزالة التكرار مع الحفاظ على الترتيب (حسب الاسم+الدور)
    seen: set[str] = set()
    items_in: list[tuple[str, str, str]] = []  # name, role_raw, role_key
    for name, role_raw in records:
        key = f"{_norm(name)}|{_norm(role_raw)}"
        if key in seen:
            continue
        seen.add(key)
        role_key = _resolve_role(role_raw)
        items_in.append((name, role_raw, role_key))
    items_in = items_in[:300]

    taken = {row[0] for row in db.query(User.username).all()}
    counters: dict[str, int] = {}

    def _next_for_prefix(prefix: str) -> str:
        start = counters.get(prefix, 1)
        i = start
        while i < 5000:
            cand = f"{prefix}{i:03d}"
            if cand not in taken:
                taken.add(cand)
                counters[prefix] = i + 1
                return cand
            i += 1
        cand = f"{prefix}-{uuid.uuid4().hex[:6]}"
        while cand in taken:
            cand = f"{prefix}-{uuid.uuid4().hex[:6]}"
        taken.add(cand)
        return cand

    def _gen_password() -> str:
        import secrets
        import string

        alphabet = string.ascii_letters + string.digits
        # استبعاد أحرف قد تسبب لبساً بصرياً
        for ch in "O0Il1":
            alphabet = alphabet.replace(ch, "")
        return "".join(secrets.choice(alphabet) for _ in range(10))

    out = []
    for name, role_raw, role_key in items_in:
        prefix = prefix_map.get(role_key, "user")
        uname = _next_for_prefix(prefix)
        role_label = rdefs[role_key].title_ar if role_key in rdefs else (role_raw or role_key or "غير محدد")
        out.append(
            {
                "name": name,
                "role": role_label,
                "username": uname,
                "temp_password": _gen_password(),
            }
        )
    return jsonify({"ok": True, "items": out})


@bp.route("/api/ai/suggest", methods=["POST"])
def api_ai_suggest():
    user = get_current_user_optional()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not (
        can_judge_exercise(user)
        or can_plan_exercises(user)
        or is_analyst(user)
        or is_control(user)
        or is_system_admin(user)
    ):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    purpose = (request.form.get("purpose") or "").strip() or "تعليمات/ملاحظات تقييم"
    context = (request.form.get("context") or "").strip()
    out = suggest_instructions_or_notes(purpose=purpose, context=context)
    return jsonify({"ok": True, "text": out})


# ——————————————————————————————————————————————————————————————
# غرف المحادثة (مرتبطة بالتمرين الحالي — إدارة النظام + الأعضاء)
# ——————————————————————————————————————————————————————————————

_CHAT_ALLOWED_SUFFIX = frozenset(
    {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".xlsx", ".xls", ".txt", ".mp4"}
)
_CHAT_MAX_UPLOAD_BYTES = 15 * 1024 * 1024

_CHAT_KIND_LABELS_AR: dict[str, str] = {
    ChatRoomKind.JUDGE_BRIGADE: "محكمي اللواء",
    ChatRoomKind.JUDGE_BN: "محكمي الكتيبة",
    ChatRoomKind.CONTROL: "هيئة السيطرة",
    ChatRoomKind.ADMIN_SUPPORT: "إدارة النظام والدعم الفني",
    ChatRoomKind.CUSTOM: "مخصصة",
}


def _chat_workspace_exercise(db, user: User) -> Exercise | None:
    if is_system_admin(user):
        return _admin_current_workspace_exercise(db, user)
    return _current_workspace_exercise(db, user)


def _chat_member_ids(db, room_id: int) -> set[int]:
    rows = db.query(ChatRoomMember.user_id).filter(ChatRoomMember.room_id == room_id).all()
    return {int(r[0]) for r in rows}


def _chat_user_can_access_room(db, user: User, room: ChatRoom) -> bool:
    if not user or room is None:
        return False
    ex = _chat_workspace_exercise(db, user)
    if ex is None or int(room.exercise_id) != int(ex.id):
        return False
    if is_system_admin(user):
        return True
    return int(user.id) in _chat_member_ids(db, int(room.id))


def _chat_file_disk_path(exercise_id: int, room_id: int, relpath: str) -> Path | None:
    if not relpath or ".." in relpath.replace("\\", "/"):
        return None
    root = CHAT_UPLOAD_DIR.resolve()
    expected_prefix = f"{int(exercise_id)}/{int(room_id)}/"
    norm = relpath.replace("\\", "/").strip().lstrip("/")
    if not norm.startswith(expected_prefix):
        return None
    out = (root / norm).resolve()
    try:
        out.relative_to(root)
    except ValueError:
        return None
    return out if out.is_file() else None


def _chat_touch_room_activity(db, room: ChatRoom) -> None:
    room.last_activity_at = datetime.utcnow()


def _chat_mark_messages_read(db, user: User, room_id: int) -> None:
    mids = [
        int(m[0])
        for m in db.query(ChatMessage.id)
        .filter(ChatMessage.room_id == room_id, ChatMessage.sender_id != int(user.id))
        .all()
    ]
    for mid in mids:
        exists = (
            db.query(ChatMessageRead)
            .filter(ChatMessageRead.message_id == mid, ChatMessageRead.user_id == int(user.id))
            .first()
        )
        if exists is None:
            db.add(ChatMessageRead(message_id=mid, user_id=int(user.id)))
    db.commit()


@bp.route("/chat-rooms", methods=["GET"])
def chat_rooms_list():
    user = get_current_user_optional()
    if not user:
        return redirect("/login?next=/chat-rooms")
    if not can_use_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _chat_workspace_exercise(db, user)
    rooms: list[ChatRoom] = []
    if ex is not None:
        if is_system_admin(user):
            rooms = (
                db.query(ChatRoom)
                .filter(ChatRoom.exercise_id == ex.id, ChatRoom.is_archived == False)
                .order_by(desc(ChatRoom.last_activity_at), desc(ChatRoom.id))
                .all()
            )
        else:
            rooms = (
                db.query(ChatRoom)
                .join(ChatRoomMember, ChatRoomMember.room_id == ChatRoom.id)
                .filter(
                    ChatRoom.exercise_id == ex.id,
                    ChatRoom.is_archived == False,
                    ChatRoomMember.user_id == int(user.id),
                )
                .order_by(desc(ChatRoom.last_activity_at), desc(ChatRoom.id))
                .all()
            )
    room_cards = []
    for r in rooms:
        uk = (r.unit_level_key or "").strip()
        room_cards.append(
            {
                "room": r,
                "unit_label": label_for_unit_level_key(uk) if uk else "",
                "kind_label": _CHAT_KIND_LABELS_AR.get((r.room_kind or "").strip(), (r.room_kind or "مخصصة")),
            }
        )
    return render_template(
        "chat_rooms_list.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            room_cards=room_cards,
            can_manage=can_manage_chat_rooms(user),
            **_hub_back_ctx_for_request_path(),
            control_link_kwargs=_role_hub_preserve_link_kwargs(),
            hub_from_form_param=_role_hub_from_form_param(),
        ),
    )


@bp.route("/chat-rooms/<int:room_id>", methods=["GET"])
def chat_room_detail(room_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/chat-rooms/{room_id}")
    if not can_use_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    room = db.get(ChatRoom, room_id)
    if room is None or room.is_archived:
        abort(404)
    if not _chat_user_can_access_room(db, user, room):
        abort(403)
    _chat_mark_messages_read(db, user, room_id)

    msgs = (
        db.query(ChatMessage)
        .filter(ChatMessage.room_id == room_id)
        .order_by(ChatMessage.created_at.asc(), ChatMessage.id.asc())
        .limit(500)
        .all()
    )
    member_ids = _chat_member_ids(db, room_id)
    uids = member_ids | {int(m.sender_id) for m in msgs}
    users_by_id: dict[int, User] = {}
    if uids:
        for u in db.query(User).filter(User.id.in_(uids)).all():
            users_by_id[int(u.id)] = u

    message_rows: list[dict] = []
    for m in msgs:
        reads = (
            db.query(ChatMessageRead)
            .filter(ChatMessageRead.message_id == m.id)
            .order_by(ChatMessageRead.read_at.asc())
            .all()
        )
        reader_ids = {int(r.user_id) for r in reads}
        others = {uid for uid in member_ids if uid != int(m.sender_id)}
        n_others = len(others)
        read_by_others = len(reader_ids & others)
        tick_level = "read" if n_others > 0 and read_by_others >= n_others else ("delivered" if read_by_others > 0 else "sent")
        sender = users_by_id.get(int(m.sender_id))
        sender_label = (
            (getattr(sender, "full_name", "") or "").strip()
            or (getattr(sender, "username", "") or "").strip()
            or f"مستخدم #{m.sender_id}"
        )
        reader_names = []
        for r in reads:
            if int(r.user_id) == int(m.sender_id):
                continue
            ru = users_by_id.get(int(r.user_id))
            reader_names.append(
                (getattr(ru, "full_name", "") or "").strip()
                or (getattr(ru, "username", "") or "").strip()
                or f"#{r.user_id}"
            )
        message_rows.append(
            {
                "msg": m,
                "sender_label": sender_label,
                "is_mine": int(m.sender_id) == int(user.id),
                "tick_level": tick_level,
                "reader_names": reader_names,
            }
        )

    ex = _chat_workspace_exercise(db, user)
    return render_template(
        "chat_room_detail.html",
        **_ctx(
            user,
            room=room,
            exercise=ex,
            message_rows=message_rows,
            can_manage=can_manage_chat_rooms(user),
            **_hub_back_ctx_for_request_path(),
            control_link_kwargs=_role_hub_preserve_link_kwargs(),
            hub_from_form_param=_role_hub_from_form_param(),
        ),
    )


@bp.route("/chat-rooms/<int:room_id>/message", methods=["POST"])
def chat_room_post_message(room_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_use_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    room = db.get(ChatRoom, room_id)
    if room is None or room.is_archived or not _chat_user_can_access_room(db, user, room):
        abort(404)
    body = (request.form.get("body") or "").strip()
    if not body or len(body) > 8000:
        return redirect(url_for("views.chat_room_detail", room_id=room_id, **_role_hub_preserve_link_kwargs()))
    m = ChatMessage(
        room_id=room.id,
        sender_id=int(user.id),
        message_type="text",
        body_text=body,
    )
    db.add(m)
    _chat_touch_room_activity(db, room)
    from app.notifications_service import notify_chat_new_message

    notify_chat_new_message(db, room=room, message=m, sender_id=int(user.id))
    db.commit()
    return redirect(url_for("views.chat_room_detail", room_id=room_id, **_control_hub_link_kwargs()))


@bp.route("/chat-rooms/<int:room_id>/upload", methods=["POST"])
def chat_room_upload(room_id: int):
    user = get_current_user_optional()
    if not user:
        abort(403)
    if not can_use_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    room = db.get(ChatRoom, room_id)
    if room is None or room.is_archived or not _chat_user_can_access_room(db, user, room):
        abort(404)
    f = request.files.get("file")
    if not f or not (f.filename or "").strip():
        return redirect(url_for("views.chat_room_detail", room_id=room_id, **_role_hub_preserve_link_kwargs()))
    raw_name = secure_filename(f.filename)
    suf = Path(raw_name).suffix.lower()
    if suf not in _CHAT_ALLOWED_SUFFIX:
        return redirect(url_for("views.chat_room_detail", room_id=room_id, **_role_hub_preserve_link_kwargs()))
    data = f.read()
    if len(data) > _CHAT_MAX_UPLOAD_BYTES:
        return redirect(url_for("views.chat_room_detail", room_id=room_id, **_role_hub_preserve_link_kwargs()))
    CHAT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    sub = CHAT_UPLOAD_DIR / str(int(room.exercise_id)) / str(int(room.id))
    sub.mkdir(parents=True, exist_ok=True)
    store_name = f"{uuid.uuid4().hex}{suf}"
    disk_path = sub / store_name
    disk_path.write_bytes(data)
    rel = f"{int(room.exercise_id)}/{int(room.id)}/{store_name}"
    mime = mimetypes.guess_type(raw_name)[0] or "application/octet-stream"
    m = ChatMessage(
        room_id=room.id,
        sender_id=int(user.id),
        message_type="file",
        body_text="",
        file_relpath=rel,
        original_filename=(f.filename or store_name)[:500],
        mime_type=(mime or "")[:200],
        file_size=len(data),
    )
    db.add(m)
    _chat_touch_room_activity(db, room)
    from app.notifications_service import notify_chat_new_message

    notify_chat_new_message(db, room=room, message=m, sender_id=int(user.id))
    db.commit()
    return redirect(url_for("views.chat_room_detail", room_id=room_id, **_control_hub_link_kwargs()))


@bp.route("/chat-rooms/<int:room_id>/files/<int:message_id>", methods=["GET"])
def chat_room_download(room_id: int, message_id: int):
    user = get_current_user_optional()
    if not user:
        return redirect(f"/login?next=/chat-rooms/{room_id}/files/{message_id}")
    if not can_use_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    room = db.get(ChatRoom, room_id)
    msg = db.get(ChatMessage, message_id)
    if room is None or msg is None or int(msg.room_id) != int(room.id):
        abort(404)
    if not _chat_user_can_access_room(db, user, room):
        abort(404)
    if (msg.message_type or "") != "file" or not (msg.file_relpath or "").strip():
        abort(404)
    path = _chat_file_disk_path(int(room.exercise_id), int(room.id), msg.file_relpath)
    if path is None:
        abort(404)
    return send_file(
        path,
        as_attachment=True,
        download_name=(msg.original_filename or path.name)[:200],
        mimetype=(msg.mime_type or None) or "application/octet-stream",
    )


@bp.route("/admin/chat-rooms", methods=["GET", "POST"])
def admin_chat_rooms():
    user = get_current_user_optional()
    if not user or not can_manage_chat_rooms(user):
        abort(403)
    from flask import g

    db = g.db
    ex = _admin_current_workspace_exercise(db, user)
    err = ""
    ok_msg = ""
    if request.method == "POST":
        act = (request.form.get("action") or "").strip()
        if ex is None:
            err = "لا يوجد تمرين حالي لإدارة الغرف."
        elif act == "create_room":
            title = (request.form.get("title") or "").strip()[:500]
            rk = (request.form.get("room_kind") or ChatRoomKind.CUSTOM).strip()
            if rk not in _CHAT_KIND_LABELS_AR:
                rk = ChatRoomKind.CUSTOM
            uk = normalize_unit_level_key(request.form.get("unit_level_key") or "")
            if not title:
                err = "أدخل عنواناً للغرفة."
            else:
                room = ChatRoom(
                    exercise_id=ex.id,
                    title=title,
                    room_kind=rk,
                    unit_level_key=uk,
                    created_by_id=int(user.id),
                    last_activity_at=datetime.utcnow(),
                )
                db.add(room)
                db.flush()
                db.add(
                    ChatRoomMember(
                        room_id=room.id,
                        user_id=int(user.id),
                        role_in_room="moderator",
                    )
                )
                db.commit()
                ok_msg = "تم إنشاء الغرفة وإضافتك كمشرف."
        elif act == "add_member" and (request.form.get("room_id") or "").strip().isdigit():
            rid = int(request.form.get("room_id"))
            uid_raw = (request.form.get("user_id") or "").strip()
            room = db.get(ChatRoom, rid)
            if room is None or int(room.exercise_id) != int(ex.id):
                err = "غرفة غير صالحة."
            elif not uid_raw.isdigit():
                err = "اختر مستخدماً."
            else:
                uid = int(uid_raw)
                exists = (
                    db.query(ChatRoomMember)
                    .filter(ChatRoomMember.room_id == rid, ChatRoomMember.user_id == uid)
                    .first()
                )
                if exists is None:
                    db.add(ChatRoomMember(room_id=rid, user_id=uid, role_in_room="member"))
                    db.commit()
                    ok_msg = "تمت إضافة العضو."
                else:
                    err = "المستخدم عضو مسبقاً."
        elif act == "remove_member" and (request.form.get("room_id") or "").strip().isdigit():
            rid = int(request.form.get("room_id"))
            uid_raw = (request.form.get("user_id") or "").strip()
            room = db.get(ChatRoom, rid)
            if room is None or int(room.exercise_id) != int(ex.id):
                err = "غرفة غير صالحة."
            elif not uid_raw.isdigit():
                err = "معرّف مستخدم غير صالح."
            else:
                uid = int(uid_raw)
                if uid == int(user.id):
                    err = "لا يمكنك إزالة نفسك بهذه الطريقة."
                else:
                    db.query(ChatRoomMember).filter(
                        ChatRoomMember.room_id == rid, ChatRoomMember.user_id == uid
                    ).delete(synchronize_session=False)
                    db.commit()
                    ok_msg = "تمت إزالة العضو."

    rooms: list[ChatRoom] = []
    all_users: list[User] = []
    if ex is not None:
        rooms = (
            db.query(ChatRoom)
            .filter(ChatRoom.exercise_id == ex.id)
            .order_by(desc(ChatRoom.last_activity_at), desc(ChatRoom.id))
            .all()
        )
        all_users = db.query(User).order_by(User.id.asc()).limit(400).all()

    room_admin_rows: list[dict] = []
    for r in rooms:
        members = (
            db.query(ChatRoomMember)
            .filter(ChatRoomMember.room_id == r.id)
            .order_by(ChatRoomMember.joined_at.asc())
            .all()
        )
        mrows = []
        for mm in members:
            u = db.get(User, mm.user_id)
            mrows.append(
                {
                    "member": mm,
                    "user_label": (
                        (getattr(u, "full_name", "") or "").strip()
                        or (getattr(u, "username", "") or "").strip()
                        or f"#{mm.user_id}"
                    ),
                }
            )
        uk = (r.unit_level_key or "").strip()
        room_admin_rows.append(
            {
                "room": r,
                "kind_label": _CHAT_KIND_LABELS_AR.get((r.room_kind or "").strip(), r.room_kind or "—"),
                "unit_label": label_for_unit_level_key(uk) if uk else "—",
                "members": mrows,
            }
        )

    return render_template(
        "admin_chat_rooms.html",
        **_ctx(
            user,
            has_exercise=ex is not None,
            exercise=ex,
            error=err,
            ok_msg=ok_msg,
            room_kind_options=_CHAT_KIND_LABELS_AR,
            unit_levels=UNIT_LEVELS,
            room_admin_rows=room_admin_rows,
            all_users=all_users,
        ),
    )

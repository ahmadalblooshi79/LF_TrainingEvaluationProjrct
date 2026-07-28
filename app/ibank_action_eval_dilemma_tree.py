# -*- coding: utf-8 -*-
"""شجرة قوائم تقييم الإجراءات: معضلة → محكم وحدة → ملف Excel → اسم المحكم → فتح."""
from __future__ import annotations

import re
from pathlib import Path

from sqlalchemy.orm import Session

from app.config import INFO_BANK_DIR
from app.ibank_dilemma_folder_import import (
    collect_linked_files_by_dilemma,
    enrich_dilemma_names,
    parse_dilemma_no_from_text,
)
from app.info_bank_tree import exercise_judge_names_by_unit
from app.models.domain import InformationBankEventFlowTable, InformationBankTreeNode
from app.planner_flow_judge_labels import (
    parse_assignee_cell_lines,
    unit_key_for_assignee_label,
    unit_label_for_assignee_label,
)

# كلمات في اسم/عنوان الملف → كلمات في صنف المحكم (المكلف)
_FILE_TO_ASSIGNEE_HINTS: list[tuple[tuple[str, ...], tuple[str, ...]]] = [
    (("طبي", "مصاب", "إسعاف", "إخلاء الطبي", "الإخلاء الطبي", "لدغة", "عقرب"), ("طبي", "طبية")),
    (("آلية", "آليات", "صيانة", "إنقاذ الآلي", "إخلاء آلية", "جنزير", "إطار", "ربدان", "عطل آلية", "تعطل"), ("صيانة",)),
    (("كمين", "رد الفعل على كمين"), ("استطلاع",)),
    (("دفاع جوي", "تهديد جوي", "مسيرة", "مسيرات", "دفاع سلبي", "رد دفاع"), ("دفاع جوي", "مدفعية", "كتيبة")),
    (("هندس", "عبوة", "ناسفة"), ("هندس", "مشاة", "استطلاع", "كتيبة", "م ق", "ك/")),
    (("صمت", "لاسلكي", "إشار", "اشار"), ("اشارة", "إشارة", "الإشارة")),
    (("حرب إلكتروني", "إعاقة إلكتروني", "إعاقة"), ("حرب إلكتروني", "إلكتروني")),
    (("تزويد", "نقل"), ("تزويد", "نقل")),
    (("هاون",), ("هاون",)),
    (("م/د", "مضاد", "لدغة عقرب م د"), ("م/د", "مضاد")),
    (("مدني", "تصوير", "تسلل", "ازدحام"), ("قيادة", "شرطة", "أمن", "لواء")),
    (("منظومة", "قائد اللواء"), ("قيادة", "لواء")),
]

_SKIP_ASSIGNEE_RE = re.compile(
    r"^(تأكد|جميع المحكمين|القوة المقابلة|مقلدات|ملاحظة|انظر)",
    re.IGNORECASE,
)

_ANALYZE_CACHE: dict[str, tuple[int, int, dict]] = {}
_DILEMMA_TREE_CACHE: dict[tuple, tuple[tuple, dict]] = {}


def invalidate_action_eval_dilemma_tree_cache() -> None:
    """إبطال كاش شجرة المعاضل بعد استيراد/تعديل ملفات تقييم الإجراءات."""
    global _DILEMMA_TREE_CACHE, _ANALYZE_CACHE
    _DILEMMA_TREE_CACHE = {}
    _ANALYZE_CACHE = {}


def _norm(s: str) -> str:
    t = (
        (s or "")
        .strip()
        .lower()
        .replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ة", "ه")
    )
    # إزالة أل التعريف لتطابق «الدفاع» مع «دفاع»
    t = re.sub(r"(^|[\s\-_/])ال", r"\1", t)
    return re.sub(r"\s+", " ", t).strip()


def _is_unit_judge_assignee(label: str) -> bool:
    """أصناف محكمي الوحدة من عمود المكلف — تجاهل التعليمات والأسطر العامة."""
    lbl = (label or "").strip()
    if not lbl or _SKIP_ASSIGNEE_RE.search(lbl):
        return False
    return lbl.startswith("محكم")


def _prefer_judge_assignees(assignees: list[str]) -> list[str]:
    judges = [a for a in assignees if _is_unit_judge_assignee(a)]
    if judges:
        return judges
    return [a for a in assignees if (a or "").strip() and not _SKIP_ASSIGNEE_RE.search(a or "")]


def analyze_action_eval_xlsx(path: Path, *, deep: bool = False) -> dict:
    """تحليل خفيف لعنوان قائمة التقييم (بدون فتح Excel إلا عند deep=True).

    المسار السريع يستخدم اسم الملف — كافٍ للربط والعرض في شجرة المعاضل.
    deep=True يقرأ الورقة (أبطأ) مع تخزين مؤقت حسب mtime/size.
    """
    out = {
        "procedure_title": "",
        "sheet_title": "",
        "is_eval_list": False,
        "error": "",
    }
    if not path.is_file():
        out["error"] = "missing"
        return out
    stem = Path(path).stem
    out["procedure_title"] = re.sub(r"^تقييم\s+", "", stem or "").strip()[:300] or stem[:300]
    out["sheet_title"] = stem[:300]
    out["is_eval_list"] = str(path).lower().endswith((".xlsx", ".xlsm"))
    if not deep:
        return out

    try:
        st = path.stat()
        cache_key = str(path.resolve())
        hit = _ANALYZE_CACHE.get(cache_key)
        if hit and hit[0] == st.st_mtime_ns and hit[1] == st.st_size:
            return dict(hit[2])
    except OSError:
        st = None
        cache_key = str(path)

    try:
        from app.evaluation_sheet_parser import read_evaluation_list_sheet

        sheet = read_evaluation_list_sheet(path)
        title = (sheet.get("sheet_title") or "").strip()
        out["sheet_title"] = title or out["sheet_title"]
        out["is_eval_list"] = bool(sheet.get("eval_structured")) and not sheet.get("error")
        proc = ""
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(max_row=5, max_col=4, values_only=True):
                for cell in row:
                    if cell and str(cell).strip():
                        cand = str(cell).strip().split("\n")[0][:300]
                        if len(cand) >= 8 and cand not in ("عناصــــــر التقييـــــم",):
                            proc = cand
                            break
                if proc:
                    break
        finally:
            wb.close()
        if not proc:
            proc = title or stem
        proc = re.sub(r"^تقييم\s+", "", proc or "").strip()
        if proc in ("قائمة التقييم", "بيانات التقييم"):
            proc = stem
        out["procedure_title"] = proc[:300]
    except Exception as exc:
        out["error"] = str(exc)[:200]
    if st is not None:
        _ANALYZE_CACHE[cache_key] = (st.st_mtime_ns, st.st_size, dict(out))
    return out


def _dilemma_tree_fingerprint(db: Session, exercise_id: int | None) -> tuple:
    from sqlalchemy import func

    n, mx_id = (
        db.query(
            func.count(InformationBankTreeNode.id),
            func.max(InformationBankTreeNode.id),
        )
        .filter(
            InformationBankTreeNode.kind == "action_eval",
            InformationBankTreeNode.is_folder.is_(False),
        )
        .one()
    )
    flow = (
        db.query(
            InformationBankEventFlowTable.id,
            InformationBankEventFlowTable.updated_at,
        )
        .order_by(InformationBankEventFlowTable.id)
        .first()
    )
    flow_fp = (
        (int(flow[0]), flow[1].isoformat() if flow[1] is not None else "")
        if flow
        else (0, "")
    )
    return (int(n or 0), int(mx_id or 0), flow_fp, int(exercise_id or 0))


def _assignees_by_dilemma_from_flow(raw_json: str) -> dict[str, dict[int, list[str]]]:
    """day_id → dilemma_no → [أصناف المحكمين من عمود المكلف]."""
    from app.action_eval_ibank_sync import _parse_flow_table_days

    days = _parse_flow_table_days(raw_json or "")
    out: dict[str, dict[int, list[str]]] = {}
    for day in days:
        day_id = str(day.get("id") or "").strip()
        if not day_id:
            continue
        bucket: dict[int, list[str]] = {}
        current_no: int | None = None
        for row in day.get("rows") or []:
            kind = (row.get("kind") or "").strip().lower()
            if kind == "dilemma":
                text = (row.get("text") or "").strip()
                dno = parse_dilemma_no_from_text(text)
                current_no = dno
                if dno is not None:
                    bucket.setdefault(dno, [])
                continue
            if kind == "event":
                current_no = None
                continue
            if kind == "row" and current_no is not None:
                for lbl in parse_assignee_cell_lines(row.get("assignee")):
                    if lbl not in bucket.setdefault(current_no, []):
                        bucket[current_no].append(lbl)
        out[day_id] = bucket
    return out


def _dilemmas_by_day_from_flow(raw_json: str) -> dict[str, list[dict]]:
    from app.action_eval_ibank_sync import _parse_flow_table_days

    days = _parse_flow_table_days(raw_json or "")
    out: dict[str, list[dict]] = {}
    for day in days:
        day_id = str(day.get("id") or "").strip()
        if not day_id:
            continue
        names: list[dict] = []
        num = 0
        for row in day.get("rows") or []:
            if (row.get("kind") or "").strip().lower() != "dilemma":
                continue
            text = (row.get("text") or "").strip()
            if not text:
                continue
            num += 1
            dno = parse_dilemma_no_from_text(text) or num
            names.append({"num": num, "text": text, "dilemma_no": dno, "files": []})
        out[day_id] = names
    return enrich_dilemma_names(out)


def _score_file_for_assignee(file_name: str, procedure_title: str, assignee: str) -> int:
    hay = _norm(f"{file_name} {procedure_title}")
    asn = _norm(assignee)
    score = 0
    for file_keys, judge_keys in _FILE_TO_ASSIGNEE_HINTS:
        if any(_norm(k) in hay for k in file_keys) and any(_norm(k) in asn for k in judge_keys):
            score += 3
    for token in re.findall(r"[\u0600-\u06ff]{4,}", asn):
        if token in hay:
            score += 1
    return score


def _match_files_to_assignees(
    assignees: list[str],
    files_meta: list[dict],
    *,
    assignee_unit_keys: dict[str, str] | None = None,
) -> dict[str, list[dict]]:
    """ربط كل ملف Excel بمحكم الوحدة — أولاً عبر مستوى الوحدة المحفوظ، ثم تلميح الاسم."""
    result: dict[str, list[dict]] = {a: [] for a in assignees}
    unmatched: list[dict] = []
    ranked = sorted(
        assignees,
        key=lambda a: (0 if (a or "").startswith("محكم") else 1, a),
    )
    uk_map = assignee_unit_keys or {}
    # unit_key → أول صنف مكلف يطابقه
    unit_to_assignee: dict[str, str] = {}
    for a in ranked:
        uk = (uk_map.get(a) or "").strip()
        if uk and uk not in unit_to_assignee:
            unit_to_assignee[uk] = a

    for meta in files_meta:
        file_uk = (meta.get("unit_key") or "").strip()
        if file_uk and file_uk in unit_to_assignee:
            result[unit_to_assignee[file_uk]].append(meta)
            continue
        best_a = None
        best_s = 0
        for a in ranked:
            s = _score_file_for_assignee(
                meta.get("name") or "", meta.get("procedure_title") or "", a
            )
            if s > best_s:
                best_s = s
                best_a = a
        # يلزم تلميح موضوعي (وليس مجرد تطابق ضعيف)
        if best_a and best_s >= 3:
            result[best_a].append(meta)
        else:
            unmatched.append(meta)
    if unmatched:
        targets = [a for a in ranked if a.startswith("محكم")] or ranked
        if not targets:
            result.setdefault("__unassigned__", []).extend(unmatched)
        else:
            for meta in unmatched:
                tgt = min(targets, key=lambda a: (len(result.get(a) or []), a))
                result[tgt].append(meta)
    return result


def build_action_eval_dilemma_judge_tree(
    db: Session,
    *,
    exercise_id: int | None = None,
) -> dict[str, list[dict]]:
    """شجرة العرض لكل يوم — المعاضل من مجرى الأحداث فقط، مع ربط ملفات الاستيراد بها."""
    fp = _dilemma_tree_fingerprint(db, exercise_id)
    cache_key = fp
    hit = _DILEMMA_TREE_CACHE.get(cache_key)
    if hit and hit[0] == fp:
        # نسخة سطحية كافية للقراءة فقط في الطلب
        return hit[1]

    out = _build_action_eval_dilemma_judge_tree_uncached(db, exercise_id=exercise_id)
    _DILEMMA_TREE_CACHE[cache_key] = (fp, out)
    return out


def _build_action_eval_dilemma_judge_tree_uncached(
    db: Session,
    *,
    exercise_id: int | None = None,
) -> dict[str, list[dict]]:
    row = (
        db.query(InformationBankEventFlowTable)
        .order_by(InformationBankEventFlowTable.id)
        .first()
    )
    raw = (getattr(row, "flow_table_json", None) or "").strip() if row else ""
    dilemmas_by_day = _dilemmas_by_day_from_flow(raw)
    assignees_by_day = _assignees_by_dilemma_from_flow(raw)
    linked = collect_linked_files_by_dilemma(db)
    judge_names = exercise_judge_names_by_unit(db, exercise_id)

    # العرض يتبع المجرى فقط: تفريغ اليوم ⇒ قائمة فارغة حتى لو بقيت ملفات مستوردة في الشجرة
    all_day_ids = sorted(
        dilemmas_by_day.keys(),
        key=lambda d: (0 if str(d).startswith("day-") else 1, str(d)),
    )

    out: dict[str, list[dict]] = {}
    for day_id in all_day_ids:
        day_files = linked.get(day_id) or {}
        day_assignees = assignees_by_day.get(day_id) or {}
        flow_rows = list(dilemmas_by_day.get(day_id) or [])
        if not flow_rows:
            out[day_id] = []
            continue
        by_no: dict[int, dict] = {}
        for d in flow_rows:
            dno = int(d.get("dilemma_no") or d.get("num") or 0)
            if dno:
                by_no[dno] = dict(d)

        day_nodes: list[dict] = []
        for dno in sorted(by_no.keys()):
            d = by_no[dno]
            file_rows = day_files.get(dno) or []
            files_meta: list[dict] = []
            for fr in file_rows:
                name = fr.get("name") or ""
                if not str(name).lower().endswith((".xlsx", ".xlsm")):
                    continue
                node_id = int(fr.get("node_id") or fr.get("id") or 0)
                node = db.get(InformationBankTreeNode, node_id) if node_id else None
                path = None
                if node and node.file_relpath:
                    path = (INFO_BANK_DIR / node.file_relpath).resolve()
                analysis = analyze_action_eval_xlsx(path) if path else {}
                unit_key = (fr.get("unit_key") or "").strip()
                if not unit_key and node is not None:
                    unit_key = (node.catalog_unit_key or "").strip()
                files_meta.append(
                    {
                        "node_id": node_id,
                        "name": name,
                        "procedure_title": analysis.get("procedure_title") or Path(name).stem,
                        "is_eval_list": bool(analysis.get("is_eval_list")),
                        "unit_key": unit_key,
                    }
                )

            assignees = _prefer_judge_assignees(list(day_assignees.get(dno) or []))
            asn_unit_keys: dict[str, str] = {}
            for asn in assignees:
                asn_unit_keys[asn] = unit_key_for_assignee_label(asn, db=db) or ""
            matched = _match_files_to_assignees(
                assignees, files_meta, assignee_unit_keys=asn_unit_keys
            )
            judge_nodes: list[dict] = []
            for asn in assignees:
                uk = asn_unit_keys.get(asn) or ""
                ul = unit_label_for_assignee_label(asn) or ""
                person = (judge_names.get(uk) or "").strip() if uk else ""
                file_nodes = []
                for meta in matched.get(asn) or []:
                    file_nodes.append(
                        {
                            **meta,
                            "judge_person_name": person or "—",
                            "open_href": f"/admin/information-bank/action-eval/view/{meta['node_id']}",
                        }
                    )
                judge_nodes.append(
                    {
                        "assignee_label": asn,
                        "unit_key": uk,
                        "unit_label": ul,
                        "judge_person_name": person,
                        "files": file_nodes,
                    }
                )
            extra = matched.get("__unassigned__") or []
            if extra or (files_meta and not assignees):
                extras = extra if assignees else files_meta
                file_nodes = []
                for meta in extras:
                    file_nodes.append(
                        {
                            **meta,
                            "judge_person_name": "—",
                            "open_href": f"/admin/information-bank/action-eval/view/{meta['node_id']}",
                        }
                    )
                if file_nodes:
                    judge_nodes.append(
                        {
                            "assignee_label": (
                                "ملفات غير مربوطة بمحكم"
                                if assignees
                                else "ملفات التقييم المستوردة"
                            ),
                            "unit_key": "",
                            "unit_label": "",
                            "judge_person_name": "",
                            "files": file_nodes,
                        }
                    )
            day_nodes.append(
                {
                    "num": d.get("num") or dno,
                    "dilemma_no": dno,
                    "text": d.get("text") or f"المعضلة/{dno}",
                    "judges": judge_nodes,
                }
            )
        out[day_id] = day_nodes
    return out

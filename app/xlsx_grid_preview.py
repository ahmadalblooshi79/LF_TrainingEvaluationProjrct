"""معاينة شبكة لملف Excel (.xlsx) في المتصفح — أسماء أعمدة من الخلايا وترقيم أسطر كما في Excel."""
from __future__ import annotations

from pathlib import Path
from typing import Any


MAX_PREVIEW_ROWS = 500
MAX_PREVIEW_COLS = 40


def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "نعم" if val else "لا"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
        return str(val)
    s = str(val).replace("\u00a0", " ")
    return s.replace("\r\n", "\n").replace("\r", "\n").strip()


def read_xlsx_sheet_preview(path: Path, *, sheet_index: int = 0) -> dict[str, Any]:
    """
    يقرأ الورقة الأولى (أو sheet_index) كسلسلة صفوف نصية.
    يعيد صفوفاً بنفس ترتيب Excel مع عدد أعمدة موحّد لعرض الجدول.
    """
    path = Path(path)
    out: dict[str, Any] = {
        "sheet_title": "",
        "grid_rows": [],
        "header_row": [],
        "body_rows": [],
        "col_labels": [],
        "ncol": 0,
        "error": None,
    }
    if not path.is_file():
        out["error"] = "الملف غير موجود."
        return out
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        out["error"] = "معاينة الجدول متاحة لملفات Excel (.xlsx) فقط."
        return out

    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError:
        out["error"] = "تعذر فتح ملف Excel لأن مكتبة openpyxl غير مثبتة. ثبّت المتطلبات ثم أعد التشغيل."
        return out

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        names = wb.sheetnames
        if not names:
            out["error"] = "لا توجد أوراق في الملف."
            return out
        si = max(0, min(sheet_index, len(names) - 1))
        ws = wb[names[si]]
        out["sheet_title"] = ws.title or ""

        mr = getattr(ws, "max_row", None) or 1
        mc = getattr(ws, "max_column", None) or 1
        max_row = min(int(mr), MAX_PREVIEW_ROWS)
        max_col = min(int(mc), MAX_PREVIEW_COLS)
        if max_row < 1:
            max_row = 1
        if max_col < 1:
            max_col = 1

        raw_rows: list[list[str]] = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            raw_rows.append([_cell_to_str(c) for c in row])

        ncol = max_col
        for r in raw_rows:
            while len(r) < ncol:
                r.append("")

        header_row = list(raw_rows[0]) if raw_rows else []
        body_rows = raw_rows[1:] if len(raw_rows) > 1 else []

        out["grid_rows"] = raw_rows
        out["header_row"] = header_row
        out["body_rows"] = body_rows
        out["ncol"] = ncol
        out["col_labels"] = []
    except Exception as exc:
        out["error"] = f"تعذر قراءة ملف Excel: {exc}"
        out["grid_rows"] = []
        out["header_row"] = []
        out["body_rows"] = []
        out["col_labels"] = []
        out["ncol"] = 0
    finally:
        wb.close()

    return out

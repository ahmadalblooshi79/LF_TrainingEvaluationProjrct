"""
استخراج نصوص الأهداف التدريبية من ملفات متعددة الصيغ (Excel, PDF, CSV, XML).
يعتمد على استخراج شامل للنص ثم تصفية عناوين/فواصل شائعة — دون الاعتماد على موضع ثابت في الملف.
"""
from __future__ import annotations

import csv
import re
import xml.etree.ElementTree as ET
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

# عناوين/تسميات شائعة في نماذج الأهداف (عربي/إنجليزي) — تُستبعد كسطر مستقل أو قصير جداً
_JUNK_EXACT = {
    "التسلسل",
    "الأهداف التدريبية",
    "الملاحظات",
    "الوحدة:",
    "الوحدة",
    "قائد الوحدة:",
    "قائد الوحدة",
    "رقم",
    "الهدف",
    "الهدف التدريبي",
    "objective",
    "objectives",
    "no",
    "no.",
    "#",
    "item",
    "description",
    "notes",
    "serial",
    "s/n",
}

_JUNK_CONTAINS_SHORT = (
    "قائمة الأهداف التدريبية",
    "الأهداف التدريبية للوحدة",
    "جدول الأهداف",
    "training objectives",
    "learning objectives",
)


def _norm(s: str) -> str:
    t = (s or "").replace("\u200f", "").replace("\u200e", "").replace("\ufeff", "").strip()
    t = re.sub(r"[\s\u00a0]+", " ", t)
    return t


def _split_bullets(s: str) -> list[str]:
    """تقسيم سطر قد يضم عدة نقاط في وورد/نص."""
    parts = re.split(r"(?:^|\n)\s*[•●○◦▪▸►\-–—]\s+", s)
    out: list[str] = []
    for p in parts:
        t = _norm(p)
        if t:
            out.append(t)
    return out if len(out) > 1 else ([_norm(s)] if _norm(s) else [])


def _is_junk_line(t: str) -> bool:
    if not t or len(t) < 4:
        return True
    if t.isdigit() or (len(t) <= 4 and re.fullmatch(r"[\d.\s\-/]+", t)):
        return True
    tl = t.casefold()
    if tl in {x.casefold() for x in _JUNK_EXACT}:
        return True
    if len(t) <= 60:
        for ph in _JUNK_CONTAINS_SHORT:
            if ph.casefold() in tl and len(t) < 80:
                return True
    # عنوان جدول قصير جداً
    if len(t) < 12 and not re.search(r"[اأإآبتثجحخدذرزسشصضطظعغفقكلمنهوي]", t):
        if re.fullmatch(r"[\w\s./\-:]+", t):
            return True
    return False


def _yield_text_chunks(text: str) -> Iterable[str]:
    for line in (text or "").splitlines():
        line = _norm(line)
        if not line:
            continue
        for chunk in _split_bullets(line):
            c = _norm(chunk)
            if c and not _is_junk_line(c):
                yield c


def _finalize(candidates: list[str], max_items: int = 200) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for raw in candidates:
        for piece in _yield_text_chunks(raw):
            if len(piece) > 2000:
                piece = piece[:2000]
            k = piece.casefold()
            if k in seen:
                continue
            seen.add(k)
            out.append(piece)
            if len(out) >= max_items:
                return out
    return out


def _from_xlsx(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), data_only=True)
    cells: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is None:
                    continue
                s = _norm(str(c))
                if s:
                    cells.append(s)
    return _finalize(cells)


def _from_csv(data: bytes) -> list[str]:
    for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            text = ""
    if not text:
        return []
    rows: list[str] = []
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    f = StringIO(text)
    reader = csv.reader(f, dialect)
    for row in reader:
        for cell in row:
            s = _norm(cell)
            if s:
                rows.append(s)
    return _finalize(rows)


def _from_xml(data: bytes) -> list[str]:
    chunks: list[str] = []
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []

    def walk(el: ET.Element) -> None:
        tag = (el.tag or "").split("}")[-1].lower()
        if el.text:
            t = _norm(el.text)
            if t:
                chunks.append(t)
        for ch in el:
            walk(ch)
        if el.tail:
            t = _norm(el.tail)
            if t:
                chunks.append(t)
        # قيم السمات القصيرة قد تحمل رموزاً أو عناوين أهداف
        for _k, v in (el.attrib or {}).items():
            tv = _norm(str(v))
            if len(tv) >= 8 and tag not in ("style", "stylesheet"):
                chunks.append(tv)

    walk(root)
    return _finalize(chunks)


def _from_pdf(data: bytes) -> list[str]:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(data))
    blobs: list[str] = []
    for page in reader.pages:
        try:
            txt = page.extract_text() or ""
        except Exception:
            txt = ""
        txt = txt.strip()
        if txt:
            blobs.append(txt)
    combined = "\n".join(blobs)
    lines = [_norm(x) for x in combined.splitlines() if _norm(x)]
    return _finalize(lines)


def _from_plain_or_txt(data: bytes) -> list[str]:
    text = data.decode("utf-8", errors="ignore")
    lines = [_norm(x) for x in text.splitlines() if _norm(x)]
    return _finalize(lines)


def extract_objectives_from_file(filename: str, data: bytes) -> list[str]:
    """
    يستخرج قائمة أهداف تدريبية من محتوى الملف حسب الامتداد.
    يعيد قائمة فارغة عند الفشل أو عدم التعرف على الصيغة.
    """
    if not data:
        return []
    ext = Path((filename or "").strip()).suffix.lower()

    def _safe(fn):
        try:
            return fn(data)
        except Exception:
            return []

    if ext in (".xlsx", ".xlsm"):
        return _safe(_from_xlsx)
    if ext in (".csv", ".tsv"):
        return _safe(_from_csv)
    if ext == ".xml":
        return _safe(_from_xml)
    if ext == ".pdf":
        return _safe(_from_pdf)
    if ext in (".txt", ".text"):
        return _safe(_from_plain_or_txt)

    head = data.lstrip()[:120]
    if head.startswith(b"<?xml") or head.startswith(b"<"):
        x = _safe(_from_xml)
        if x:
            return x
    return _safe(_from_plain_or_txt)


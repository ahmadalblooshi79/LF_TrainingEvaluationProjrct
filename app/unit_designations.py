"""كشف المسميات والدلالات الرئيسية للوحدات — تحميل وحل المسميات البديلة."""
from __future__ import annotations

import re
from pathlib import Path

from sqlalchemy.orm import Session

from app.models.domain import UnitDesignation, UnitDesignationAlias

_DATA_XLSX = Path(__file__).resolve().parent / "data" / "unit_designations.xlsx"
_REPO_XLSX = Path(__file__).resolve().parents[1] / "كشف المسميات والدلالات الرئيسية للوحدات.xlsx"

# كاش في الذاكرة: norm(alias|canonical) → unit_id
_ALIAS_NORM_TO_UNIT: dict[str, str] = {}
_UNIT_BY_ID: dict[str, UnitDesignation] = {}
_LOADED = False


def normalize_designation_text(s: str) -> str:
    """توحيد نص للمقارنة (همزات، مسافات، شرطة)."""
    t = (s or "").strip()
    t = (
        t.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ة", "ه")
        .replace("ى", "ي")
    )
    t = t.replace("الـ", "ال").replace("الـ ", "ال")
    t = re.sub(r"(\d+)من", r"\1 من", t)
    t = re.sub(r"(كتيبة)\s+(\d+)\b", r"\1/\2", t)
    t = re.sub(r"(ك)\s+(\d+)\b", r"\1/\2", t)
    # توحيد «م/د» و«م د» و«مـ د» الشائعة في أسماء المجلدات
    t = re.sub(r"م\s*/\s*د", "م/د", t)
    t = re.sub(r"م\s*ـ\s*د", "م/د", t)
    t = re.sub(r"م\s+د\b", "م/د", t)
    t = re.sub(r"\s*/\s*", "/", t)
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def _xlsx_candidates() -> list[Path]:
    """يفضّل ملف الجذر المحدَّث ثم النسخة في app/data."""
    out: list[Path] = []
    for p in (_REPO_XLSX, _DATA_XLSX):
        if p.is_file() and p not in out:
            out.append(p)
    return out


def _xlsx_master_count(path: Path) -> int:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb["Units_Master"] if "Units_Master" in wb.sheetnames else wb[wb.sheetnames[0]]
            n = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:
                    n += 1
            return n
        finally:
            wb.close()
    except Exception:
        return 0


def seed_unit_designations_from_xlsx(db: Session, *, force: bool = False) -> dict[str, int]:
    """تحميل/تحديث الدلالات والمسميات من ملف Excel إلى قاعدة البيانات."""
    paths = _xlsx_candidates()
    if not paths:
        reload_unit_designation_cache(db)
        return {"masters": 0, "aliases": 0, "skipped": 1}

    existing = db.query(UnitDesignation).count()
    path = paths[0]
    xlsx_count = _xlsx_master_count(path)
    # أعد المزامنة تلقائياً عند إضافة وحدات جديدة في الكشف
    if existing and not force and xlsx_count > 0 and xlsx_count != existing:
        force = True
    if existing and not force:
        reload_unit_designation_cache(db)
        return {"masters": existing, "aliases": db.query(UnitDesignationAlias).count(), "cached": 1}

    from openpyxl import load_workbook

    path = paths[0]
    wb = load_workbook(path, data_only=True)
    try:
        ws_m = wb["Units_Master"] if "Units_Master" in wb.sheetnames else wb[wb.sheetnames[0]]
        ws_a = wb["Units_Alias"] if "Units_Alias" in wb.sheetnames else None

        masters = 0
        for i, row in enumerate(ws_m.iter_rows(min_row=2, values_only=True), start=1):
            if not row or not row[0]:
                continue
            uid = str(row[0]).strip()
            label = str(row[1] or "").strip()
            utype = str(row[2] or "").strip()
            desc = str(row[3] or "").strip() if len(row) > 3 else ""
            status = str(row[4] or "").strip() if len(row) > 4 else "فعال"
            if not uid or not label:
                continue
            rec = db.get(UnitDesignation, uid)
            if rec is None:
                rec = UnitDesignation(unit_id=uid)
                db.add(rec)
            rec.canonical_label = label
            rec.unit_type = utype
            rec.description = desc
            rec.is_active = status in ("فعال", "active", "1", "true", "True", "")
            rec.sort_order = i
            masters += 1

        aliases = 0
        seen_norms: set[str] = set()
        if ws_a is not None:
            for row in ws_a.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                aid = str(row[0]).strip()
                uid = str(row[1] or "").strip()
                alias = str(row[2] or "").strip()
                notes = str(row[3] or "").strip() if len(row) > 3 else ""
                if not aid or not uid or not alias:
                    continue
                if db.get(UnitDesignation, uid) is None:
                    continue
                norm = normalize_designation_text(alias)
                if not norm:
                    continue
                if norm in seen_norms:
                    continue
                existing_alias = (
                    db.query(UnitDesignationAlias)
                    .filter(UnitDesignationAlias.alias_label_norm == norm)
                    .first()
                )
                if existing_alias is not None:
                    existing_alias.unit_id = uid
                    existing_alias.alias_label = alias
                    if notes:
                        existing_alias.notes = notes
                    seen_norms.add(norm)
                    aliases += 1
                    continue
                rec = db.get(UnitDesignationAlias, aid)
                if rec is None:
                    rec = UnitDesignationAlias(alias_id=aid)
                    db.add(rec)
                rec.unit_id = uid
                rec.alias_label = alias
                rec.alias_label_norm = norm
                rec.notes = notes
                seen_norms.add(norm)
                aliases += 1

        # أضف الدلالة الرئيسية نفسها كمسمى للبحث المباشر
        for m in db.query(UnitDesignation).all():
            label = (m.canonical_label or "").strip()
            if not label:
                continue
            norm = normalize_designation_text(label)
            if not norm or norm in seen_norms:
                continue
            exists = (
                db.query(UnitDesignationAlias)
                .filter(UnitDesignationAlias.alias_label_norm == norm)
                .first()
            )
            if exists is None:
                syn_id = f"S_{m.unit_id}"
                if db.get(UnitDesignationAlias, syn_id) is None:
                    db.add(
                        UnitDesignationAlias(
                            alias_id=syn_id,
                            unit_id=m.unit_id,
                            alias_label=label,
                            alias_label_norm=norm,
                            notes="دلالة رئيسية",
                        )
                    )
                    aliases += 1
                    seen_norms.add(norm)
            else:
                seen_norms.add(norm)

        # توليد مسميات «محكم السرية/N من …» للسرايا في الكشف إن لم تُدرج في الورقة
        company_re = re.compile(r"^(.+?)\s*-\s*السرية/(\d+)\s*$")
        for m in db.query(UnitDesignation).all():
            label = (m.canonical_label or "").strip()
            cm = company_re.match(label)
            if not cm:
                continue
            parent, n = cm.group(1).strip(), cm.group(2)
            candidates = [f"محكم السرية/{n} من {parent}"]
            if parent.startswith("قيادة "):
                candidates.append(f"محكم السرية/{n} من {parent[len('قيادة '):]}")
            for i, cand in enumerate(candidates):
                norm = normalize_designation_text(cand)
                if not norm or norm in seen_norms:
                    continue
                exists = (
                    db.query(UnitDesignationAlias)
                    .filter(UnitDesignationAlias.alias_label_norm == norm)
                    .first()
                )
                if exists is not None:
                    seen_norms.add(norm)
                    continue
                syn_id = f"SC{i}_{m.unit_id}"
                if db.get(UnitDesignationAlias, syn_id) is not None:
                    continue
                db.add(
                    UnitDesignationAlias(
                        alias_id=syn_id,
                        unit_id=m.unit_id,
                        alias_label=cand,
                        alias_label_norm=norm,
                        notes="مولَّد من الدلالة الرئيسية",
                    )
                )
                aliases += 1
                seen_norms.add(norm)

        db.commit()
    finally:
        wb.close()

    reload_unit_designation_cache(db)
    return {"masters": masters, "aliases": aliases, "source": str(path)}


def reload_unit_designation_cache(db: Session | None = None) -> None:
    """إعادة بناء كاش الذاكرة من قاعدة البيانات."""
    global _LOADED
    _ALIAS_NORM_TO_UNIT.clear()
    _UNIT_BY_ID.clear()

    close = False
    if db is None:
        from app.database import SessionLocal

        db = SessionLocal()
        close = True
    try:
        for m in db.query(UnitDesignation).filter(UnitDesignation.is_active.is_(True)).all():
            _UNIT_BY_ID[m.unit_id] = m
            n = normalize_designation_text(m.canonical_label or "")
            if n:
                _ALIAS_NORM_TO_UNIT[n] = m.unit_id
        for a in db.query(UnitDesignationAlias).all():
            if a.unit_id not in _UNIT_BY_ID:
                continue
            n = (a.alias_label_norm or "").strip() or normalize_designation_text(
                a.alias_label or ""
            )
            if n:
                _ALIAS_NORM_TO_UNIT[n] = a.unit_id
        _LOADED = True
    finally:
        if close:
            db.close()


def ensure_unit_designations_loaded(db: Session | None = None) -> None:
    if _LOADED and _UNIT_BY_ID:
        return
    close = False
    if db is None:
        from app.database import SessionLocal

        db = SessionLocal()
        close = True
    try:
        if db.query(UnitDesignation).count() == 0:
            seed_unit_designations_from_xlsx(db, force=True)
        else:
            reload_unit_designation_cache(db)
    finally:
        if close:
            db.close()


def resolve_unit_id_for_assignee(assignee_label: str) -> str:
    """يعيد Unit_ID (مثل U001) من مسمى بديل أو دلالة رئيسية."""
    ensure_unit_designations_loaded()
    raw = re.sub(r"^[\s•·\-–]+", "", (assignee_label or "").strip()).strip()
    if not raw:
        return ""
    stripped_num = re.sub(r"^[\d\s.\-–]+", "", raw).strip()
    candidates = [
        raw,
        stripped_num,
        re.sub(r"(\d+)من", r"\1 من", raw),
        re.sub(r"\s+", " ", raw),
        re.sub(r"\s+", " ", stripped_num),
    ]
    seen: set[str] = set()
    for cand in candidates:
        cand = (cand or "").strip()
        if not cand or cand in seen:
            continue
        seen.add(cand)
        n = normalize_designation_text(cand)
        uid = _ALIAS_NORM_TO_UNIT.get(n)
        if uid:
            return uid
    # أطول دلالة/مسمى محتواة في النص (مثل «كتيبة المدفعية» ↔ «قيادة كتيبة المدفعية»)
    norm_raw = normalize_designation_text(stripped_num or raw)
    if not norm_raw:
        return ""
    best_uid = ""
    best_score = 0
    for alias_norm, uid in _ALIAS_NORM_TO_UNIT.items():
        if not alias_norm or not uid:
            continue
        if alias_norm in norm_raw:
            score = len(alias_norm) + 1000
        elif len(norm_raw) >= 8 and norm_raw in alias_norm:
            score = len(norm_raw)
        else:
            continue
        if score > best_score:
            best_score = score
            best_uid = uid
    return best_uid


def canonical_label_for_unit_id(unit_id: str) -> str:
    ensure_unit_designations_loaded()
    rec = _UNIT_BY_ID.get((unit_id or "").strip())
    return (rec.canonical_label if rec else "") or ""


def canonical_label_for_assignee(assignee_label: str) -> str:
    """الدلالة الرئيسية للوحدة من أي مسمى بديل في عمود المكلف."""
    uid = resolve_unit_id_for_assignee(assignee_label)
    return canonical_label_for_unit_id(uid) if uid else ""


def list_designations_for_ibank(db: Session) -> list[dict]:
    """قائمة الدلالات والمسميات لواجهة بنك المعلومات (مصدر عام بلا تمرين)."""
    ensure_unit_designations_loaded(db)
    rows = (
        db.query(UnitDesignation)
        .order_by(UnitDesignation.sort_order, UnitDesignation.unit_id)
        .all()
    )
    aliases_by_unit: dict[str, list[dict[str, str]]] = {}
    for a in db.query(UnitDesignationAlias).order_by(UnitDesignationAlias.alias_id).all():
        uid = (a.unit_id or "").strip()
        if not uid:
            continue
        aliases_by_unit.setdefault(uid, []).append(
            {
                "alias_id": a.alias_id,
                "alias_label": (a.alias_label or "").strip(),
                "notes": (a.notes or "").strip(),
            }
        )
    out: list[dict] = []
    for i, m in enumerate(rows, start=1):
        uid = (m.unit_id or "").strip()
        out.append(
            {
                "unit_id": uid,
                "num": i,
                "canonical_label": (m.canonical_label or "").strip(),
                "unit_type": (m.unit_type or "").strip(),
                "is_active": bool(m.is_active),
                "aliases": aliases_by_unit.get(uid, []),
            }
        )
    return out


def _next_prefixed_id(db: Session, *, model, field: str, prefix: str) -> str:
    nums: list[int] = []
    rx = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.I)
    for (raw,) in db.query(getattr(model, field)).all():
        m = rx.match((raw or "").strip())
        if m:
            nums.append(int(m.group(1)))
    n = (max(nums) if nums else 0) + 1
    return f"{prefix}{n:03d}"


def next_unit_id(db: Session) -> str:
    return _next_prefixed_id(db, model=UnitDesignation, field="unit_id", prefix="U")


def next_alias_id(db: Session) -> str:
    return _next_prefixed_id(db, model=UnitDesignationAlias, field="alias_id", prefix="A")


def ensure_canonical_alias(db: Session, *, unit_id: str, label: str) -> None:
    label = (label or "").strip()
    if not label:
        return
    norm = normalize_designation_text(label)
    if not norm:
        return
    exists = (
        db.query(UnitDesignationAlias)
        .filter(UnitDesignationAlias.alias_label_norm == norm)
        .first()
    )
    if exists is not None:
        exists.unit_id = unit_id
        exists.alias_label = label
        return
    syn_id = f"S_{unit_id}"
    rec = db.get(UnitDesignationAlias, syn_id)
    if rec is None:
        rec = UnitDesignationAlias(alias_id=syn_id)
        db.add(rec)
    rec.unit_id = unit_id
    rec.alias_label = label
    rec.alias_label_norm = norm
    rec.notes = "دلالة رئيسية"


def apply_canonical_label_to_organization(db: Session, *, old_label: str, new_label: str) -> int:
    """يحدّث تسمية التنظيم إن طابقت الدلالة السابقة — بلا ربط بتمرين."""
    from app.models.domain import InformationBankTreeNode, InformationBankUnitLevel

    old_l = (old_label or "").strip()
    new_l = (new_label or "").strip()
    if not old_l or not new_l or old_l == new_l:
        return 0
    n = 0
    for row in db.query(InformationBankUnitLevel).filter_by(label=old_l).all():
        row.label = new_l
        n += 1
        for node in (
            db.query(InformationBankTreeNode)
            .filter(
                InformationBankTreeNode.catalog_unit_key == row.key,
                InformationBankTreeNode.is_folder.is_(True),
            )
            .all()
        ):
            if (node.name or "").strip() == old_l:
                node.name = new_l
    return n

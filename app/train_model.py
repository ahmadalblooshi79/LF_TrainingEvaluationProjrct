"""
تدريب نموذج تصنيف أوفلاين من ملفات Excel (.xlsx) — بدون إنترنت بعد تثبيت الحزم.

الاستخدام من جذر المشروع:
  .venv\\Scripts\\python.exe -m app.train_model
  .venv\\Scripts\\python.exe app\\train_model.py
  الوضع الافتراضي ``auto``: إن لم تتوفر عينات حقيقية يُستخدم تدريب تركيبي تلقائياً (مع تنبيه في الطرفية).
  .venv\\Scripts\\python.exe -m app.train_model --demo
  .venv\\Scripts\\python.exe -m app.train_model --data-dir ml_training/raw --model-out ml_training/models/eval_classifier.joblib

ضع ملفات التدريب في المجلد.
وضع ``--format rubric`` لقوالب «قائمة التقييم» (صف القصوى/المكتسبة): يُستخرج كل بند عليه علامة،
والهدف هو نطاق الأداء من نسبة (مكتسب÷قصوى).
الوضع ``flat`` (افتراضي سابقاً): صف أول = رؤوس، آخر عمود = تصنيف ما لم يُحدّد الاسم.
``auto``: قاعدة البيانات (نتائج محفوظة) ثم ملفات قائمة التقييم المعبأة — دون الجدول المسطح (تجنّباً لسوء تفسير القوالب).
``saved``: من ``evaluation_list_saved_results`` فقط (ما حُفظ من صفحة قائمة التقييم).
``flat``: جدول حقيقي؛ غالباً مع ``--label-name``.
"""
from __future__ import annotations

import sys
from pathlib import Path

# التشغيل المباشر ‎python app/train_model.py‎ لا يضع جذر المشروع على المسار؛ نصلح ذلك قبل استيراد ‎app‎.
_REPO_ROOT = Path(__file__).resolve().parent.parent
_rs = str(_REPO_ROOT)
if _rs not in sys.path:
    sys.path.insert(0, _rs)

import argparse
import json
from collections import Counter
from typing import Any

import joblib
import numpy as np
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier
from sklearn.impute import SimpleImputer
from sklearn.metrics import accuracy_score, classification_report
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OrdinalEncoder, StandardScaler

from app.evaluation_list_columns import parse_max_cell
from app.evaluation_sheet_parser import (
    extract_rubric_classification_dataset,
    ratio_to_performance_band,
)


def build_demo_synthetic_training_set(*, n_samples: int = 120) -> tuple[
    np.ndarray,
    np.ndarray,
    list[str],
    list[int],
    list[str],
    list[int],
    list[int],
    str,
]:
    """
    بيانات تركيبية بنفس شكل تدريب قائمة التقييم (بند + قصوى → نطاق أداء)
    حتى يعمل التدريب دون إكسل أو قاعدة بيانات (للاختبار فقط).
    """
    rng = np.random.default_rng(42)
    bands = ["راسب", "مقبول", "جيد", "جيد_جدا", "ممتاز"]
    elements = ["أ.", "ب.", "جـ.", "1", "2", "د.", "هـ."]
    xs: list[list[Any]] = []
    ys: list[str] = []
    for _ in range(int(n_samples)):
        mx = float(rng.integers(5, 21))
        ratio = float(rng.uniform(0.05, 0.99))
        aq = round(ratio * mx * 4) / 4.0
        aq = max(0.0, min(mx, aq))
        el = str(rng.choice(elements))
        xs.append([el, mx])
        ys.append(ratio_to_performance_band(aq / mx if mx > 0 else 0.0))
    X = np.asarray(xs, dtype=object)
    y = np.asarray(ys)
    fnames = ["بند", "القصوى"]
    fi = list(range(len(fnames)))
    classes = sorted(set(ys))
    training_mode = "demo_synthetic"
    numeric_idx = [1]
    categorical_idx = [0]
    return X, y, fnames, fi, classes, numeric_idx, categorical_idx, training_mode


def extract_classification_from_saved_db() -> tuple[list[list[Any]], list[str], list[str], str]:
    """يبني عينات التصنيف من ``evaluation_list_saved_results.payload_json`` (صفوف التقييم المحفوظة في الويب)."""
    feat_names = ["بند", "القصوى"]
    xs: list[list[Any]] = []
    ys: list[str] = []
    try:
        from sqlalchemy import text

        from app.database import engine
    except Exception as exc:
        return [], [], feat_names, f"تعذر الاتصال بقاعدة البيانات: {exc}"

    sql = text(
        """
        SELECT payload_json FROM evaluation_list_saved_results
        WHERE payload_json IS NOT NULL AND length(trim(payload_json)) > 4
        """
    )
    try:
        with engine.connect() as conn:
            rows_db = conn.execute(sql).fetchall()
    except Exception as exc:
        return [], [], feat_names, f"قراءة النتائج المحفوظة فشلت: {exc}"

    for (raw,) in rows_db:
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if not isinstance(data, dict):
            continue
        for row in data.get("rows") or []:
            if not isinstance(row, dict):
                continue
            aq_raw = (str(row.get("acquired") or "")).strip()
            if not aq_raw or aq_raw.lower() == "na":
                continue
            try:
                aq_f = float(aq_raw.replace(",", ".").replace("٫", "."))
            except ValueError:
                continue
            mx = parse_max_cell(row.get("max_val"))
            if mx is None or float(mx) <= 0:
                continue
            mx_f = float(mx)
            aq_f = max(0.0, min(mx_f, aq_f))
            ratio = aq_f / mx_f
            el = (str(row.get("element") or "")).strip() or "___فارغ___"
            xs.append([el, mx_f])
            ys.append(ratio_to_performance_band(ratio))

    if len(ys) < 5:
        return (
            [],
            [],
            feat_names,
            "لا توجد عينات كافية من نتائج التقييم المحفوظة في قاعدة البيانات (الحد الأدنى 5 بنود بعلامات). احفظ تقييمات من صفحة قائمة التقييم أولاً.",
        )
    return xs, ys, feat_names, ""


def _repo_root() -> Path:
    return _REPO_ROOT


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "نعم" if val else "لا"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
    return str(val).strip()


def _to_float(s: str) -> float | None:
    t = (s or "").strip().replace("\u00a0", " ").replace(",", ".").replace("٫", ".")
    if not t or t.lower() in ("na", "n/a", "-", "—"):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def load_xlsx_matrix(
    path: Path, *, sheet_index: int = 0, max_rows: int = 5000, max_cols: int = 200
) -> tuple[list[str], list[list[str]], str | None]:
    """يعيد (رؤوس الأعمدة من الصف الأول، صفوف الجسم، رسالة خطأ)."""
    path = Path(path)
    if not path.is_file():
        return [], [], "ملف غير موجود"
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return [], [], "امتداد غير مدعوم (استخدم .xlsx)"

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return [], [], "ثبّت openpyxl: pip install openpyxl"

    err: str | None = None
    headers: list[str] = []
    body: list[list[str]] = []
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        try:
            names = wb.sheetnames
            if not names:
                return [], [], "لا توجد أوراق"
            si = max(0, min(int(sheet_index), len(names) - 1))
            ws = wb[names[si]]
            mr = int(getattr(ws, "max_row", None) or 1)
            mc = int(getattr(ws, "max_column", None) or 1)
            max_r = min(mr, max_rows)
            max_c = min(mc, max_cols)
            raw: list[list[str]] = []
            for row in ws.iter_rows(
                min_row=1,
                max_row=max_r,
                min_col=1,
                max_col=max_c,
                values_only=True,
            ):
                raw.append([_cell_str(c) for c in row])
            if not raw:
                return [], [], "ورقة فارغة"
            ncol = max_c
            for r in raw:
                while len(r) < ncol:
                    r.append("")
            headers = list(raw[0])
            body = raw[1:] if len(raw) > 1 else []
        finally:
            wb.close()
    except Exception as exc:
        err = str(exc)
        return [], [], err
    return headers, body, err


def _column_kind(samples: list[str]) -> str:
    """numeric إذا غالبية القيم أرقاماً، وإلا categorical."""
    vals = [s for s in samples if (s or "").strip()]
    if not vals:
        return "categorical"
    n_ok = sum(1 for s in vals if _to_float(s) is not None)
    return "numeric" if (n_ok / len(vals)) >= 0.75 else "categorical"


def build_xy(
    paths: list[Path],
    *,
    sheet_index: int,
    label_name: str | None,
    label_col_index: int | None,
) -> tuple[np.ndarray, np.ndarray, list[str], list[int], list[str], str]:
    """
    يجمع كل الصفوف من الملفات.
    يعيد: X (object array للخلط بين رقمي ونصي لاحقاً), y, أسماء الميزات, فهارس الأعمدة المستخدمة, التصنيفات, خطأ.
    """
    all_rows: list[list[str]] = []
    headers_ref: list[str] | None = None

    for p in paths:
        h, body, err = load_xlsx_matrix(p, sheet_index=sheet_index)
        if err:
            return np.array([]), np.array([]), [], [], [], err
        if not h or not body:
            continue
        if headers_ref is None:
            headers_ref = h
        elif [x.strip() for x in h] != [x.strip() for x in headers_ref]:
            return (
                np.array([]),
                np.array([]),
                [],
                [],
                [],
                f"رؤوس الأعمدة تختلف عن الملف الأول: {p.name}",
            )
        all_rows.extend(body)

    if headers_ref is None or not all_rows:
        return np.array([]), np.array([]), [], [], [], "لا توجد صفوف بيانات في الملفات"

    ncols = len(headers_ref)
    # عمود التصنيف
    if label_name and label_name.strip():
        key = label_name.strip()
        try:
            y_idx = next(i for i, name in enumerate(headers_ref) if (name or "").strip() == key)
        except StopIteration:
            return np.array([]), np.array([]), [], [], [], f"عمود التصنيف غير موجود: {key}"
    elif label_col_index is not None:
        y_idx = max(0, min(int(label_col_index), ncols - 1))
    else:
        y_idx = ncols - 1

    feat_indices = [j for j in range(ncols) if j != y_idx]
    if not feat_indices:
        return np.array([]), np.array([]), [], [], [], "لا توجد أعمدة ميزات بعد حجب عمود التصنيف"

    # عيّن نوع كل عمود ميزة من عيّنة عشوائية
    col_samples: dict[int, list[str]] = {j: [] for j in feat_indices}
    for row in all_rows[: min(500, len(all_rows))]:
        for j in feat_indices:
            if j < len(row):
                col_samples[j].append(row[j])

    kinds = {j: _column_kind(col_samples[j]) for j in feat_indices}
    feature_names = [headers_ref[j].strip() or f"col_{j}" for j in feat_indices]

    X_list: list[list[Any]] = []
    y_list: list[str] = []
    for row in all_rows:
        while len(row) < ncols:
            row.append("")
        y_raw = (row[y_idx] or "").strip()
        if not y_raw:
            continue
        feats: list[Any] = []
        for j in feat_indices:
            cell = (row[j] if j < len(row) else "") or ""
            if kinds[j] == "numeric":
                v = _to_float(cell)
                feats.append(v if v is not None else np.nan)
            else:
                feats.append(cell.strip() if cell.strip() else "___empty___")
        X_list.append(feats)
        y_list.append(y_raw)

    if len(y_list) < 5:
        return (
            np.array([]),
            np.array([]),
            [],
            [],
            [],
            f"عدد الأمثلة بعد إزالة الصفوف بلا تصنيف قليل جداً: {len(y_list)} (الحد الأدنى المقترح 5)",
        )

    X = np.asarray(X_list, dtype=object)
    y = np.asarray(y_list)
    classes = sorted(set(y_list))
    return X, y, feature_names, feat_indices, classes, ""


def _prune_numeric_columns_without_values(X: np.ndarray, numeric_idx: list[int]) -> list[int]:
    """يمنع أعمدة رقمية فارغة كلياً من كسر SimpleImputer/StandardScaler."""
    keep: list[int] = []
    for j in numeric_idx:
        col = X[:, j]
        any_ok = False
        for v in col:
            if isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v)):
                any_ok = True
                break
            if isinstance(v, str) and _to_float(v) is not None:
                any_ok = True
                break
        if any_ok:
            keep.append(j)
    return keep


def make_pipeline(numeric_cols: list[int], categorical_cols: list[int]) -> Pipeline:
    """numeric_cols / categorical_cols: فهارس داخل مصفوفة الميزات فقط (0..n_feat-1)."""
    transformers: list[tuple[str, Pipeline, list[int]]] = []
    if numeric_cols:
        transformers.append(
            (
                "num",
                Pipeline(
                    steps=[
                        ("imputer", SimpleImputer(strategy="median")),
                        ("scaler", StandardScaler()),
                    ]
                ),
                numeric_cols,
            )
        )
    if categorical_cols:
        transformers.append(
            (
                "cat",
                Pipeline(
                    steps=[
                        (
                            "enc",
                            OrdinalEncoder(
                                handle_unknown="use_encoded_value",
                                unknown_value=-1,
                            ),
                        ),
                    ]
                ),
                categorical_cols,
            )
        )
    pre = ColumnTransformer(transformers=transformers, remainder="drop", sparse_threshold=0)
    clf = RandomForestClassifier(
        n_estimators=200,
        max_depth=None,
        min_samples_leaf=2,
        class_weight="balanced_subsample",
        random_state=42,
        n_jobs=-1,
    )
    return Pipeline(steps=[("prep", pre), ("clf", clf)])


def main(argv: list[str] | None = None) -> int:
    root = _repo_root()
    default_data = root / "ml_training" / "raw"
    default_model = root / "ml_training" / "models" / "eval_classifier.joblib"
    default_meta = root / "ml_training" / "models" / "eval_classifier.meta.json"

    p = argparse.ArgumentParser(description="تدريب تصنيف أوفلاين من ملفات Excel")
    p.add_argument(
        "--data-dir",
        type=Path,
        default=default_data,
        help="مجلد يحتوي ملفات .xlsx للتدريب",
    )
    p.add_argument(
        "--model-out",
        type=Path,
        default=default_model,
        help="مسار حفظ النموذج (.joblib)",
    )
    p.add_argument(
        "--meta-out",
        type=Path,
        default=default_meta,
        help="مسار ملف JSON لوصف النموذج والأعمدة",
    )
    p.add_argument("--sheet", type=int, default=0, help="فهرس الورقة (0 = الأولى)")
    p.add_argument(
        "--label-name",
        default="",
        help='اسم عمود التصنيف بالضبط كما في الصف الأول؛ إن تُرك فارغاً يُستخدم آخر عمود',
    )
    p.add_argument(
        "--label-col",
        type=int,
        default=-1,
        help="فهرس عمود التصنيف (0-based). القيمة -1 تعني آخر عمود إذا لم يُمرَّر --label-name",
    )
    p.add_argument("--test-size", type=float, default=0.2, help="نسبة مجموعة الاختبار")
    p.add_argument(
        "--format",
        choices=("auto", "flat", "rubric", "saved"),
        default="auto",
        help="saved=قاعدة البيانات؛ rubric=إكسل قائمة تقييم؛ flat=جدول مسطح؛ auto=ترتيب ذكي",
    )
    p.add_argument(
        "--demo",
        action="store_true",
        help="تدريب تجريبي ببيانات تركيبية (يُحفظ النموذج دون الحاجة إلى إكسل أو قاعدة بيانات)",
    )
    args = p.parse_args(argv)

    if sys.platform == "win32":
        for _stream in (sys.stdout, sys.stderr):
            try:
                _stream.reconfigure(encoding="utf-8")
            except Exception:
                pass

    data_dir = Path(args.data_dir).resolve()
    paths: list[Path] = []

    label_name = (args.label_name or "").strip() or None
    label_idx = None if label_name else (int(args.label_col) if int(args.label_col) >= 0 else None)

    fmt = str(args.format or "auto").strip().lower()
    training_mode = "flat"
    X: np.ndarray | None = None
    y: np.ndarray | None = None
    feature_names: list[str] = []
    feat_indices: list[int] = []
    classes: list[str] = []
    err_msg = ""
    numeric_idx: list[int] = []
    categorical_idx: list[int] = []

    if args.demo:
        X, y, feature_names, feat_indices, classes, numeric_idx, categorical_idx, training_mode = (
            build_demo_synthetic_training_set(n_samples=120)
        )
        paths = []
        if not data_dir.is_dir():
            data_dir.mkdir(parents=True, exist_ok=True)

    if X is None:
        paths = []
        if fmt in ("flat", "rubric"):
            if not data_dir.is_dir():
                print(f"خطأ: المجلد غير موجود: {data_dir}", file=sys.stderr)
                print("أنشئ المجلد وضع ملفات .xlsx بداخله.", file=sys.stderr)
                return 1
            paths = sorted(data_dir.glob("*.xlsx")) + sorted(data_dir.glob("*.xlsm"))
            if not paths:
                print(f"لا توجد ملفات .xlsx في: {data_dir}", file=sys.stderr)
                print('جرّب: .venv\\Scripts\\python.exe -m app.train_model --demo', file=sys.stderr)
                return 1
        elif fmt in ("auto", "saved"):
            if data_dir.is_dir():
                paths = sorted(data_dir.glob("*.xlsx")) + sorted(data_dir.glob("*.xlsm"))

    def _apply_rubric_matrix(xs: list, ys: list, fnames: list[str], mode: str) -> None:
        nonlocal X, y, feature_names, feat_indices, classes, training_mode
        nonlocal numeric_idx, categorical_idx
        X = np.asarray(xs, dtype=object)
        y = np.asarray(ys)
        feature_names = fnames
        feat_indices = list(range(len(fnames)))
        classes = sorted(set(np.asarray(ys).tolist()))
        training_mode = mode
        numeric_idx = [1]
        categorical_idx = [0]

    if not args.demo:
        if fmt == "saved":
            xs, ys, fnames, err_db = extract_classification_from_saved_db()
            if err_db or len(ys) < 5:
                print(err_db or "لا توجد عينات كافية.", file=sys.stderr)
                return 1
            _apply_rubric_matrix(xs, ys, fnames, "saved_db")

        elif fmt == "auto":
            xs, ys, fnames, err_db = extract_classification_from_saved_db()
            if not err_db and len(ys) >= 5:
                _apply_rubric_matrix(xs, ys, fnames, "saved_db")

        if X is None and fmt in ("auto", "rubric"):
            xs, ys, fnames, err_r = extract_rubric_classification_dataset(
                paths, sheet_index=int(args.sheet)
            )
            rubric_ok = not err_r and len(ys) >= 5
            if rubric_ok:
                _apply_rubric_matrix(xs, ys, fnames, "rubric")
            elif fmt == "rubric":
                print(err_r or "لا توجد عينات كافية من بنود التقييم المعبّأة (الحد الأدنى 5 بنود).", file=sys.stderr)
                return 1

        if X is None and fmt == "flat":
            X, y, feature_names, feat_indices, classes, err_msg = build_xy(
                paths,
                sheet_index=int(args.sheet),
                label_name=label_name,
                label_col_index=label_idx,
            )
            training_mode = "flat"
            if err_msg:
                print(err_msg, file=sys.stderr)
                return 1
            n_feat = X.shape[1]
            numeric_idx = []
            categorical_idx = []
            for j in range(n_feat):
                col_vals = X[:, j]
                n_float = sum(
                    1
                    for v in col_vals
                    if isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v))
                )
                n_str = sum(1 for v in col_vals if isinstance(v, str))
                if n_float >= n_str:
                    numeric_idx.append(j)
                else:
                    categorical_idx.append(j)

        if X is None and fmt == "auto":
            print(
                "لم يتوفر مصدر تدريب حقيقي (قاعدة بيانات أو إكسل معبّأ بما يكفي). "
                "سيتم استخدام بيانات تركيبية للاختبار فقط.",
                file=sys.stderr,
            )
            print(
                "[train_model] Using synthetic demo — no real samples. "
                "For production: save evaluations in the app (--format saved), "
                "fill «المكتسبة» in ml_training/raw (--format rubric), "
                "or run with --demo explicitly.",
                file=sys.stderr,
            )
            (
                X,
                y,
                feature_names,
                feat_indices,
                classes,
                numeric_idx,
                categorical_idx,
                training_mode,
            ) = build_demo_synthetic_training_set(n_samples=120)
            paths = []
            if not data_dir.is_dir():
                try:
                    data_dir.mkdir(parents=True, exist_ok=True)
                except OSError:
                    pass

    if X is None or y is None:
        print("لم يُحمّل أي جدول تدريب.", file=sys.stderr)
        return 1

    numeric_idx = _prune_numeric_columns_without_values(X, numeric_idx)

    if not numeric_idx and not categorical_idx:
        print(
            "خطأ: لا توجد أعمدة صالحة للتدريب (لا رقمية ولا نصية بعد التصفية). "
            "إما احفظ تقييمات من الويب ثم شغّل بـ --format saved، أو ضع ملفات إكسل معبأة في ml_training/raw.",
            file=sys.stderr,
        )
        return 1

    # OrdinalEncoder يتوقع مصفوفة ثنائية الأبعاد n×1 لكل عمود
    class_counts = Counter(y.tolist())
    stratify = (
        y
        if len(class_counts) > 1 and all(c >= 2 for c in class_counts.values())
        else None
    )
    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=float(args.test_size),
        random_state=42,
        stratify=stratify,
    )

    pipe = make_pipeline(numeric_idx, categorical_idx)
    try:
        pipe.fit(X_train, y_train)
    except ValueError as exc:
        print(
            f"تعذر إكمال التدريب على شكل البيانات الحالي: {exc}\n"
            "• للقوالب الفارغة في ml_training/raw: عبّئ عمود المكتسبة أو احفظ النتائج من صفحة قائمة التقييم في النظام ثم نفّذ: "
            ".venv\\Scripts\\python.exe -m app.train_model --format saved",
            file=sys.stderr,
        )
        return 1
    y_pred = pipe.predict(X_test)
    acc = accuracy_score(y_test, y_pred)

    model_out = Path(args.model_out).resolve()
    model_out.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(pipe, model_out)

    meta = {
        "training_format": training_mode,
        "data_dir": str(data_dir),
        "files": [str(x.name) for x in paths],
        "sheet_index": int(args.sheet),
        "label_column": (
            "rubric_ratio_band"
            if training_mode in ("rubric", "saved_db")
            else (label_name or ("last" if label_idx is None else label_idx))
        ),
        "feature_names": feature_names,
        "feature_indices_in_sheet": feat_indices,
        "numeric_feature_indices": numeric_idx,
        "categorical_feature_indices": categorical_idx,
        "classes": classes,
        "n_samples": int(len(y)),
        "test_accuracy": round(float(acc), 4),
        "classification_report": classification_report(y_test, y_pred, zero_division=0),
    }
    meta_out = Path(args.meta_out).resolve()
    meta_out.parent.mkdir(parents=True, exist_ok=True)
    meta_out.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    src_hint = (
        "قاعدة البيانات / النتائج المحفوظة"
        if training_mode == "saved_db"
        else f"{len(paths)} ملفاً في المجلد"
    )
    print(f"تم التدريب ({training_mode}) على {len(y)} صفاً — المصدر: {src_hint}.")
    print(f"دقة الاختبار (تقريبية): {acc:.4f}")
    print(f"النموذج: {model_out}")
    print(f"الوصف: {meta_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

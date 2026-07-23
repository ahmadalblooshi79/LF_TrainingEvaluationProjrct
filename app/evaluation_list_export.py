"""تصدير صفحة قائمة التقييم إلى ملف Excel مطابق للقالب العسكري."""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from app.evaluation_list_columns import (
    EVAL_IMPORT_COL_ACQUIRED,
    EVAL_IMPORT_COL_MAX,
    EVAL_IMPORT_COL_NOTES,
    grade_label_from_percent,
    is_evaluation_import_footer_stop_row,
    normalize_ar_header,
    parse_max_cell,
    should_skip_evaluation_import_row,
)
from app.evaluation_sheet_parser import _find_rubric_subheader_row_index, _pad_grid
from app.xlsx_grid_preview import _cell_to_str


def export_download_filename(item_title: str | None, fallback: str = "قائمة_التقييم.xlsx") -> str:
    """اسم ملف التنزيل من عنوان القائمة الظاهر في الصفحة."""
    name = (item_title or "").strip() or fallback
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = name.strip(" .")
    if not name:
        name = fallback
    low = name.lower()
    if not (low.endswith(".xlsx") or low.endswith(".xlsm") or low.endswith(".xls")):
        name = f"{name}.xlsx"
    return name


def export_doc_title_from_list_page(item_title: str | None, *, fallback: str = "") -> str:
    """
    عنوان صف B1:J1 عند التصدير — يُقتبس من عنوان قائمة التقييم في صفحة النظام
    (مثل «01 تقييم رفع الحالة.xlsx») مع إزالة امتداد الملف للعرض داخل Excel.
    """
    raw = normalize_ar_header(item_title or "")
    if not raw:
        return normalize_ar_header(fallback or "")
    title = re.sub(r"\.(xlsx|xlsm|xls)$", "", raw, flags=re.I).strip()
    return title or raw


def _sheet_grid(ws, max_row: int, max_col: int) -> list[list[str]]:
    grid: list[list[str]] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        grid.append([_cell_to_str(c) for c in row])
    return _pad_grid(grid, max_col)


def _acquired_export_value(raw: Any) -> Any:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    if s == "na":
        return "لا ينطبق"
    try:
        v = float(s.replace(",", ".").replace("٫", "."))
    except ValueError:
        return s
    if v == int(v):
        return int(v)
    return round(v, 2)


def _footer_label_key(s: str) -> str:
    return (
        normalize_ar_header(s or "")
        .replace("ـ", "")
        .replace(":", "")
        .replace("：", "")
        .replace(".", "")
        .strip()
    )


def _is_name_placeholder(value: Any) -> bool:
    s = str(value or "").strip()
    if not s:
        return True
    return "أدخل" in s


def _writable_cell(ws, row: int, col: int):
    """يعيد خلية قابلة للكتابة (أصل الدمج إن وُجد)."""
    cell = ws.cell(row, col)
    try:
        from openpyxl.cell.cell import MergedCell  # type: ignore
    except Exception:
        MergedCell = ()  # type: ignore
    if MergedCell and isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col)
    return cell


def _set_cell_value(ws, row: int, col: int, value: Any) -> None:
    _writable_cell(ws, row, col).value = value


def _fill_footer_judge_name(ws, judge_name: str, *, max_row: int, max_col: int) -> None:
    """
    يملأ صف تذييل «المحكم» باسم المحكم من صفحة قائمة التقييم.
    القالب الشائع: E=«المحكم:» و G=«أدخل اسم المحكم» (أو خلية مجاورة).
    """
    name = normalize_ar_header(judge_name or "")
    if not name or name == "—":
        return

    start = max(1, max_row - 30)
    for r in range(max_row, start - 1, -1):
        for c in range(1, min(max_col, 12) + 1):
            cell = _writable_cell(ws, r, c)
            raw = _cell_to_str(cell.value)
            if not raw:
                continue
            key = _footer_label_key(raw)

            # الخلية نفسها نص placeholder لاسم المحكم
            if "أدخل" in raw and "محكم" in raw.replace("ـ", ""):
                cell.value = name
                return

            # تسمية المحكم فقط (ليست ملاحظات طويلة ولا صف الترويسة العلوي)
            if key != "المحكم" and not (key.startswith("المحكم") and len(key) <= 12):
                continue
            if len(raw) > 40:
                continue

            # اكتب في أقرب خلية اسم (غالباً G أو F أو العمود التالي)
            candidates = [c + 2, c + 1, 7, 6, 3]
            for nc in candidates:
                if nc < 1 or nc > max_col or nc == c:
                    continue
                target = _writable_cell(ws, r, nc)
                if _is_name_placeholder(target.value) or (
                    target.value is not None and "أدخل" in str(target.value)
                ):
                    target.value = name
                    return
            # احتياطي: العمود G إن وُجد
            _set_cell_value(ws, r, 7 if max_col >= 7 else min(c + 2, max_col), name)
            return


def build_evaluation_list_xlsx_bytes(
    source_path: Path,
    *,
    doc_title: str,
    unit_label: str,
    date_str: str,
    commander_name: str,
    judge_name: str,
    eval_rows: list[dict[str, Any]] | None,
    saved_rows: list[dict[str, Any]] | None,
) -> bytes:
    """
    ينسخ ملف المصدر، يحدّث العنوان والبيانات الوصفية وعلامات المحكم،
    ويحذف أي ورقة إضافية غير ورقة التقييم.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError("مكتبة openpyxl غير مثبتة") from exc

    path = Path(source_path)
    if not path.is_file():
        raise FileNotFoundError(str(path))

    wb = load_workbook(filename=str(path))
    try:
        # ورقة واحدة فقط
        keep = None
        for name in list(wb.sheetnames):
            if name == "قائمة التقييم" or keep is None:
                keep = name
        for name in list(wb.sheetnames):
            if name != keep:
                del wb[name]
        ws = wb[keep] if keep else wb.active

        title = normalize_ar_header(doc_title or "")
        if title:
            _set_cell_value(ws, 1, 2, title)  # B1

        if unit_label and unit_label != "—":
            _set_cell_value(ws, 2, 3, unit_label)  # C2
        if date_str:
            _set_cell_value(ws, 2, 6, date_str)  # F2
        if commander_name and commander_name != "—":
            _set_cell_value(ws, 3, 3, commander_name)  # C3
        if judge_name and judge_name != "—":
            _set_cell_value(ws, 3, 6, judge_name)  # F3

        mr = int(getattr(ws, "max_row", None) or 1)
        mc = int(getattr(ws, "max_column", None) or 1)
        grid = _sheet_grid(ws, mr, mc)
        rubric_i = _find_rubric_subheader_row_index(grid)
        start_excel = (rubric_i + 2) if rubric_i is not None else 2

        template_rows = eval_rows or []
        saved = saved_rows or []
        ti = 0
        for excel_r in range(start_excel, mr + 1):
            cells = list(grid[excel_r - 1]) if excel_r - 1 < len(grid) else []
            while len(cells) < mc:
                cells.append("")
            if is_evaluation_import_footer_stop_row(cells):
                break
            if should_skip_evaluation_import_row(cells, excel_row_1based=excel_r):
                continue
            if ti >= len(template_rows):
                break
            trow = template_rows[ti]
            srow = saved[ti] if ti < len(saved) and isinstance(saved[ti], dict) else {}
            ti += 1

            if (trow.get("row_kind") or "score") == "section":
                continue

            aq = None
            if isinstance(srow, dict) and "acquired" in srow:
                aq = _acquired_export_value(srow.get("acquired"))
            elif trow.get("acquired_initial"):
                aq = _acquired_export_value(trow.get("acquired_initial"))
            if aq is not None:
                _set_cell_value(ws, excel_r, EVAL_IMPORT_COL_ACQUIRED + 1, aq)

            notes = ""
            if isinstance(srow, dict):
                notes = normalize_ar_header(str(srow.get("notes") or ""))
            if not notes:
                notes = normalize_ar_header(str(trow.get("notes_initial") or ""))
            if notes:
                _set_cell_value(ws, excel_r, EVAL_IMPORT_COL_NOTES + 1, notes)

            # قيم اختيارية للنسبة/النتيجة إن فُرغت الصيغ
            mx = parse_max_cell(
                cells[EVAL_IMPORT_COL_MAX] if len(cells) > EVAL_IMPORT_COL_MAX else ""
            )
            if aq is not None and aq != "لا ينطبق" and mx is not None and float(mx) > 0:
                try:
                    aq_f = float(aq)
                    pct = (aq_f / float(mx)) * 100.0
                    g_cell = _writable_cell(ws, excel_r, 7)
                    h_cell = _writable_cell(ws, excel_r, 8)
                    # أبقِ الصيغ إن وُجدت؛ وإلا اكتب قيماً
                    if g_cell.value is None or not str(g_cell.value).startswith("="):
                        g_cell.value = round(pct / 100.0, 4)
                    if h_cell.value is None or not str(h_cell.value).startswith("="):
                        h_cell.value = grade_label_from_percent(pct)
                except (TypeError, ValueError):
                    pass

        # صف التذييل «المحكم» — اسم المحكم من صفحة قائمة التقييم
        _fill_footer_judge_name(ws, judge_name, max_row=mr, max_col=mc)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    finally:
        try:
            wb.close()
        except Exception:
            pass

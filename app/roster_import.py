"""استخراج صفوف قائمة الوحدة (رقم عسكري، رتبة، اسم، منصب) من ملف مرفوع."""
from __future__ import annotations

import csv
import io
from typing import Any


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_roster_rows_from_upload(
    file_storage,
    *,
    max_rows: int = 500,
) -> list[tuple[str, str, str, str]]:
    """
    يعيد قائمة (رقم_عسكري، رتبة، اسم، منصب) لكل صف.
    يدعم: CSV/TXT (فاصلة أو تاب)، Excel .xlsx (أول ورقة، أربعة أعمدة).
    """
    if not file_storage or not getattr(file_storage, "filename", ""):
        return []
    fn = (file_storage.filename or "").lower()
    raw = file_storage.read() or b""
    if not raw:
        return []

    out: list[tuple[str, str, str, str]] = []

    if fn.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        bio = io.BytesIO(raw)
        wb = load_workbook(filename=bio, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1, max_row=max_rows + 5, values_only=True):
                cells = [_cell_str(c) for c in (row or ())[:4]]
                while len(cells) < 4:
                    cells.append("")
                if not any(cells):
                    continue
                if _looks_like_header_row(cells):
                    continue
                out.append(tuple(cells))
                if len(out) >= max_rows:
                    break
        finally:
            wb.close()
        return out

    # CSV / نص
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delim = "\t" if "\t" in sample else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    for cells in reader:
        cells = [(c or "").strip() for c in cells[:4]]
        while len(cells) < 4:
            cells.append("")
        if not any(cells):
            continue
        if _looks_like_header_row(cells):
            continue
        out.append((cells[0], cells[1], cells[2], cells[3]))
        if len(out) >= max_rows:
            break
    return out


def _looks_like_header_row(cells: list[str]) -> bool:
    joined = " ".join(cells).lower()
    hints = (
        "رقم",
        "عسكري",
        "military",
        "رتبة",
        "rank",
        "اسم",
        "name",
        "منصب",
        "position",
        "الترقيم",
        "تسلسل",
    )
    hits = sum(1 for h in hints if h in joined)
    return hits >= 2 and len(joined) < 120
